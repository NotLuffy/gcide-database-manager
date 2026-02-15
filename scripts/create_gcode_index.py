import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paths
folders = {
    'Kevin Revised': r'I:\My Drive\NC Master\Kevins USB Check REVISED',
    'Maurice': r'I:\My Drive\NC Master\Maurices USB Check',
    'Combined': r'I:\My Drive\NC Master\Combined Folder',
}

extensions = {'.nc', '.txt', '.ngc', '.gcode', '.tap', '.mpf', ''}

def extract_file_info(filepath):
    """Extract all relevant info from a G-code file"""
    info = {
        'filename': os.path.basename(filepath),
        'subfolder': '',
        'prog_num': '',
        'title': '',
        'od': '',           # Outside diameter (round size)
        'id_mm': '',        # Inside diameter in mm
        'id_inch': '',      # Inside diameter in inches
        'thickness': '',    # Part thickness
        'hub': '',          # Hub thickness
        'cb': '',           # Counter bore
        'step': '',         # Step dimension
        'ring': '',         # Ring dimension
        'cut': '',          # Cut to dimension
        'type': '',         # Type (2PC, STEEL, STUD, LUG, HC, etc.)
    }

    try:
        with open(filepath, 'r', errors='ignore') as f:
            content = f.read()
    except:
        return info

    # Get program number and title
    title_match = re.search(r'^[Oo](\d+)\s*\(([^)]+)\)', content, re.MULTILINE)
    if title_match:
        info['prog_num'] = 'O' + title_match.group(1)
        info['title'] = title_match.group(2).strip()
        title = info['title'].upper()

        # Extract OD (round size) - look for patterns like 7IN, 7.5IN, 13.0, etc.
        od_match = re.search(r'^([\d.]+)\s*(?:IN(?:CH)?)?(?:\s*DIA)?', title)
        if od_match:
            info['od'] = od_match.group(1)

        # Extract ID in mm - look for patterns like 74MM, 107MM, 116.7MM
        id_mm_match = re.search(r'([\d.]+)\s*MM\s*(?:ID)?', title)
        if id_mm_match:
            info['id_mm'] = id_mm_match.group(1)

        # Also check for pattern like 116.7/220MM
        id_mm_match2 = re.search(r'([\d.]+)/([\d.]+)\s*MM', title)
        if id_mm_match2:
            info['id_mm'] = id_mm_match2.group(1)

        # Extract ID in inches - look for patterns like 5.47IN ID
        id_inch_match = re.search(r'([\d.]+)\s*IN\s*ID', title)
        if id_inch_match:
            info['id_inch'] = id_inch_match.group(1)

        # Extract thickness - look for patterns like .75 THK, 1.0 THK, etc.
        thk_match = re.search(r'([\d.]+)\s*(?:THK|THICK)', title)
        if thk_match:
            info['thickness'] = thk_match.group(1)
        else:
            # Also look for standalone numbers that might be thickness
            # Pattern: after MM ID, a number before type indicators
            thk_match2 = re.search(r'MM\s*(?:ID\s*)?([\d.]+)(?:\s|$|--)', title)
            if thk_match2:
                info['thickness'] = thk_match2.group(1)

        # Extract HC (hub counterbore) - look for patterns like 1.5 HC, .75HC
        hc_match = re.search(r'([\d.]+)\s*HC', title)
        if hc_match:
            info['cb'] = hc_match.group(1)

        # Extract type indicators
        types = []
        if '2PC' in title or '2 PC' in title or '--2PC' in title:
            types.append('2PC')
        if 'STEEL' in title or 'STL' in title:
            types.append('STEEL')
        if 'STUD' in title:
            types.append('STUD')
        if 'LUG' in title:
            types.append('LUG')
        if 'STEP' in title or 'DEEP' in title or 'B/C' in title:  # DEEP and B/C indicate step program
            types.append('STEP')
        if 'RNG' in title or 'RING' in title:
            types.append('RING')
        if 'HC' in title and 'HCXX' not in title and 'DEEP' not in title:  # Exclude DEEP from HC
            if 'HC' not in types:
                types.append('HC')
        if 'HCXX' in title:
            types.append('HCXX')
        info['type'] = ', '.join(types) if types else ''

    # Extract from comments in the code
    # Hub thickness
    hub_match = re.search(r'HUB\s*(?:IS)?\s*([\d.]+)', content, re.IGNORECASE)
    if hub_match:
        info['hub'] = hub_match.group(1)

    # Cut dimension
    cut_match = re.search(r'CUT\s*(?:PART\s*TO\s*)?([\d.]+)', content, re.IGNORECASE)
    if cut_match:
        info['cut'] = cut_match.group(1)

    # Ring dimension
    ring_match = re.search(r'\(([\d.]+)\s*(?:IN)?\s*RING\)', content, re.IGNORECASE)
    if ring_match:
        info['ring'] = ring_match.group(1)

    # Step dimension
    step_match = re.search(r'STEP\s*([\d.]+)', content, re.IGNORECASE)
    if step_match:
        info['step'] = step_match.group(1)

    return info

def process_folder(base_path, exclude_folder=None):
    """Process all files in a folder and return list of file info"""
    all_files = []

    for root, dirs, files in os.walk(base_path):
        if exclude_folder and exclude_folder in dirs:
            dirs.remove(exclude_folder)

        rel_folder = os.path.relpath(root, base_path)
        if rel_folder == '.':
            rel_folder = ''

        for f in files:
            ext = os.path.splitext(f)[1].lower()
            if ext not in extensions:
                continue

            filepath = os.path.join(root, f)
            info = extract_file_info(filepath)
            info['subfolder'] = rel_folder
            all_files.append(info)

    return all_files

# Create workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Styles
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

headers = ['Filename', 'Subfolder', 'Program #', 'Title', 'OD', 'ID (mm)', 'ID (inch)',
           'Thickness', 'Hub', 'CB', 'Step', 'Ring', 'Cut', 'Type']

for sheet_name, folder_path in folders.items():
    print(f'Processing {sheet_name}...', flush=True)

    exclude = 'MASTER' if 'Maurice' in sheet_name else None
    files = process_folder(folder_path, exclude)

    print(f'  Found {len(files)} files', flush=True)

    # Create sheet
    ws = wb.create_sheet(title=sheet_name)

    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Add data
    for row, info in enumerate(files, 2):
        ws.cell(row=row, column=1, value=info['filename']).border = thin_border
        ws.cell(row=row, column=2, value=info['subfolder']).border = thin_border
        ws.cell(row=row, column=3, value=info['prog_num']).border = thin_border
        ws.cell(row=row, column=4, value=info['title']).border = thin_border
        ws.cell(row=row, column=5, value=info['od']).border = thin_border
        ws.cell(row=row, column=6, value=info['id_mm']).border = thin_border
        ws.cell(row=row, column=7, value=info['id_inch']).border = thin_border
        ws.cell(row=row, column=8, value=info['thickness']).border = thin_border
        ws.cell(row=row, column=9, value=info['hub']).border = thin_border
        ws.cell(row=row, column=10, value=info['cb']).border = thin_border
        ws.cell(row=row, column=11, value=info['step']).border = thin_border
        ws.cell(row=row, column=12, value=info['ring']).border = thin_border
        ws.cell(row=row, column=13, value=info['cut']).border = thin_border
        ws.cell(row=row, column=14, value=info['type']).border = thin_border

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = len(headers[col-1])
        for row in range(2, min(len(files) + 2, 100)):  # Sample first 100 rows
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Add autofilter
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(files) + 1}'

# Save workbook
output_path = r'I:\My Drive\NC Master\GCode_File_Index.xlsx'
wb.save(output_path)

print()
print('=' * 60)
print('EXCEL INDEX CREATED')
print('=' * 60)
print(f'Saved to: {output_path}')
print()
print('Sheets created:')
for sheet_name in folders.keys():
    print(f'  - {sheet_name}')
