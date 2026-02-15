"""
Tool Home Position Scanner
Scans G-code files for G53 Z tool home positions and flags:
- WARNING: Z values from -13.001 to -15.999 (over Z-13)
- CRITICAL: Z values at -16 or beyond (Z-16 and further negative)

Exports results to Excel with:
- Column A: File Name
- Column B: File Location (folder path)
- Column C: Line details (line number, full line, Z value found)
"""

import os
import re
from datetime import datetime
from typing import List, Dict, Tuple
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl")


class ToolHomeScanner:
    """Scans G-code files for tool home position Z values"""

    # G-code file extensions to scan
    GCODE_EXTENSIONS = {'.nc', '.txt', '.ngc', '.gcode', '.tap', '.mpf', ''}

    # Z thresholds
    WARNING_THRESHOLD = -13  # Anything more negative than Z-13 is WARNING
    CRITICAL_THRESHOLD = -16  # Z-16 or more negative is CRITICAL

    def __init__(self, scan_path: str):
        self.scan_path = scan_path
        self.results: List[Dict] = []

        # Pattern to match G53 lines with Z coordinates
        # Examples: G53 X0 Z-13, G53 X0. Z-11., G53X0Z-9, G53 Z-16
        self.g53_pattern = re.compile(
            r'G53\s*(?:X[\d.\-]*\s*)?Z([\-\d.]+)',
            re.IGNORECASE
        )

    def is_gcode_file(self, filepath: str) -> bool:
        """Check if file is likely a G-code file"""
        ext = os.path.splitext(filepath)[1].lower()

        # Check extension
        if ext in self.GCODE_EXTENSIONS:
            # For files with no extension or .txt, peek at content
            if ext in ('', '.txt'):
                try:
                    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                        first_lines = f.read(500)
                        # Look for G-code indicators
                        if re.search(r'[OoPp]\d{4,}|G[0-9]|M[0-9]|N\d+', first_lines):
                            return True
                        return False
                except:
                    return False
            return True
        return False

    def scan_file(self, filepath: str) -> List[Dict]:
        """Scan a single file for G53 Z positions"""
        file_results = []

        try:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
        except Exception as e:
            print(f"Error reading {filepath}: {e}")
            return file_results

        filename = os.path.basename(filepath)
        folder_path = os.path.dirname(filepath)

        for line_num, line in enumerate(lines, 1):
            line_stripped = line.strip()

            # Skip pure comments
            if line_stripped.startswith('(') or line_stripped.startswith(';'):
                continue

            match = self.g53_pattern.search(line_stripped)
            if match:
                try:
                    z_value = float(match.group(1))
                except ValueError:
                    continue

                # Determine severity
                severity = None
                if z_value <= self.CRITICAL_THRESHOLD:  # Z-16 or more negative
                    severity = 'CRITICAL'
                elif z_value < self.WARNING_THRESHOLD:  # More negative than Z-13 but not Z-16
                    severity = 'WARNING'

                # Only record if it's a warning or critical
                if severity:
                    file_results.append({
                        'filename': filename,
                        'folder': folder_path,
                        'line_num': line_num,
                        'line_content': line_stripped,
                        'z_value': z_value,
                        'severity': severity
                    })

        return file_results

    def scan_directory(self, progress_callback=None) -> Tuple[int, int]:
        """
        Scan entire directory tree for G-code files
        Returns (files_scanned, issues_found)
        """
        self.results = []
        files_scanned = 0

        print(f"\nScanning: {self.scan_path}")
        print("-" * 60)

        # Collect all files first for progress tracking
        all_files = []
        for root, dirs, files in os.walk(self.scan_path):
            for filename in files:
                filepath = os.path.join(root, filename)
                if self.is_gcode_file(filepath):
                    all_files.append(filepath)

        total_files = len(all_files)
        print(f"Found {total_files} G-code files to scan\n")

        for i, filepath in enumerate(all_files):
            files_scanned += 1

            # Progress update every 50 files
            if progress_callback:
                progress_callback(i + 1, total_files, filepath)
            elif (i + 1) % 50 == 0 or (i + 1) == total_files:
                print(f"  Scanned {i + 1}/{total_files} files... ({len(self.results)} issues found)", flush=True)

            file_results = self.scan_file(filepath)
            self.results.extend(file_results)

        # Sort results by folder, then filename, then line number
        self.results.sort(key=lambda x: (x['folder'].lower(), x['filename'].lower(), x['line_num']))

        return files_scanned, len(self.results)

    def export_to_excel(self, output_path: str = None) -> str:
        """Export results to Excel file"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        if not self.results:
            print("No issues found to export.")
            return None

        # Generate output filename if not provided
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(
                os.path.dirname(self.scan_path) if os.path.isdir(self.scan_path) else os.getcwd(),
                f"tool_home_scan_results_{timestamp}.xlsx"
            )

        # Create workbook
        wb = openpyxl.Workbook()

        # ========== Summary Sheet ==========
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")

        critical_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        warning_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Summary statistics
        critical_count = sum(1 for r in self.results if r['severity'] == 'CRITICAL')
        warning_count = sum(1 for r in self.results if r['severity'] == 'WARNING')
        unique_files = len(set(r['filename'] for r in self.results))
        unique_folders = len(set(r['folder'] for r in self.results))

        ws_summary['A1'] = "Tool Home Position Scan Results"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:C1')

        ws_summary['A3'] = "Scan Location:"
        ws_summary['B3'] = self.scan_path
        ws_summary['A4'] = "Scan Date:"
        ws_summary['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        ws_summary['A6'] = "Statistics"
        ws_summary['A6'].font = header_font
        ws_summary['A7'] = "Total Issues Found:"
        ws_summary['B7'] = len(self.results)
        ws_summary['A8'] = "CRITICAL (Z-16 or beyond):"
        ws_summary['B8'] = critical_count
        ws_summary['B8'].fill = critical_fill
        ws_summary['A9'] = "WARNING (over Z-13):"
        ws_summary['B9'] = warning_count
        ws_summary['B9'].fill = warning_fill
        ws_summary['A10'] = "Files Affected:"
        ws_summary['B10'] = unique_files
        ws_summary['A11'] = "Folders Affected:"
        ws_summary['B11'] = unique_folders

        # Thresholds explanation
        ws_summary['A13'] = "Threshold Definitions"
        ws_summary['A13'].font = header_font
        ws_summary['A14'] = "WARNING:"
        ws_summary['B14'] = "Z values more negative than Z-13 (e.g., Z-13.5, Z-14, Z-15)"
        ws_summary['A15'] = "CRITICAL:"
        ws_summary['B15'] = "Z values at Z-16 or more negative (e.g., Z-16, Z-17, Z-20)"

        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 60

        # ========== Detailed Results Sheet ==========
        ws_details = wb.create_sheet("Detailed Results")

        # Headers
        headers = ["Severity", "File Name", "File Location", "Line #", "Line Content", "Z Value"]
        for col, header in enumerate(headers, 1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_num, result in enumerate(self.results, 2):
            ws_details.cell(row=row_num, column=1, value=result['severity'])
            ws_details.cell(row=row_num, column=2, value=result['filename'])
            ws_details.cell(row=row_num, column=3, value=result['folder'])
            ws_details.cell(row=row_num, column=4, value=result['line_num'])
            ws_details.cell(row=row_num, column=5, value=result['line_content'])
            ws_details.cell(row=row_num, column=6, value=result['z_value'])

            # Apply severity coloring to entire row
            fill = critical_fill if result['severity'] == 'CRITICAL' else warning_fill
            for col in range(1, 7):
                cell = ws_details.cell(row=row_num, column=col)
                cell.fill = fill
                cell.border = thin_border

        # Adjust column widths
        ws_details.column_dimensions['A'].width = 12  # Severity
        ws_details.column_dimensions['B'].width = 25  # File Name
        ws_details.column_dimensions['C'].width = 60  # File Location
        ws_details.column_dimensions['D'].width = 8   # Line #
        ws_details.column_dimensions['E'].width = 50  # Line Content
        ws_details.column_dimensions['F'].width = 10  # Z Value

        # Freeze header row
        ws_details.freeze_panes = 'A2'

        # ========== By Folder Sheet (grouped) ==========
        ws_folders = wb.create_sheet("By Folder")

        # Group by folder
        folders_data = {}
        for result in self.results:
            folder = result['folder']
            if folder not in folders_data:
                folders_data[folder] = []
            folders_data[folder].append(result)

        # Sort folders
        sorted_folders = sorted(folders_data.keys(), key=str.lower)

        row = 1
        for folder in sorted_folders:
            folder_results = folders_data[folder]

            # Folder header
            cell = ws_folders.cell(row=row, column=1, value=f"Folder: {folder}")
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            ws_folders.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            row += 1

            # Column headers for this section
            sub_headers = ["Severity", "File Name", "Line #", "Line Content", "Z Value"]
            for col, header in enumerate(sub_headers, 1):
                cell = ws_folders.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.border = thin_border
            row += 1

            # Data for this folder
            for result in sorted(folder_results, key=lambda x: (x['filename'].lower(), x['line_num'])):
                ws_folders.cell(row=row, column=1, value=result['severity'])
                ws_folders.cell(row=row, column=2, value=result['filename'])
                ws_folders.cell(row=row, column=3, value=result['line_num'])
                ws_folders.cell(row=row, column=4, value=result['line_content'])
                ws_folders.cell(row=row, column=5, value=result['z_value'])

                fill = critical_fill if result['severity'] == 'CRITICAL' else warning_fill
                for col in range(1, 6):
                    cell = ws_folders.cell(row=row, column=col)
                    cell.fill = fill
                    cell.border = thin_border
                row += 1

            row += 1  # Blank row between folders

        # Adjust column widths
        ws_folders.column_dimensions['A'].width = 12
        ws_folders.column_dimensions['B'].width = 25
        ws_folders.column_dimensions['C'].width = 8
        ws_folders.column_dimensions['D'].width = 50
        ws_folders.column_dimensions['E'].width = 10

        # Save workbook
        wb.save(output_path)
        print(f"\nResults exported to: {output_path}")

        return output_path


def main():
    """Main entry point for the scanner"""
    import argparse

    parser = argparse.ArgumentParser(description='Scan G-code files for tool home position Z values')
    parser.add_argument('path', nargs='?', default=r"I:\My Drive\NC Master\Kevins USB Check",
                        help='Path to scan (default: I:\\My Drive\\NC Master\\Kevins USB Check)')
    parser.add_argument('-o', '--output', help='Output Excel file path (optional)')

    args = parser.parse_args()

    scan_path = args.path

    if not os.path.exists(scan_path):
        print(f"Error: Path does not exist: {scan_path}")
        return 1

    # Create scanner and run
    scanner = ToolHomeScanner(scan_path)
    files_scanned, issues_found = scanner.scan_directory()

    print(f"\n{'='*60}")
    print(f"Scan Complete!")
    print(f"  Files Scanned: {files_scanned}")
    print(f"  Issues Found: {issues_found}")

    if issues_found > 0:
        critical_count = sum(1 for r in scanner.results if r['severity'] == 'CRITICAL')
        warning_count = sum(1 for r in scanner.results if r['severity'] == 'WARNING')
        print(f"    - CRITICAL (Z-16+): {critical_count}")
        print(f"    - WARNING (over Z-13): {warning_count}")

        # Export to Excel
        output_path = scanner.export_to_excel(args.output)

        # Print first few results as preview
        print(f"\nFirst 10 issues found:")
        print("-" * 60)
        for i, result in enumerate(scanner.results[:10]):
            print(f"  [{result['severity']}] {result['filename']} Line {result['line_num']}: Z{result['z_value']}")
    else:
        print("\nNo tool home position issues found!")

    return 0


if __name__ == "__main__":
    exit(main())
