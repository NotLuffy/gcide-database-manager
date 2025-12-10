import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import sqlite3
import os
import re
from datetime import datetime
from dataclasses import dataclass
from typing import List, Optional, Tuple, Set
import json
import sys
import shutil

# Import the improved parser
from improved_gcode_parser import ImprovedGCodeParser, GCodeParseResult


class MultiSelectCombobox(ttk.Frame):
    """Custom multi-select combobox widget"""
    def __init__(self, parent, values, bg_color, fg_color, input_bg, button_bg, width=15):
        super().__init__(parent, style='Dark.TFrame')
        self.values = values
        self.selected = set()
        self.bg_color = bg_color
        self.fg_color = fg_color
        self.input_bg = input_bg
        self.button_bg = button_bg

        # Display label showing selected count
        self.display_label = tk.Label(self, text="All", bg=input_bg, fg=fg_color,
                                     width=width, anchor='w', relief='sunken', cursor='hand2')
        self.display_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.display_label.bind('<Button-1>', self.show_popup)

        self.popup = None

    def show_popup(self, event=None):
        """Show multi-select popup"""
        if self.popup:
            self.popup.destroy()
            self.popup = None
            return

        # Create popup window
        self.popup = tk.Toplevel(self.master)
        self.popup.title("Select Items")
        self.popup.geometry("320x400")
        self.popup.configure(bg=self.bg_color)

        # Position below the button
        x = self.display_label.winfo_rootx()
        y = self.display_label.winfo_rooty() + self.display_label.winfo_height()
        self.popup.geometry(f"+{x}+{y}")

        # Make it a transient window
        self.popup.transient(self.master)
        self.popup.grab_set()

        # Frame for checkboxes
        frame = tk.Frame(self.popup, bg=self.bg_color)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Add scrollbar
        canvas = tk.Canvas(frame, bg=self.bg_color, highlightthickness=0)
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=self.bg_color)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Checkboxes for each value
        self.check_vars = {}
        for value in self.values:
            var = tk.BooleanVar(value=value in self.selected)
            self.check_vars[value] = var

            cb = tk.Checkbutton(scrollable_frame, text=value, variable=var,
                              bg=self.bg_color, fg=self.fg_color,
                              selectcolor=self.input_bg, activebackground=self.bg_color,
                              activeforeground=self.fg_color,
                              font=('Arial', 10), pady=5, padx=5,
                              cursor='hand2')
            cb.pack(anchor='w', pady=3, padx=5, fill=tk.X)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Buttons
        button_frame = tk.Frame(self.popup, bg=self.bg_color)
        button_frame.pack(fill=tk.X, padx=10, pady=10)

        btn_all = tk.Button(button_frame, text="Select All", command=self.select_all,
                          bg=self.button_bg, fg=self.fg_color,
                          font=('Arial', 10, 'bold'), width=12, height=2)
        btn_all.pack(side=tk.LEFT, padx=3, fill=tk.X, expand=True)

        btn_none = tk.Button(button_frame, text="Clear All", command=self.select_none,
                           bg=self.button_bg, fg=self.fg_color,
                           font=('Arial', 10, 'bold'), width=12, height=2)
        btn_none.pack(side=tk.LEFT, padx=3, fill=tk.X, expand=True)

        btn_apply = tk.Button(button_frame, text="✓ Apply", command=self.apply_selection,
                            bg='#4a90e2', fg=self.fg_color,
                            font=('Arial', 11, 'bold'), width=12, height=2)
        btn_apply.pack(side=tk.LEFT, padx=3, fill=tk.X, expand=True)

        # Close on click outside (bind to popup destroy)
        self.popup.bind('<FocusOut>', lambda e: self.close_popup())

    def select_all(self):
        """Select all items"""
        for var in self.check_vars.values():
            var.set(True)

    def select_none(self):
        """Clear all selections"""
        for var in self.check_vars.values():
            var.set(False)

    def apply_selection(self):
        """Apply the selection and close popup"""
        self.selected = {value for value, var in self.check_vars.items() if var.get()}
        self.update_display()
        self.close_popup()

    def close_popup(self):
        """Close the popup"""
        if self.popup:
            self.popup.destroy()
            self.popup = None

    def update_display(self):
        """Update the display label"""
        if not self.selected or len(self.selected) == len(self.values):
            self.display_label.config(text="All")
        elif len(self.selected) == 1:
            self.display_label.config(text=list(self.selected)[0])
        else:
            self.display_label.config(text=f"{len(self.selected)} selected")

    def get_selected(self) -> Set[str]:
        """Get selected values"""
        return self.selected if self.selected else set(self.values)

    def clear(self):
        """Clear all selections"""
        self.selected = set()
        self.update_display()


@dataclass
class ProgramRecord:
    """Represents a gcode program record"""
    program_number: str
    title: Optional[str]  # Raw title from G-code (content in parentheses)
    spacer_type: str
    outer_diameter: Optional[float]
    thickness: Optional[float]
    thickness_display: Optional[str]  # Display format: "10MM" or "0.75"
    center_bore: Optional[float]
    hub_height: Optional[float]
    hub_diameter: Optional[float]
    counter_bore_diameter: Optional[float]
    counter_bore_depth: Optional[float]
    paired_program: Optional[str]
    material: Optional[str]
    notes: Optional[str]
    date_created: Optional[str]
    last_modified: Optional[str]
    file_path: str
    # Validation fields
    detection_confidence: Optional[str] = None
    detection_method: Optional[str] = None
    validation_status: Optional[str] = None  # 'CRITICAL', 'DIMENSIONAL', 'BORE_WARNING', 'WARNING', 'PASS'
    validation_issues: Optional[str] = None  # JSON list - CRITICAL errors (RED)
    validation_warnings: Optional[str] = None  # JSON list - General warnings (YELLOW)
    bore_warnings: Optional[str] = None  # JSON list - Bore warnings (ORANGE)
    dimensional_issues: Optional[str] = None  # JSON list - P-code/thickness issues (PURPLE)
    cb_from_gcode: Optional[float] = None
    ob_from_gcode: Optional[float] = None
    lathe: Optional[str] = None  # 'L1', 'L2', 'L3', 'L2/L3'

class GCodeDatabaseGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("G-Code Database Manager - Wheel Spacer Programs")
        self.root.geometry("1600x900")
        
        # Dark mode colors
        self.bg_color = "#2b2b2b"
        self.fg_color = "#ffffff"
        self.input_bg = "#3c3c3c"
        self.button_bg = "#4a4a4a"
        self.accent_color = "#4a90e2"
        
        self.root.configure(bg=self.bg_color)
        
        # Database setup
        self.db_path = "gcode_database.db"
        self.init_database()

        # User session (will be set after login, for now use default admin)
        self.current_user = None
        self.current_user_id = None
        self.current_username = "admin"  # Default to admin for now
        self.current_user_role = "admin"

        # Configuration
        self.config_file = "gcode_manager_config.json"
        self.load_config()

        # Initialize repository system
        self.init_repository()

        # Initialize improved parser
        self.parser = ImprovedGCodeParser()

        # Configure ttk dark theme
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Dark.TFrame', background=self.bg_color)

        # Get available values from database for filters
        self.available_types = self.get_available_values("spacer_type")
        self.available_materials = self.get_available_values("material")
        self.available_statuses = self.get_available_values("validation_status")

        # Build GUI
        self.setup_gui()

        # Enable drag and drop
        self.setup_drag_drop()

        self.refresh_results()
        
    def init_database(self):
        """Initialize SQLite database with schema"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS programs (
                program_number TEXT PRIMARY KEY,
                title TEXT,
                spacer_type TEXT NOT NULL,
                outer_diameter REAL,
                thickness REAL,
                thickness_display TEXT,
                center_bore REAL,
                hub_height REAL,
                hub_diameter REAL,
                counter_bore_diameter REAL,
                counter_bore_depth REAL,
                paired_program TEXT,
                material TEXT,
                notes TEXT,
                date_created TEXT,
                last_modified TEXT,
                file_path TEXT,
                detection_confidence TEXT,
                detection_method TEXT,
                validation_status TEXT,
                validation_issues TEXT,
                validation_warnings TEXT,
                bore_warnings TEXT,
                dimensional_issues TEXT,
                cb_from_gcode REAL,
                ob_from_gcode REAL,
                lathe TEXT
            )
        ''')

        # Upgrade existing database if needed
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN title TEXT")
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN detection_confidence TEXT")
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN detection_method TEXT")
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN validation_status TEXT")
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN validation_issues TEXT")
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN validation_warnings TEXT")
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN cb_from_gcode REAL")
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN ob_from_gcode REAL")
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN bore_warnings TEXT")
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN dimensional_issues TEXT")
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN thickness_display TEXT")
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN lathe TEXT")
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN duplicate_type TEXT")  # 'SOLID', 'NAME_COLLISION', 'CONTENT_DUP', NULL
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN parent_file TEXT")  # Reference to parent program_number
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN duplicate_group TEXT")  # Group ID for related duplicates
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN current_version INTEGER DEFAULT 1")  # Current version number
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN modified_by TEXT")  # Last user who modified
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN is_managed INTEGER DEFAULT 0")  # 1 if in repository, 0 if external
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN round_size REAL")  # Detected round size (e.g., 6.25, 10.5)
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN round_size_confidence TEXT")  # 'HIGH', 'MEDIUM', 'LOW', 'NONE'
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN round_size_source TEXT")  # 'TITLE', 'GCODE', 'DIMENSION', 'MANUAL'
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN in_correct_range INTEGER DEFAULT 1")  # 1 if program number matches round size range
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN legacy_names TEXT")  # JSON array of previous program numbers
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN last_renamed_date TEXT")  # ISO timestamp of last rename
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN rename_reason TEXT")  # Why it was renamed
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN tools_used TEXT")  # JSON list of tool numbers (e.g., ["T101", "T121", "T202"])
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN tool_sequence TEXT")  # JSON ordered list of tools in sequence
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN tool_validation_status TEXT")  # 'PASS', 'WARNING', 'ERROR'
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN tool_validation_issues TEXT")  # JSON list of tool issues
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN safety_blocks_status TEXT")  # 'PASS', 'WARNING', 'MISSING'
        except:
            pass
        try:
            cursor.execute("ALTER TABLE programs ADD COLUMN safety_blocks_issues TEXT")  # JSON list of missing safety blocks
        except:
            pass

        # Create users table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                user_id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT,
                full_name TEXT,
                role TEXT DEFAULT 'viewer',
                email TEXT,
                date_created TEXT,
                last_login TEXT,
                is_active INTEGER DEFAULT 1
            )
        ''')

        # Create default admin user if no users exist
        cursor.execute("SELECT COUNT(*) FROM users")
        if cursor.fetchone()[0] == 0:
            from datetime import datetime
            cursor.execute("""
                INSERT INTO users (username, password_hash, full_name, role, date_created, is_active)
                VALUES ('admin', NULL, 'Administrator', 'admin', ?, 1)
            """, (datetime.now().isoformat(),))

        # Create program_versions table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS program_versions (
                version_id INTEGER PRIMARY KEY AUTOINCREMENT,
                program_number TEXT NOT NULL,
                version_number TEXT NOT NULL,
                version_tag TEXT,
                file_content TEXT,
                file_hash TEXT,
                file_path TEXT,
                date_created TEXT,
                created_by TEXT,
                change_summary TEXT,
                dimensions_snapshot TEXT,
                FOREIGN KEY (program_number) REFERENCES programs(program_number)
            )
        ''')

        # Add file_path column to existing program_versions table
        try:
            cursor.execute("ALTER TABLE program_versions ADD COLUMN file_path TEXT")
        except:
            pass

        # Create activity_log table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS activity_log (
                log_id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                username TEXT,
                action_type TEXT,
                program_number TEXT,
                details TEXT,
                timestamp TEXT,
                FOREIGN KEY (user_id) REFERENCES users(user_id)
            )
        ''')

        # Create edit_locks table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS edit_locks (
                lock_id INTEGER PRIMARY KEY AUTOINCREMENT,
                program_number TEXT UNIQUE,
                locked_by INTEGER,
                locked_by_username TEXT,
                locked_at TEXT,
                FOREIGN KEY (locked_by) REFERENCES users(user_id)
            )
        ''')

        # Create comments table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS comments (
                comment_id INTEGER PRIMARY KEY AUTOINCREMENT,
                program_number TEXT,
                version_id INTEGER,
                user_id INTEGER,
                username TEXT,
                comment_text TEXT,
                parent_comment_id INTEGER,
                timestamp TEXT,
                FOREIGN KEY (user_id) REFERENCES users(user_id),
                FOREIGN KEY (version_id) REFERENCES program_versions(version_id)
            )
        ''')

        # Create production_tracking table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS production_tracking (
                track_id INTEGER PRIMARY KEY AUTOINCREMENT,
                program_number TEXT,
                version_id INTEGER,
                production_status TEXT,
                run_count INTEGER DEFAULT 0,
                last_used TEXT,
                quality_notes TEXT,
                operator TEXT,
                FOREIGN KEY (version_id) REFERENCES program_versions(version_id)
            )
        ''')

        # Create program_number_registry table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS program_number_registry (
                program_number TEXT PRIMARY KEY,
                round_size REAL,
                range_start INTEGER,
                range_end INTEGER,
                status TEXT DEFAULT 'AVAILABLE',
                file_path TEXT,
                duplicate_count INTEGER DEFAULT 0,
                last_checked TEXT,
                notes TEXT
            )
        ''')

        # Create duplicate_resolutions table for tracking resolution history
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS duplicate_resolutions (
                resolution_id INTEGER PRIMARY KEY AUTOINCREMENT,
                resolution_date TEXT,
                duplicate_type TEXT,
                program_numbers TEXT,
                action_taken TEXT,
                files_affected TEXT,
                old_values TEXT,
                new_values TEXT,
                user_override INTEGER DEFAULT 0,
                notes TEXT
            )
        ''')

        conn.commit()
        conn.close()

    @staticmethod
    def format_program_number(number):
        """
        Format a program number with proper leading zeros.

        Args:
            number: Integer or string program number (with or without 'o' prefix)

        Returns:
            str: Formatted program number (e.g., 'o00001', 'o01000', 'o12345')

        Examples:
            format_program_number(1) -> 'o00001'
            format_program_number(100) -> 'o00100'
            format_program_number('o1000') -> 'o01000'
            format_program_number('1000') -> 'o01000'
        """
        # Convert to string and remove any 'o' or 'O' prefix
        num_str = str(number).replace('o', '').replace('O', '')

        # Convert to integer and back to string (removes leading zeros if any)
        try:
            num_int = int(num_str)
            # Format with leading zeros (5 digits total)
            return f"o{num_int:05d}"
        except ValueError:
            # If conversion fails, return as-is with 'o' prefix
            return f"o{num_str}"

    def init_repository(self):
        """Initialize managed file repository structure"""
        # Get base path (where this script is located)
        base_path = os.path.dirname(os.path.abspath(__file__))

        # Define repository paths
        self.repository_path = os.path.join(base_path, 'repository')
        self.versions_path = os.path.join(base_path, 'versions')
        self.backups_path = os.path.join(base_path, 'backups')
        self.deleted_path = os.path.join(base_path, 'deleted')

        # Create directories if they don't exist
        os.makedirs(self.repository_path, exist_ok=True)
        os.makedirs(self.versions_path, exist_ok=True)
        os.makedirs(self.backups_path, exist_ok=True)
        os.makedirs(self.deleted_path, exist_ok=True)

        print(f"[Repository] Initialized at: {self.repository_path}")
        print(f"[Versions] Initialized at: {self.versions_path}")
        print(f"[Backups] Initialized at: {self.backups_path}")
        print(f"[Deleted] Initialized at: {self.deleted_path}")

    def is_managed_file(self, file_path):
        """Check if a file is in the managed repository"""
        if not file_path:
            return False
        abs_path = os.path.abspath(file_path)
        repo_path = os.path.abspath(self.repository_path)
        return abs_path.startswith(repo_path)

    def import_to_repository(self, source_file, program_number=None):
        """
        Import a file into the managed repository with automatic archiving.
        New file keeps standard name, old file gets version suffix and moves to archive.

        Returns the new path in the repository.
        """
        try:
            # Initialize repository manager if not already done
            if not hasattr(self, 'repo_manager'):
                from repository_manager import RepositoryManager
                self.repo_manager = RepositoryManager(self.db_path, self.repository_path)

            # If already in repository, return as-is
            if self.is_managed_file(source_file):
                return source_file

            # Extract program number if not provided
            if not program_number:
                filename = os.path.basename(source_file)
                program_number = os.path.splitext(filename)[0].lower()

            # Use repository manager's import with archive
            # This handles: archiving old version, importing new file with standard name
            dest_path = self.repo_manager.import_with_archive(source_file, program_number)

            if dest_path:
                return dest_path

            # Fallback to old behavior if repo manager fails
            print(f"[Repository] Warning: Archive system failed, using legacy import")

            filename = os.path.basename(source_file)
            dest_path = os.path.join(self.repository_path, filename)

            # If file already exists, handle collision
            if os.path.exists(dest_path):
                # Check if it's the same file (same content)
                with open(source_file, 'rb') as f1:
                    source_content = f1.read()
                with open(dest_path, 'rb') as f2:
                    dest_content = f2.read()

                if source_content == dest_content:
                    # Same file, just return existing path
                    print(f"[Repository] File already exists (identical): {filename}")
                    return dest_path
                else:
                    # Different file with same name - create unique name
                    base, ext = os.path.splitext(filename)
                    counter = 1
                    while os.path.exists(dest_path):
                        dest_path = os.path.join(self.repository_path, f"{base}_{counter}{ext}")
                        counter += 1
                    print(f"[Repository] Collision detected, using: {os.path.basename(dest_path)}")

            # Copy file to repository
            shutil.copy2(source_file, dest_path)
            print(f"[Repository] Imported: {filename}")

            return dest_path

        except Exception as e:
            print(f"[Repository] Error importing file: {e}")
            import traceback
            traceback.print_exc()
            return None

    def create_version_file(self, program_number, version_number, source_file_path):
        """
        Save a version as a physical file in the versions folder.
        Returns the path to the version file.
        """
        try:
            # Create version folder for this program
            version_folder = os.path.join(self.versions_path, program_number)
            os.makedirs(version_folder, exist_ok=True)

            # Get file extension
            ext = os.path.splitext(source_file_path)[1] or '.nc'

            # Version filename
            version_filename = f"{version_number}{ext}"
            version_path = os.path.join(version_folder, version_filename)

            # Copy file to version folder
            if os.path.exists(source_file_path):
                shutil.copy2(source_file_path, version_path)
                print(f"[Version] Saved version file: {version_filename}")
                return version_path
            else:
                print(f"[Version] Source file not found: {source_file_path}")
                return None

        except Exception as e:
            print(f"[Version] Error creating version file: {e}")
            return None

    def get_version_file_path(self, program_number, version_number):
        """Get the file path for a specific version"""
        version_folder = os.path.join(self.versions_path, program_number)

        if not os.path.exists(version_folder):
            return None

        # Try common extensions
        for ext in ['.nc', '.gcode', '.txt', '']:
            version_path = os.path.join(version_folder, f"{version_number}{ext}")
            if os.path.exists(version_path):
                return version_path

        return None

    def migrate_file_to_repository(self, program_number):
        """
        Migrate a single file from external location to managed repository.
        Returns True if successful, False otherwise.
        """
        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            # Get current file path and managed status
            cursor.execute("SELECT file_path, is_managed FROM programs WHERE program_number = ?", (program_number,))
            result = cursor.fetchone()

            if not result:
                conn.close()
                return False

            current_path, is_managed = result

            # Already managed
            if is_managed:
                print(f"[Migration] {program_number} already in repository")
                conn.close()
                return True

            # Check if file exists
            if not current_path or not os.path.exists(current_path):
                print(f"[Migration] File not found: {current_path}")
                conn.close()
                return False

            # Import to repository
            new_path = self.import_to_repository(current_path, program_number)

            if new_path:
                # Update database
                cursor.execute("""
                    UPDATE programs SET file_path = ?, is_managed = 1 WHERE program_number = ?
                """, (new_path, program_number))

                conn.commit()
                print(f"[Migration] Successfully migrated {program_number}")

                # Log activity
                self.log_activity('migrate_to_repository', program_number, {
                    'old_path': current_path,
                    'new_path': new_path
                })

                conn.close()
                return True
            else:
                conn.close()
                return False

        except Exception as e:
            print(f"[Migration] Error migrating {program_number}: {e}")
            return False

    def migrate_all_to_repository(self):
        """
        Migrate all external files to the managed repository.
        Returns (success_count, error_count)
        """
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Get all programs with external files
        cursor.execute("""
            SELECT program_number FROM programs
            WHERE file_path IS NOT NULL AND (is_managed IS NULL OR is_managed = 0)
        """)

        programs = [row[0] for row in cursor.fetchall()]
        conn.close()

        print(f"[Migration] Found {len(programs)} external files to migrate")

        success = 0
        errors = 0

        for prog_num in programs:
            if self.migrate_file_to_repository(prog_num):
                success += 1
            else:
                errors += 1

        print(f"[Migration] Complete: {success} successful, {errors} errors")
        return success, errors

    def get_repository_stats(self):
        """Get statistics about repository usage"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        stats = {}

        # Total programs
        cursor.execute("SELECT COUNT(*) FROM programs WHERE file_path IS NOT NULL")
        stats['total_programs'] = cursor.fetchone()[0]

        # Managed files
        cursor.execute("SELECT COUNT(*) FROM programs WHERE is_managed = 1")
        stats['managed_files'] = cursor.fetchone()[0]

        # External files
        stats['external_files'] = stats['total_programs'] - stats['managed_files']

        # Total versions
        cursor.execute("SELECT COUNT(*) FROM program_versions")
        stats['total_versions'] = cursor.fetchone()[0]

        # Repository size
        repo_size = 0
        if os.path.exists(self.repository_path):
            for file in os.listdir(self.repository_path):
                file_path = os.path.join(self.repository_path, file)
                if os.path.isfile(file_path):
                    repo_size += os.path.getsize(file_path)

        stats['repository_size_mb'] = repo_size / (1024 * 1024)

        # Versions size
        versions_size = 0
        if os.path.exists(self.versions_path):
            for root, _, files in os.walk(self.versions_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    versions_size += os.path.getsize(file_path)

        stats['versions_size_mb'] = versions_size / (1024 * 1024)

        conn.close()
        return stats

    def log_activity(self, action_type, program_number=None, details=None):
        """Log user activity to the activity_log table"""
        try:
            # Use timeout to prevent database locked errors
            conn = sqlite3.connect(self.db_path, timeout=10.0)
            cursor = conn.cursor()

            from datetime import datetime
            timestamp = datetime.now().isoformat()

            # Convert details to JSON if it's a dict
            if isinstance(details, dict):
                import json
                details = json.dumps(details)

            cursor.execute("""
                INSERT INTO activity_log (user_id, username, action_type, program_number, details, timestamp)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (self.current_user_id, self.current_username, action_type, program_number, details, timestamp))

            conn.commit()
            conn.close()
        except sqlite3.OperationalError as e:
            if "locked" in str(e).lower():
                print(f"[Activity Log] Database temporarily locked, activity not logged: {action_type}")
            else:
                print(f"[Activity Log] Error logging activity: {e}")
        except Exception as e:
            print(f"[Activity Log] Error logging activity: {e}")

    def create_version(self, program_number, change_summary=None):
        """Create a new version of a program"""
        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            # Get current program data
            cursor.execute("SELECT file_path, current_version FROM programs WHERE program_number = ?", (program_number,))
            result = cursor.fetchone()

            if not result:
                conn.close()
                return None

            file_path, current_version = result

            if not file_path or not os.path.exists(file_path):
                conn.close()
                return None

            # Read file content
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()

            # Calculate file hash
            import hashlib
            file_hash = hashlib.sha256(content.encode()).hexdigest()

            # Get dimensions snapshot
            cursor.execute("""
                SELECT outer_diameter, thickness, center_bore, hub_height, hub_diameter,
                       counter_bore_diameter, material, spacer_type
                FROM programs WHERE program_number = ?
            """, (program_number,))
            dims = cursor.fetchone()

            import json
            dimensions_snapshot = json.dumps({
                'outer_diameter': dims[0],
                'thickness': dims[1],
                'center_bore': dims[2],
                'hub_height': dims[3],
                'hub_diameter': dims[4],
                'counter_bore_diameter': dims[5],
                'material': dims[6],
                'spacer_type': dims[7]
            }) if dims else None

            # Increment version
            new_version = (current_version or 0) + 1
            version_number = f"v{new_version}.0"

            # Create version file (save physical copy)
            version_file_path = self.create_version_file(program_number, version_number, file_path)

            # Create version record
            from datetime import datetime
            cursor.execute("""
                INSERT INTO program_versions (
                    program_number, version_number, file_content, file_hash, file_path,
                    date_created, created_by, change_summary, dimensions_snapshot
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (program_number, version_number, content, file_hash, version_file_path,
                  datetime.now().isoformat(), self.current_username, change_summary, dimensions_snapshot))

            version_id = cursor.lastrowid

            # Update program's current version
            cursor.execute("""
                UPDATE programs SET current_version = ?, modified_by = ? WHERE program_number = ?
            """, (new_version, self.current_username, program_number))

            conn.commit()
            conn.close()

            # Log activity
            self.log_activity('create_version', program_number, {
                'version_number': version_number,
                'change_summary': change_summary,
                'version_file': version_file_path
            })

            return version_id

        except Exception as e:
            print(f"[Version] Error creating version: {e}")
            return None

    def get_version_history(self, program_number):
        """Get all versions of a program"""
        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            cursor.execute("""
                SELECT version_id, version_number, version_tag, date_created,
                       created_by, change_summary
                FROM program_versions
                WHERE program_number = ?
                ORDER BY date_created DESC
            """, (program_number,))

            versions = cursor.fetchall()
            conn.close()

            return versions

        except Exception as e:
            print(f"[Version] Error getting version history: {e}")
            return []

    def compare_versions(self, version_id1, version_id2):
        """Compare two versions of a program"""
        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            cursor.execute("SELECT file_content, version_number FROM program_versions WHERE version_id = ?", (version_id1,))
            v1 = cursor.fetchone()

            cursor.execute("SELECT file_content, version_number FROM program_versions WHERE version_id = ?", (version_id2,))
            v2 = cursor.fetchone()

            conn.close()

            if not v1 or not v2:
                return None

            return {
                'version1': {'content': v1[0], 'version_number': v1[1]},
                'version2': {'content': v2[0], 'version_number': v2[1]}
            }

        except Exception as e:
            print(f"[Version] Error comparing versions: {e}")
            return None

    # ==================== ROUND SIZE DETECTION & RANGE MANAGEMENT ====================

    def get_round_size_ranges(self):
        """Return dictionary of round size to program number ranges"""
        return {
            10.25: (10000, 12999, "10.25 & 10.50"),
            10.50: (10000, 12999, "10.25 & 10.50"),
            13.0:  (13000, 13999, "13.0"),
            5.75:  (50000, 59999, "5.75"),
            6.0:   (60000, 62499, "6.0"),
            6.25:  (62500, 64999, "6.25"),
            6.5:   (65000, 69999, "6.5"),
            7.0:   (70000, 74999, "7.0"),
            7.5:   (75000, 79000, "7.5"),
            8.0:   (80000, 84999, "8.0"),
            8.5:   (85000, 89999, "8.5"),
            9.5:   (90000, 99999, "9.5"),
            # Free ranges (use when specific range is full)
            0.0:   (1000, 9999, "Free Range 1"),
            -1.0:  (14000, 49999, "Free Range 2")
        }

    def get_range_for_round_size(self, round_size):
        """Get program number range for a round size"""
        ranges = self.get_round_size_ranges()

        # Exact match
        if round_size in ranges:
            return ranges[round_size][:2]  # Return (start, end)

        # Find closest match (for slight variations like 6.24 → 6.25)
        # Only consider positive round sizes (exclude free ranges)
        positive_ranges = {k: v for k, v in ranges.items() if k > 0}

        if not positive_ranges:
            return None

        closest_size = min(positive_ranges.keys(), key=lambda x: abs(x - round_size))

        # Tight tolerance for very close matches (6.24 → 6.25)
        tight_tolerance = 0.1
        if abs(closest_size - round_size) <= tight_tolerance:
            return ranges[closest_size][:2]

        # Smart fallback for orphaned round sizes (like 5.0, 5.5, etc.)
        # Use a more generous tolerance to find the nearest logical range
        smart_fallback_tolerance = 1.0
        if abs(closest_size - round_size) <= smart_fallback_tolerance:
            return ranges[closest_size][:2]

        # If still no match, try to find the nearest range boundary
        # Example: 5.0" → use 5.75" range (smallest available)
        # Example: 11.0" → use 10.25/10.50" range (nearest)
        if round_size > 0:
            # Find the nearest range by distance
            nearest = min(positive_ranges.items(),
                         key=lambda x: abs(x[0] - round_size))
            return nearest[1][:2]

        return None

    def detect_round_size_from_title(self, title):
        """Extract round size from title string"""
        if not title:
            return None

        import re

        # Pattern: Look for numbers followed by optional decimal and "OD" or "rnd"
        # Examples: "6.25 OD", "10.5 rnd", "7.0OD", "625OD"
        patterns = [
            r'(\d+\.?\d*)\s*(?:OD|od|rnd|RND|round)',  # 6.25 OD, 10.5 rnd
            r'(\d+)\.(\d+)',  # 6.25, 10.50
        ]

        for pattern in patterns:
            match = re.search(pattern, title)
            if match:
                try:
                    if len(match.groups()) == 2:
                        # Decimal number split into groups
                        round_size = float(f"{match.group(1)}.{match.group(2)}")
                    else:
                        round_size = float(match.group(1))

                    # Validate it's in a reasonable range
                    if 5.0 <= round_size <= 15.0:
                        return round_size
                except:
                    continue

        return None

    def detect_round_size_from_gcode(self, program_number):
        """Get round size from ob_from_gcode field"""
        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            cursor.execute("SELECT ob_from_gcode FROM programs WHERE program_number = ?",
                          (program_number,))
            result = cursor.fetchone()
            conn.close()

            if result and result[0]:
                ob_value = result[0]
                # ob_from_gcode is the outer bore/diameter
                if 5.0 <= ob_value <= 15.0:
                    return ob_value
        except Exception as e:
            print(f"[RoundSize] Error getting from gcode: {e}")

        return None

    def detect_round_size_from_dimension(self, program_number):
        """Get round size from outer_diameter field"""
        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            cursor.execute("SELECT outer_diameter FROM programs WHERE program_number = ?",
                          (program_number,))
            result = cursor.fetchone()
            conn.close()

            if result and result[0]:
                od_value = result[0]
                if 5.0 <= od_value <= 15.0:
                    return od_value
        except Exception as e:
            print(f"[RoundSize] Error getting from dimension: {e}")

        return None

    def detect_round_size(self, program_number, title=None):
        """
        Detect round size using multiple methods with priority order.
        Returns: (round_size, confidence, source)
        confidence: 'HIGH', 'MEDIUM', 'LOW', 'NONE'
        source: 'TITLE', 'GCODE', 'DIMENSION', 'MANUAL'
        """
        # Method 1: Parse title (most reliable if present)
        if title:
            title_match = self.detect_round_size_from_title(title)
            if title_match:
                return (title_match, 'HIGH', 'TITLE')

        # Method 2: Get from G-code OB (high confidence)
        gcode_match = self.detect_round_size_from_gcode(program_number)
        if gcode_match:
            return (gcode_match, 'HIGH', 'GCODE')

        # Method 3: Get from database dimension (medium confidence)
        dimension_match = self.detect_round_size_from_dimension(program_number)
        if dimension_match:
            return (dimension_match, 'MEDIUM', 'DIMENSION')

        # Method 4: Manual required
        return (None, 'NONE', 'MANUAL')

    def is_in_correct_range(self, program_number, round_size):
        """Check if program number is in correct range for its round size"""
        if not round_size:
            return True  # Can't validate without round size

        # Extract numeric part of program number
        try:
            prog_num = int(str(program_number).replace('o', '').replace('O', ''))
        except:
            return False

        # Get range for this round size
        range_info = self.get_range_for_round_size(round_size)
        if not range_info:
            return False

        range_start, range_end = range_info
        return range_start <= prog_num <= range_end

    def update_round_size_for_program(self, program_number, round_size=None, confidence=None,
                                     source=None, manual_override=False):
        """Update round size fields for a program"""
        try:
            # Auto-detect if not provided
            if round_size is None and not manual_override:
                round_size, confidence, source = self.detect_round_size(program_number)

            # Check if in correct range
            in_correct_range = 1 if self.is_in_correct_range(program_number, round_size) else 0

            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute("""
                UPDATE programs
                SET round_size = ?,
                    round_size_confidence = ?,
                    round_size_source = ?,
                    in_correct_range = ?
                WHERE program_number = ?
            """, (round_size, confidence, source, in_correct_range, program_number))

            conn.commit()
            conn.close()

            return True
        except Exception as e:
            print(f"[RoundSize] Error updating: {e}")
            return False

    def batch_detect_round_sizes(self, program_numbers=None):
        """Detect and update round sizes for multiple programs"""
        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            # Get programs to process
            if program_numbers:
                placeholders = ','.join('?' * len(program_numbers))
                query = f"SELECT program_number, title FROM programs WHERE program_number IN ({placeholders})"
                cursor.execute(query, program_numbers)
            else:
                cursor.execute("SELECT program_number, title FROM programs")

            programs = cursor.fetchall()
            conn.close()

            results = {
                'processed': 0,
                'detected': 0,
                'failed': 0,
                'manual_needed': 0
            }

            for program_number, title in programs:
                round_size, confidence, source = self.detect_round_size(program_number, title)

                if round_size:
                    if self.update_round_size_for_program(program_number, round_size,
                                                         confidence, source):
                        results['detected'] += 1
                    else:
                        results['failed'] += 1
                else:
                    results['manual_needed'] += 1

                results['processed'] += 1

            return results
        except Exception as e:
            print(f"[RoundSize] Batch detection error: {e}")
            return None

    def populate_program_registry(self):
        """
        Populate the program_number_registry table with all 97,001 program numbers.
        Marks existing programs as 'IN_USE' and tracks duplicates.

        Returns:
            dict: Statistics about registry population
        """
        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            # Get all round size ranges
            ranges = self.get_round_size_ranges()

            # Clear existing registry (fresh start)
            cursor.execute("DELETE FROM program_number_registry")

            # Get all existing programs from database
            cursor.execute("SELECT program_number, file_path FROM programs")
            existing_programs = {row[0]: row[1] for row in cursor.fetchall()}

            # Track statistics
            stats = {
                'total_generated': 0,
                'in_use': 0,
                'available': 0,
                'duplicates': 0,
                'by_range': {}
            }

            # Generate all program numbers for each range
            # Track which ranges we've already processed to avoid duplicates (e.g., 10.25 and 10.50 share same range)
            processed_ranges = set()

            for round_size, (range_start, range_end, range_name) in ranges.items():
                # Skip if we've already processed this range
                range_key = (range_start, range_end)
                if range_key in processed_ranges:
                    continue
                processed_ranges.add(range_key)

                range_stats = {
                    'total': 0,
                    'in_use': 0,
                    'available': 0
                }

                for prog_num in range(range_start, range_end + 1):
                    program_number = self.format_program_number(prog_num)

                    # Check if this program exists in database
                    if program_number in existing_programs:
                        status = 'IN_USE'
                        file_path = existing_programs[program_number]
                        stats['in_use'] += 1
                        range_stats['in_use'] += 1
                    else:
                        status = 'AVAILABLE'
                        file_path = None
                        stats['available'] += 1
                        range_stats['available'] += 1

                    # Insert into registry
                    cursor.execute("""
                        INSERT INTO program_number_registry
                        (program_number, round_size, range_start, range_end, status, file_path, last_checked)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (program_number, round_size, range_start, range_end, status, file_path,
                          datetime.now().isoformat()))

                    stats['total_generated'] += 1
                    range_stats['total'] += 1

                stats['by_range'][range_name] = range_stats

            # Find duplicates (programs with same number but different content)
            cursor.execute("""
                SELECT program_number, COUNT(*) as count
                FROM programs
                GROUP BY program_number
                HAVING count > 1
            """)

            duplicate_programs = cursor.fetchall()
            for prog_num, count in duplicate_programs:
                cursor.execute("""
                    UPDATE program_number_registry
                    SET duplicate_count = ?,
                        notes = 'WARNING: Multiple files with this program number'
                    WHERE program_number = ?
                """, (count, prog_num))
                stats['duplicates'] += 1

            conn.commit()
            conn.close()

            return stats

        except Exception as e:
            messagebox.showerror("Registry Error", f"Failed to populate program registry:\n{str(e)}")
            return None

    def find_next_available_number(self, round_size, preferred_number=None):
        """
        Find the next available program number for a given round size.

        Args:
            round_size: The round size (e.g., 6.25, 10.5)
            preferred_number: Optional preferred number to try first

        Returns:
            str: Next available program number (e.g., 'o62500') or None if range full
        """
        try:
            # Get the range for this round size
            range_info = self.get_range_for_round_size(round_size)
            if not range_info:
                return None

            range_start, range_end = range_info

            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # If preferred number provided, check if it's available
            if preferred_number:
                try:
                    pref_num = int(str(preferred_number).replace('o', '').replace('O', ''))
                    if range_start <= pref_num <= range_end:
                        cursor.execute("""
                            SELECT status FROM program_number_registry
                            WHERE program_number = ?
                        """, (self.format_program_number(pref_num),))
                        result = cursor.fetchone()
                        if result and result[0] == 'AVAILABLE':
                            conn.close()
                            return self.format_program_number(pref_num)
                except:
                    pass

            # Find first available number in range
            # Query by numeric range, not round_size (more reliable)
            cursor.execute("""
                SELECT program_number
                FROM program_number_registry
                WHERE CAST(REPLACE(program_number, 'o', '') AS INTEGER) BETWEEN ? AND ?
                AND status = 'AVAILABLE'
                ORDER BY CAST(REPLACE(program_number, 'o', '') AS INTEGER)
                LIMIT 1
            """, (range_start, range_end))

            result = cursor.fetchone()
            conn.close()

            if result:
                return result[0]
            else:
                return None  # Range is full

        except Exception as e:
            messagebox.showerror("Registry Error", f"Failed to find available number:\n{str(e)}")
            return None

    def get_registry_statistics(self):
        """
        Get statistics about the program number registry.

        Returns:
            dict: Statistics about each range and overall usage
        """
        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            stats = {
                'total_numbers': 0,
                'in_use': 0,
                'available': 0,
                'reserved': 0,
                'duplicates': 0,
                'by_range': {}
            }

            # Get overall statistics
            cursor.execute("""
                SELECT status, COUNT(*) as count
                FROM program_number_registry
                GROUP BY status
            """)

            for status, count in cursor.fetchall():
                stats['total_numbers'] += count
                if status == 'IN_USE':
                    stats['in_use'] = count
                elif status == 'AVAILABLE':
                    stats['available'] = count
                elif status == 'RESERVED':
                    stats['reserved'] = count

            # Get duplicate count
            cursor.execute("""
                SELECT COUNT(*) FROM program_number_registry
                WHERE duplicate_count > 0
            """)
            stats['duplicates'] = cursor.fetchone()[0]

            # Get statistics by range
            ranges = self.get_round_size_ranges()
            for round_size, (range_start, range_end, range_name) in ranges.items():
                cursor.execute("""
                    SELECT
                        COUNT(*) as total,
                        SUM(CASE WHEN status = 'IN_USE' THEN 1 ELSE 0 END) as in_use,
                        SUM(CASE WHEN status = 'AVAILABLE' THEN 1 ELSE 0 END) as available,
                        SUM(CASE WHEN duplicate_count > 0 THEN 1 ELSE 0 END) as duplicates
                    FROM program_number_registry
                    WHERE round_size = ?
                """, (round_size,))

                row = cursor.fetchone()
                stats['by_range'][range_name] = {
                    'round_size': round_size,
                    'range': f"o{range_start}-o{range_end}",
                    'total': row[0] or 0,
                    'in_use': row[1] or 0,
                    'available': row[2] or 0,
                    'duplicates': row[3] or 0,
                    'usage_percent': (row[1] or 0) / (row[0] or 1) * 100
                }

            conn.close()
            return stats

        except Exception as e:
            messagebox.showerror("Registry Error", f"Failed to get registry statistics:\n{str(e)}")
            return None

    def get_out_of_range_programs(self):
        """
        Get all programs that are in the wrong range for their round size.

        Returns:
            list: List of tuples (program_number, round_size, current_range, correct_range, title)
        """
        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            cursor.execute("""
                SELECT program_number, round_size, round_size_source, title
                FROM programs
                WHERE in_correct_range = 0
                AND round_size IS NOT NULL
                ORDER BY round_size, program_number
            """)

            out_of_range = []
            for prog_num, round_size, source, title in cursor.fetchall():
                # Get current range
                try:
                    num = int(str(prog_num).replace('o', '').replace('O', ''))
                    current_range = None
                    ranges = self.get_round_size_ranges()
                    for rs, (start, end, name) in ranges.items():
                        if start <= num <= end:
                            current_range = f"o{start}-o{end} ({name})"
                            break
                    if not current_range:
                        current_range = "Unknown"
                except:
                    current_range = "Invalid"

                # Get correct range
                correct_range_info = self.get_range_for_round_size(round_size)
                if correct_range_info:
                    correct_range = f"o{correct_range_info[0]}-o{correct_range_info[1]}"
                else:
                    correct_range = "No range defined"

                out_of_range.append((prog_num, round_size, current_range, correct_range, title))

            conn.close()
            return out_of_range

        except Exception as e:
            messagebox.showerror("Registry Error", f"Failed to get out-of-range programs:\n{str(e)}")
            return []

    def rename_to_correct_range(self, program_number, dry_run=False):
        """
        Rename a program to the correct range for its round size.
        This is the core function for Type 1 duplicate resolution.

        Args:
            program_number: Program to rename (e.g., 'o62000')
            dry_run: If True, only simulate the rename without making changes

        Returns:
            dict: Result with keys:
                - success: bool
                - old_number: str
                - new_number: str
                - round_size: float
                - file_path: str
                - legacy_name_added: bool
                - error: str (if failed)
        """
        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            # Get program info
            cursor.execute("""
                SELECT round_size, file_path, title, legacy_names
                FROM programs
                WHERE program_number = ?
            """, (program_number,))

            result = cursor.fetchone()
            if not result:
                return {'success': False, 'error': f'Program {program_number} not found'}

            round_size, file_path, title, legacy_names = result

            if not round_size:
                return {'success': False, 'error': f'Program {program_number} has no detected round size'}

            # Check if already in correct range
            if self.is_in_correct_range(program_number, round_size):
                return {'success': False, 'error': f'Program {program_number} is already in correct range'}

            # Find next available number in correct range
            new_number = self.find_next_available_number(round_size)
            if not new_number:
                return {'success': False, 'error': f'No available numbers in range for round size {round_size}'}

            if dry_run:
                conn.close()
                return {
                    'success': True,
                    'dry_run': True,
                    'old_number': program_number,
                    'new_number': new_number,
                    'round_size': round_size,
                    'file_path': file_path,
                    'title': title
                }

            # Update legacy names
            if legacy_names:
                try:
                    legacy_list = json.loads(legacy_names)
                except:
                    legacy_list = []
            else:
                legacy_list = []

            legacy_list.append({
                'old_number': program_number,
                'renamed_date': datetime.now().isoformat(),
                'reason': 'Out of range - moved to correct range'
            })

            # Read file content
            if not os.path.exists(file_path):
                conn.close()
                return {'success': False, 'error': f'File not found: {file_path}'}

            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()

            # Update program number in file content
            import re

            # Strip any suffix from old program number for matching internal content
            # Database might have o00002(1) or o00002_1, but file content has O00002
            old_num_plain = program_number.replace('o', '').replace('O', '')
            # Remove any suffix patterns: (1), (2), _1, _2, etc.
            old_num_base = re.sub(r'[(_]\d+[\)]?$', '', old_num_plain)

            new_num_plain = new_number.replace('o', '').replace('O', '')

            # Replace program number (common patterns)
            updated_content = content

            # Pattern 1: O12345 or o12345 at start of line
            # Use the BASE number without suffix to match file content
            updated_content = re.sub(
                rf'^[oO]{old_num_base}\b',
                new_number.upper(),
                updated_content,
                flags=re.MULTILINE
            )

            # Pattern 2: In program number comments
            # Use the BASE number without suffix to match file content
            updated_content = re.sub(
                rf'\b[oO]{old_num_base}\b',
                new_number.upper(),
                updated_content
            )

            # Add legacy comment at top of file (after first line if it's a program number)
            lines = updated_content.split('\n')
            legacy_comment = f"(RENAMED FROM {program_number.upper()} ON {datetime.now().strftime('%Y-%m-%d')} - OUT OF RANGE)"

            # Insert after the program number line
            if len(lines) > 0:
                lines.insert(1, legacy_comment)
            else:
                lines.insert(0, legacy_comment)

            updated_content = '\n'.join(lines)

            # Generate new file path with new program number
            old_dir = os.path.dirname(file_path)
            new_filename = f"{new_number}.nc"
            new_file_path = os.path.join(old_dir, new_filename)

            # Write updated content to new file
            with open(new_file_path, 'w', encoding='utf-8') as f:
                f.write(updated_content)

            # Delete old file if new file was created successfully
            if os.path.exists(new_file_path) and os.path.exists(file_path):
                os.remove(file_path)

            # Update database - programs table with new file path
            cursor.execute("""
                UPDATE programs
                SET program_number = ?,
                    file_path = ?,
                    legacy_names = ?,
                    last_renamed_date = ?,
                    rename_reason = 'Out of range correction',
                    in_correct_range = 1
                WHERE program_number = ?
            """, (new_number, new_file_path, json.dumps(legacy_list), datetime.now().isoformat(), program_number))

            # Update registry - mark old number as available
            cursor.execute("""
                UPDATE program_number_registry
                SET status = 'AVAILABLE',
                    file_path = NULL
                WHERE program_number = ?
            """, (program_number,))

            # Update registry - mark new number as in use with new file path
            cursor.execute("""
                UPDATE program_number_registry
                SET status = 'IN_USE',
                    file_path = ?
                WHERE program_number = ?
            """, (new_file_path, new_number))

            # Log resolution in duplicate_resolutions table
            cursor.execute("""
                INSERT INTO duplicate_resolutions
                (resolution_date, duplicate_type, program_numbers, action_taken,
                 files_affected, old_values, new_values, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                datetime.now().isoformat(),
                'TYPE_1_OUT_OF_RANGE',
                json.dumps([program_number, new_number]),
                'RENAME',
                json.dumps([{'old': file_path, 'new': new_file_path}]),
                json.dumps({'program_number': program_number, 'round_size': round_size, 'old_file': file_path}),
                json.dumps({'program_number': new_number, 'round_size': round_size, 'new_file': new_file_path}),
                f'Renamed from {program_number} to {new_number} - file renamed from {os.path.basename(file_path)} to {new_filename}'
            ))

            conn.commit()
            conn.close()

            return {
                'success': True,
                'old_number': program_number,
                'new_number': new_number,
                'round_size': round_size,
                'file_path': new_file_path,
                'old_file_path': file_path,
                'new_file_path': new_file_path,
                'title': title,
                'legacy_name_added': True
            }

        except Exception as e:
            return {'success': False, 'error': f'Rename failed: {str(e)}'}

    def batch_resolve_out_of_range(self, program_numbers=None, dry_run=False, progress_callback=None):
        """
        Batch rename programs that are out of range.

        Args:
            program_numbers: List of specific programs to rename, or None for all out-of-range
            dry_run: If True, simulate without making changes
            progress_callback: Function to call with progress updates

        Returns:
            dict: Statistics about the batch operation
        """
        try:
            # Get programs to process
            if program_numbers:
                programs_to_process = program_numbers
            else:
                out_of_range = self.get_out_of_range_programs()
                programs_to_process = [prog[0] for prog in out_of_range]  # Extract program numbers

            total = len(programs_to_process)
            stats = {
                'total': total,
                'successful': 0,
                'failed': 0,
                'skipped': 0,
                'errors': [],
                'renames': []
            }

            for i, prog_num in enumerate(programs_to_process):
                if progress_callback:
                    progress_callback(i + 1, total, prog_num)

                result = self.rename_to_correct_range(prog_num, dry_run=dry_run)

                if result['success']:
                    stats['successful'] += 1
                    stats['renames'].append({
                        'old': result['old_number'],
                        'new': result['new_number'],
                        'round_size': result['round_size'],
                        'file': result.get('file_path', '')
                    })
                elif 'already in correct range' in result.get('error', ''):
                    stats['skipped'] += 1
                else:
                    stats['failed'] += 1
                    stats['errors'].append({
                        'program': prog_num,
                        'error': result.get('error', 'Unknown error')
                    })

            return stats

        except Exception as e:
            return {
                'total': 0,
                'successful': 0,
                'failed': 0,
                'skipped': 0,
                'errors': [{'program': 'BATCH', 'error': str(e)}],
                'renames': []
            }

    def preview_rename_plan(self, limit=None):
        """
        Preview what would happen if we renamed all out-of-range programs.
        This is a dry-run of the batch resolution.

        Args:
            limit: Maximum number of programs to preview (None = all)

        Returns:
            list: Preview data with old number, new number, round size, title
        """
        try:
            out_of_range = self.get_out_of_range_programs()

            if limit:
                out_of_range = out_of_range[:limit]

            preview = []
            # Track which numbers we've already assigned in this batch
            # to prevent conflicts where multiple programs get assigned the same number
            assigned_in_batch = set()

            for prog_num, round_size, current_range, correct_range, title in out_of_range:
                # Find what the new number would be
                new_number = self.find_next_available_number(round_size)

                # If this number was already assigned in this batch, find the next one
                while new_number and new_number in assigned_in_batch:
                    # Get the numeric part of the new_number
                    try:
                        current_num = int(new_number.replace('o', '').replace('O', ''))
                        range_info = self.get_range_for_round_size(round_size)
                        if not range_info:
                            new_number = None
                            break

                        range_start, range_end = range_info

                        # Try the next sequential number
                        next_num = current_num + 1
                        if next_num > range_end:
                            # Range is full
                            new_number = None
                            break

                        # Check if this next number is available in the registry
                        conn = sqlite3.connect(self.db_path)
                        cursor = conn.cursor()
                        cursor.execute("""
                            SELECT status FROM program_number_registry
                            WHERE program_number = ?
                        """, (self.format_program_number(next_num),))
                        result = cursor.fetchone()
                        conn.close()

                        if result and result[0] == 'AVAILABLE':
                            new_number = self.format_program_number(next_num)
                        else:
                            # This number is also taken, keep searching
                            current_num = next_num
                            new_number = self.format_program_number(current_num)
                    except Exception:
                        new_number = None
                        break

                if new_number:
                    # Mark this number as assigned in this batch
                    assigned_in_batch.add(new_number)

                    preview.append({
                        'old_number': prog_num,
                        'new_number': new_number,
                        'round_size': round_size,
                        'current_range': current_range,
                        'correct_range': correct_range,
                        'title': title,
                        'status': 'Ready'
                    })
                else:
                    preview.append({
                        'old_number': prog_num,
                        'new_number': 'NO SPACE',
                        'round_size': round_size,
                        'current_range': current_range,
                        'correct_range': correct_range,
                        'title': title,
                        'status': 'Error: No available numbers'
                    })

            return preview

        except Exception as e:
            messagebox.showerror("Preview Error", f"Failed to generate preview:\n{str(e)}")
            return []

    def load_config(self):
        """Load configuration from file"""
        default_config = {
            "last_folder": "",
            "material_list": ["6061-T6", "Steel", "Stainless", "Other"],
            "spacer_types": ["standard", "hub_centric", "steel_ring", "2pc_part1", "2pc_part2"]
        }
        
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r') as f:
                    self.config = json.load(f)
            except:
                self.config = default_config
        else:
            self.config = default_config
            
    def save_config(self):
        """Save configuration to file"""
        with open(self.config_file, 'w') as f:
            json.dump(self.config, f, indent=2)

    def backup_database(self) -> bool:
        """Create a backup of the database in a special backup folder

        Returns:
            bool: True if backup was successful, False otherwise
        """
        try:
            # Create backup folder if it doesn't exist
            backup_folder = "Database_Backups"
            if not os.path.exists(backup_folder):
                os.makedirs(backup_folder)

            # Generate backup filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_filename = f"gcode_database_backup_{timestamp}.db"
            backup_path = os.path.join(backup_folder, backup_filename)

            # Copy database file
            shutil.copy2(self.db_path, backup_path)

            # Keep only last 10 backups to prevent folder bloat
            self.cleanup_old_backups(backup_folder, keep_count=10)

            return True
        except Exception as e:
            messagebox.showerror("Backup Error", f"Failed to create database backup:\n{str(e)}")
            return False

    def cleanup_old_backups(self, backup_folder: str, keep_count: int = 10):
        """Keep only the most recent N backups

        Args:
            backup_folder: Path to backup folder
            keep_count: Number of recent backups to keep
        """
        try:
            # Get all backup files
            backup_files = [f for f in os.listdir(backup_folder)
                           if f.startswith("gcode_database_backup_") and f.endswith(".db")]

            # Sort by modification time (newest first)
            backup_files.sort(key=lambda x: os.path.getmtime(os.path.join(backup_folder, x)), reverse=True)

            # Delete older backups
            for old_backup in backup_files[keep_count:]:
                old_path = os.path.join(backup_folder, old_backup)
                os.remove(old_path)
        except Exception as e:
            # Don't show error for cleanup failures, just log it
            print(f"Backup cleanup warning: {str(e)}")

    def get_available_values(self, column: str) -> List[str]:
        """Get distinct values from database column for filter dropdowns"""
        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()
            cursor.execute(f"SELECT DISTINCT {column} FROM programs WHERE {column} IS NOT NULL ORDER BY {column}")
            values = [row[0] for row in cursor.fetchall()]
            conn.close()
            return values
        except:
            return []
            
    def setup_gui(self):
        """Setup the main GUI"""
        # Create main container
        main_container = tk.Frame(self.root, bg=self.bg_color)
        main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Top section - Title and Ribbon Tabs
        top_frame = tk.Frame(main_container, bg=self.bg_color)
        top_frame.pack(fill=tk.X, pady=(0, 10))

        self.create_top_section(top_frame)

        # Ribbon tabs section
        ribbon_frame = tk.Frame(main_container, bg=self.bg_color, relief=tk.RAISED, borderwidth=1)
        ribbon_frame.pack(fill=tk.X, pady=(0, 10))

        self.create_ribbon_tabs(ribbon_frame)

        # Create dropdown selector for view mode instead of tabs
        view_selector_frame = tk.Frame(main_container, bg=self.bg_color)
        view_selector_frame.pack(fill=tk.X, pady=(0, 5))

        tk.Label(view_selector_frame, text="View:", bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)

        self.view_mode_var = tk.StringVar(value="all")
        view_dropdown = ttk.Combobox(view_selector_frame, textvariable=self.view_mode_var,
                                     values=["all", "repository", "external"],
                                     state="readonly", width=20, font=("Arial", 10))
        view_dropdown.pack(side=tk.LEFT, padx=5)
        view_dropdown.bind('<<ComboboxSelected>>', self.on_view_mode_change)

        # Create action button frames (hidden/shown based on view mode)
        action_buttons_frame = tk.Frame(main_container, bg=self.bg_color)
        action_buttons_frame.pack(fill=tk.X, pady=(0, 5))

        # Create separate frames for each view mode
        self.all_programs_tab = tk.Frame(action_buttons_frame, bg=self.bg_color)
        self.repository_tab = tk.Frame(action_buttons_frame, bg=self.bg_color)
        self.external_tab = tk.Frame(action_buttons_frame, bg=self.bg_color)

        # Store reference to action buttons frame
        self.action_buttons_container = action_buttons_frame

        # Setup tab-specific action buttons
        self.setup_all_programs_tab()
        self.setup_repository_tab()
        self.setup_external_tab()

        # Show "all" tab by default
        self.all_programs_tab.pack(fill=tk.X)

        # Create shared filter and results section (below tabs)
        # Middle section - Filters with toggle capability
        self.filter_outer_frame = tk.Frame(main_container, bg=self.bg_color)
        self.filter_outer_frame.pack(fill=tk.X, pady=(0, 10))

        # Toggle bar - always visible
        toggle_bar = tk.Frame(self.filter_outer_frame, bg=self.bg_color)
        toggle_bar.pack(fill=tk.X)

        # Toggle button
        self.filter_expanded = True
        self.toggle_filter_btn = tk.Button(toggle_bar, text="▼ Search & Filter",
                                          command=self.toggle_filter_section,
                                          bg=self.accent_color, fg=self.fg_color,
                                          font=("Arial", 10, "bold"), width=20, anchor="w")
        self.toggle_filter_btn.pack(side=tk.LEFT, padx=5, pady=2)

        tk.Label(toggle_bar, text="(Click to minimize/expand)",
                bg=self.bg_color, fg="#888888", font=("Arial", 8, "italic")).pack(side=tk.LEFT, padx=5)

        # Collapsible filter frame
        filter_frame = tk.Frame(self.filter_outer_frame, bg=self.bg_color)
        filter_frame.pack(fill=tk.X, pady=(5, 0))
        self.filter_collapsible_frame = filter_frame

        self.create_filter_section(filter_frame)

        # Results header with counter
        results_header = tk.Frame(main_container, bg=self.bg_color)
        results_header.pack(fill=tk.X, pady=(5, 0))

        tk.Label(results_header, text="Results", bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=10)

        # Results counter label - visible and prominent
        self.results_counter_label = tk.Label(results_header, text="", bg=self.bg_color,
                                              fg="#4CAF50", font=("Arial", 10, "bold"))
        self.results_counter_label.pack(side=tk.LEFT, padx=10)

        # Bottom section - Results
        results_frame = tk.Frame(main_container, bg=self.bg_color, relief=tk.SUNKEN, borderwidth=1)
        results_frame.pack(fill=tk.BOTH, expand=True)

        self.create_results_section(results_frame)

    def setup_all_programs_tab(self):
        """Setup the All Programs tab - shows all files (repository + external)"""
        # Info and action buttons
        info_frame = tk.Frame(self.all_programs_tab, bg=self.bg_color)
        info_frame.pack(fill=tk.X, pady=5, padx=10)

        # Info label
        tk.Label(info_frame, text="Viewing all programs (repository + external files)",
                bg=self.bg_color, fg=self.fg_color, font=("Arial", 10, "italic")).pack(side=tk.LEFT)

        # Stats button
        tk.Button(info_frame, text="📊 Stats", command=self.show_all_programs_stats,
                 bg=self.accent_color, fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=10, height=1).pack(side=tk.RIGHT, padx=5)

    def setup_repository_tab(self):
        """Setup the Repository tab (managed files only)"""
        # Info and action buttons
        info_frame = tk.Frame(self.repository_tab, bg=self.bg_color)
        info_frame.pack(fill=tk.X, pady=5, padx=10)

        # Info label
        tk.Label(info_frame, text="Repository: Managed files in repository/ folder",
                bg=self.bg_color, fg=self.fg_color, font=("Arial", 10, "italic")).pack(side=tk.LEFT)

        # Repository stats button
        tk.Button(info_frame, text="📊 Stats", command=self.show_repository_stats,
                 bg=self.accent_color, fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=10, height=1).pack(side=tk.RIGHT, padx=5)

        # Repository management buttons - Row 1 (Basic operations)
        repo_buttons1 = tk.Frame(self.repository_tab, bg=self.bg_color)
        repo_buttons1.pack(fill=tk.X, pady=5, padx=10)

        tk.Button(repo_buttons1, text="🗑️ Delete", command=self.delete_from_repository,
                 bg="#D32F2F", fg=self.fg_color, font=("Arial", 8, "bold"),
                 width=12).pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)

        tk.Button(repo_buttons1, text="📤 Export", command=self.export_selected_file,
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 8, "bold"),
                 width=12).pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)

        tk.Button(repo_buttons1, text="🔍 Manage Duplicates",
                 command=self.manage_duplicates,
                 bg="#FF6B00", fg=self.fg_color, font=("Arial", 8, "bold"),
                 width=18).pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)

        tk.Button(repo_buttons1, text="🔄 Sync Filenames",
                 command=self.sync_filenames_with_database,
                 bg="#00BCD4", fg=self.fg_color, font=("Arial", 8, "bold"),
                 width=16).pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)

        # Repository management buttons - Row 2 (Program Number Management)
        repo_buttons2 = tk.Frame(self.repository_tab, bg=self.bg_color)
        repo_buttons2.pack(fill=tk.X, pady=2, padx=10)

        tk.Button(repo_buttons2, text="📋 Registry",
                 command=self.show_registry_window,
                 bg="#6B5B93", fg=self.fg_color, font=("Arial", 8, "bold"),
                 width=12).pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)

        tk.Button(repo_buttons2, text="⚠️ Out-of-Range",
                 command=self.show_out_of_range_window,
                 bg="#C41E3A", fg=self.fg_color, font=("Arial", 8, "bold"),
                 width=14).pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)

        tk.Button(repo_buttons2, text="🔧 Batch Rename",
                 command=self.show_batch_rename_window,
                 bg="#9B59B6", fg=self.fg_color, font=("Arial", 8, "bold"),
                 width=14).pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)

        tk.Button(repo_buttons2, text="🎯 Move to Range",
                 command=self.move_to_correct_range,
                 bg="#E91E63", fg=self.fg_color, font=("Arial", 8, "bold"),
                 width=14).pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)

        # Repository management buttons - Row 3 (Utilities)
        repo_buttons3 = tk.Frame(self.repository_tab, bg=self.bg_color)
        repo_buttons3.pack(fill=tk.X, pady=2, padx=10)

        tk.Button(repo_buttons3, text="📦 Export by Round Size",
                 command=self.export_repository_by_round_size,
                 bg="#4CAF50", fg=self.fg_color, font=("Arial", 8, "bold"),
                 width=20).pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)

        tk.Button(repo_buttons3, text="🔧 Repair Paths",
                 command=self.repair_file_paths,
                 bg="#FF9800", fg=self.fg_color, font=("Arial", 8, "bold"),
                 width=14).pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)

        tk.Button(repo_buttons3, text="🔢 Fix Prog# Format",
                 command=self.fix_program_number_formatting,
                 bg="#9C27B0", fg=self.fg_color, font=("Arial", 8, "bold"),
                 width=16).pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)

    def setup_external_tab(self):
        """Setup the External/Scanned tab (external files only)"""
        # Info and action buttons
        info_frame = tk.Frame(self.external_tab, bg=self.bg_color)
        info_frame.pack(fill=tk.X, pady=5, padx=10)

        tk.Label(info_frame, text="External: Scanned files NOT in repository (temporary view)",
                bg=self.bg_color, fg=self.fg_color, font=("Arial", 10, "italic")).pack(side=tk.LEFT)

        # Stats button
        tk.Button(info_frame, text="📊 Stats", command=self.show_external_stats,
                 bg=self.accent_color, fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=10, height=1).pack(side=tk.RIGHT, padx=5)

        # External file management buttons
        external_buttons = tk.Frame(self.external_tab, bg=self.bg_color)
        external_buttons.pack(fill=tk.X, pady=5, padx=10)

        tk.Button(external_buttons, text="➕ Add to Repository", command=self.add_selected_to_repository,
                 bg="#388E3C", fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=18, height=1).pack(side=tk.LEFT, padx=3, expand=True, fill=tk.X)

        tk.Button(external_buttons, text="🗑️ Remove from DB", command=self.remove_from_database,
                 bg="#D32F2F", fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=15, height=1).pack(side=tk.LEFT, padx=3, expand=True, fill=tk.X)

        tk.Button(external_buttons, text="🔄 Refresh", command=lambda: self.refresh_results(view_mode='external'),
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=12, height=1).pack(side=tk.LEFT, padx=3, expand=True, fill=tk.X)

    def on_view_mode_change(self, event=None):
        """Handle view mode dropdown change"""
        view_mode = self.view_mode_var.get()

        # Hide all tab frames
        self.all_programs_tab.pack_forget()
        self.repository_tab.pack_forget()
        self.external_tab.pack_forget()

        # Show the selected tab frame
        if view_mode == 'all':
            self.all_programs_tab.pack(fill=tk.X)
            self.refresh_results(view_mode='all')
        elif view_mode == 'repository':
            self.repository_tab.pack(fill=tk.X)
            self.refresh_results(view_mode='repository')
        elif view_mode == 'external':
            self.external_tab.pack(fill=tk.X)
            self.refresh_results(view_mode='external')

    def toggle_filter_section(self):
        """Toggle the visibility of the filter section"""
        if self.filter_expanded:
            # Collapse the filter section
            self.filter_collapsible_frame.pack_forget()
            self.toggle_filter_btn.config(text="▶ Search & Filter (Click to expand)")
            self.filter_expanded = False
        else:
            # Expand the filter section
            self.filter_collapsible_frame.pack(fill=tk.X, pady=(5, 0))
            self.toggle_filter_btn.config(text="▼ Search & Filter")
            self.filter_expanded = True

    def create_top_section(self, parent):
        """Create top action buttons"""
        # Title
        title = tk.Label(parent, text="G-Code Database Manager",
                        font=("Arial", 16, "bold"),
                        bg=self.bg_color, fg=self.fg_color)
        title.pack(side=tk.LEFT, padx=10)

    def create_ribbon_tabs(self, parent):
        """Create ribbon-style tab interface for organizing buttons"""
        # Create notebook (tabbed interface)
        style = ttk.Style()
        style.configure('Ribbon.TNotebook', background=self.bg_color, borderwidth=0)
        style.configure('Ribbon.TNotebook.Tab', padding=[20, 10], font=('Arial', 10, 'bold'))

        ribbon = ttk.Notebook(parent, style='Ribbon.TNotebook')
        ribbon.pack(fill=tk.X, padx=5, pady=5)

        # Tab 1: File Management
        tab_files = tk.Frame(ribbon, bg=self.bg_color)
        ribbon.add(tab_files, text='📂 Files')

        files_group = tk.Frame(tab_files, bg=self.bg_color)
        files_group.pack(fill=tk.X, padx=5, pady=5)

        tk.Button(files_group, text="📁 Scan Folder", command=self.scan_folder,
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(files_group, text="🆕 Scan New Only", command=self.scan_for_new_files,
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(files_group, text="🔄 Rescan Database", command=self.rescan_database,
                 bg="#FF6F00", fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(files_group, text="⚡ Rescan Changed", command=self.rescan_changed_files,
                 bg="#2E7D32", fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(files_group, text="📁 Organize by OD", command=self.organize_files_by_od,
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(files_group, text="📋 Copy Filtered", command=self.copy_filtered_view,
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(files_group, text="➕ Add Entry", command=self.add_entry,
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        # Tab 2: Duplicates
        tab_duplicates = tk.Frame(ribbon, bg=self.bg_color)
        ribbon.add(tab_duplicates, text='🔍 Duplicates')

        dup_group = tk.Frame(tab_duplicates, bg=self.bg_color)
        dup_group.pack(fill=tk.X, padx=5, pady=5)

        tk.Button(dup_group, text="🔍 Find Repeats", command=self.find_and_mark_repeats,
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(dup_group, text="⚖️ Compare Files", command=self.compare_files,
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(dup_group, text="📝 Rename Duplicates", command=self.rename_duplicate_files,
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(dup_group, text="🔧 Fix Program #s", command=self.fix_program_numbers,
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(dup_group, text="🗑️ Delete Duplicates", command=self.delete_duplicates,
                 bg="#D32F2F", fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        # Tab 3: Reports
        tab_reports = tk.Frame(ribbon, bg=self.bg_color)
        ribbon.add(tab_reports, text='📊 Reports')

        reports_group = tk.Frame(tab_reports, bg=self.bg_color)
        reports_group.pack(fill=tk.X, padx=5, pady=5)

        tk.Button(reports_group, text="📊 Export Excel", command=self.export_csv,
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(reports_group, text="📈 Google Sheets", command=self.export_google_sheets,
                 bg="#34A853", fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(reports_group, text="📋 Unused #s", command=self.export_unused_numbers,
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(reports_group, text="📊 Statistics", command=self.show_statistics,
                 bg="#1976D2", fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(reports_group, text="🔧 Tool Analysis", command=self.view_tool_statistics,
                 bg="#9C27B0", fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(reports_group, text="❓ Help & Workflow", command=self.show_legend,
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        # Tab 4: Backup & Database
        tab_database = tk.Frame(ribbon, bg=self.bg_color)
        ribbon.add(tab_database, text='💾 Backup')

        db_group = tk.Frame(tab_database, bg=self.bg_color)
        db_group.pack(fill=tk.X, padx=5, pady=5)

        tk.Button(db_group, text="💾 Backup DB", command=self.create_manual_backup,
                 bg="#1976D2", fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(db_group, text="📦 Full Backup", command=self.create_full_backup,
                 bg="#2E7D32", fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(db_group, text="📂 Restore Backup", command=self.restore_from_backup,
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(db_group, text="📋 View Backups", command=self.view_backups,
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(db_group, text="💾 Save Profile", command=self.save_database_profile,
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(db_group, text="📂 Load Profile", command=self.load_database_profile,
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(db_group, text="📋 Manage Profiles", command=self.manage_database_profiles,
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        # Tab 5: Workflow
        tab_workflow = tk.Frame(ribbon, bg=self.bg_color)
        ribbon.add(tab_workflow, text='🔄 Workflow')

        workflow_group = tk.Frame(tab_workflow, bg=self.bg_color)
        workflow_group.pack(fill=tk.X, padx=5, pady=5)

        tk.Button(workflow_group, text="🔄 Sync Registry", command=self.sync_registry_ui,
                 bg="#9C27B0", fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(workflow_group, text="🎯 Detect Round Sizes", command=self.detect_round_sizes_ui,
                 bg="#673AB7", fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(workflow_group, text="📊 Round Size Stats", command=self.show_round_size_stats,
                 bg="#3F51B5", fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(workflow_group, text="📘 Workflow Guide", command=self.show_workflow_guide,
                 bg="#2196F3", fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        # Tab 6: Maintenance
        tab_maint = tk.Frame(ribbon, bg=self.bg_color)
        ribbon.add(tab_maint, text='⚙️ Maintenance')

        maint_group = tk.Frame(tab_maint, bg=self.bg_color)
        maint_group.pack(fill=tk.X, padx=5, pady=5)

        tk.Button(maint_group, text="🗑️ Clear Database", command=self.clear_database,
                 bg="#D32F2F", fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

        tk.Button(maint_group, text="❌ Delete Filtered", command=self.delete_filtered_view,
                 bg="#D32F2F", fg=self.fg_color, font=("Arial", 9, "bold"),
                 width=14, height=2).pack(side=tk.LEFT, padx=3)

    def create_filter_section(self, parent):
        """Create filter controls"""
        filter_container = tk.Frame(parent, bg=self.bg_color)
        filter_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Row 0 - Title Search Bar (prominent)
        row0 = tk.Frame(filter_container, bg=self.bg_color)
        row0.pack(fill=tk.X, pady=(5, 10))

        tk.Label(row0, text="🔍 Title Search:", bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
        self.filter_title = tk.Entry(row0, bg=self.input_bg, fg=self.fg_color,
                                     font=("Arial", 10), width=60)
        self.filter_title.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # Bind Enter key to trigger search
        self.filter_title.bind('<Return>', lambda _: self.refresh_results())

        # Clear title search button
        btn_clear_title = tk.Button(row0, text="✕", command=self.clear_title_search,
                                   bg=self.button_bg, fg=self.fg_color,
                                   font=("Arial", 9, "bold"), width=3)
        btn_clear_title.pack(side=tk.LEFT, padx=2)

        tk.Label(row0, text="(Use + for multi-term search, e.g., \"lug + 1.25 + 74\". Press Enter to search)",
                bg=self.bg_color, fg="#888888", font=("Arial", 8, "italic")).pack(side=tk.LEFT, padx=5)

        # Row 1
        row1 = tk.Frame(filter_container, bg=self.bg_color)
        row1.pack(fill=tk.X, pady=5)

        # Program Number
        tk.Label(row1, text="Program #:", bg=self.bg_color, fg=self.fg_color).pack(side=tk.LEFT, padx=5)
        self.filter_program = tk.Entry(row1, bg=self.input_bg, fg=self.fg_color, width=15)
        self.filter_program.pack(side=tk.LEFT, padx=5)
        
        # Spacer Type
        tk.Label(row1, text="Type:", bg=self.bg_color, fg=self.fg_color).pack(side=tk.LEFT, padx=5)
        type_values = self.available_types if self.available_types else self.config["spacer_types"]
        self.filter_type = MultiSelectCombobox(row1, type_values, self.bg_color, self.fg_color,
                                              self.input_bg, self.button_bg, width=15)
        self.filter_type.pack(side=tk.LEFT, padx=5)

        # Material
        tk.Label(row1, text="Material:", bg=self.bg_color, fg=self.fg_color).pack(side=tk.LEFT, padx=5)
        material_values = self.available_materials if self.available_materials else self.config["material_list"]
        self.filter_material = MultiSelectCombobox(row1, material_values, self.bg_color, self.fg_color,
                                                  self.input_bg, self.button_bg, width=15)
        self.filter_material.pack(side=tk.LEFT, padx=5)

        # Validation Status
        tk.Label(row1, text="Status:", bg=self.bg_color, fg=self.fg_color).pack(side=tk.LEFT, padx=5)
        status_values = self.available_statuses if self.available_statuses else ["CRITICAL", "SAFETY_ERROR", "TOOL_ERROR", "BORE_WARNING", "DIMENSIONAL", "TOOL_WARNING", "WARNING", "PASS"]
        self.filter_status = MultiSelectCombobox(row1, status_values, self.bg_color, self.fg_color,
                                                self.input_bg, self.button_bg, width=18)
        self.filter_status.pack(side=tk.LEFT, padx=5)

        # Duplicate Type
        tk.Label(row1, text="Dup Type:", bg=self.bg_color, fg=self.fg_color).pack(side=tk.LEFT, padx=5)
        dup_type_values = ["SOLID", "NAME_COLLISION", "CONTENT_DUP", "None"]
        self.filter_dup_type = MultiSelectCombobox(row1, dup_type_values, self.bg_color, self.fg_color,
                                                   self.input_bg, self.button_bg, width=15)
        self.filter_dup_type.pack(side=tk.LEFT, padx=5)

        # Row 2 - Dimensional filters
        row2 = tk.Frame(filter_container, bg=self.bg_color)
        row2.pack(fill=tk.X, pady=5)
        
        # Outer Diameter
        # OD Range - Common standard sizes
        tk.Label(row2, text="OD Range:", bg=self.bg_color, fg=self.fg_color).pack(side=tk.LEFT, padx=5)
        od_values = ["", "5.75", "6.0", "6.25", "6.5", "7.0", "7.5", "8.0", "8.5", "9.5", "10.25", "10.5", "13.0"]
        self.filter_od_min = ttk.Combobox(row2, values=od_values, width=8)
        self.filter_od_min.pack(side=tk.LEFT, padx=2)
        tk.Label(row2, text="to", bg=self.bg_color, fg=self.fg_color).pack(side=tk.LEFT, padx=2)
        self.filter_od_max = ttk.Combobox(row2, values=od_values, width=8)
        self.filter_od_max.pack(side=tk.LEFT, padx=2)

        # Thickness - Common values
        tk.Label(row2, text="Thick:", bg=self.bg_color, fg=self.fg_color).pack(side=tk.LEFT, padx=5)
        thick_values = ["", "0.25", "0.38", "0.5", "0.62", "0.75", "1.0", "1.25", "1.5", "1.75", "2.0", "2.5", "3.0"]
        self.filter_thickness_min = ttk.Combobox(row2, values=thick_values, width=8)
        self.filter_thickness_min.pack(side=tk.LEFT, padx=2)
        tk.Label(row2, text="to", bg=self.bg_color, fg=self.fg_color).pack(side=tk.LEFT, padx=2)
        self.filter_thickness_max = ttk.Combobox(row2, values=thick_values, width=8)
        self.filter_thickness_max.pack(side=tk.LEFT, padx=2)

        # Center Bore - Common CB sizes (mm)
        tk.Label(row2, text="CB:", bg=self.bg_color, fg=self.fg_color).pack(side=tk.LEFT, padx=5)
        cb_values = ["", "38", "40", "50", "54", "56", "60", "63", "66", "70", "77", "78", "84", "93", "100", "106", "108", "110", "125", "130", "150", "170", "220"]
        self.filter_cb_min = ttk.Combobox(row2, values=cb_values, width=8)
        self.filter_cb_min.pack(side=tk.LEFT, padx=2)
        tk.Label(row2, text="to", bg=self.bg_color, fg=self.fg_color).pack(side=tk.LEFT, padx=2)
        self.filter_cb_max = ttk.Combobox(row2, values=cb_values, width=8)
        self.filter_cb_max.pack(side=tk.LEFT, padx=2)

        # Row 2.5 - Additional dimensional filters (Hub Dia, Hub H, Step)
        row2_5 = tk.Frame(filter_container, bg=self.bg_color)
        row2_5.pack(fill=tk.X, pady=5)

        # Hub Diameter - Common OB/Hub sizes (mm)
        tk.Label(row2_5, text="Hub Dia:", bg=self.bg_color, fg=self.fg_color).pack(side=tk.LEFT, padx=5)
        hub_dia_values = ["", "54", "56", "57", "59", "60", "63", "66", "70", "73", "74", "77", "78", "84", "87", "93", "95", "100", "106", "108", "110"]
        self.filter_hub_dia_min = ttk.Combobox(row2_5, values=hub_dia_values, width=8)
        self.filter_hub_dia_min.pack(side=tk.LEFT, padx=2)
        tk.Label(row2_5, text="to", bg=self.bg_color, fg=self.fg_color).pack(side=tk.LEFT, padx=2)
        self.filter_hub_dia_max = ttk.Combobox(row2_5, values=hub_dia_values, width=8)
        self.filter_hub_dia_max.pack(side=tk.LEFT, padx=2)

        # Hub Height - Common hub heights (inches)
        tk.Label(row2_5, text="Hub H:", bg=self.bg_color, fg=self.fg_color).pack(side=tk.LEFT, padx=5)
        hub_h_values = ["", "0.25", "0.33", "0.38", "0.44", "0.47", "0.5", "0.55", "0.6", "0.65", "0.7", "0.75", "1.0", "1.25", "1.5"]
        self.filter_hub_h_min = ttk.Combobox(row2_5, values=hub_h_values, width=8)
        self.filter_hub_h_min.pack(side=tk.LEFT, padx=2)
        tk.Label(row2_5, text="to", bg=self.bg_color, fg=self.fg_color).pack(side=tk.LEFT, padx=2)
        self.filter_hub_h_max = ttk.Combobox(row2_5, values=hub_h_values, width=8)
        self.filter_hub_h_max.pack(side=tk.LEFT, padx=2)

        # Step Diameter - Common step/shelf sizes (mm)
        tk.Label(row2_5, text="Step D:", bg=self.bg_color, fg=self.fg_color).pack(side=tk.LEFT, padx=5)
        step_d_values = ["", "64", "70", "74", "78", "82", "84", "85", "87", "90", "93", "95", "100", "106", "108", "110", "125", "130"]
        self.filter_step_d_min = ttk.Combobox(row2_5, values=step_d_values, width=8)
        self.filter_step_d_min.pack(side=tk.LEFT, padx=2)
        tk.Label(row2_5, text="to", bg=self.bg_color, fg=self.fg_color).pack(side=tk.LEFT, padx=2)
        self.filter_step_d_max = ttk.Combobox(row2_5, values=step_d_values, width=8)
        self.filter_step_d_max.pack(side=tk.LEFT, padx=2)

        # Row 2.6 - Error type filter
        row2_6 = tk.Frame(filter_container, bg=self.bg_color)
        row2_6.pack(fill=tk.X, pady=5)

        tk.Label(row2_6, text="Error Contains:", bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=5)

        # Combobox with common error types (editable for custom searches)
        common_errors = [
            "",  # Empty for "all"
            "CB TOO LARGE",
            "CB TOO SMALL",
            "THICKNESS ERROR",
            "OB TOO LARGE",
            "OB TOO SMALL",
            "OD MISMATCH",
            "DRILL",
            "CHAMFER",
            "STEP",
            "HUB",
            "BORE"
        ]
        self.filter_error_text = ttk.Combobox(row2_6, values=common_errors, width=47)
        self.filter_error_text.pack(side=tk.LEFT, padx=2)
        self.filter_error_text.set("")  # Default to empty

        tk.Label(row2_6, text="(Select common error or type custom search)",
                bg=self.bg_color, fg="#888888", font=("Arial", 8, "italic")).pack(side=tk.LEFT, padx=5)

        # Row 3 - Sort options
        row3 = tk.Frame(filter_container, bg=self.bg_color)
        row3.pack(fill=tk.X, pady=5)

        tk.Label(row3, text="Sort by:", bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=5)

        # Sort columns available
        sort_columns = ["", "Program #", "Dup", "Type", "Lathe", "OD", "Thick", "CB", "Hub H", "Hub D",
                       "CB Bore", "Step D", "Material", "Status"]

        # Sort 1
        self.sort1_col = ttk.Combobox(row3, values=sort_columns, state="readonly", width=10)
        self.sort1_col.pack(side=tk.LEFT, padx=2)
        self.sort1_col.set("CB")
        self.sort1_dir = ttk.Combobox(row3, values=["Low→High", "High→Low"], state="readonly", width=9)
        self.sort1_dir.pack(side=tk.LEFT, padx=2)
        self.sort1_dir.set("Low→High")

        tk.Label(row3, text="then", bg=self.bg_color, fg=self.fg_color).pack(side=tk.LEFT, padx=3)

        # Sort 2
        self.sort2_col = ttk.Combobox(row3, values=sort_columns, state="readonly", width=10)
        self.sort2_col.pack(side=tk.LEFT, padx=2)
        self.sort2_col.set("OD")
        self.sort2_dir = ttk.Combobox(row3, values=["Low→High", "High→Low"], state="readonly", width=9)
        self.sort2_dir.pack(side=tk.LEFT, padx=2)
        self.sort2_dir.set("Low→High")

        tk.Label(row3, text="then", bg=self.bg_color, fg=self.fg_color).pack(side=tk.LEFT, padx=3)

        # Sort 3
        self.sort3_col = ttk.Combobox(row3, values=sort_columns, state="readonly", width=10)
        self.sort3_col.pack(side=tk.LEFT, padx=2)
        self.sort3_col.set("")
        self.sort3_dir = ttk.Combobox(row3, values=["Low→High", "High→Low"], state="readonly", width=9)
        self.sort3_dir.pack(side=tk.LEFT, padx=2)
        self.sort3_dir.set("Low→High")

        btn_sort = tk.Button(row3, text="↕️ Sort", command=self.apply_multi_sort,
                            bg=self.button_bg, fg=self.fg_color,
                            font=("Arial", 9, "bold"), width=8)
        btn_sort.pack(side=tk.LEFT, padx=5)

        # Row 4 - Action buttons
        row4 = tk.Frame(filter_container, bg=self.bg_color)
        row4.pack(fill=tk.X, pady=5)

        btn_search = tk.Button(row4, text="🔍 Search", command=self.refresh_results,
                              bg=self.accent_color, fg=self.fg_color,
                              font=("Arial", 10, "bold"), width=12, height=1)
        btn_search.pack(side=tk.LEFT, padx=5)

        btn_refresh = tk.Button(row4, text="🔄 Refresh", command=self.refresh_results,
                               bg="#2E7D32", fg=self.fg_color,
                               font=("Arial", 10, "bold"), width=12, height=1)
        btn_refresh.pack(side=tk.LEFT, padx=5)

        btn_clear = tk.Button(row4, text="❌ Clear Filters", command=self.clear_filters,
                             bg=self.button_bg, fg=self.fg_color,
                             font=("Arial", 10, "bold"), width=12, height=1)
        btn_clear.pack(side=tk.LEFT, padx=5)

        btn_batch = tk.Button(row4, text="📦 Batch Operations", command=self.show_batch_operations,
                             bg="#FF6F00", fg=self.fg_color,
                             font=("Arial", 10, "bold"), width=15, height=1)
        btn_batch.pack(side=tk.LEFT, padx=5)

        # Duplicates only checkbox
        self.filter_duplicates = tk.BooleanVar()
        dup_check = tk.Checkbutton(row4, text="Duplicates Only", variable=self.filter_duplicates,
                                   bg=self.bg_color, fg=self.fg_color, selectcolor=self.input_bg,
                                   activebackground=self.bg_color, activeforeground=self.fg_color,
                                   font=("Arial", 9))
        dup_check.pack(side=tk.LEFT, padx=10)

        # Missing file path checkbox
        self.filter_missing_file_path = tk.BooleanVar()
        missing_path_check = tk.Checkbutton(row4, text="Missing File Path", variable=self.filter_missing_file_path,
                                           bg=self.bg_color, fg=self.fg_color, selectcolor=self.input_bg,
                                           activebackground=self.bg_color, activeforeground=self.fg_color,
                                           font=("Arial", 9))
        missing_path_check.pack(side=tk.LEFT, padx=10)

        # Results count label
        self.results_label = tk.Label(row4, text="", bg=self.bg_color, fg=self.fg_color,
                                     font=("Arial", 10))
        self.results_label.pack(side=tk.RIGHT, padx=10)
        
    def create_results_section(self, parent):
        """Create results table"""
        # Create treeview with scrollbars
        tree_frame = tk.Frame(parent, bg=self.bg_color)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        # Treeview
        columns = ("Program #", "Dup", "Title", "Type", "Lathe", "OD", "Thick", "CB", "Hub H", "Hub D",
                  "CB Bore", "Step D", "Material", "Status", "Warning Details", "File")

        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings",
                                yscrollcommand=vsb.set, xscrollcommand=hsb.set,
                                selectmode='extended')  # Enable multi-select with Ctrl/Shift

        # Configure tags for color coding (severity-based)
        self.tree.tag_configure('critical', background='#4d1f1f', foreground='#ff6b6b')         # RED - Critical errors
        self.tree.tag_configure('safety_error', background='#3d1515', foreground='#ff4444')    # DARK RED - Safety blocks missing
        self.tree.tag_configure('tool_error', background='#4d2515', foreground='#ff7744')      # ORANGE-RED - Tool errors
        self.tree.tag_configure('bore_warning', background='#4d3520', foreground='#ffa500')    # ORANGE - Bore warnings
        self.tree.tag_configure('dimensional', background='#3d1f4d', foreground='#da77f2')     # PURPLE - P-code/thickness
        self.tree.tag_configure('tool_warning', background='#4d4015', foreground='#ffcc44')    # AMBER - Tool suggestions
        self.tree.tag_configure('warning', background='#4d3d1f', foreground='#ffd43b')         # YELLOW - General warnings
        self.tree.tag_configure('repeat', background='#3d3d3d', foreground='#909090')          # GRAY - Repeat files
        self.tree.tag_configure('pass', background='#1f4d2e', foreground='#69db7c')            # GREEN - Pass
        
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        
        # Configure columns
        column_widths = {
            "Program #": 100,
            "Dup": 40,
            "Title": 250,
            "Type": 120,
            "Lathe": 60,
            "OD": 80,
            "Thick": 80,
            "CB": 80,
            "Hub H": 80,
            "Hub D": 80,
            "CB Bore": 80,
            "Step D": 70,
            "Material": 100,
            "Status": 90,
            "Warning Details": 300,
            "File": 200
        }
        
        for col in columns:
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_column(c))
            # Warning Details and Title should be left-aligned for readability
            if col in ["Warning Details", "Title", "File"]:
                self.tree.column(col, width=column_widths[col], anchor="w")
            else:
                self.tree.column(col, width=column_widths[col], anchor="center")
        
        # Layout
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Context menu
        self.tree.bind("<Button-3>", self.show_context_menu)
        self.tree.bind("<Double-1>", self.open_file)
        
        # Bottom action buttons - made more compact
        action_frame = tk.Frame(parent, bg=self.bg_color)
        action_frame.pack(fill=tk.X, padx=10, pady=(0, 5))

        btn_open = tk.Button(action_frame, text="📄 Open", command=self.open_file,
                            bg=self.button_bg, fg=self.fg_color,
                            font=("Arial", 8, "bold"), width=10)
        btn_open.pack(side=tk.LEFT, padx=3)

        btn_edit = tk.Button(action_frame, text="✏️ Edit", command=self.edit_entry,
                            bg=self.button_bg, fg=self.fg_color,
                            font=("Arial", 8, "bold"), width=10)
        btn_edit.pack(side=tk.LEFT, padx=3)

        btn_delete = tk.Button(action_frame, text="🗑️ Delete", command=self.delete_entry,
                              bg=self.button_bg, fg=self.fg_color,
                              font=("Arial", 8, "bold"), width=10)
        btn_delete.pack(side=tk.LEFT, padx=3)

        btn_view = tk.Button(action_frame, text="👁️ Details", command=self.view_details,
                            bg=self.button_bg, fg=self.fg_color,
                            font=("Arial", 8, "bold"), width=10)
        btn_view.pack(side=tk.LEFT, padx=3)

        btn_compare = tk.Button(action_frame, text="🔄 Compare", command=self.compare_files,
                               bg="#FF9800", fg=self.fg_color,
                               font=("Arial", 8, "bold"), width=12)
        btn_compare.pack(side=tk.LEFT, padx=3)

        # Status label for drag-drop feedback (separate from results counter)
        self.dragdrop_label = tk.Label(action_frame, text="",
                                     bg=self.bg_color, fg=self.fg_color,
                                     font=("Arial", 9))
        self.dragdrop_label.pack(side=tk.RIGHT, padx=10)

    def setup_drag_drop(self):
        """Setup drag and drop for the main window"""
        try:
            from tkinterdnd2 import DND_FILES

            # Check if root is already a TkinterDnD.Tk instance
            # If not, we'll use event bindings as fallback
            try:
                self.root.drop_target_register(DND_FILES)
                self.root.dnd_bind('<<Drop>>', self.on_drop)
                self.root.dnd_bind('<<DragEnter>>', self.on_drag_enter)
                self.root.dnd_bind('<<DragLeave>>', self.on_drag_leave)
                print("[Drag & Drop] TkinterDnD2 enabled successfully")
            except Exception:
                print("[Drag & Drop] TkinterDnD2 not available, using fallback method")
                self._setup_drag_drop_fallback()
        except ImportError:
            print("[Drag & Drop] TkinterDnD2 not installed, using fallback method")
            self._setup_drag_drop_fallback()

    def _setup_drag_drop_fallback(self):
        """Fallback drag-drop using standard tk file dialog"""
        # Add a button to manually add files if drag-drop not available
        pass  # Manual add file functionality already exists via scan_folder

    def on_drag_enter(self, _):
        """Visual feedback when files are dragged over window"""
        self.root.configure(bg="#3a4a5a")  # Slightly lighter bg
        if hasattr(self, 'dragdrop_label'):
            self.dragdrop_label.config(text="Drop files here to import...")

    def on_drag_leave(self, _):
        """Restore appearance when drag leaves"""
        self.root.configure(bg=self.bg_color)
        if hasattr(self, 'dragdrop_label'):
            self.dragdrop_label.config(text="")

    def on_drop(self, event):
        """Handle dropped files"""
        self.root.configure(bg=self.bg_color)

        # Parse the file paths from the drop event
        files = self._parse_drop_files(event.data)

        if not files:
            if hasattr(self, 'dragdrop_label'):
                self.dragdrop_label.config(text="No valid files dropped")
            return

        # Filter for gcode files only
        gcode_files = [f for f in files if self._is_gcode_file(f)]

        if not gcode_files:
            messagebox.showwarning("No G-Code Files",
                                 f"No G-Code files found in {len(files)} dropped file(s).\n\n"
                                 "G-Code files must contain pattern: o##### (4+ digits)")
            if hasattr(self, 'dragdrop_label'):
                self.dragdrop_label.config(text="")
            return

        # Process the files
        self.process_dropped_files(gcode_files)

    def _parse_drop_files(self, data):
        """Parse file paths from drag-drop event data"""
        # Handle different formats from different systems
        if isinstance(data, list):
            return data

        # Windows/Unix paths can be space or newline separated
        # Wrapped in {} or not
        files = []
        data = str(data)

        # Remove outer braces if present
        if data.startswith('{') and data.endswith('}'):
            data = data[1:-1]

        # Split by spaces, but respect paths with spaces
        import shlex
        try:
            files = shlex.split(data)
        except:
            # Fallback: simple split
            files = data.split()

        return [f.strip() for f in files if f.strip()]

    def _is_gcode_file(self, filepath):
        """Check if file is a G-code file based on naming pattern"""
        if not os.path.isfile(filepath):
            return False

        filename = os.path.basename(filepath)
        # Match files with o##### pattern (4+ digits)
        return bool(re.search(r'[oO]\d{4,}', filename))

    def process_dropped_files(self, filepaths):
        """
        Process dropped files with smart duplicate detection and warnings.

        Features:
        - Duplicate detection (exact content match)
        - Filename collision warnings with auto-rename suggestions
        - Program number mismatch warnings
        """
        # Get existing files from database
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute("SELECT program_number, file_path FROM programs WHERE file_path IS NOT NULL")
        db_files = cursor.fetchall()

        # Build lookup structures
        db_programs = {}  # {program_number: file_path}
        db_filenames = {}  # {filename_lower: (program_number, file_path)}
        db_contents = {}  # {content_hash: (program_number, file_path)}

        for prog_num, file_path in db_files:
            db_programs[prog_num] = file_path

            if file_path and os.path.exists(file_path):
                filename_lower = os.path.basename(file_path).lower()
                db_filenames[filename_lower] = (prog_num, file_path)

                # Read content for exact duplicate detection
                try:
                    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                        content = f.read()
                        content_hash = hash(content)
                        db_contents[content_hash] = (prog_num, file_path)
                except Exception:
                    pass

        conn.close()

        # Analyze each dropped file
        files_to_add = []
        warnings = []

        for filepath in filepaths:
            filename = os.path.basename(filepath)
            filename_lower = filename.lower()

            # Extract program number from filename
            match = re.search(r'[oO](\d{4,})', filename)
            if not match:
                warnings.append(f"⚠️ Skipped {filename}: No program number found")
                continue

            file_prog_num = 'o' + match.group(1)

            # Read file content
            try:
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                    content_hash = hash(content)
            except Exception as e:
                warnings.append(f"❌ Error reading {filename}: {str(e)}")
                continue

            # Check 1: Exact content duplicate
            if content_hash in db_contents:
                existing_prog, existing_path = db_contents[content_hash]
                existing_name = os.path.basename(existing_path)
                warnings.append(f"🔁 Duplicate: {filename} is identical to {existing_name} (already in database)")
                continue

            # Check 2: Filename collision (same name, different content)
            if filename_lower in db_filenames:
                existing_prog, existing_path = db_filenames[filename_lower]

                # Offer auto-rename suggestion
                suggestion = self._suggest_unique_filename(filename, db_filenames)

                result = messagebox.askyesnocancel(
                    "Filename Collision Detected",
                    f"File already exists: {filename}\n"
                    f"In database as: {existing_prog}\n\n"
                    f"The file you're adding has DIFFERENT content.\n\n"
                    f"Options:\n"
                    f"• YES: Save with suggested name: {suggestion}\n"
                    f"• NO: Skip this file\n"
                    f"• CANCEL: Stop importing\n\n"
                    f"Suggested name will also update internal program number."
                )

                if result is None:  # Cancel
                    break
                elif result:  # Yes - rename
                    # Add with new name
                    files_to_add.append((filepath, suggestion, content))
                    warnings.append(f"✏️ Renamed: {filename} → {suggestion}")
                else:  # No - skip
                    warnings.append(f"⏭️ Skipped: {filename} (filename collision)")
                continue

            # Check 3: Program number mismatch (filename != internal program number)
            # Parse the file to get internal program number
            internal_prog_num = self._extract_program_number_from_content(content)

            if internal_prog_num and internal_prog_num != file_prog_num:
                result = messagebox.askyesno(
                    "Program Number Mismatch",
                    f"Warning: Filename and internal program number don't match!\n\n"
                    f"Filename: {filename} ({file_prog_num})\n"
                    f"Internal: {internal_prog_num}\n\n"
                    f"This file may have been renamed incorrectly.\n\n"
                    f"Add it anyway?\n"
                    f"(It will be stored as {file_prog_num} based on filename)"
                )

                if not result:
                    warnings.append(f"⏭️ Skipped: {filename} (program number mismatch)")
                    continue
                else:
                    warnings.append(f"⚠️ Added with mismatch: {filename} (internal: {internal_prog_num})")

            # Check 4: Program number already exists (different file)
            if file_prog_num in db_programs:
                existing_path = db_programs[file_prog_num]
                existing_name = os.path.basename(existing_path)

                if existing_name.lower() != filename_lower:
                    result = messagebox.askyesno(
                        "Program Number Exists",
                        f"Program number {file_prog_num} already exists!\n\n"
                        f"Existing: {existing_name}\n"
                        f"New file: {filename}\n\n"
                        f"These appear to be different files with the same program number.\n\n"
                        f"Add the new file anyway?\n"
                        f"(Will replace the existing entry)"
                    )

                    if not result:
                        warnings.append(f"⏭️ Skipped: {filename} (program number exists)")
                        continue

            # All checks passed - add file
            files_to_add.append((filepath, filename, content))

        # Process files to add
        if files_to_add:
            self._import_files(files_to_add, warnings)

        # Show summary
        if warnings:
            summary = "\n".join(warnings)
            messagebox.showinfo("Import Summary",
                              f"Processed {len(filepaths)} file(s)\n"
                              f"Added: {len(files_to_add)}\n\n"
                              f"Details:\n{summary[:500]}")

        # Refresh the display
        self.refresh_results()

        if hasattr(self, 'dragdrop_label'):
            self.dragdrop_label.config(text=f"Imported {len(files_to_add)} file(s)")

    def _suggest_unique_filename(self, original_filename, existing_filenames):
        """Suggest a unique filename by appending _v2, _v3, etc."""
        base, ext = os.path.splitext(original_filename)

        # Extract program number
        match = re.search(r'([oO]\d{4,})', base)
        if not match:
            return original_filename

        prog_prefix = match.group(1)

        counter = 2
        while True:
            # Update program number
            new_prog_num = f"{prog_prefix[0]}{int(prog_prefix[1:]) + counter - 1}"
            new_filename = base.replace(prog_prefix, new_prog_num) + ext

            if new_filename.lower() not in existing_filenames:
                return new_filename

            counter += 1

            if counter > 1000:  # Safety limit
                return f"{base}_UNIQUE{ext}"

    def _extract_program_number_from_content(self, content):
        """Extract program number from G-code content"""
        # Look for O##### pattern in first few lines
        lines = content.split('\n')[:50]  # Check first 50 lines

        for line in lines:
            match = re.search(r'[oO](\d{4,})', line)
            if match:
                return 'o' + match.group(1)

        return None

    def _import_files(self, files_to_add, warnings):
        """Import files into the database"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        target_folder = self.config.get("target_folder", "")
        if not target_folder or not os.path.exists(target_folder):
            messagebox.showerror("No Target Folder",
                               "Please set a target folder in settings before importing files.")
            conn.close()
            return

        for original_path, suggested_filename, _ in files_to_add:
            try:
                # Determine destination path
                dest_path = os.path.join(target_folder, suggested_filename)

                # Copy file to target folder
                if original_path != dest_path:
                    shutil.copy2(original_path, dest_path)

                # Parse the file
                record = self.parse_gcode_file(dest_path)
                if not record:
                    warnings.append(f"⚠️ Failed to parse: {suggested_filename}")
                    continue

                # Insert into database (or update if exists)
                cursor.execute("""
                    INSERT OR REPLACE INTO programs (
                        program_number, title, spacer_type, outer_diameter, thickness, thickness_display,
                        center_bore, hub_height, hub_diameter, counter_bore_diameter, counter_bore_depth,
                        paired_program, material, notes, date_created, last_modified, file_path,
                        detection_confidence, detection_method, validation_status, validation_issues,
                        validation_warnings, cb_from_gcode, ob_from_gcode, bore_warnings, dimensional_issues,
                        lathe, duplicate_type
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    record.program_number,
                    record.title,
                    record.spacer_type,
                    record.outer_diameter,
                    record.thickness,
                    record.thickness_display,
                    record.center_bore,
                    record.hub_height,
                    record.hub_diameter,
                    record.counter_bore_diameter,
                    record.counter_bore_depth,
                    record.paired_program,
                    record.material,
                    record.notes,
                    record.date_created,
                    record.last_modified,
                    record.file_path,
                    record.detection_confidence,
                    record.detection_method,
                    record.validation_status,
                    record.validation_issues,
                    record.validation_warnings,
                    record.cb_from_gcode,
                    record.ob_from_gcode,
                    record.bore_warnings,
                    record.dimensional_issues,
                    record.lathe,
                    record.duplicate_type
                ))

                warnings.append(f"✅ Added: {suggested_filename}")

            except Exception as e:
                warnings.append(f"❌ Error importing {suggested_filename}: {str(e)}")

        conn.commit()
        conn.close()

    def parse_gcode_file(self, filepath: str) -> Optional[ProgramRecord]:
        """Parse a gcode file using the improved parser"""
        try:
            # Use improved parser
            result = self.parser.parse_file(filepath)
            if not result:
                return None

            # Determine validation status (prioritized by severity)
            validation_status = "PASS"
            if result.validation_issues:
                validation_status = "CRITICAL"  # RED - Critical errors
            elif result.bore_warnings:
                validation_status = "BORE_WARNING"  # ORANGE - Bore dimension warnings
            elif result.dimensional_issues:
                validation_status = "DIMENSIONAL"  # PURPLE - P-code/thickness mismatches
            elif result.validation_warnings:
                validation_status = "WARNING"  # YELLOW - General warnings

            # Convert to ProgramRecord
            return ProgramRecord(
                program_number=result.program_number,
                title=result.title,
                spacer_type=result.spacer_type,
                outer_diameter=result.outer_diameter,
                thickness=result.thickness,
                thickness_display=result.thickness_display,
                center_bore=result.center_bore,
                hub_height=result.hub_height,
                hub_diameter=result.hub_diameter,
                counter_bore_diameter=result.counter_bore_diameter,
                counter_bore_depth=result.counter_bore_depth,
                paired_program=None,
                material=result.material,
                notes=None,
                date_created=result.date_created,
                last_modified=result.last_modified,
                file_path=result.file_path,
                detection_confidence=result.detection_confidence,
                detection_method=result.detection_method,
                validation_status=validation_status,
                validation_issues=json.dumps(result.validation_issues) if result.validation_issues else None,
                validation_warnings=json.dumps(result.validation_warnings) if result.validation_warnings else None,
                bore_warnings=json.dumps(result.bore_warnings) if result.bore_warnings else None,
                dimensional_issues=json.dumps(result.dimensional_issues) if result.dimensional_issues else None,
                cb_from_gcode=result.cb_from_gcode,
                ob_from_gcode=result.ob_from_gcode,
                lathe=result.lathe
            )

        except Exception as e:
            print(f"Error parsing {filepath}: {e}")
            import traceback
            traceback.print_exc()
            return None
            
    def extract_dimension(self, text: str, patterns: List[str]) -> Optional[float]:
        """Extract dimension using multiple regex patterns"""
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                try:
                    return float(match.group(1))
                except:
                    continue
        return None

    def update_gcode_program_number(self, file_path, new_program_number):
        """Update the internal O-number in a G-code file

        Args:
            file_path: Path to the G-code file
            new_program_number: New program number (e.g., "o57001" or "57001")

        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Ensure program number has 'o' prefix
            if not new_program_number.lower().startswith('o'):
                new_program_number = 'o' + new_program_number

            # Read file content
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()

            # Find and update the program number line (usually first line starting with O)
            updated = False
            for i, line in enumerate(lines):
                stripped = line.strip().upper()
                # Match O##### at start of line
                if stripped.startswith('O') and re.match(r'^O\d{4,}', stripped):
                    # Replace the O-number
                    lines[i] = new_program_number.upper() + '\n'
                    updated = True
                    break

            if updated:
                # Write back to file
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.writelines(lines)
                return True
            else:
                # No O-number found in file, add it at the beginning
                lines.insert(0, new_program_number.upper() + '\n')
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.writelines(lines)
                return True

        except Exception as e:
            print(f"Error updating G-code program number: {e}")
            return False

    def scan_folder(self):
        """Scan a folder for gcode files and import them"""
        folder = filedialog.askdirectory(title="Select Folder with G-Code Files",
                                        initialdir=self.config.get("last_folder", ""))

        if not folder:
            return

        # Ask user: Import to repository or keep external?
        import_mode_dialog = tk.Toplevel(self.root)
        import_mode_dialog.title("Import Mode")
        import_mode_dialog.geometry("550x300")
        import_mode_dialog.configure(bg=self.bg_color)
        import_mode_dialog.transient(self.root)
        import_mode_dialog.grab_set()

        # Title
        tk.Label(import_mode_dialog, text="📁 Choose Import Mode",
                font=("Arial", 14, "bold"), bg=self.bg_color, fg=self.fg_color).pack(pady=15)

        # Description
        desc = ("How would you like to handle the scanned files?\n\n"
                "• Repository: Copy files to managed repository folder\n"
                "  (Files under your control, organized, tracked)\n\n"
                "• External: Reference files in their current location\n"
                "  (Files stay where they are, temporary view)")
        tk.Label(import_mode_dialog, text=desc,
                font=("Arial", 10), bg=self.bg_color, fg=self.fg_color,
                justify=tk.LEFT).pack(padx=20, pady=10)

        # Store result
        import_to_repo = [False]  # Use list to allow modification in nested function

        def choose_repository():
            import_to_repo[0] = True
            import_mode_dialog.destroy()

        def choose_external():
            import_to_repo[0] = False
            import_mode_dialog.destroy()

        # Buttons
        btn_frame = tk.Frame(import_mode_dialog, bg=self.bg_color)
        btn_frame.pack(pady=15)

        tk.Button(btn_frame, text="📁 Repository\n(Copy to Repository)",
                 command=choose_repository,
                 bg="#388E3C", fg=self.fg_color, font=("Arial", 10, "bold"),
                 width=20, height=3).pack(side=tk.LEFT, padx=10)

        tk.Button(btn_frame, text="🔍 External\n(Keep in Place)",
                 command=choose_external,
                 bg=self.accent_color, fg=self.fg_color, font=("Arial", 10, "bold"),
                 width=20, height=3).pack(side=tk.LEFT, padx=10)

        # Wait for user choice
        self.root.wait_window(import_mode_dialog)

        # Store the choice for this scan
        add_to_repository = import_to_repo[0]

        self.config["last_folder"] = folder
        self.save_config()
        
        # Show progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Scanning Files...")
        progress_window.geometry("700x500")
        progress_window.configure(bg=self.bg_color)

        progress_label = tk.Label(progress_window, text="Scanning folder...",
                                 bg=self.bg_color, fg=self.fg_color,
                                 font=("Arial", 12))
        progress_label.pack(pady=20)

        progress_text = scrolledtext.ScrolledText(progress_window,
                                                 bg=self.input_bg, fg=self.fg_color,
                                                 font=("Courier", 9),
                                                 width=80, height=20)
        progress_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        self.root.update()

        # Scan for gcode files (with or without extension)
        all_scanned_files = []
        file_count = 0
        for root, dirs, files in os.walk(folder):
            for file in files:
                # Match any file containing o##### pattern (4+ digits)
                # This includes: o57000, o57000.nc, o57000.gcode, etc.
                if re.search(r'[oO]\d{4,}', file):
                    all_scanned_files.append(os.path.join(root, file))
                    file_count += 1
                    # Update progress every 50 files
                    if file_count % 50 == 0:
                        progress_label.config(text=f"Scanning... found {file_count} files so far")
                        self.root.update()

        progress_label.config(text=f"Analyzing {len(all_scanned_files)} files...")
        progress_text.insert(tk.END, f"Found {len(all_scanned_files)} total files. Analyzing...\n\n")
        progress_text.see(tk.END)
        self.root.update()

        # Get existing files from database
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT file_path FROM programs WHERE file_path IS NOT NULL")
        existing_files_data = {}  # {filename_lower: [file_paths]}

        for row in cursor.fetchall():
            file_path = row[0]
            if file_path and os.path.exists(file_path):
                basename = os.path.basename(file_path).lower()
                if basename not in existing_files_data:
                    existing_files_data[basename] = []
                existing_files_data[basename].append(file_path)

        # Detect duplicates within the scan AND against database
        files_by_name = {}  # Group scanned files by filename
        for filepath in all_scanned_files:
            basename_lower = os.path.basename(filepath).lower()
            if basename_lower not in files_by_name:
                files_by_name[basename_lower] = []
            files_by_name[basename_lower].append(filepath)

        # Categorize files
        files_to_process = []
        exact_duplicates_db = []  # Same name + content as database
        exact_duplicates_scan = []  # Same name + content within scan
        name_collisions_db = []  # Same name, different content vs database
        name_collisions_scan = []  # Same name, different content within scan

        analyzed_count = 0
        total_unique_names = len(files_by_name)

        for basename_lower, filepaths in files_by_name.items():
            analyzed_count += 1
            # Update progress every 10 unique filenames
            if analyzed_count % 10 == 0 or analyzed_count == total_unique_names:
                progress_label.config(text=f"Analyzing duplicates... {analyzed_count}/{total_unique_names}")
                self.root.update()
            # Check against database first
            in_database = basename_lower in existing_files_data

            if len(filepaths) == 1:
                # Single file with this name in scan
                filepath = filepaths[0]

                if in_database:
                    # Compare content with database file(s)
                    try:
                        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                            new_content = f.read()

                        is_exact_match = False
                        for db_path in existing_files_data[basename_lower]:
                            try:
                                with open(db_path, 'r', encoding='utf-8', errors='ignore') as f:
                                    db_content = f.read()
                                if new_content == db_content:
                                    is_exact_match = True
                                    break
                            except:
                                continue

                        if is_exact_match:
                            exact_duplicates_db.append(filepath)
                        else:
                            name_collisions_db.append((filepath, existing_files_data[basename_lower][0]))
                    except:
                        name_collisions_db.append((filepath, existing_files_data[basename_lower][0]))
                else:
                    # New file, safe to add
                    files_to_process.append(filepath)

            else:
                # Multiple files with same name in scan
                # Read all contents and find unique ones
                file_contents = {}
                for filepath in filepaths:
                    try:
                        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                            content = f.read()
                        if content not in file_contents:
                            file_contents[content] = []
                        file_contents[content].append(filepath)
                    except:
                        # If can't read, treat as unique
                        file_contents[filepath] = [filepath]

                # For each unique content, keep first file
                for content, paths in file_contents.items():
                    first_file = paths[0]
                    duplicates_in_scan = paths[1:]

                    # Check if this content matches database
                    if in_database:
                        try:
                            is_exact_match = False
                            for db_path in existing_files_data[basename_lower]:
                                try:
                                    with open(db_path, 'r', encoding='utf-8', errors='ignore') as f:
                                        db_content = f.read()
                                    if content == db_content:
                                        is_exact_match = True
                                        break
                                except:
                                    continue

                            if is_exact_match:
                                exact_duplicates_db.append(first_file)
                            else:
                                name_collisions_db.append((first_file, existing_files_data[basename_lower][0]))
                        except:
                            name_collisions_db.append((first_file, existing_files_data[basename_lower][0]))
                    else:
                        # New unique content, add first one
                        files_to_process.append(first_file)

                    # Mark duplicates within scan
                    for dup in duplicates_in_scan:
                        exact_duplicates_scan.append((dup, first_file))

        # Display analysis results
        progress_text.insert(tk.END, f"=== SCAN ANALYSIS ===\n")
        progress_text.insert(tk.END, f"Files to process: {len(files_to_process)}\n")
        progress_text.insert(tk.END, f"Exact duplicates (vs database): {len(exact_duplicates_db)}\n")
        progress_text.insert(tk.END, f"Exact duplicates (within scan): {len(exact_duplicates_scan)}\n")
        progress_text.insert(tk.END, f"Name collisions (vs database): {len(name_collisions_db)}\n")
        progress_text.insert(tk.END, f"Name collisions (within scan): {len(name_collisions_scan)}\n\n")

        # Show details
        if exact_duplicates_db:
            progress_text.insert(tk.END, f"=== SKIPPING - Already in Database (Exact Match) ===\n")
            for dup in exact_duplicates_db[:10]:
                progress_text.insert(tk.END, f"  SKIP: {os.path.basename(dup)}\n")
            if len(exact_duplicates_db) > 10:
                progress_text.insert(tk.END, f"  ... and {len(exact_duplicates_db) - 10} more\n")
            progress_text.insert(tk.END, "\n")

        if exact_duplicates_scan:
            progress_text.insert(tk.END, f"=== SKIPPING - Duplicates Within Scan ===\n")
            for dup, kept in exact_duplicates_scan[:10]:
                progress_text.insert(tk.END, f"  SKIP: {os.path.basename(dup)} (same as {os.path.basename(kept)})\n")
            if len(exact_duplicates_scan) > 10:
                progress_text.insert(tk.END, f"  ... and {len(exact_duplicates_scan) - 10} more\n")
            progress_text.insert(tk.END, "\n")

        if name_collisions_db:
            progress_text.insert(tk.END, f"⚠️  WARNING - Name Collisions vs Database ===\n")
            progress_text.insert(tk.END, f"These files will be SKIPPED. Rename them first!\n\n")
            for new_file, db_file in name_collisions_db[:5]:
                progress_text.insert(tk.END, f"  ⚠️  {os.path.basename(new_file)}\n")
                progress_text.insert(tk.END, f"     Scan: {new_file}\n")
                progress_text.insert(tk.END, f"     Database: {db_file}\n\n")
            if len(name_collisions_db) > 5:
                progress_text.insert(tk.END, f"  ... and {len(name_collisions_db) - 5} more\n")
            progress_text.insert(tk.END, "\n")

        progress_text.insert(tk.END, f"Processing {len(files_to_process)} files...\n\n")
        progress_text.see(tk.END)
        self.root.update()

        # Process only the unique files
        added = 0
        updated = 0
        errors = 0
        duplicates_within_processing = 0  # Track duplicates found during processing

        # Track which program numbers we've seen during processing
        seen_in_scan = {}  # program_number -> filepath
        gcode_files = files_to_process  # Use filtered list
        total_to_process = len(gcode_files)

        for idx, filepath in enumerate(gcode_files, 1):
            filename = os.path.basename(filepath)
            progress_label.config(text=f"Processing {idx}/{total_to_process}: {filename}")
            progress_text.insert(tk.END, f"[{idx}/{total_to_process}] Processing: {filename}\n")
            progress_text.see(tk.END)
            self.root.update()
            
            try:
                record = self.parse_gcode_file(filepath)
            except Exception as e:
                errors += 1
                progress_text.insert(tk.END, f"  PARSE EXCEPTION: {str(e)[:100]}\n")
                progress_text.see(tk.END)
                continue

            if record:
                try:
                    # Check for duplicate in this scan and assign unique suffix
                    original_prog_num = record.program_number
                    if record.program_number in seen_in_scan:
                        # Find next available suffix
                        suffix = 1
                        while f"{original_prog_num}({suffix})" in seen_in_scan:
                            suffix += 1
                        record.program_number = f"{original_prog_num}({suffix})"
                        progress_text.insert(tk.END, f"  DUPLICATE: {original_prog_num} -> saved as {record.program_number}\n")
                        progress_text.see(tk.END)
                        duplicates_within_processing += 1

                    # Track this file with its (possibly modified) program number
                    seen_in_scan[record.program_number] = filepath

                    # Check if exists in database
                    cursor.execute("SELECT program_number FROM programs WHERE program_number = ?",
                                 (record.program_number,))
                    exists = cursor.fetchone()

                    if exists:
                        # Update existing
                        cursor.execute('''
                            UPDATE programs SET
                                title = ?, spacer_type = ?, outer_diameter = ?, thickness = ?, thickness_display = ?,
                                center_bore = ?, hub_height = ?, hub_diameter = ?,
                                counter_bore_diameter = ?, counter_bore_depth = ?,
                                material = ?, last_modified = ?, file_path = ?,
                                detection_confidence = ?, detection_method = ?,
                                validation_status = ?, validation_issues = ?,
                                validation_warnings = ?, bore_warnings = ?, dimensional_issues = ?,
                                cb_from_gcode = ?, ob_from_gcode = ?, lathe = ?
                            WHERE program_number = ?
                        ''', (record.title, record.spacer_type, record.outer_diameter, record.thickness, record.thickness_display,
                             record.center_bore, record.hub_height, record.hub_diameter,
                             record.counter_bore_diameter, record.counter_bore_depth,
                             record.material, record.last_modified, record.file_path,
                             record.detection_confidence, record.detection_method,
                             record.validation_status, record.validation_issues,
                             record.validation_warnings, record.bore_warnings, record.dimensional_issues,
                             record.cb_from_gcode, record.ob_from_gcode, record.lathe,
                             record.program_number))
                        updated += 1
                    else:
                        # If user chose repository mode, import the file first
                        final_file_path = record.file_path
                        is_managed_file = 0

                        if add_to_repository and record.file_path and os.path.exists(record.file_path):
                            imported_path = self.import_to_repository(record.file_path, record.program_number)
                            if imported_path:
                                final_file_path = imported_path
                                is_managed_file = 1
                                progress_text.insert(tk.END, f"  → Imported to repository\n")
                                progress_text.see(tk.END)

                        # Insert new
                        cursor.execute('''
                            INSERT INTO programs VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (record.program_number, record.title, record.spacer_type, record.outer_diameter,
                             record.thickness, record.thickness_display, record.center_bore, record.hub_height,
                             record.hub_diameter, record.counter_bore_diameter,
                             record.counter_bore_depth, record.paired_program,
                             record.material, record.notes, record.date_created,
                             record.last_modified, final_file_path, record.detection_confidence,
                             record.detection_method, record.validation_status,
                             record.validation_issues, record.validation_warnings,
                             record.bore_warnings, record.dimensional_issues,
                             record.cb_from_gcode, record.ob_from_gcode, record.lathe,
                             None, None, None,  # duplicate_type, parent_file, duplicate_group
                             None, self.current_username, is_managed_file,  # current_version, modified_by, is_managed
                             None, None, None, None,  # round_size, round_size_confidence, round_size_source, in_correct_range
                             None, None, None,  # legacy_names, last_renamed_date, rename_reason
                             None, None, None, None, None, None))  # tools_used, tool_sequence, tool_validation_status, tool_validation_issues, safety_blocks_status, safety_blocks_issues
                        added += 1
                        
                except Exception as e:
                    errors += 1
                    progress_text.insert(tk.END, f"  DATABASE ERROR: {str(e)[:100]}\n")
                    progress_text.see(tk.END)
            else:
                errors += 1
                progress_text.insert(tk.END, f"  PARSE ERROR: Could not extract data\n")
                progress_text.see(tk.END)
        
        conn.commit()
        conn.close()
        
        # Calculate total duplicates (both exact duplicates and name collisions)
        total_duplicates = len(exact_duplicates_db) + len(exact_duplicates_scan) + len(name_collisions_db) + len(name_collisions_scan)

        # Show results
        progress_label.config(text="Complete!")
        progress_text.insert(tk.END, f"\n{'='*50}\n")
        progress_text.insert(tk.END, f"SCAN SUMMARY\n")
        progress_text.insert(tk.END, f"{'='*50}\n")
        progress_text.insert(tk.END, f"Total files scanned: {len(all_scanned_files)}\n")
        progress_text.insert(tk.END, f"Files to process: {len(files_to_process)}\n")
        progress_text.insert(tk.END, f"Duplicates skipped: {total_duplicates}\n")
        progress_text.insert(tk.END, f"  - Exact match (DB): {len(exact_duplicates_db)}\n")
        progress_text.insert(tk.END, f"  - Exact match (scan): {len(exact_duplicates_scan)}\n")
        progress_text.insert(tk.END, f"  - Name collision (DB): {len(name_collisions_db)}\n")
        progress_text.insert(tk.END, f"  - Name collision (scan): {len(name_collisions_scan)}\n")
        progress_text.insert(tk.END, f"Added: {added}\n")
        progress_text.insert(tk.END, f"Updated: {updated}\n")
        progress_text.insert(tk.END, f"Errors: {errors}\n")
        if duplicates_within_processing > 0:
            progress_text.insert(tk.END, f"Program number conflicts (saved with suffix): {duplicates_within_processing}\n")
        progress_text.insert(tk.END, f"Unique programs: {len(seen_in_scan)}\n")
        progress_text.see(tk.END)
        
        close_btn = tk.Button(progress_window, text="Close", 
                             command=progress_window.destroy,
                             bg=self.button_bg, fg=self.fg_color,
                             font=("Arial", 10, "bold"))
        close_btn.pack(pady=10)

        # Refresh filter dropdowns with new values
        self.refresh_filter_values()
        self.refresh_results()

    def scan_for_new_files(self):
        """Scan a folder for gcode files and import only NEW files not already in database"""
        folder = filedialog.askdirectory(title="Select Folder to Scan for New Files",
                                        initialdir=self.config.get("last_folder", ""))

        if not folder:
            return

        self.config["last_folder"] = folder
        self.save_config()

        # Show progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Scanning for New Files...")
        progress_window.geometry("600x400")
        progress_window.configure(bg=self.bg_color)

        progress_label = tk.Label(progress_window, text="Scanning...",
                                 bg=self.bg_color, fg=self.fg_color,
                                 font=("Arial", 12))
        progress_label.pack(pady=20)

        progress_text = scrolledtext.ScrolledText(progress_window,
                                                 bg=self.input_bg, fg=self.fg_color,
                                                 width=70, height=15)
        progress_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        self.root.update()

        # Get all existing files from database (filename and content hash)
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Scan for gcode files (with or without extension)
        gcode_files = []

        # Scan for new files is always in external mode (not adding to repository)
        # Check for duplicates to avoid re-importing existing files
        if True:  # External mode
            cursor.execute("SELECT file_path FROM programs WHERE file_path IS NOT NULL")
            existing_files_data = {}  # {filename_lower: [file_paths]}

            for row in cursor.fetchall():
                file_path = row[0]
                if file_path and os.path.exists(file_path):
                    basename = os.path.basename(file_path).lower()
                    if basename not in existing_files_data:
                        existing_files_data[basename] = []
                    existing_files_data[basename].append(file_path)

            name_collisions = []  # Files with same name but need content check

            for root, dirs, files in os.walk(folder):
                for file in files:
                    # Match any file containing o##### pattern (4+ digits)
                    if re.search(r'[oO]\d{4,}', file):
                        filepath = os.path.join(root, file)
                        basename_lower = file.lower()

                        # Check if filename exists in database
                        if basename_lower in existing_files_data:
                            # Need to check content to see if it's exact duplicate
                            name_collisions.append((filepath, existing_files_data[basename_lower]))
                        else:
                            # New filename, safe to add
                            gcode_files.append(filepath)

            # Check name collisions for exact duplicates vs different content
            # Use file size + hash comparison for better performance
            exact_duplicates = []
            different_content = []

            import hashlib

            def calculate_file_hash(filepath):
                """Calculate SHA256 hash of a file efficiently"""
                sha256_hash = hashlib.sha256()
                try:
                    with open(filepath, 'rb') as f:
                        # Read in 4KB chunks to handle large files
                        for byte_block in iter(lambda: f.read(4096), b""):
                            sha256_hash.update(byte_block)
                    return sha256_hash.hexdigest()
                except:
                    return None

            for new_file, existing_paths in name_collisions:
                try:
                    # Step 1: Quick file size check
                    new_size = os.path.getsize(new_file)

                    # Compare with existing file(s)
                    is_exact_match = False
                    for existing_path in existing_paths:
                        try:
                            existing_size = os.path.getsize(existing_path)

                            # If sizes are different, definitely different files
                            if new_size != existing_size:
                                continue

                            # Step 2: Sizes match - calculate hashes for accurate comparison
                            new_hash = calculate_file_hash(new_file)
                            existing_hash = calculate_file_hash(existing_path)

                            if new_hash and existing_hash and new_hash == existing_hash:
                                is_exact_match = True
                                break
                        except:
                            continue

                    if is_exact_match:
                        exact_duplicates.append(new_file)
                    else:
                        different_content.append((new_file, existing_paths[0]))
                except:
                    # If can't read, treat as different content (to be safe)
                    different_content.append((new_file, existing_paths[0]))

            progress_label.config(text=f"Analyzing files...")
            progress_text.insert(tk.END, f"Total files scanned: {len(gcode_files) + len(name_collisions)}\n")
            progress_text.insert(tk.END, f"New files (unique names): {len(gcode_files)}\n")
            progress_text.insert(tk.END, f"Exact duplicates (ignored): {len(exact_duplicates)}\n")
            progress_text.insert(tk.END, f"Name collisions (different content): {len(different_content)}\n\n")

            # Show exact duplicates being skipped
            if exact_duplicates:
                progress_text.insert(tk.END, f"=== SKIPPING EXACT DUPLICATES ===\n")
                for dup_file in exact_duplicates[:10]:
                    progress_text.insert(tk.END, f"  SKIP: {os.path.basename(dup_file)} (exact match already in database)\n")
                if len(exact_duplicates) > 10:
                    progress_text.insert(tk.END, f"  ... and {len(exact_duplicates) - 10} more\n")
                progress_text.insert(tk.END, f"\n")

            # Warn about name collisions
            if different_content:
                progress_text.insert(tk.END, f"⚠️  WARNING - NAME COLLISIONS DETECTED ===\n")
                progress_text.insert(tk.END, f"The following files have the same name as files in the database\n")
                progress_text.insert(tk.END, f"but DIFFERENT CONTENT. You must rename them before adding!\n\n")
                for new_file, existing_file in different_content[:10]:
                    progress_text.insert(tk.END, f"  ⚠️  {os.path.basename(new_file)}\n")
                    progress_text.insert(tk.END, f"     New: {new_file}\n")
                    progress_text.insert(tk.END, f"     Existing: {existing_file}\n\n")
                if len(different_content) > 10:
                    progress_text.insert(tk.END, f"  ... and {len(different_content) - 10} more\n")
                progress_text.insert(tk.END, f"\nPlease rename these files and scan again.\n\n")
        else:
            # Repository mode - skip duplicate checking, just add all files
            progress_text.insert(tk.END, "Repository mode: Skipping duplicate checks for faster import...\n\n")
            for root, dirs, files in os.walk(folder):
                for file in files:
                    # Match any file containing o##### pattern (4+ digits)
                    if re.search(r'[oO]\d{4,}', file):
                        filepath = os.path.join(root, file)
                        gcode_files.append(filepath)

        progress_text.insert(tk.END, f"Processing {len(gcode_files)} new files...\n\n")
        progress_text.see(tk.END)
        self.root.update()

        if len(gcode_files) == 0:
            progress_label.config(text="No new files found!")
            close_btn = tk.Button(progress_window, text="Close",
                                 command=progress_window.destroy,
                                 bg=self.button_bg, fg=self.fg_color,
                                 font=("Arial", 10, "bold"))
            close_btn.pack(pady=10)
            conn.close()  # Close the database connection
            return

        # Process files
        added = 0
        errors = 0
        duplicates = 0
        duplicates_within_processing = 0  # Track duplicates found during processing

        # Track which program numbers we've seen in this scan
        seen_in_scan = {}  # program_number -> filepath

        for idx, filepath in enumerate(gcode_files, 1):
            filename = os.path.basename(filepath)

            # Check file size
            try:
                file_size = os.path.getsize(filepath)
                size_kb = file_size / 1024
            except:
                size_kb = 0

            # Update UI only every 10 files or on first/last file for better performance
            update_ui = (idx % 10 == 0) or (idx == 1) or (idx == len(gcode_files))

            if update_ui:
                progress_label.config(text=f"Processing {idx}/{len(gcode_files)}: {filename}")

            progress_text.insert(tk.END, f"[{idx}/{len(gcode_files)}] Processing: {filename} ({size_kb:.1f} KB)\n")

            try:
                record = self.parse_gcode_file(filepath)
            except Exception as e:
                errors += 1
                progress_text.insert(tk.END, f"  PARSE EXCEPTION: {str(e)[:100]}\n")
                if update_ui:
                    progress_text.see(tk.END)
                    self.root.update()
                continue

            if record:
                try:
                    # Check for duplicate in this scan and assign unique suffix
                    original_prog_num = record.program_number
                    if record.program_number in seen_in_scan:
                        # Find next available suffix
                        suffix = 1
                        while f"{original_prog_num}({suffix})" in seen_in_scan:
                            suffix += 1
                        record.program_number = f"{original_prog_num}({suffix})"
                        progress_text.insert(tk.END, f"  DUPLICATE: {original_prog_num} -> saved as {record.program_number}\n")
                        duplicates_within_processing += 1

                    # Track this file with its (possibly modified) program number
                    seen_in_scan[record.program_number] = filepath

                    # Insert new record
                    cursor.execute('''
                        INSERT INTO programs (
                            program_number, title, spacer_type, outer_diameter, thickness, thickness_display,
                            center_bore, hub_height, hub_diameter,
                            counter_bore_diameter, counter_bore_depth,
                            paired_program, material, notes, date_created, last_modified, file_path,
                            detection_confidence, detection_method,
                            validation_status, validation_issues,
                            validation_warnings, bore_warnings, dimensional_issues,
                            cb_from_gcode, ob_from_gcode, lathe,
                            duplicate_type, parent_file, duplicate_group
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        record.program_number,
                        record.title,
                        record.spacer_type,
                        record.outer_diameter,
                        record.thickness,
                        record.thickness_display,
                        record.center_bore,
                        record.hub_height,
                        record.hub_diameter,
                        record.counter_bore_diameter,
                        record.counter_bore_depth,
                        record.paired_program,
                        record.material,
                        record.notes,
                        datetime.now().isoformat(),
                        datetime.now().isoformat(),
                        filepath,
                        record.detection_confidence,
                        record.detection_method,
                        record.validation_status,
                        record.validation_issues,
                        record.validation_warnings,
                        record.bore_warnings,
                        record.dimensional_issues,
                        record.cb_from_gcode,
                        record.ob_from_gcode,
                        record.lathe,
                        None,  # duplicate_type
                        None,  # parent_file
                        None   # duplicate_group
                    ))
                    added += 1
                    progress_text.insert(tk.END, f"  ADDED: {record.program_number}\n")

                    # Batch commit every 100 files for much better performance
                    if added % 100 == 0:
                        conn.commit()
                        progress_text.insert(tk.END, f"\n[Database] Committed {added} files...\n\n")

                    # Update UI every 10 files for better performance
                    if update_ui:
                        progress_text.see(tk.END)
                        self.root.update()
                except sqlite3.Error as e:
                    errors += 1
                    progress_text.insert(tk.END, f"  DATABASE ERROR: {str(e)[:100]}\n")
                    if update_ui:
                        progress_text.see(tk.END)
                        self.root.update()
            else:
                errors += 1
                progress_text.insert(tk.END, f"  PARSE ERROR: Could not extract data\n")
                if update_ui:
                    progress_text.see(tk.END)
                    self.root.update()

        # Final commit for remaining files
        conn.commit()
        conn.close()

        # Show results
        progress_label.config(text="Complete!")
        progress_text.insert(tk.END, f"\n{'='*50}\n")
        progress_text.insert(tk.END, f"New files found: {len(gcode_files)}\n")
        progress_text.insert(tk.END, f"Duplicates: {duplicates}\n")
        if duplicates_within_processing > 0:
            progress_text.insert(tk.END, f"Program number conflicts (saved with suffix): {duplicates_within_processing}\n")
        progress_text.insert(tk.END, f"Added: {added}\n")
        progress_text.insert(tk.END, f"Errors: {errors}\n")
        progress_text.see(tk.END)

        close_btn = tk.Button(progress_window, text="Close",
                             command=progress_window.destroy,
                             bg=self.button_bg, fg=self.fg_color,
                             font=("Arial", 10, "bold"))
        close_btn.pack(pady=10)

        # Refresh filter dropdowns with new values
        self.refresh_filter_values()
        self.refresh_results()

    def _ml_predict_dimension(self, ml_extractor, dimension, parse_result, program_number):
        """Helper method to predict a single dimension using ML"""
        try:
            import numpy as np

            # Extract features from title
            title = parse_result.title if parse_result.title else ""
            features = ml_extractor.extract_title_features(title)

            # Add G-code features
            features['cb_from_gcode'] = parse_result.cb_from_gcode if parse_result.cb_from_gcode else 0
            features['ob_from_gcode'] = parse_result.ob_from_gcode if parse_result.ob_from_gcode else 0

            # Add known dimensions for cross-prediction
            features['known_od'] = parse_result.outer_diameter if (dimension != 'outer_diameter' and parse_result.outer_diameter) else 0
            features['known_cb'] = parse_result.center_bore if (dimension != 'center_bore' and parse_result.center_bore) else 0
            features['known_thickness'] = parse_result.thickness if (dimension != 'thickness' and parse_result.thickness) else 0

            # Create feature vector matching training
            feature_vector = [features.get(fn, 0) for fn in ml_extractor.feature_names]
            feature_vector = np.array(feature_vector).reshape(1, -1)

            # Predict
            model = ml_extractor.models.get(dimension)
            if model:
                prediction = model.predict(feature_vector)[0]
                return prediction
            return None
        except Exception as e:
            return None

    def rescan_database(self):
        """Re-parse all files already in database to refresh with latest parser improvements"""
        # Create custom dialog with ML fallback option
        dialog = tk.Toplevel(self.root)
        dialog.title("Rescan Database")
        dialog.geometry("500x300")
        dialog.configure(bg=self.bg_color)
        dialog.transient(self.root)
        dialog.grab_set()

        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (500 // 2)
        y = (dialog.winfo_screenheight() // 2) - (300 // 2)
        dialog.geometry(f"500x300+{x}+{y}")

        # Message
        msg_frame = tk.Frame(dialog, bg=self.bg_color)
        msg_frame.pack(pady=20, padx=20, fill=tk.BOTH, expand=True)

        msg_label = tk.Label(msg_frame,
                            text="This will re-parse ALL files in the database to refresh their data.\n\n"
                                 "This is useful after parser improvements.\n\n"
                                 "Files will be re-analyzed but NOT moved or renamed.",
                            bg=self.bg_color, fg=self.fg_color,
                            font=("Arial", 10),
                            justify=tk.LEFT,
                            wraplength=450)
        msg_label.pack(anchor='w')

        # ML Fallback checkbox
        use_ml_var = tk.BooleanVar(value=False)  # Default to OFF for speed

        ml_frame = tk.Frame(dialog, bg=self.bg_color)
        ml_frame.pack(pady=10, padx=20, fill=tk.X)

        ml_check = tk.Checkbutton(ml_frame,
                                 text="Use ML Fallback for missing dimensions",
                                 variable=use_ml_var,
                                 bg=self.bg_color, fg=self.fg_color,
                                 selectcolor=self.input_bg,
                                 activebackground=self.bg_color,
                                 activeforeground=self.fg_color,
                                 font=("Arial", 10, "bold"),
                                 cursor='hand2')
        ml_check.pack(anchor='w')

        ml_note = tk.Label(ml_frame,
                          text="(Slower but fills all missing dimensions automatically)",
                          bg=self.bg_color, fg=self.fg_color,
                          font=("Arial", 8, "italic"),
                          justify=tk.LEFT)
        ml_note.pack(anchor='w', padx=25)

        # Buttons
        button_frame = tk.Frame(dialog, bg=self.bg_color)
        button_frame.pack(pady=20)

        result = {'proceed': False, 'use_ml': False}

        def on_proceed():
            result['proceed'] = True
            result['use_ml'] = use_ml_var.get()
            dialog.destroy()

        def on_cancel():
            result['proceed'] = False
            dialog.destroy()

        proceed_btn = tk.Button(button_frame, text="Proceed",
                               command=on_proceed,
                               bg=self.button_bg, fg=self.fg_color,
                               font=("Arial", 10, "bold"),
                               width=15, height=2)
        proceed_btn.pack(side=tk.LEFT, padx=10)

        cancel_btn = tk.Button(button_frame, text="Cancel",
                              command=on_cancel,
                              bg=self.bg_color, fg=self.fg_color,
                              font=("Arial", 10),
                              width=15, height=2)
        cancel_btn.pack(side=tk.LEFT, padx=10)

        # Wait for dialog to close
        self.root.wait_window(dialog)

        if not result['proceed']:
            return

        use_ml_fallback = result['use_ml']

        # Show progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Rescanning Database...")
        progress_window.geometry("700x500")
        progress_window.configure(bg=self.bg_color)

        progress_label = tk.Label(progress_window, text="Rescanning files...",
                                 bg=self.bg_color, fg=self.fg_color,
                                 font=("Arial", 12))
        progress_label.pack(pady=20)

        progress_text = scrolledtext.ScrolledText(progress_window,
                                                 bg=self.input_bg, fg=self.fg_color,
                                                 font=("Courier", 9),
                                                 width=80, height=20)
        progress_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        self.root.update()

        # Get all files from database
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT program_number, file_path FROM programs WHERE file_path IS NOT NULL")
        all_files = cursor.fetchall()

        total_files = len(all_files)
        progress_text.insert(tk.END, f"Found {total_files} files in database\n")
        progress_text.insert(tk.END, f"Starting re-scan...\n\n")
        progress_text.see(tk.END)
        self.root.update()

        # Import parser
        from improved_gcode_parser import ImprovedGCodeParser
        parser = ImprovedGCodeParser()

        # Try to initialize ML fallback (only if user enabled it)
        ml_extractor = None
        ml_available = False
        if use_ml_fallback:
            try:
                from ml_dimension_extractor import MLDimensionExtractor
                ml_extractor = MLDimensionExtractor(self.db_path)
                if ml_extractor.load_models():
                    ml_available = True
                    progress_text.insert(tk.END, "[ML Fallback] Loaded ML models for missing dimensions\n\n")
                else:
                    progress_text.insert(tk.END, "[ML Fallback] Training ML models...\n")
                    ml_extractor.load_data()
                    ml_extractor.train_all_models()
                    ml_extractor.save_models()
                    ml_available = True
                    progress_text.insert(tk.END, "[ML Fallback] ML models trained successfully\n\n")
                progress_text.see(tk.END)
                self.root.update()
            except ImportError:
                progress_text.insert(tk.END, "[ML Fallback] ML libraries not installed - skipping ML predictions\n")
                progress_text.insert(tk.END, "              Install with: pip install pandas scikit-learn numpy\n\n")
                progress_text.see(tk.END)
                self.root.update()
            except Exception as e:
                progress_text.insert(tk.END, f"[ML Fallback] Error initializing ML: {str(e)[:80]}\n\n")
                progress_text.see(tk.END)
                self.root.update()
        else:
            progress_text.insert(tk.END, "ML Fallback disabled (for faster scanning)\n\n")
            progress_text.see(tk.END)
            self.root.update()

        updated = 0
        skipped = 0
        errors = 0
        ml_predictions = 0

        for idx, (prog_num, file_path) in enumerate(all_files, 1):
            filename = os.path.basename(file_path)

            # Update progress
            if idx % 50 == 0 or idx == total_files:
                progress_label.config(text=f"Rescanning... {idx}/{total_files} files")
                self.root.update()

            # Check if file exists
            if not os.path.exists(file_path):
                skipped += 1
                if idx <= 10 or idx % 100 == 0:  # Show first 10 and every 100th
                    progress_text.insert(tk.END, f"[{idx}/{total_files}] {filename} - SKIP (file not found)\n")
                    progress_text.see(tk.END)
                    self.root.update()
                continue

            # Parse the file
            try:
                parse_result = parser.parse_file(file_path)

                if not parse_result:
                    errors += 1
                    if idx <= 10 or idx % 100 == 0:
                        progress_text.insert(tk.END, f"[{idx}/{total_files}] {filename} - ERROR (parse failed)\n")
                        progress_text.see(tk.END)
                        self.root.update()
                    continue

                # Apply ML fallback for missing dimensions
                if ml_available and ml_extractor:
                    # Check and predict Outer Diameter
                    if not parse_result.outer_diameter:
                        ml_od = self._ml_predict_dimension(ml_extractor, 'outer_diameter', parse_result, prog_num)
                        if ml_od:
                            parse_result.outer_diameter = ml_od
                            if parse_result.detection_confidence != 'ML_FALLBACK':
                                parse_result.detection_confidence = 'ML_FALLBACK'
                            parse_result.detection_notes.append(f'OD from ML: {ml_od:.2f}"')
                            ml_predictions += 1

                    # Check and predict Thickness
                    if not parse_result.thickness:
                        ml_thickness = self._ml_predict_dimension(ml_extractor, 'thickness', parse_result, prog_num)
                        if ml_thickness:
                            parse_result.thickness = ml_thickness
                            if parse_result.detection_confidence != 'ML_FALLBACK':
                                parse_result.detection_confidence = 'ML_FALLBACK'
                            parse_result.detection_notes.append(f'Thickness from ML: {ml_thickness:.3f}"')
                            ml_predictions += 1

                    # Check and predict Center Bore
                    if not parse_result.center_bore:
                        ml_cb = self._ml_predict_dimension(ml_extractor, 'center_bore', parse_result, prog_num)
                        if ml_cb:
                            parse_result.center_bore = ml_cb
                            if parse_result.detection_confidence != 'ML_FALLBACK':
                                parse_result.detection_confidence = 'ML_FALLBACK'
                            parse_result.detection_notes.append(f'CB from ML: {ml_cb:.1f}mm')
                            ml_predictions += 1

                # Calculate validation status (prioritized by severity)
                # NOTE: Safety and tool validation disabled for now - too many false positives
                validation_status = "PASS"
                if parse_result.validation_issues:
                    validation_status = "CRITICAL"  # RED - Critical errors
                # elif parse_result.safety_blocks_status == "MISSING":
                #     validation_status = "SAFETY_ERROR"  # DARK RED - Missing safety blocks (DISABLED - too strict)
                # elif parse_result.tool_validation_status == "ERROR":
                #     validation_status = "TOOL_ERROR"  # ORANGE-RED - Wrong/missing tools (DISABLED - needs tuning)
                elif parse_result.bore_warnings:
                    validation_status = "BORE_WARNING"  # ORANGE - Bore dimension warnings
                elif parse_result.dimensional_issues:
                    validation_status = "DIMENSIONAL"  # PURPLE - P-code/thickness mismatches
                # elif parse_result.tool_validation_status == "WARNING":
                #     validation_status = "TOOL_WARNING"  # AMBER - Tool suggestions (DISABLED - needs tuning)
                elif parse_result.validation_warnings:
                    validation_status = "WARNING"  # YELLOW - General warnings

                # Update database with refreshed data
                # Note: paired_program not updated (not in parser result)
                notes = '|'.join(parse_result.detection_notes) if parse_result.detection_notes else None

                cursor.execute("""
                    UPDATE programs
                    SET title = ?,
                        spacer_type = ?,
                        outer_diameter = ?,
                        thickness = ?,
                        thickness_display = ?,
                        center_bore = ?,
                        hub_height = ?,
                        hub_diameter = ?,
                        counter_bore_diameter = ?,
                        counter_bore_depth = ?,
                        material = ?,
                        detection_confidence = ?,
                        detection_method = ?,
                        validation_status = ?,
                        validation_issues = ?,
                        validation_warnings = ?,
                        cb_from_gcode = ?,
                        ob_from_gcode = ?,
                        bore_warnings = ?,
                        dimensional_issues = ?,
                        lathe = ?,
                        notes = ?,
                        tools_used = ?,
                        tool_sequence = ?,
                        tool_validation_status = ?,
                        tool_validation_issues = ?,
                        safety_blocks_status = ?,
                        safety_blocks_issues = ?
                    WHERE program_number = ?
                """, (
                    parse_result.title,
                    parse_result.spacer_type,
                    parse_result.outer_diameter,
                    parse_result.thickness,
                    parse_result.thickness_display,
                    parse_result.center_bore,
                    parse_result.hub_height,
                    parse_result.hub_diameter,
                    parse_result.counter_bore_diameter,
                    parse_result.counter_bore_depth,
                    parse_result.material,
                    parse_result.detection_confidence,
                    parse_result.detection_method,
                    validation_status,
                    '|'.join(parse_result.validation_issues) if parse_result.validation_issues else None,
                    '|'.join(parse_result.validation_warnings) if parse_result.validation_warnings else None,
                    parse_result.cb_from_gcode,
                    parse_result.ob_from_gcode,
                    '|'.join(parse_result.bore_warnings) if parse_result.bore_warnings else None,
                    '|'.join(parse_result.dimensional_issues) if parse_result.dimensional_issues else None,
                    parse_result.lathe,
                    notes,
                    json.dumps(parse_result.tools_used) if parse_result.tools_used else None,
                    json.dumps(parse_result.tool_sequence) if parse_result.tool_sequence else None,
                    None,  # tool_validation_status - DISABLED
                    None,  # tool_validation_issues - DISABLED
                    None,  # safety_blocks_status - DISABLED
                    None,  # safety_blocks_issues - DISABLED
                    prog_num
                ))

                updated += 1

                # Show first 10, then every 100th, and last 10
                if idx <= 10 or idx % 100 == 0 or idx > total_files - 10:
                    progress_text.insert(tk.END, f"[{idx}/{total_files}] {filename} - ✓ Updated\n")
                    progress_text.see(tk.END)
                    self.root.update()

            except Exception as e:
                errors += 1
                if idx <= 10 or idx % 100 == 0:
                    progress_text.insert(tk.END, f"[{idx}/{total_files}] {filename} - ERROR: {str(e)}\n")
                    progress_text.see(tk.END)
                    self.root.update()

        # Commit all changes
        conn.commit()
        conn.close()

        # Summary
        progress_label.config(text="Rescan Complete!")
        progress_text.insert(tk.END, f"\n{'='*60}\n")
        progress_text.insert(tk.END, f"RESCAN COMPLETE\n")
        progress_text.insert(tk.END, f"{'='*60}\n")
        progress_text.insert(tk.END, f"Total files: {total_files}\n")
        progress_text.insert(tk.END, f"Updated: {updated}\n")
        progress_text.insert(tk.END, f"Skipped (not found): {skipped}\n")
        if ml_available and ml_predictions > 0:
            progress_text.insert(tk.END, f"\n[ML Fallback] {ml_predictions} dimensions predicted by ML\n")
            progress_text.insert(tk.END, f"              (Programs marked as 'ML_FALLBACK')\n")
        progress_text.insert(tk.END, f"Errors: {errors}\n")
        progress_text.see(tk.END)

        # Close button
        close_btn = tk.Button(progress_window, text="Close",
                             command=progress_window.destroy,
                             bg=self.button_bg, fg=self.fg_color,
                             font=("Arial", 10, "bold"))
        close_btn.pack(pady=10)

        # Refresh the display
        self.refresh_results()

    def rescan_changed_files(self):
        """Re-scan only files that have been modified since last database update"""
        # Confirmation dialog
        if not messagebox.askyesno("Rescan Changed Files",
                                   "This will re-parse only files that have been modified\n"
                                   "since their last database update.\n\n"
                                   "This is much faster than full rescan.\n\n"
                                   "Continue?"):
            return

        # Create backup first
        if not self.backup_database():
            messagebox.showerror("Backup Failed", "Could not create backup before rescan.\nOperation cancelled.")
            return

        # Show progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Rescanning Changed Files...")
        progress_window.geometry("700x500")
        progress_window.configure(bg=self.bg_color)

        progress_label = tk.Label(progress_window, text="Checking for modified files...",
                                 bg=self.bg_color, fg=self.fg_color,
                                 font=("Arial", 12))
        progress_label.pack(pady=20)

        progress_text = scrolledtext.ScrolledText(progress_window,
                                                 bg=self.input_bg, fg=self.fg_color,
                                                 font=("Courier", 9),
                                                 width=80, height=20)
        progress_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        self.root.update()

        # Get all files from database with last_modified timestamps
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT program_number, file_path, last_modified FROM programs WHERE file_path IS NOT NULL")
        all_files = cursor.fetchall()

        total_files = len(all_files)
        progress_text.insert(tk.END, f"Checking {total_files} files for modifications...\n\n")
        progress_text.see(tk.END)
        self.root.update()

        # Import parser
        from improved_gcode_parser import ImprovedGCodeParser
        parser = ImprovedGCodeParser()

        updated = 0
        skipped = 0
        errors = 0
        not_found = 0
        modified_files = []

        # First pass: identify modified files
        for idx, (prog_num, file_path, db_modified) in enumerate(all_files, 1):
            if idx % 100 == 0:
                progress_label.config(text=f"Checking... {idx}/{total_files} files")
                self.root.update()

            # Check if file exists
            if not os.path.exists(file_path):
                not_found += 1
                continue

            # Get file modification time
            try:
                file_mtime = datetime.fromtimestamp(os.path.getmtime(file_path)).isoformat()

                # Compare timestamps - only re-parse if file is newer than DB record
                if db_modified is None or file_mtime > db_modified:
                    modified_files.append((prog_num, file_path))
            except Exception as e:
                errors += 1
                continue

        # Report findings
        progress_text.insert(tk.END, f"Found {len(modified_files)} modified files (out of {total_files} total)\n")
        if not_found > 0:
            progress_text.insert(tk.END, f"  {not_found} files not found (skipped)\n")
        progress_text.insert(tk.END, f"\n")

        if len(modified_files) == 0:
            progress_text.insert(tk.END, "✓ All files are up to date!\n")
            progress_label.config(text="No changes detected")
        else:
            progress_text.insert(tk.END, f"Re-parsing {len(modified_files)} modified files...\n\n")
            progress_text.see(tk.END)
            self.root.update()

            # Second pass: re-parse modified files
            for idx, (prog_num, file_path) in enumerate(modified_files, 1):
                filename = os.path.basename(file_path)

                # Update progress
                if idx % 10 == 0 or idx == len(modified_files):
                    progress_label.config(text=f"Updating... {idx}/{len(modified_files)} modified files")
                    self.root.update()

                # Parse the file
                try:
                    parse_result = parser.parse_file(file_path)

                    if not parse_result:
                        errors += 1
                        if idx <= 5 or idx % 50 == 0:
                            progress_text.insert(tk.END, f"[{idx}/{len(modified_files)}] {filename} - ERROR (parse failed)\n")
                            progress_text.see(tk.END)
                            self.root.update()
                        continue

                    # Calculate validation status
                    # NOTE: Safety and tool validation disabled for now - too many false positives
                    validation_status = "PASS"
                    if parse_result.validation_issues:
                        validation_status = "CRITICAL"
                    # elif parse_result.safety_blocks_status == "MISSING":
                    #     validation_status = "SAFETY_ERROR"  # DISABLED - too strict
                    # elif parse_result.tool_validation_status == "ERROR":
                    #     validation_status = "TOOL_ERROR"  # DISABLED - needs tuning
                    elif parse_result.bore_warnings:
                        validation_status = "BORE_WARNING"
                    elif parse_result.dimensional_issues:
                        validation_status = "DIMENSIONAL"
                    # elif parse_result.tool_validation_status == "WARNING":
                    #     validation_status = "TOOL_WARNING"  # DISABLED - needs tuning
                    elif parse_result.validation_warnings:
                        validation_status = "WARNING"

                    # Update database with refreshed data
                    notes = '|'.join(parse_result.detection_notes) if parse_result.detection_notes else None

                    cursor.execute("""
                        UPDATE programs
                        SET title = ?,
                            spacer_type = ?,
                            outer_diameter = ?,
                            thickness = ?,
                            thickness_display = ?,
                            center_bore = ?,
                            hub_height = ?,
                            hub_diameter = ?,
                            counter_bore_diameter = ?,
                            counter_bore_depth = ?,
                            material = ?,
                            detection_confidence = ?,
                            detection_method = ?,
                            validation_status = ?,
                            validation_issues = ?,
                            validation_warnings = ?,
                            cb_from_gcode = ?,
                            ob_from_gcode = ?,
                            bore_warnings = ?,
                            dimensional_issues = ?,
                            lathe = ?,
                            notes = ?,
                            tools_used = ?,
                            tool_sequence = ?,
                            tool_validation_status = ?,
                            tool_validation_issues = ?,
                            safety_blocks_status = ?,
                            safety_blocks_issues = ?,
                            last_modified = ?
                        WHERE program_number = ?
                    """, (
                        parse_result.title,
                        parse_result.spacer_type,
                        parse_result.outer_diameter,
                        parse_result.thickness,
                        parse_result.thickness_display,
                        parse_result.center_bore,
                        parse_result.hub_height,
                        parse_result.hub_diameter,
                        parse_result.counter_bore_diameter,
                        parse_result.counter_bore_depth,
                        parse_result.material,
                        parse_result.detection_confidence,
                        parse_result.detection_method,
                        validation_status,
                        '|'.join(parse_result.validation_issues) if parse_result.validation_issues else None,
                        '|'.join(parse_result.validation_warnings) if parse_result.validation_warnings else None,
                        parse_result.cb_from_gcode,
                        parse_result.ob_from_gcode,
                        '|'.join(parse_result.bore_warnings) if parse_result.bore_warnings else None,
                        '|'.join(parse_result.dimensional_issues) if parse_result.dimensional_issues else None,
                        parse_result.lathe,
                        notes,
                        json.dumps(parse_result.tools_used) if parse_result.tools_used else None,
                        json.dumps(parse_result.tool_sequence) if parse_result.tool_sequence else None,
                        None,  # tool_validation_status - DISABLED
                        None,  # tool_validation_issues - DISABLED
                        None,  # safety_blocks_status - DISABLED
                        None,  # safety_blocks_issues - DISABLED
                        datetime.now().isoformat(),
                        prog_num
                    ))

                    updated += 1

                    # Show progress for first few and periodically
                    if idx <= 5 or idx % 50 == 0 or idx == len(modified_files):
                        progress_text.insert(tk.END, f"[{idx}/{len(modified_files)}] {filename} - ✓ Updated\n")
                        progress_text.see(tk.END)
                        self.root.update()

                except Exception as e:
                    errors += 1
                    if idx <= 5 or idx % 50 == 0:
                        progress_text.insert(tk.END, f"[{idx}/{len(modified_files)}] {filename} - ERROR: {str(e)[:60]}\n")
                        progress_text.see(tk.END)
                        self.root.update()

        # Commit all changes
        conn.commit()
        conn.close()

        # Summary
        progress_label.config(text="Scan Complete!")
        progress_text.insert(tk.END, f"\n{'='*60}\n")
        progress_text.insert(tk.END, f"RESCAN CHANGED FILES COMPLETE\n")
        progress_text.insert(tk.END, f"{'='*60}\n")
        progress_text.insert(tk.END, f"Total files in database: {total_files}\n")
        progress_text.insert(tk.END, f"Modified files found: {len(modified_files)}\n")
        progress_text.insert(tk.END, f"Successfully updated: {updated}\n")
        progress_text.insert(tk.END, f"Skipped (unchanged): {skipped}\n")
        if not_found > 0:
            progress_text.insert(tk.END, f"Not found: {not_found}\n")
        if errors > 0:
            progress_text.insert(tk.END, f"Errors: {errors}\n")
        progress_text.insert(tk.END, f"\n✓ Much faster than full rescan!\n")
        progress_text.see(tk.END)

        # Close button
        close_btn = tk.Button(progress_window, text="Close",
                             command=progress_window.destroy,
                             bg=self.button_bg, fg=self.fg_color,
                             font=("Arial", 10, "bold"))
        close_btn.pack(pady=10)

        # Refresh the display
        self.refresh_results()

    def view_tool_statistics(self):
        """Display tool usage statistics across all programs"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Get all programs with tool data
        cursor.execute("""
            SELECT program_number, spacer_type, outer_diameter, tools_used, tool_sequence,
                   tool_validation_status, tool_validation_issues, safety_blocks_status, safety_blocks_issues
            FROM programs
            WHERE tools_used IS NOT NULL AND tools_used != 'null'
        """)
        results = cursor.fetchall()

        if not results:
            messagebox.showinfo("No Tool Data",
                              "No tool usage data found.\n\n"
                              "Run a rescan to extract tool information from your G-code files.")
            conn.close()
            return

        # Create statistics window
        stats_window = tk.Toplevel(self.root)
        stats_window.title("Tool Usage Statistics")
        stats_window.geometry("1000x700")
        stats_window.configure(bg=self.bg_color)

        # Title
        title_label = tk.Label(stats_window, text="🔧 Tool Usage Statistics & Safety Analysis",
                              bg=self.bg_color, fg=self.fg_color,
                              font=("Arial", 14, "bold"))
        title_label.pack(pady=10)

        # Create notebook for tabs
        notebook = ttk.Notebook(stats_window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Tab 1: Tool Usage Summary
        summary_frame = tk.Frame(notebook, bg=self.bg_color)
        notebook.add(summary_frame, text="Tool Summary")

        summary_text = scrolledtext.ScrolledText(summary_frame,
                                                 bg=self.input_bg, fg=self.fg_color,
                                                 font=("Courier", 10),
                                                 wrap=tk.WORD)
        summary_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Calculate statistics
        tool_counts = {}
        tool_by_part_type = {}
        tool_sequences_common = {}
        tool_issues_count = 0
        safety_issues_count = 0

        for row in results:
            prog_num, spacer_type, od, tools_json, sequence_json, tool_status, tool_issues_json, safety_status, safety_issues_json = row

            # Parse JSON
            try:
                tools = json.loads(tools_json) if tools_json else []
                sequence = json.loads(sequence_json) if sequence_json else []
                tool_issues = json.loads(tool_issues_json) if tool_issues_json else []
                safety_issues = json.loads(safety_issues_json) if safety_issues_json else []

                # Count tools
                for tool in tools:
                    tool_counts[tool] = tool_counts.get(tool, 0) + 1

                    # By part type
                    if spacer_type not in tool_by_part_type:
                        tool_by_part_type[spacer_type] = {}
                    tool_by_part_type[spacer_type][tool] = tool_by_part_type[spacer_type].get(tool, 0) + 1

                # Count sequences
                seq_str = " → ".join(sequence)
                if seq_str:
                    tool_sequences_common[seq_str] = tool_sequences_common.get(seq_str, 0) + 1

                # Count issues
                if tool_issues:
                    tool_issues_count += 1
                if safety_issues:
                    safety_issues_count += 1

            except:
                continue

        # Write summary
        summary_text.insert(tk.END, f"{'='*80}\n")
        summary_text.insert(tk.END, f"TOOL USAGE SUMMARY\n")
        summary_text.insert(tk.END, f"{'='*80}\n\n")

        summary_text.insert(tk.END, f"Total programs analyzed: {len(results)}\n")
        summary_text.insert(tk.END, f"Programs with tool issues: {tool_issues_count}\n")
        summary_text.insert(tk.END, f"Programs with safety issues: {safety_issues_count}\n\n")

        summary_text.insert(tk.END, f"{'='*80}\n")
        summary_text.insert(tk.END, f"MOST COMMON TOOLS (All Parts)\n")
        summary_text.insert(tk.END, f"{'='*80}\n\n")

        for tool, count in sorted(tool_counts.items(), key=lambda x: x[1], reverse=True):
            pct = (count / len(results)) * 100
            summary_text.insert(tk.END, f"{tool:6s}: {count:5d} programs ({pct:5.1f}%)\n")

        summary_text.insert(tk.END, f"\n{'='*80}\n")
        summary_text.insert(tk.END, f"TOOLS BY PART TYPE\n")
        summary_text.insert(tk.END, f"{'='*80}\n\n")

        for part_type in sorted(tool_by_part_type.keys()):
            summary_text.insert(tk.END, f"{part_type}:\n")
            for tool, count in sorted(tool_by_part_type[part_type].items(), key=lambda x: x[1], reverse=True)[:5]:
                summary_text.insert(tk.END, f"  {tool}: {count} programs\n")
            summary_text.insert(tk.END, "\n")

        summary_text.insert(tk.END, f"{'='*80}\n")
        summary_text.insert(tk.END, f"MOST COMMON TOOL SEQUENCES (Top 10)\n")
        summary_text.insert(tk.END, f"{'='*80}\n\n")

        for seq, count in sorted(tool_sequences_common.items(), key=lambda x: x[1], reverse=True)[:10]:
            summary_text.insert(tk.END, f"{count:4d}x: {seq}\n")

        summary_text.config(state=tk.DISABLED)

        # Tab 2: Tool Issues
        issues_frame = tk.Frame(notebook, bg=self.bg_color)
        notebook.add(issues_frame, text="Tool Issues")

        issues_text = scrolledtext.ScrolledText(issues_frame,
                                                bg=self.input_bg, fg=self.fg_color,
                                                font=("Courier", 9),
                                                wrap=tk.WORD)
        issues_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        issues_text.insert(tk.END, f"{'='*80}\n")
        issues_text.insert(tk.END, f"TOOL VALIDATION ISSUES\n")
        issues_text.insert(tk.END, f"{'='*80}\n\n")

        for row in results:
            prog_num, spacer_type, od, tools_json, sequence_json, tool_status, tool_issues_json, safety_status, safety_issues_json = row

            try:
                tool_issues = json.loads(tool_issues_json) if tool_issues_json else []

                if tool_issues:
                    issues_text.insert(tk.END, f"{prog_num} ({spacer_type}):\n")
                    for issue in tool_issues:
                        issues_text.insert(tk.END, f"  ⚠ {issue}\n")
                    issues_text.insert(tk.END, "\n")
            except:
                continue

        if tool_issues_count == 0:
            issues_text.insert(tk.END, "✓ No tool issues found!\n")

        issues_text.config(state=tk.DISABLED)

        # Tab 3: Safety Issues
        safety_frame = tk.Frame(notebook, bg=self.bg_color)
        notebook.add(safety_frame, text="Safety Issues")

        safety_text = scrolledtext.ScrolledText(safety_frame,
                                                bg=self.input_bg, fg=self.fg_color,
                                                font=("Courier", 9),
                                                wrap=tk.WORD)
        safety_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        safety_text.insert(tk.END, f"{'='*80}\n")
        safety_text.insert(tk.END, f"SAFETY BLOCK VALIDATION ISSUES\n")
        safety_text.insert(tk.END, f"{'='*80}\n\n")

        for row in results:
            prog_num, spacer_type, od, tools_json, sequence_json, tool_status, tool_issues_json, safety_status, safety_issues_json = row

            try:
                safety_issues = json.loads(safety_issues_json) if safety_issues_json else []

                if safety_issues:
                    safety_text.insert(tk.END, f"{prog_num} ({spacer_type}):\n")
                    for issue in safety_issues:
                        safety_text.insert(tk.END, f"  ⚠ {issue}\n")
                    safety_text.insert(tk.END, "\n")
            except:
                continue

        if safety_issues_count == 0:
            safety_text.insert(tk.END, "✓ No safety issues found!\n")

        safety_text.config(state=tk.DISABLED)

        # Close button
        close_btn = tk.Button(stats_window, text="Close",
                             command=stats_window.destroy,
                             bg=self.button_bg, fg=self.fg_color,
                             font=("Arial", 10, "bold"))
        close_btn.pack(pady=10)

        conn.close()

    def fix_program_numbers(self):
        """Update internal O-numbers to match filenames - FILTERED VIEW ONLY"""
        # Get currently displayed items (respects filters)
        displayed_items = self.tree.get_children()

        if not displayed_items:
            messagebox.showwarning("No Files",
                "No files in current view.\n\n"
                "Apply filters to show the files you want to fix.")
            return

        # Get program numbers and file info from filtered view
        filtered_files = []
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        for item in displayed_items:
            values = self.tree.item(item)['values']
            if values:
                prog_num = values[0]
                # Get file path from database
                cursor.execute("""
                    SELECT program_number, file_path
                    FROM programs
                    WHERE program_number = ?
                """, (prog_num,))
                row = cursor.fetchone()
                if row and row[1]:  # Has file_path
                    filtered_files.append(row)

        conn.close()

        if not filtered_files:
            messagebox.showwarning("No Files with Paths",
                "No files in filtered view have physical file paths.\n\n"
                "Only files with linked physical files can be fixed.")
            return

        # Confirm operation
        msg = f"Fix program numbers for {len(filtered_files)} file(s) in filtered view?\n\n"
        msg += "This will:\n"
        msg += "• Read the O-number from each filename\n"
        msg += "• Update the internal O-number in the G-code file to match\n"
        msg += "• Update database program_number to match filename\n\n"
        msg += "Use this AFTER renaming files with 'Rename Duplicates'\n"
        msg += "to synchronize internal content with filenames.\n\n"
        msg += "Database backup will be created automatically."

        result = messagebox.askyesno("Confirm Fix Program Numbers", msg)
        if not result:
            return

        # Create backup
        if not self.backup_database():
            messagebox.showerror("Backup Failed",
                "Database backup failed. Operation canceled for safety.")
            return

        # Show progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Fixing Program Numbers (Filtered View)")
        progress_window.geometry("700x600")
        progress_window.configure(bg=self.bg_color)

        progress_label = tk.Label(progress_window, text="Processing filtered files...",
                                 bg=self.bg_color, fg=self.fg_color,
                                 font=("Arial", 12))
        progress_label.pack(pady=20)

        progress_text = scrolledtext.ScrolledText(progress_window,
                                                 bg=self.input_bg, fg=self.fg_color,
                                                 font=("Courier", 9),
                                                 wrap=tk.WORD)
        progress_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.root.update()

        progress_text.insert(tk.END, f"Processing {len(filtered_files)} file(s) from filtered view...\n\n")

        # Process files from filtered view
        fixed = 0
        skipped = 0
        errors = 0
        total_files = len(filtered_files)

        # Open single database connection for all updates (prevents locking issues)
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        for file_idx, file_info in enumerate(filtered_files, 1):
            old_prog_num, filepath = file_info
            filename = os.path.basename(filepath)

            progress_label.config(text=f"Processing {file_idx}/{total_files}: {filename}")
            progress_text.insert(tk.END, f"[{file_idx}/{total_files}] {filename}\n")
            progress_text.see(tk.END)
            self.root.update()

            try:
                # Check if file exists
                if not os.path.exists(filepath):
                    progress_text.insert(tk.END, f"  SKIP: File not found\n\n")
                    skipped += 1
                    continue

                # Extract O-number from filename (e.g., o80556.txt -> O80556)
                match = re.search(r'[oO](\d+)', filename)
                if not match:
                    progress_text.insert(tk.END, f"  SKIP: Cannot extract O-number from filename\n\n")
                    skipped += 1
                    continue

                file_onumber = int(match.group(1))
                new_prog_str = f"O{file_onumber:05d}"  # Always 5 digits with leading zeros
                new_prog_str_lower = new_prog_str.lower()

                # Check if already correct
                if old_prog_num.lower() == new_prog_str_lower:
                    progress_text.insert(tk.END, f"  ✓ Already correct: {new_prog_str_lower}\n\n")
                    skipped += 1
                    continue

                # Read file content
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    lines = f.readlines()

                # Update internal program number (first line with O#####)
                updated = False
                for i, line in enumerate(lines):
                    stripped = line.strip()
                    match_line = re.match(r'^([oO]\d{4,})\s*(\(.*)?$', stripped)
                    if match_line:
                        title_part = match_line.group(2) if match_line.group(2) else ""
                        if title_part:
                            lines[i] = f"{new_prog_str} {title_part}\n"
                        else:
                            lines[i] = f"{new_prog_str}\n"
                        updated = True
                        break

                if not updated:
                    progress_text.insert(tk.END, f"  ⚠️  WARNING: No O-number line found in file to update\n")

                # Write updated content
                with open(filepath, 'w', encoding='utf-8') as f:
                    f.writelines(lines)

                # Update database with new program number
                cursor.execute("""
                    UPDATE programs
                    SET program_number = ?
                    WHERE program_number = ?
                """, (new_prog_str_lower, old_prog_num))
                conn.commit()

                progress_text.insert(tk.END, f"  ✓ {old_prog_num} → {new_prog_str_lower}\n\n")
                fixed += 1
                progress_text.see(tk.END)

            except Exception as e:
                progress_text.insert(tk.END, f"  ❌ ERROR: {str(e)[:100]}\n\n")
                progress_text.see(tk.END)
                errors += 1

        # Close database connection
        conn.close()

        # Show results
        progress_label.config(text="Complete!")
        progress_text.insert(tk.END, f"\n{'='*50}\n")
        progress_text.insert(tk.END, f"FIX PROGRAM NUMBERS - SUMMARY\n")
        progress_text.insert(tk.END, f"{'='*50}\n")
        progress_text.insert(tk.END, f"Files processed: {total_files}\n")
        progress_text.insert(tk.END, f"Successfully fixed: {fixed}\n")
        progress_text.insert(tk.END, f"Already correct/Skipped: {skipped}\n")
        progress_text.insert(tk.END, f"Errors: {errors}\n\n")
        progress_text.insert(tk.END, f"Internal O-numbers now match filenames.\n")
        progress_text.insert(tk.END, f"Database updated.\n")
        progress_text.see(tk.END)

        # Refresh the view to show new program numbers
        self.refresh_filter_values()
        self.refresh_results()

        close_btn = tk.Button(progress_window, text="Close",
                             command=progress_window.destroy,
                             bg=self.button_bg, fg=self.fg_color,
                             font=("Arial", 10, "bold"))
        close_btn.pack(pady=10)

    def clear_database(self):
        """Clear all records from the database"""
        # Confirm with user
        result = messagebox.askyesno(
            "Clear Database",
            "Are you sure you want to delete ALL records from the database?\n\n"
            "This action cannot be undone!",
            icon='warning'
        )

        if not result:
            return

        # Double confirm
        result2 = messagebox.askyesno(
            "Confirm Clear",
            "This will permanently delete all program records.\n\n"
            "Are you absolutely sure?",
            icon='warning'
        )

        if not result2:
            return

        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            # Get count before deletion
            cursor.execute("SELECT COUNT(*) FROM programs")
            count = cursor.fetchone()[0]

            # Delete all records
            cursor.execute("DELETE FROM programs")
            conn.commit()
            conn.close()

            # Refresh the display
            self.refresh_filter_values()
            self.refresh_results()

            messagebox.showinfo(
                "Database Cleared",
                f"Successfully deleted {count} records from the database.\n\n"
                "The database is now empty."
            )

        except Exception as e:
            messagebox.showerror("Error", f"Failed to clear database:\n{str(e)}")

    def save_database_profile(self):
        """Save current database as a named profile"""
        import shutil
        from datetime import datetime

        # Create profiles directory if it doesn't exist
        profiles_dir = os.path.join(os.path.dirname(self.db_path), "database_profiles")
        os.makedirs(profiles_dir, exist_ok=True)

        # Prompt for profile name
        profile_window = tk.Toplevel(self.root)
        profile_window.title("Save Database Profile")
        profile_window.geometry("450x200")
        profile_window.configure(bg=self.bg_color)
        profile_window.transient(self.root)
        profile_window.grab_set()

        tk.Label(profile_window,
                text="Save current database state as a profile",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 12, "bold")).pack(pady=15)

        tk.Label(profile_window,
                text="Profile Name:",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 10)).pack(pady=5)

        # Suggest default name with timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        default_name = f"Profile_{timestamp}"

        entry = tk.Entry(profile_window, bg=self.input_bg, fg=self.fg_color,
                        font=("Arial", 11), width=40)
        entry.insert(0, default_name)
        entry.pack(pady=10)
        entry.focus()
        entry.select_range(0, tk.END)

        result = [None]

        def save():
            profile_name = entry.get().strip()
            if not profile_name:
                messagebox.showwarning("Invalid Name", "Please enter a profile name.")
                return

            # Remove any invalid filename characters
            profile_name = "".join(c for c in profile_name if c.isalnum() or c in (' ', '-', '_'))

            profile_path = os.path.join(profiles_dir, f"{profile_name}.db")

            # Check if profile already exists
            if os.path.exists(profile_path):
                overwrite = messagebox.askyesno(
                    "Profile Exists",
                    f"Profile '{profile_name}' already exists.\n\nOverwrite?"
                )
                if not overwrite:
                    return

            try:
                # Copy current database to profile
                shutil.copy2(self.db_path, profile_path)

                # Get record count
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM programs")
                count = cursor.fetchone()[0]
                conn.close()

                result[0] = profile_name
                messagebox.showinfo(
                    "Profile Saved",
                    f"Database profile '{profile_name}' saved successfully!\n\n"
                    f"Records: {count}\n"
                    f"Location: {profile_path}"
                )
                profile_window.destroy()

            except Exception as e:
                messagebox.showerror("Error", f"Failed to save profile:\n{str(e)}")

        def cancel():
            profile_window.destroy()

        btn_frame = tk.Frame(profile_window, bg=self.bg_color)
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="💾 Save", command=save,
                 bg=self.accent_color, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=12).pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="❌ Cancel", command=cancel,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=12).pack(side=tk.LEFT, padx=5)

        self.root.wait_window(profile_window)

    def load_database_profile(self):
        """Load a saved database profile"""
        import shutil

        profiles_dir = os.path.join(os.path.dirname(self.db_path), "database_profiles")

        # Check if profiles directory exists
        if not os.path.exists(profiles_dir):
            messagebox.showinfo(
                "No Profiles",
                "No saved profiles found.\n\n"
                "Use 'Save Profile' to create your first database profile."
            )
            return

        # Get list of profile files
        profile_files = [f for f in os.listdir(profiles_dir) if f.endswith('.db')]

        if not profile_files:
            messagebox.showinfo(
                "No Profiles",
                "No saved profiles found.\n\n"
                "Use 'Save Profile' to create your first database profile."
            )
            return

        # Create selection window
        select_window = tk.Toplevel(self.root)
        select_window.title("Load Database Profile")
        select_window.geometry("600x500")
        select_window.configure(bg=self.bg_color)
        select_window.transient(self.root)
        select_window.grab_set()

        tk.Label(select_window,
                text="Select a profile to load",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 12, "bold")).pack(pady=15)

        # Listbox with profile info
        list_frame = tk.Frame(select_window, bg=self.bg_color)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        listbox = tk.Listbox(list_frame, bg=self.input_bg, fg=self.fg_color,
                            font=("Courier", 10), yscrollcommand=scrollbar.set,
                            selectmode=tk.SINGLE)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=listbox.yview)

        # Populate listbox with profile info
        profile_data = []
        for profile_file in sorted(profile_files, reverse=True):
            profile_path = os.path.join(profiles_dir, profile_file)
            profile_name = profile_file[:-3]  # Remove .db extension

            # Get file stats
            stat = os.stat(profile_path)
            size_mb = stat.st_size / (1024 * 1024)
            modified = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M")

            # Get record count
            try:
                conn = sqlite3.connect(profile_path)
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM programs")
                count = cursor.fetchone()[0]
                conn.close()
            except:
                count = "?"

            display = f"{profile_name:<35} | {count:>6} records | {modified} | {size_mb:.1f} MB"
            listbox.insert(tk.END, display)
            profile_data.append((profile_name, profile_path, count))

        def load_selected():
            selection = listbox.curselection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a profile to load.")
                return

            idx = selection[0]
            profile_name, profile_path, count = profile_data[idx]

            # Confirm load
            result = messagebox.askyesno(
                "Confirm Load Profile",
                f"Load profile '{profile_name}'?\n\n"
                f"This will replace your current database with:\n"
                f"  • {count} records\n"
                f"  • From: {os.path.basename(profile_path)}\n\n"
                f"⚠️  Your current database will be backed up first.",
                icon='warning'
            )

            if not result:
                return

            try:
                # Backup current database
                backup_name = f"before_load_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
                backup_path = os.path.join(os.path.dirname(self.db_path), "backups", backup_name)
                os.makedirs(os.path.dirname(backup_path), exist_ok=True)
                shutil.copy2(self.db_path, backup_path)

                # Load profile
                shutil.copy2(profile_path, self.db_path)

                # Refresh display
                self.refresh_filter_values()
                self.refresh_results()

                messagebox.showinfo(
                    "Profile Loaded",
                    f"Successfully loaded profile '{profile_name}'!\n\n"
                    f"Records loaded: {count}\n"
                    f"Previous database backed up to:\n{backup_name}"
                )

                select_window.destroy()

            except Exception as e:
                messagebox.showerror("Error", f"Failed to load profile:\n{str(e)}")

        def cancel():
            select_window.destroy()

        btn_frame = tk.Frame(select_window, bg=self.bg_color)
        btn_frame.pack(pady=15)

        tk.Button(btn_frame, text="📂 Load Selected", command=load_selected,
                 bg=self.accent_color, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=15).pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="❌ Cancel", command=cancel,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=15).pack(side=tk.LEFT, padx=5)

    def manage_database_profiles(self):
        """Manage saved database profiles - view, delete, rename"""
        import shutil
        from datetime import datetime

        profiles_dir = os.path.join(os.path.dirname(self.db_path), "database_profiles")

        # Check if profiles directory exists
        if not os.path.exists(profiles_dir):
            os.makedirs(profiles_dir, exist_ok=True)

        # Create management window
        manage_window = tk.Toplevel(self.root)
        manage_window.title("Manage Database Profiles")
        manage_window.geometry("800x600")
        manage_window.configure(bg=self.bg_color)

        tk.Label(manage_window,
                text="Database Profile Manager",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 14, "bold")).pack(pady=15)

        # Treeview for profile list
        tree_frame = tk.Frame(manage_window, bg=self.bg_color)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        columns = ('name', 'records', 'date', 'size')
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)

        tree.heading('name', text='Profile Name')
        tree.heading('records', text='Records')
        tree.heading('date', text='Date Modified')
        tree.heading('size', text='Size (MB)')

        tree.column('name', width=300)
        tree.column('records', width=100)
        tree.column('date', width=200)
        tree.column('size', width=100)

        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscroll=scrollbar.set)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        def refresh_profile_list():
            # Clear existing items
            for item in tree.get_children():
                tree.delete(item)

            # Get profile files
            profile_files = [f for f in os.listdir(profiles_dir) if f.endswith('.db')]

            if not profile_files:
                tree.insert('', tk.END, values=("No profiles found - use 'Save Profile' to create one", "", "", ""))
                return

            for profile_file in sorted(profile_files, reverse=True):
                profile_path = os.path.join(profiles_dir, profile_file)
                profile_name = profile_file[:-3]

                # Get stats
                stat = os.stat(profile_path)
                size_mb = f"{stat.st_size / (1024 * 1024):.2f}"
                modified = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")

                # Get record count
                try:
                    conn = sqlite3.connect(profile_path)
                    cursor = conn.cursor()
                    cursor.execute("SELECT COUNT(*) FROM programs")
                    count = cursor.fetchone()[0]
                    conn.close()
                except:
                    count = "Error"

                tree.insert('', tk.END, values=(profile_name, count, modified, size_mb))

        def delete_profile():
            selection = tree.selection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a profile to delete.")
                return

            item = selection[0]
            profile_name = tree.item(item)['values'][0]

            if profile_name == "No profiles found - use 'Save Profile' to create one":
                return

            result = messagebox.askyesno(
                "Confirm Delete",
                f"Delete profile '{profile_name}'?\n\n"
                f"This cannot be undone!",
                icon='warning'
            )

            if result:
                try:
                    profile_path = os.path.join(profiles_dir, f"{profile_name}.db")
                    os.remove(profile_path)
                    refresh_profile_list()
                    messagebox.showinfo("Deleted", f"Profile '{profile_name}' deleted successfully.")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to delete profile:\n{str(e)}")

        def export_profile():
            selection = tree.selection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a profile to export.")
                return

            item = selection[0]
            profile_name = tree.item(item)['values'][0]

            if profile_name == "No profiles found - use 'Save Profile' to create one":
                return

            # Ask where to save
            export_path = filedialog.asksaveasfilename(
                defaultextension=".db",
                initialfile=f"{profile_name}.db",
                filetypes=[("Database files", "*.db"), ("All files", "*.*")],
                title="Export Profile"
            )

            if export_path:
                try:
                    profile_path = os.path.join(profiles_dir, f"{profile_name}.db")
                    shutil.copy2(profile_path, export_path)
                    messagebox.showinfo("Exported", f"Profile exported to:\n{export_path}")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to export profile:\n{str(e)}")

        # Initial load
        refresh_profile_list()

        # Button frame
        btn_frame = tk.Frame(manage_window, bg=self.bg_color)
        btn_frame.pack(pady=15)

        tk.Button(btn_frame, text="🗑️ Delete", command=delete_profile,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=12).pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="📤 Export", command=export_profile,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=12).pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="🔄 Refresh", command=refresh_profile_list,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=12).pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="❌ Close", command=manage_window.destroy,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=12).pack(side=tk.LEFT, padx=5)

    def fix_duplicates(self):
        """Fix duplicate program numbers by assigning new unique o##### values based on OD"""
        # Find duplicates in database (entries with (##) suffix)
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Find all programs with duplicate suffix pattern
        cursor.execute("""
            SELECT program_number, file_path, outer_diameter
            FROM programs
            WHERE program_number LIKE '%(%)'
            ORDER BY program_number
        """)
        duplicates = cursor.fetchall()

        if not duplicates:
            messagebox.showinfo("No Duplicates", "No duplicate programs found in the database.\n\n"
                              "Duplicates have a (1), (2), etc. suffix.")
            conn.close()
            return

        # Confirm with user
        result = messagebox.askyesno(
            "Fix Duplicates",
            f"Found {len(duplicates)} duplicate programs.\n\n"
            "This will:\n"
            "1. Assign new unique o##### program numbers based on OD\n"
            "2. Rename the files\n"
            "3. Update the internal program number in each file\n"
            "4. Update the database\n\n"
            "Proceed?"
        )

        if not result:
            conn.close()
            return

        # Get all existing program numbers (to avoid collisions)
        cursor.execute("SELECT program_number FROM programs")
        existing_programs = set(row[0] for row in cursor.fetchall())

        # Also check the filesystem for any o##### files
        # Get unique directories from duplicates
        directories = set()
        for _, file_path, _ in duplicates:
            if file_path:
                directories.add(os.path.dirname(file_path))

        # Scan directories for existing program numbers
        for directory in directories:
            if os.path.exists(directory):
                for file in os.listdir(directory):
                    match = re.search(r'[oO](\d{4,})', file)
                    if match:
                        existing_programs.add(f"o{match.group(1)}")

        conn.close()

        # Show progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Fixing Duplicates...")
        progress_window.geometry("600x500")
        progress_window.configure(bg=self.bg_color)

        progress_label = tk.Label(progress_window, text="Processing duplicates...",
                                 bg=self.bg_color, fg=self.fg_color,
                                 font=("Arial", 12))
        progress_label.pack(pady=20)

        progress_text = scrolledtext.ScrolledText(progress_window,
                                                 bg=self.input_bg, fg=self.fg_color,
                                                 width=70, height=20)
        progress_text.pack(padx=10, pady=10)

        self.root.update()

        # Process each duplicate
        fixed = 0
        errors = 0

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        for old_program_number, file_path, outer_diameter in duplicates:
            progress_text.insert(tk.END, f"\nProcessing: {old_program_number}\n")
            progress_text.see(tk.END)
            self.root.update()

            # Determine OD range for new program number
            new_program_number = self._get_next_available_program_number(
                outer_diameter, existing_programs
            )

            if not new_program_number:
                progress_text.insert(tk.END, f"  ERROR: Could not find available program number\n")
                errors += 1
                continue

            # Add to existing set to prevent reuse
            existing_programs.add(new_program_number)

            progress_text.insert(tk.END, f"  New program number: {new_program_number}\n")

            # Update file if it exists
            if file_path and os.path.exists(file_path):
                try:
                    # Read file content (preserve original encoding and line endings)
                    with open(file_path, 'rb') as f:
                        content_bytes = f.read()

                    # Detect line ending style
                    if b'\r\n' in content_bytes:
                        line_ending = '\r\n'
                    elif b'\r' in content_bytes:
                        line_ending = '\r'
                    else:
                        line_ending = '\n'

                    # Decode content
                    try:
                        content = content_bytes.decode('utf-8')
                    except UnicodeDecodeError:
                        content = content_bytes.decode('latin-1')

                    # Split into lines (normalize first, then restore)
                    lines = content.replace('\r\n', '\n').replace('\r', '\n').split('\n')

                    # Update internal program number
                    new_lines = []
                    internal_updated = False
                    for line in lines:
                        stripped = line.strip()
                        # Check if line is a program number (O followed by digits)
                        if re.match(r'^[oO]\d{4,}\s*$', stripped):
                            new_lines.append(f"O{new_program_number[1:]}")  # Use uppercase O
                            internal_updated = True
                        else:
                            new_lines.append(line)

                    # Write back to file with original line endings
                    with open(file_path, 'w', newline='') as f:
                        f.write(line_ending.join(new_lines))

                    if internal_updated:
                        progress_text.insert(tk.END, f"  Updated internal program number\n")

                    # Rename file
                    directory = os.path.dirname(file_path)
                    old_filename = os.path.basename(file_path)

                    # Determine new filename (preserve extension if any)
                    if '.' in old_filename:
                        ext = old_filename[old_filename.rfind('.'):]
                        new_filename = f"{new_program_number}{ext}"
                    else:
                        new_filename = new_program_number

                    new_file_path = os.path.join(directory, new_filename)

                    # Check if target file already exists
                    if os.path.exists(new_file_path):
                        progress_text.insert(tk.END, f"  WARNING: Target file already exists: {new_filename}\n")
                        errors += 1
                        continue

                    os.rename(file_path, new_file_path)
                    progress_text.insert(tk.END, f"  Renamed: {old_filename} -> {new_filename}\n")

                    # Update database
                    cursor.execute("""
                        UPDATE programs
                        SET program_number = ?, file_path = ?
                        WHERE program_number = ?
                    """, (new_program_number, new_file_path, old_program_number))

                    progress_text.insert(tk.END, f"  Database updated\n")
                    fixed += 1

                except Exception as e:
                    progress_text.insert(tk.END, f"  ERROR: {str(e)}\n")
                    errors += 1
            else:
                # File doesn't exist, just update database
                progress_text.insert(tk.END, f"  WARNING: File not found, updating database only\n")
                cursor.execute("""
                    UPDATE programs
                    SET program_number = ?
                    WHERE program_number = ?
                """, (new_program_number, old_program_number))
                fixed += 1

            progress_text.see(tk.END)
            self.root.update()

        conn.commit()
        conn.close()

        # Show results
        progress_label.config(text="Complete!")
        progress_text.insert(tk.END, f"\n{'='*50}\n")
        progress_text.insert(tk.END, f"Fixed: {fixed}\n")
        progress_text.insert(tk.END, f"Errors: {errors}\n")
        progress_text.see(tk.END)

        close_btn = tk.Button(progress_window, text="Close",
                             command=progress_window.destroy,
                             bg=self.button_bg, fg=self.fg_color,
                             font=("Arial", 10, "bold"))
        close_btn.pack(pady=10)

        # Refresh display
        self.refresh_filter_values()
        self.refresh_results()

    def _get_next_available_program_number(self, outer_diameter: float, existing: set) -> str:
        """
        Get next available program number based on OD.

        OD-based ranges:
        - 7.0", 7.5", 8.0", 8.5" -> o7####
        - 6.0", 6.25", 6.5" -> o6####
        - 5.75" -> o5####
        - All others -> any available
        """
        # Determine preferred range based on OD
        if outer_diameter:
            if outer_diameter >= 7.0:
                preferred_ranges = [(70000, 79999), (80000, 89999), (60000, 69999), (50000, 59999)]
            elif outer_diameter >= 6.0:
                preferred_ranges = [(60000, 69999), (70000, 79999), (50000, 59999), (80000, 89999)]
            elif outer_diameter >= 5.5:
                preferred_ranges = [(50000, 59999), (60000, 69999), (70000, 79999), (80000, 89999)]
            else:
                preferred_ranges = [(50000, 59999), (60000, 69999), (70000, 79999), (80000, 89999)]
        else:
            # No OD info, try all ranges
            preferred_ranges = [(50000, 59999), (60000, 69999), (70000, 79999), (80000, 89999)]

        # Search for available number in preferred order
        for start, end in preferred_ranges:
            for num in range(start, end + 1):
                candidate = f"o{num}"
                if candidate not in existing and candidate.upper() not in existing:
                    return candidate

        # If all preferred ranges are full, search extended ranges
        for num in range(10000, 99999):
            candidate = f"o{num}"
            if candidate not in existing and candidate.upper() not in existing:
                return candidate

        return None

    def refresh_filter_values(self):
        """Refresh available filter values from database"""
        # Update available values from database
        new_types = self.get_available_values("spacer_type")
        new_materials = self.get_available_values("material")
        new_statuses = self.get_available_values("validation_status")

        # Update filter widgets if values changed
        if new_types != self.available_types:
            self.available_types = new_types
            self.filter_type.values = new_types
            self.filter_type.clear()

        if new_materials != self.available_materials:
            self.available_materials = new_materials
            self.filter_material.values = new_materials
            self.filter_material.clear()

        if new_statuses != self.available_statuses:
            self.available_statuses = new_statuses
            self.filter_status.values = new_statuses
            self.filter_status.clear()

    def refresh_results(self, view_mode='all', external_only=False):
        """Refresh the results table based on current filters

        Args:
            view_mode: 'all' (default), 'repository', or 'external'
            external_only: Deprecated, use view_mode='external' instead
        """
        # Handle deprecated parameter
        if external_only:
            view_mode = 'external'

        # Clear existing
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Build query
        query = "SELECT * FROM programs WHERE 1=1"
        params = []

        # Add view mode filter
        if view_mode == 'repository':
            # Only show managed files (in repository)
            query += " AND is_managed = 1"
        elif view_mode == 'external':
            # Only show external files (NOT in repository)
            query += " AND (is_managed = 0 OR is_managed IS NULL)"
        # 'all' mode shows everything (no additional filter)

        # Title search filter - supports multiple terms with +
        if self.filter_title.get():
            search_text = self.filter_title.get().strip()

            # Check if using + operator for multi-term search
            if '+' in search_text:
                # Split by + and strip whitespace from each term
                search_terms = [term.strip() for term in search_text.split('+') if term.strip()]

                # Each term must be present (AND logic)
                for term in search_terms:
                    query += " AND title LIKE ?"
                    params.append(f"%{term}%")
            else:
                # Single term search
                query += " AND title LIKE ?"
                params.append(f"%{search_text}%")

        # Program number filter
        if self.filter_program.get():
            query += " AND program_number LIKE ?"
            params.append(f"%{self.filter_program.get()}%")
        
        # Type filter (multi-select)
        selected_types = self.filter_type.get_selected()
        if selected_types and len(selected_types) < len(self.filter_type.values):
            placeholders = ','.join('?' * len(selected_types))
            query += f" AND spacer_type IN ({placeholders})"
            params.extend(selected_types)

        # Material filter (multi-select)
        selected_materials = self.filter_material.get_selected()
        if selected_materials and len(selected_materials) < len(self.filter_material.values):
            placeholders = ','.join('?' * len(selected_materials))
            query += f" AND material IN ({placeholders})"
            params.extend(selected_materials)

        # Validation Status filter (multi-select)
        selected_statuses = self.filter_status.get_selected()
        if selected_statuses and len(selected_statuses) < len(self.filter_status.values):
            placeholders = ','.join('?' * len(selected_statuses))
            query += f" AND validation_status IN ({placeholders})"
            params.extend(selected_statuses)

        # Duplicate Type filter (multi-select)
        selected_dup_types = self.filter_dup_type.get_selected()
        if selected_dup_types and len(selected_dup_types) < len(self.filter_dup_type.values):
            # Handle "None" selection (files with no duplicate_type)
            if "None" in selected_dup_types:
                other_types = [t for t in selected_dup_types if t != "None"]
                if other_types:
                    placeholders = ','.join('?' * len(other_types))
                    query += f" AND (duplicate_type IN ({placeholders}) OR duplicate_type IS NULL)"
                    params.extend(other_types)
                else:
                    query += " AND duplicate_type IS NULL"
            else:
                placeholders = ','.join('?' * len(selected_dup_types))
                query += f" AND duplicate_type IN ({placeholders})"
                params.extend(selected_dup_types)

        # OD range
        if self.filter_od_min.get():
            query += " AND outer_diameter >= ?"
            params.append(float(self.filter_od_min.get()))
        if self.filter_od_max.get():
            query += " AND outer_diameter <= ?"
            params.append(float(self.filter_od_max.get()))
        
        # Thickness range
        if self.filter_thickness_min.get():
            query += " AND thickness >= ?"
            params.append(float(self.filter_thickness_min.get()))
        if self.filter_thickness_max.get():
            query += " AND thickness <= ?"
            params.append(float(self.filter_thickness_max.get()))
        
        # CB range
        if self.filter_cb_min.get():
            query += " AND center_bore >= ?"
            params.append(float(self.filter_cb_min.get()))
        if self.filter_cb_max.get():
            query += " AND center_bore <= ?"
            params.append(float(self.filter_cb_max.get()))

        # Hub Diameter range
        if self.filter_hub_dia_min.get():
            query += " AND hub_diameter >= ?"
            params.append(float(self.filter_hub_dia_min.get()))
        if self.filter_hub_dia_max.get():
            query += " AND hub_diameter <= ?"
            params.append(float(self.filter_hub_dia_max.get()))

        # Hub Height range
        if self.filter_hub_h_min.get():
            query += " AND hub_height >= ?"
            params.append(float(self.filter_hub_h_min.get()))
        if self.filter_hub_h_max.get():
            query += " AND hub_height <= ?"
            params.append(float(self.filter_hub_h_max.get()))

        # Step Diameter range
        if self.filter_step_d_min.get():
            query += " AND counter_bore_diameter >= ?"
            params.append(float(self.filter_step_d_min.get()))
        if self.filter_step_d_max.get():
            query += " AND counter_bore_diameter <= ?"
            params.append(float(self.filter_step_d_max.get()))

        # Error text filter (searches in validation_issues, bore_warnings, dimensional_issues)
        if self.filter_error_text.get():
            error_search = f"%{self.filter_error_text.get()}%"
            query += " AND (validation_issues LIKE ? OR bore_warnings LIKE ? OR dimensional_issues LIKE ?)"
            params.extend([error_search, error_search, error_search])

        # Missing file path filter
        if self.filter_missing_file_path.get():
            query += " AND (file_path IS NULL OR file_path = '')"

        query += " ORDER BY program_number"

        # Note: Duplicates filter is applied after query in the display logic
        # since it requires checking for duplicate filenames across all results
        
        # Execute query
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute(query, params)
        results = cursor.fetchall()
        conn.close()

        # Populate tree
        # Column indices (based on database schema):
        # 0:program_number, 1:title, 2:spacer_type, 3:outer_diameter, 4:thickness, 5:thickness_display,
        # 6:center_bore, 7:hub_height, 8:hub_diameter, 9:counter_bore_diameter, 10:counter_bore_depth,
        # 11:paired_program, 12:material, 13:notes, 14:date_created, 15:last_modified, 16:file_path,
        # 17:detection_confidence, 18:detection_method, 19:validation_status, ...

        # Build set of duplicate filenames (exact filenames that appear more than once)
        filename_counts = {}
        for row in results:
            if row[16]:  # file_path
                filename = os.path.basename(row[16]).lower()  # Case-insensitive
                filename_counts[filename] = filename_counts.get(filename, 0) + 1

        # Filenames that appear more than once are duplicates
        duplicate_filenames = {fn for fn, count in filename_counts.items() if count > 1}

        # Assign occurrence numbers to duplicates
        occurrence_tracker = {}
        filename_occurrences = {}  # Maps file_path to its occurrence number
        for row in results:
            if row[16]:  # file_path
                filename = os.path.basename(row[16]).lower()
                if filename in duplicate_filenames:
                    occurrence_tracker[filename] = occurrence_tracker.get(filename, 0) + 1
                    filename_occurrences[row[16]] = occurrence_tracker[filename]

        # Count status breakdown
        status_counts = {
            'CRITICAL': 0,
            'SAFETY_ERROR': 0,
            'TOOL_ERROR': 0,
            'BORE_WARNING': 0,
            'DIMENSIONAL': 0,
            'TOOL_WARNING': 0,
            'WARNING': 0,
            'PASS': 0,
            'REPEAT': 0
        }

        for row in results:
            program_number = row[0]
            # Check if this filename is a duplicate and get its occurrence number
            is_dup = ""
            dup_num = ""
            if row[16]:  # file_path
                filename = os.path.basename(row[16]).lower()
                if filename in duplicate_filenames:
                    dup_num = filename_occurrences.get(row[16], 1)
                    is_dup = f"({dup_num})"

            # Apply duplicates filter - skip non-duplicates if filter is checked
            if self.filter_duplicates.get() and not is_dup:
                continue

            title = row[1] if row[1] else "-"  # NEW: Title from G-code
            spacer_type = row[2]  # Shifted from row[1]
            od = f"{row[3]:.3f}" if row[3] else "-"  # Shifted from row[2]

            # Use thickness_display (row[5]) if available, otherwise fall back to formatted thickness (row[4])
            thick = row[5] if row[5] else (f"{row[4]:.3f}" if row[4] else "-")  # Shifted from row[4]/row[3]

            cb = f"{row[6]:.1f}" if row[6] else "-"  # Shifted from row[5]

            # Hub Height - only applicable for hub_centric
            if spacer_type == 'hub_centric':
                hub_h = f"{row[7]:.2f}" if row[7] else "-"  # Shifted from row[6]
            else:
                hub_h = "N/A"

            # Hub Diameter (OB) - only applicable for hub_centric
            if spacer_type == 'hub_centric':
                hub_d = f"{row[8]:.1f}" if row[8] else "-"  # Shifted from row[7]
            else:
                hub_d = "N/A"

            # Counter Bore - only applicable for STEP parts
            if spacer_type == 'step':
                cb_bore = f"{row[9]:.1f}" if row[9] else "-"  # Shifted from row[8]
            else:
                cb_bore = "N/A"

            # Step Depth - only applicable for STEP parts (row[10] = counter_bore_depth)
            if spacer_type == 'step':
                step_d = f"{row[10]:.2f}" if row[10] else "-"
            else:
                step_d = "N/A"

            material = row[12] if row[12] else "-"  # Shifted from row[11]
            filename = os.path.basename(row[16]) if row[16] else "-"  # Shifted from row[15]

            # Lathe (index 26)
            lathe = row[26] if len(row) > 26 and row[26] else "-"

            # Validation status (index 19 - validation_status)
            validation_status = row[19] if len(row) > 19 and row[19] else "N/A"  # Shifted from row[18]

            # Extract warning details based on status type
            warning_details = "-"
            if validation_status == 'CRITICAL' and len(row) > 20 and row[20]:  # validation_issues
                # Parse CRITICAL issues
                try:
                    import json
                    if isinstance(row[20], str):
                        issues = json.loads(row[20])
                    else:
                        issues = row[20] if isinstance(row[20], list) else []
                    warning_details = "; ".join(str(x) for x in issues[:2]) if issues else ""
                except:
                    try:
                        issues = [i.strip() for i in str(row[20]).split('|') if i.strip()]
                        warning_details = "; ".join(issues[:2]) if issues else str(row[20])[:100]
                    except:
                        warning_details = str(row[20])[:100] if row[20] else ""
            elif validation_status == 'BORE_WARNING' and len(row) > 24 and row[24]:  # bore_warnings
                # Parse BORE warnings
                try:
                    import json
                    if isinstance(row[24], str):
                        warns = json.loads(row[24])
                    else:
                        warns = row[24] if isinstance(row[24], list) else []
                    warning_details = "; ".join(str(x) for x in warns[:2]) if warns else ""
                except:
                    try:
                        warns = [i.strip() for i in str(row[24]).split('|') if i.strip()]
                        warning_details = "; ".join(warns[:2]) if warns else str(row[24])[:100]
                    except:
                        warning_details = str(row[24])[:100] if row[24] else ""
            elif validation_status == 'DIMENSIONAL' and len(row) > 25 and row[25]:  # dimensional_issues
                # Parse DIMENSIONAL issues
                try:
                    import json
                    if isinstance(row[25], str):
                        dim_issues = json.loads(row[25])
                    else:
                        dim_issues = row[25] if isinstance(row[25], list) else []
                    warning_details = "; ".join(str(x) for x in dim_issues[:2]) if dim_issues else ""
                except:
                    try:
                        dim_issues = [i.strip() for i in str(row[25]).split('|') if i.strip()]
                        warning_details = "; ".join(dim_issues[:2]) if dim_issues else str(row[25])[:100]
                    except:
                        warning_details = str(row[25])[:100] if row[25] else ""
            elif validation_status == 'SAFETY_ERROR' and len(row) > 33 and row[33]:  # safety_blocks_issues
                # Parse SAFETY_ERROR issues
                try:
                    import json
                    if isinstance(row[33], str):
                        issues = json.loads(row[33])
                    else:
                        issues = row[33] if isinstance(row[33], list) else []
                    warning_details = "; ".join(str(x) for x in issues[:2]) if issues else ""
                except:
                    try:
                        issues = [i.strip() for i in str(row[33]).split('|') if i.strip()]
                        warning_details = "; ".join(issues[:2]) if issues else str(row[33])[:100]
                    except:
                        warning_details = str(row[33])[:100] if row[33] else ""
            elif validation_status in ('TOOL_ERROR', 'TOOL_WARNING') and len(row) > 31 and row[31]:  # tool_validation_issues
                # Parse TOOL_ERROR and TOOL_WARNING issues
                try:
                    import json
                    if isinstance(row[31], str):
                        issues = json.loads(row[31])
                    else:
                        issues = row[31] if isinstance(row[31], list) else []
                    warning_details = "; ".join(str(x) for x in issues[:2]) if issues else ""
                except:
                    try:
                        issues = [i.strip() for i in str(row[31]).split('|') if i.strip()]
                        warning_details = "; ".join(issues[:2]) if issues else str(row[31])[:100]
                    except:
                        warning_details = str(row[31])[:100] if row[31] else ""
            elif validation_status == 'WARNING' and len(row) > 21 and row[21]:  # validation_warnings
                # Parse general WARNING
                try:
                    import json
                    if isinstance(row[21], str):
                        warns = json.loads(row[21])
                    else:
                        warns = row[21] if isinstance(row[21], list) else []
                    warning_details = "; ".join(str(x) for x in warns[:2]) if warns else ""
                except:
                    try:
                        warns = [i.strip() for i in str(row[21]).split('|') if i.strip()]
                        warning_details = "; ".join(warns[:2]) if warns else str(row[21])[:100]
                    except:
                        warning_details = str(row[21])[:100] if row[21] else ""

            # Count status
            if validation_status in status_counts:
                status_counts[validation_status] += 1
            elif validation_status == 'ERROR':  # Old status name
                status_counts['CRITICAL'] += 1

            # Determine color tag (prioritized by severity)
            tag = ''
            if validation_status == 'CRITICAL':
                tag = 'critical'  # RED - Critical errors (CB/OB way off)
            elif validation_status == 'SAFETY_ERROR':
                tag = 'safety_error'  # DARK RED - Missing safety blocks
            elif validation_status == 'TOOL_ERROR':
                tag = 'tool_error'  # ORANGE-RED - Wrong/missing tools
            elif validation_status == 'BORE_WARNING':
                tag = 'bore_warning'  # ORANGE - Bore dimensions at tolerance limit
            elif validation_status == 'DIMENSIONAL':
                tag = 'dimensional'  # PURPLE - P-code/thickness mismatches
            elif validation_status == 'TOOL_WARNING':
                tag = 'tool_warning'  # AMBER - Tool suggestions
            elif validation_status == 'WARNING':
                tag = 'warning'  # YELLOW - General warnings
            elif validation_status == 'REPEAT':
                tag = 'repeat'  # GRAY - Duplicate files
            elif validation_status == 'PASS':
                tag = 'pass'  # GREEN - Pass
            # Old status names for backward compatibility
            elif validation_status == 'ERROR':
                tag = 'critical'

            self.tree.insert("", "end", values=(
                program_number, is_dup, title, spacer_type, lathe, od, thick, cb,
                hub_h, hub_d, cb_bore, step_d, material, validation_status, warning_details, filename
            ), tags=(tag,))

        # Update count with status breakdown and view mode indicator
        # Add view mode label
        if view_mode == 'repository':
            view_label = "Repository"
        elif view_mode == 'external':
            view_label = "External"
        else:
            view_label = "All Programs"

        status_text = f"{view_label}: {len(results)} programs  |  "
        status_parts = []
        if status_counts['CRITICAL'] > 0:
            status_parts.append(f"CRITICAL: {status_counts['CRITICAL']}")
        if status_counts['SAFETY_ERROR'] > 0:
            status_parts.append(f"SAFETY: {status_counts['SAFETY_ERROR']}")
        if status_counts['TOOL_ERROR'] > 0:
            status_parts.append(f"TOOL_ERR: {status_counts['TOOL_ERROR']}")
        if status_counts['BORE_WARNING'] > 0:
            status_parts.append(f"BORE: {status_counts['BORE_WARNING']}")
        if status_counts['DIMENSIONAL'] > 0:
            status_parts.append(f"DIM: {status_counts['DIMENSIONAL']}")
        if status_counts['TOOL_WARNING'] > 0:
            status_parts.append(f"TOOL_WARN: {status_counts['TOOL_WARNING']}")
        if status_counts['WARNING'] > 0:
            status_parts.append(f"WARN: {status_counts['WARNING']}")
        if status_counts['REPEAT'] > 0:
            status_parts.append(f"REPEAT: {status_counts['REPEAT']}")
        if status_counts['PASS'] > 0:
            status_parts.append(f"PASS: {status_counts['PASS']}")

        status_text += "  ".join(status_parts) if status_parts else "No status data"

        # Update both labels (one in filter section, one in results header)
        if hasattr(self, 'results_label'):
            self.results_label.config(text=status_text)
        if hasattr(self, 'results_counter_label'):
            self.results_counter_label.config(text=status_text)
        
    def clear_title_search(self):
        """Clear just the title search field"""
        self.filter_title.delete(0, tk.END)
        self.refresh_results()

    def clear_filters(self):
        """Clear all filter fields"""
        self.filter_title.delete(0, tk.END)
        self.filter_program.delete(0, tk.END)
        self.filter_type.clear()
        self.filter_material.clear()
        self.filter_status.clear()
        self.filter_dup_type.clear()
        # All dimensional filters are now comboboxes - use .set("")
        self.filter_od_min.set("")
        self.filter_od_max.set("")
        self.filter_thickness_min.set("")
        self.filter_thickness_max.set("")
        self.filter_cb_min.set("")
        self.filter_cb_max.set("")
        self.filter_hub_dia_min.set("")
        self.filter_hub_dia_max.set("")
        self.filter_hub_h_min.set("")
        self.filter_hub_h_max.set("")
        self.filter_step_d_min.set("")
        self.filter_step_d_max.set("")
        self.filter_error_text.set("")
        self.filter_duplicates.set(False)
        self.filter_missing_file_path.set(False)
        self.refresh_results()
        
    def sort_column(self, col):
        """Sort treeview by column with toggle support"""
        # Initialize sort state if not exists
        if not hasattr(self, '_sort_state'):
            self._sort_state = {'column': None, 'reverse': False}

        # Toggle direction if same column, otherwise new sort
        if self._sort_state['column'] == col:
            self._sort_state['reverse'] = not self._sort_state['reverse']
        else:
            self._sort_state['column'] = col
            self._sort_state['reverse'] = False

        # Get all data
        children = self.tree.get_children('')

        def get_sort_value(child):
            """Get sortable value from column"""
            val = self.tree.set(child, col)
            # Handle empty/N/A values - put at end
            if val in ('-', 'N/A', ''):
                return (1, 0)  # (is_empty, value) - empty values sort last
            try:
                return (0, float(val))
            except:
                return (0, val.lower() if isinstance(val, str) else val)

        # Sort data
        data = sorted(
            children,
            key=get_sort_value,
            reverse=self._sort_state['reverse']
        )

        # Rearrange
        for index, child in enumerate(data):
            self.tree.move(child, '', index)

        # Update column header to show sort direction
        direction = "▼" if self._sort_state['reverse'] else "▲"
        for column in self.tree['columns']:
            # Reset all headers
            base_name = column.replace(" ▲", "").replace(" ▼", "")
            if column == col or column.replace(" ▲", "").replace(" ▼", "") == col:
                self.tree.heading(column, text=f"{base_name} {direction}",
                                 command=lambda c=base_name: self.sort_column(c))
            else:
                self.tree.heading(column, text=base_name,
                                 command=lambda c=base_name: self.sort_column(c))

    def apply_multi_sort(self):
        """Apply multi-column sorting based on dropdown selections"""
        # Get sort configurations
        sort_configs = []

        for col_combo, dir_combo in [(self.sort1_col, self.sort1_dir),
                                     (self.sort2_col, self.sort2_dir),
                                     (self.sort3_col, self.sort3_dir)]:
            col = col_combo.get()
            if col:  # Only add if column is selected
                reverse = dir_combo.get() == "High→Low"
                sort_configs.append((col, reverse))

        if not sort_configs:
            return

        # Get all children
        children = list(self.tree.get_children(''))

        def get_sort_value(child, column):
            """Get sortable value from column"""
            val = self.tree.set(child, column)
            # Handle empty/N/A values - put at end
            if val in ('-', 'N/A', ''):
                return (1, 0)  # (is_empty, value) - empty values sort last
            try:
                return (0, float(val))
            except:
                return (0, val.lower() if isinstance(val, str) else val)

        # Sort using stable sort - apply in reverse order of priority
        # (last sort first, then earlier sorts override while preserving order)
        sorted_children = children

        # Apply sorts in reverse order (last sort first, then override with earlier sorts)
        for col, reverse in reversed(sort_configs):
            sorted_children = sorted(
                sorted_children,
                key=lambda child: get_sort_value(child, col),
                reverse=reverse
            )

        # Rearrange
        for index, child in enumerate(sorted_children):
            self.tree.move(child, '', index)

    def open_file(self, event=None):
        """Open selected gcode file"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select a program to open")
            return
        
        program_number = self.tree.item(selected[0])['values'][0]
        
        # Get file path from database
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT file_path FROM programs WHERE program_number = ?", (program_number,))
        result = cursor.fetchone()
        conn.close()
        
        if result and result[0]:
            filepath = result[0]
            if os.path.exists(filepath):
                # Open with default application
                if os.name == 'nt':  # Windows
                    os.startfile(filepath)
                elif os.name == 'posix':  # Mac/Linux
                    os.system(f'open "{filepath}"' if sys.platform == 'darwin' else f'xdg-open "{filepath}"')
            else:
                messagebox.showerror("File Not Found", f"File not found:\n{filepath}")
        else:
            messagebox.showerror("Error", "No file path in database")
            
    def add_entry(self):
        """Add a new entry manually"""
        EditEntryWindow(self, None, self.refresh_results)
        
    def edit_entry(self):
        """Edit selected entry"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select a program to edit")
            return
        
        program_number = self.tree.item(selected[0])['values'][0]
        EditEntryWindow(self, program_number, self.refresh_results)
        
    def delete_entry(self):
        """Delete selected entry"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select a program to delete")
            return
        
        program_number = self.tree.item(selected[0])['values'][0]
        
        if messagebox.askyesno("Confirm Delete", 
                              f"Delete program {program_number}?\n\nThis will only remove the database entry, not the file."):
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM programs WHERE program_number = ?", (program_number,))
            conn.commit()
            conn.close()
            
            self.refresh_results()
            messagebox.showinfo("Deleted", "Entry deleted successfully")
            
    def view_details(self):
        """View full details of selected entry"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select a program to view")
            return

        program_number = self.tree.item(selected[0])['values'][0]

        # Get full record
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM programs WHERE program_number = ?", (program_number,))
        result = cursor.fetchone()
        conn.close()

        if result:
            DetailsWindow(self, result)

    def show_version_history_window(self):
        """Show version history window for selected program"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select a program to view version history")
            return

        program_number = self.tree.item(selected[0])['values'][0]

        # Get file path for the program
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT file_path FROM programs WHERE program_number = ?", (program_number,))
        result = cursor.fetchone()
        conn.close()

        if not result or not result[0]:
            messagebox.showerror("Error", "File path not found for this program.")
            return

        file_path = result[0]

        # Open version history window
        VersionHistoryWindow(self.root, self, program_number, file_path)

    def export_csv(self):
        """Export database to Excel with separate sheets for each round size"""
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Missing Library",
                "openpyxl is required for Excel export.\n\n"
                "Install it with:\npip install openpyxl")
            return

        from datetime import datetime
        default_filename = f"GCode_Programs_{datetime.now().strftime('%Y%m%d')}.xlsx"

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Excel Workbook", "*.xlsx"), ("All files", "*.*")],
            title="Export to Excel - By Round Size"
        )

        if filepath:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # Define organized column order - dimensions and key info only
            export_columns = [
                # Identity
                'program_number',
                'title',

                # Dimensions (sorted by OD, Type, CB, Thickness)
                'outer_diameter',
                'spacer_type',
                'center_bore',
                'thickness',
                'hub_diameter',
                'hub_height',
                'counter_bore_diameter',
                'counter_bore_depth',

                # Classification
                'material',
                'lathe',

                # Status (summary only)
                'validation_status',
                'detection_confidence',

                # Metadata
                'last_modified'
            ]

            # Build query with ordering: OD, Type, CB, Thickness
            # NULLS LAST ensures NULL values don't cause sorting issues
            columns_str = ', '.join(export_columns)
            cursor.execute(f"""
                SELECT {columns_str} FROM programs
                ORDER BY
                    CASE WHEN outer_diameter IS NULL THEN 1 ELSE 0 END,
                    outer_diameter,
                    spacer_type,
                    CASE WHEN center_bore IS NULL THEN 1 ELSE 0 END,
                    center_bore,
                    CASE WHEN thickness IS NULL THEN 1 ELSE 0 END,
                    thickness
            """)
            results = cursor.fetchall()

            # Create Excel workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Define standard round sizes
            round_sizes = [
                ('5.75"', 5.75),
                ('6.00"', 6.00),
                ('6.25"', 6.25),
                ('6.50"', 6.50),
                ('7.00"', 7.00),
                ('7.50"', 7.50),
                ('8.00"', 8.00),
                ('9.00"', 9.00),
                ('9.50"', 9.50),
                ('10.00"', 10.00),
                ('13.00"', 13.00),
            ]

            # Group programs by OD
            programs_by_od = {}
            for row in results:
                od = row[2]  # outer_diameter column (index 2: program_number, title, outer_diameter)
                if od not in programs_by_od:
                    programs_by_od[od] = []
                programs_by_od[od].append(row)

            # Create sheet for each round size
            for sheet_name, od_value in round_sizes:
                if od_value in programs_by_od:
                    ws = wb.create_sheet(title=sheet_name)
                    self._write_excel_sheet(ws, programs_by_od[od_value])

            # Create "Other Sizes" sheet for non-standard sizes
            other_ods = [od for od in programs_by_od.keys()
                        if od not in [size[1] for size in round_sizes]]
            if other_ods:
                ws = wb.create_sheet(title="Other Sizes")
                other_programs = []
                # Sort ODs, handling None values (put them last)
                sorted_ods = sorted([od for od in other_ods if od is not None])
                none_ods = [od for od in other_ods if od is None]
                for od in sorted_ods + none_ods:
                    other_programs.extend(programs_by_od[od])
                self._write_excel_sheet(ws, other_programs)

            # Save workbook
            wb.save(filepath)
            conn.close()

            total_sheets = len(wb.sheetnames)
            messagebox.showinfo("Export Complete",
                f"Exported {len(results)} records to {total_sheets} sheets:\n{filepath}")

    def _write_excel_sheet(self, ws, data):
        """Write data to an Excel sheet with formatting"""
        from openpyxl.styles import Font, PatternFill, Alignment

        # Header
        headers = [
            'Program #',
            'Title',
            'OD (in)',
            'Type',
            'CB (mm)',
            'Thickness (in)',
            'Hub (mm)',
            'Hub Height (in)',
            'Counterbore (mm)',
            'CB Depth (in)',
            'Material',
            'Lathe',
            'Status',
            'Confidence',
            'Last Modified'
        ]

        # Write header with formatting
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)

                if value is None:
                    cell.value = ""
                elif isinstance(value, float):
                    # Format floats to 2 decimal places
                    cell.value = round(value, 2)
                    cell.number_format = '0.00'
                else:
                    cell.value = str(value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

    def export_google_sheets(self):
        """Export database to Excel optimized for Google Sheets import"""
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Missing Library",
                "openpyxl is required for Google Sheets export.\n\n"
                "Install it with:\npip install openpyxl")
            return

        from datetime import datetime
        default_filename = f"GCode_GoogleSheets_{datetime.now().strftime('%Y%m%d')}.xlsx"

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Excel Workbook", "*.xlsx"), ("All files", "*.*")],
            title="Export for Google Sheets"
        )

        if filepath:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # Define organized column order - same as Excel export
            export_columns = [
                'program_number', 'title',
                'outer_diameter', 'spacer_type', 'center_bore', 'thickness',
                'hub_diameter', 'hub_height', 'counter_bore_diameter', 'counter_bore_depth',
                'material', 'lathe',
                'validation_status', 'detection_confidence',
                'last_modified'
            ]

            # Build query with ordering: OD, Type, CB, Thickness
            # NULLS LAST ensures NULL values don't cause sorting issues
            columns_str = ', '.join(export_columns)
            cursor.execute(f"""
                SELECT {columns_str} FROM programs
                ORDER BY
                    CASE WHEN outer_diameter IS NULL THEN 1 ELSE 0 END,
                    outer_diameter,
                    spacer_type,
                    CASE WHEN center_bore IS NULL THEN 1 ELSE 0 END,
                    center_bore,
                    CASE WHEN thickness IS NULL THEN 1 ELSE 0 END,
                    thickness
            """)
            results = cursor.fetchall()

            # Create Excel workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Define standard round sizes - combine some to reduce sheet count for Google Sheets
            # Google Sheets has better performance with fewer sheets
            round_size_groups = [
                ('5.75" - 6.00"', [5.75, 6.00]),
                ('6.25"', [6.25]),
                ('6.50"', [6.50]),
                ('7.00"', [7.00]),
                ('7.50"', [7.50]),
                ('8.00"', [8.00]),
                ('9.00" - 9.50"', [9.00, 9.50]),
                ('10.00"', [10.00]),
                ('13.00"', [13.00]),
            ]

            # Group programs by OD
            programs_by_od = {}
            for row in results:
                od = row[2]  # outer_diameter column (index 2: program_number, title, outer_diameter)
                if od not in programs_by_od:
                    programs_by_od[od] = []
                programs_by_od[od].append(row)

            # Create sheet for each round size group
            for sheet_name, od_values in round_size_groups:
                group_programs = []
                for od_value in od_values:
                    if od_value in programs_by_od:
                        group_programs.extend(programs_by_od[od_value])

                if group_programs:
                    ws = wb.create_sheet(title=sheet_name)
                    self._write_google_sheet(ws, group_programs)

            # Create "Other Sizes" sheet for non-standard sizes
            all_standard_ods = []
            for _, od_list in round_size_groups:
                all_standard_ods.extend(od_list)

            other_ods = [od for od in programs_by_od.keys()
                        if od not in all_standard_ods]
            if other_ods:
                ws = wb.create_sheet(title="Other Sizes")
                other_programs = []
                # Sort ODs, handling None values (put them last)
                sorted_ods = sorted([od for od in other_ods if od is not None])
                none_ods = [od for od in other_ods if od is None]
                for od in sorted_ods + none_ods:
                    other_programs.extend(programs_by_od[od])
                self._write_google_sheet(ws, other_programs)

            # Save workbook
            wb.save(filepath)
            conn.close()

            total_sheets = len(wb.sheetnames)
            messagebox.showinfo("Google Sheets Export Complete",
                f"Exported {len(results)} records to {total_sheets} sheets.\n\n"
                f"To import to Google Sheets:\n"
                f"1. Go to sheets.google.com\n"
                f"2. File → Import → Upload\n"
                f"3. Select: {filepath}\n"
                f"4. Import location: Replace spreadsheet\n\n"
                f"Optimized for Google Sheets with combined round sizes.")

    def _write_google_sheet(self, ws, data):
        """Write data to an Excel sheet optimized for Google Sheets"""
        from openpyxl.styles import Font, PatternFill, Alignment

        # Header - same as Excel export
        headers = [
            'Program #', 'Title',
            'OD (in)', 'Type', 'CB (mm)', 'Thickness (in)',
            'Hub (mm)', 'Hub Height (in)', 'Counterbore (mm)', 'CB Depth (in)',
            'Material', 'Lathe',
            'Status', 'Confidence', 'Last Modified'
        ]

        # Write header with Google Sheets-friendly formatting
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            # Google green color for headers
            cell.fill = PatternFill(start_color="34A853", end_color="34A853", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)

                if value is None:
                    cell.value = ""
                elif isinstance(value, float):
                    # Format floats to 2 decimal places
                    cell.value = round(value, 2)
                    cell.number_format = '0.00'
                else:
                    cell.value = str(value)

        # Auto-adjust column widths (Google Sheets respects this)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

    def export_unused_numbers(self):
        """Export a CSV of unused program numbers within standard ranges"""
        import csv

        # Define ranges - full range from 00001 to 99999
        ranges = [
            ("o00001-o09999", 1, 9999),
            ("o10000-o19999", 10000, 19999),
            ("o20000-o29999", 20000, 29999),
            ("o30000-o39999", 30000, 39999),
            ("o40000-o49999", 40000, 49999),
            ("o50000-o59999", 50000, 59999),
            ("o60000-o69999", 60000, 69999),
            ("o70000-o79999", 70000, 79999),
            ("o80000-o89999", 80000, 89999),
            ("o90000-o99999", 90000, 99999),
        ]

        # Get all existing program numbers from database
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT program_number FROM programs")
        existing = set()
        for row in cursor.fetchall():
            prog = row[0]
            # Extract numeric part (handle duplicates like o70000(1))
            match = re.search(r'[oO]?(\d+)', str(prog))
            if match:
                existing.add(int(match.group(1)))
        conn.close()

        # Ask user for save location
        filepath = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            title="Export Unused Program Numbers"
        )

        if not filepath:
            return

        # Generate unused numbers for each range
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["Range", "Program Number", "Available"])

            total_unused = 0
            for range_name, start, end in ranges:
                unused_in_range = []
                for num in range(start, end + 1):
                    if num not in existing:
                        unused_in_range.append(num)

                # Write unused numbers for this range
                for num in unused_in_range:
                    writer.writerow([range_name, f"o{num}", "Yes"])
                    total_unused += 1

        messagebox.showinfo(
            "Export Complete",
            f"Exported {total_unused} unused program numbers to:\n{filepath}\n\n"
            f"These are available numbers within standard OD ranges."
        )

    def find_and_mark_repeats(self):
        """Enhanced duplicate detection with parent/child relationships and classification"""
        import uuid
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Show progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Finding & Classifying Duplicates...")
        progress_window.geometry("700x500")
        progress_window.configure(bg=self.bg_color)

        progress_label = tk.Label(progress_window, text="Analyzing database...",
                                 bg=self.bg_color, fg=self.fg_color,
                                 font=("Arial", 12))
        progress_label.pack(pady=20)

        progress_text = scrolledtext.ScrolledText(progress_window,
                                                 bg=self.input_bg, fg=self.fg_color,
                                                 width=80, height=20)
        progress_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        self.root.update()

        # Get all files from database with full info
        cursor.execute('''
            SELECT program_number, title, spacer_type, outer_diameter, thickness,
                   center_bore, hub_height, hub_diameter, counter_bore_diameter,
                   counter_bore_depth, last_modified, file_path, detection_confidence,
                   validation_status, date_created
            FROM programs
        ''')
        all_files = cursor.fetchall()

        progress_text.insert(tk.END, f"Found {len(all_files)} total files in database.\n")
        progress_text.insert(tk.END, f"Classifying duplicates...\n\n")
        progress_text.see(tk.END)
        self.root.update()

        # Build filename and content maps
        filename_map = {}  # filename -> list of files
        content_map = {}   # (title+dims) -> list of files

        for file_data in all_files:
            prog_num, title, stype, od, thick, cb, hub_h, hub_d, cb_d, cb_dep, modified, fpath, confidence, val_status, created = file_data

            filename = os.path.basename(fpath).lower() if fpath else ""

            # Add to filename map
            if filename:
                if filename not in filename_map:
                    filename_map[filename] = []
                filename_map[filename].append(file_data)

            # Create content key
            content_key = (
                title.strip().upper() if title else "",
                stype,
                round(od, 3) if od else None,
                round(thick, 3) if thick else None,
                round(cb, 3) if cb else None,
                round(hub_h, 3) if hub_h else None,
                round(hub_d, 3) if hub_d else None,
                round(cb_d, 3) if cb_d else None,
                round(cb_dep, 3) if cb_dep else None
            )

            if content_key not in content_map:
                content_map[content_key] = []
            content_map[content_key].append(file_data)

        # Classify duplicates
        solid_dups = 0
        name_collisions = 0
        content_dups = 0

        # Process SOLID DUPLICATES (same filename AND same content)
        for filename, files in filename_map.items():
            if len(files) > 1:
                # Check if they also have same content
                first_content = (files[0][1], files[0][2], files[0][3], files[0][4], files[0][5])
                all_same_content = all(
                    (f[1], f[2], f[3], f[4], f[5]) == first_content for f in files
                )

                if all_same_content:
                    # SOLID DUPLICATE - same filename, same content
                    group_id = str(uuid.uuid4())[:8]

                    # Choose parent: oldest file with best validation
                    def sort_key(f):
                        # Priority: 1) Validation status, 2) Confidence, 3) Oldest date
                        val_priority = {'PASS': 0, 'WARNING': 1, 'DIMENSIONAL': 2, 'BORE_WARNING': 3, 'CRITICAL': 4, 'REPEAT': 5}
                        conf_priority = {'HIGH': 0, 'MEDIUM': 1, 'LOW': 2}
                        date = f[14] if f[14] else f[10] if f[10] else '9999-12-31'  # created or modified
                        return (val_priority.get(f[13], 9), conf_priority.get(f[12], 9), date)

                    sorted_files = sorted(files, key=sort_key)
                    parent = sorted_files[0]
                    children = sorted_files[1:]

                    progress_text.insert(tk.END, f"SOLID DUP: {filename}\n")
                    progress_text.insert(tk.END, f"  ✓ Parent: {parent[0]} (oldest with best validation)\n")

                    # Mark children as REPEAT
                    for child in children:
                        cursor.execute('''
                            UPDATE programs
                            SET validation_status = 'REPEAT',
                                duplicate_type = 'SOLID',
                                parent_file = ?,
                                duplicate_group = ?
                            WHERE program_number = ?
                        ''', (parent[0], group_id, child[0]))
                        progress_text.insert(tk.END, f"  ✗ Child: {child[0]}\n")
                        solid_dups += 1

                    progress_text.insert(tk.END, "\n")
                    progress_text.see(tk.END)
                    self.root.update()
                else:
                    # NAME COLLISION - same filename, different content
                    group_id = str(uuid.uuid4())[:8]
                    progress_text.insert(tk.END, f"NAME COLLISION: {filename}\n")
                    for f in files:
                        cursor.execute('''
                            UPDATE programs
                            SET duplicate_type = 'NAME_COLLISION',
                                duplicate_group = ?
                            WHERE program_number = ?
                        ''', (group_id, f[0]))
                        progress_text.insert(tk.END, f"  ! {f[0]} - Different content, needs rename\n")
                        name_collisions += 1
                    progress_text.insert(tk.END, "\n")
                    progress_text.see(tk.END)
                    self.root.update()

        # Process CONTENT DUPLICATES (same content, different filenames)
        for content_key, files in content_map.items():
            if len(files) > 1:
                # Get unique filenames in this group
                filenames = set(os.path.basename(f[11]).lower() if f[11] else "" for f in files)

                if len(filenames) > 1:  # Different filenames but same content
                    group_id = str(uuid.uuid4())[:8]

                    # Choose parent: oldest file with best validation
                    def sort_key(f):
                        val_priority = {'PASS': 0, 'WARNING': 1, 'DIMENSIONAL': 2, 'BORE_WARNING': 3, 'CRITICAL': 4, 'REPEAT': 5}
                        conf_priority = {'HIGH': 0, 'MEDIUM': 1, 'LOW': 2}
                        date = f[14] if f[14] else f[10] if f[10] else '9999-12-31'
                        return (val_priority.get(f[13], 9), conf_priority.get(f[12], 9), date)

                    sorted_files = sorted(files, key=sort_key)
                    parent = sorted_files[0]
                    children = sorted_files[1:]

                    progress_text.insert(tk.END, f"CONTENT DUP: {content_key[0][:50]}\n")
                    progress_text.insert(tk.END, f"  ✓ Parent: {parent[0]} (oldest with best validation)\n")

                    for child in children:
                        cursor.execute('''
                            UPDATE programs
                            SET validation_status = 'REPEAT',
                                duplicate_type = 'CONTENT_DUP',
                                parent_file = ?,
                                duplicate_group = ?
                            WHERE program_number = ?
                        ''', (parent[0], group_id, child[0]))
                        progress_text.insert(tk.END, f"  ✗ Child: {child[0]}\n")
                        content_dups += 1

                    progress_text.insert(tk.END, "\n")
                    progress_text.see(tk.END)
                    self.root.update()

        conn.commit()
        conn.close()

        # Show results
        progress_label.config(text="Complete!")
        progress_text.insert(tk.END, f"{'='*70}\n")
        progress_text.insert(tk.END, f"SOLID Duplicates (same file+content): {solid_dups}\n")
        progress_text.insert(tk.END, f"NAME Collisions (same name, diff content): {name_collisions}\n")
        progress_text.insert(tk.END, f"CONTENT Duplicates (diff name, same content): {content_dups}\n")
        progress_text.insert(tk.END, f"\nTotal duplicates found: {solid_dups + name_collisions + content_dups}\n")
        progress_text.see(tk.END)

        close_btn = tk.Button(progress_window, text="Close",
                             command=progress_window.destroy,
                             bg=self.button_bg, fg=self.fg_color,
                             font=("Arial", 10, "bold"))
        close_btn.pack(pady=10)

    def delete_duplicates(self):
        """Delete all duplicate files (REPEAT status) keeping only parent files"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Count duplicates to be deleted
        cursor.execute("""
            SELECT COUNT(*) FROM programs
            WHERE validation_status = 'REPEAT'
        """)
        dup_count = cursor.fetchone()[0]

        if dup_count == 0:
            conn.close()
            messagebox.showinfo("No Duplicates",
                "No duplicate files found.\n\n"
                "Use 'Find Repeats' first to identify duplicates.")
            return

        # Get details about what will be deleted
        cursor.execute("""
            SELECT program_number, duplicate_type, parent_file
            FROM programs
            WHERE validation_status = 'REPEAT'
            ORDER BY duplicate_type, parent_file
        """)
        duplicates = cursor.fetchall()
        conn.close()

        # Confirmation dialog
        result = messagebox.askyesno(
            "Delete All Duplicates",
            f"This will DELETE {dup_count} duplicate record(s) from the DATABASE.\n\n"
            f"✓ Parent files (best versions) will be KEPT\n"
            f"✗ Child files (duplicates) will be DELETED\n\n"
            f"⚠️  The actual files on disk will NOT be deleted.\n"
            f"⚠️  Only database records will be removed.\n\n"
            f"Do you want to continue?",
            icon='warning'
        )

        if not result:
            return

        # Second confirmation - type "DELETE DUPLICATES"
        confirm_window = tk.Toplevel(self.root)
        confirm_window.title("Final Confirmation")
        confirm_window.geometry("500x250")
        confirm_window.configure(bg=self.bg_color)
        confirm_window.transient(self.root)
        confirm_window.grab_set()

        tk.Label(confirm_window,
                text=f"⚠️  You are about to delete {dup_count} duplicate records",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 12, "bold")).pack(pady=10)

        tk.Label(confirm_window,
                text='Type "DELETE DUPLICATES" to confirm:',
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 10)).pack(pady=10)

        confirm_entry = tk.Entry(confirm_window, bg=self.input_bg, fg=self.fg_color,
                                font=("Arial", 12), width=25)
        confirm_entry.pack(pady=10)
        confirm_entry.focus()

        confirmed = [False]

        def check_confirmation():
            if confirm_entry.get() == "DELETE DUPLICATES":
                confirmed[0] = True
                confirm_window.destroy()
            else:
                messagebox.showerror("Invalid", 'You must type "DELETE DUPLICATES" exactly to confirm.')

        def cancel_delete():
            confirm_window.destroy()

        btn_frame = tk.Frame(confirm_window, bg=self.bg_color)
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="Confirm", command=check_confirmation,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=10).pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="Cancel", command=cancel_delete,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=10).pack(side=tk.LEFT, padx=5)

        # Bind Enter key to confirm
        confirm_entry.bind('<Return>', lambda e: check_confirmation())

        # Wait for window to close
        self.root.wait_window(confirm_window)

        if not confirmed[0]:
            return

        # Create database backup before deleting
        if not self.backup_database():
            messagebox.showerror("Backup Failed",
                "Database backup failed. Delete operation canceled for safety.\n\n"
                "Please check the error and try again.")
            return

        # Show progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Deleting Duplicates")
        progress_window.geometry("700x500")
        progress_window.configure(bg=self.bg_color)

        tk.Label(progress_window,
                text="Deleting duplicate records...",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 12, "bold")).pack(pady=10)

        progress_text = scrolledtext.ScrolledText(progress_window,
                                                  bg=self.input_bg, fg=self.fg_color,
                                                  font=("Courier", 9),
                                                  wrap=tk.WORD)
        progress_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Group duplicates by type for reporting
        by_type = {}
        for prog_num, dup_type, parent in duplicates:
            if dup_type not in by_type:
                by_type[dup_type] = []
            by_type[dup_type].append((prog_num, parent))

        # Show what's being deleted
        progress_text.insert(tk.END, f"Deleting {dup_count} duplicate records:\n\n")

        for dup_type, items in by_type.items():
            progress_text.insert(tk.END, f"--- {dup_type or 'Unknown Type'} ({len(items)} records) ---\n")
            for prog_num, parent in items[:10]:  # Show first 10
                progress_text.insert(tk.END, f"  DELETE: {prog_num} (parent: {parent or 'N/A'})\n")
            if len(items) > 10:
                progress_text.insert(tk.END, f"  ... and {len(items) - 10} more\n")
            progress_text.insert(tk.END, "\n")
            self.root.update()

        # Delete from database
        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            cursor.execute("""
                DELETE FROM programs
                WHERE validation_status = 'REPEAT'
            """)

            deleted_count = cursor.rowcount
            conn.commit()
            conn.close()

            progress_text.insert(tk.END, f"{'='*60}\n")
            progress_text.insert(tk.END, f"✓ Successfully deleted {deleted_count} duplicate record(s)\n")
            progress_text.insert(tk.END, f"✓ Parent files have been preserved\n")
            progress_text.insert(tk.END, f"⚠️  Files on disk were NOT deleted\n")

            # Refresh the display
            self.refresh_filter_values()
            self.refresh_results()

        except Exception as e:
            progress_text.insert(tk.END, f"\n\n❌ ERROR: {str(e)}\n")
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")

        # Add close button
        close_btn = tk.Button(progress_window, text="Close",
                             command=progress_window.destroy,
                             bg=self.button_bg, fg=self.fg_color,
                             font=("Arial", 10, "bold"))
        close_btn.pack(pady=10)

        # Refresh the display
        self.refresh_results()

    def compare_files(self):
        """Compare selected files side-by-side with difference highlighting"""
        # Get selected items from treeview
        selected_items = self.tree.selection()

        if len(selected_items) < 2:
            messagebox.showwarning("Selection Required",
                "Please select 2 or more files to compare.\n\n"
                "Hold Ctrl and click to select multiple files.")
            return

        if len(selected_items) > 4:
            result = messagebox.askyesno("Many Files Selected",
                f"You selected {len(selected_items)} files.\n\n"
                f"Comparing many files may be difficult to view.\n\n"
                f"Continue anyway?")
            if not result:
                return

        # Extract program numbers
        program_numbers = []
        for item in selected_items:
            values = self.tree.item(item)['values']
            if values:
                program_numbers.append(values[0])

        # Get file details from database
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        placeholders = ','.join('?' * len(program_numbers))
        cursor.execute(f'''
            SELECT program_number, title, file_path, spacer_type, outer_diameter, thickness,
                   center_bore, hub_height, hub_diameter, counter_bore_diameter,
                   counter_bore_depth, validation_status, detection_confidence,
                   duplicate_type, parent_file, last_modified, date_created
            FROM programs
            WHERE program_number IN ({placeholders})
        ''', program_numbers)

        files_data = cursor.fetchall()
        conn.close()

        if len(files_data) < 2:
            messagebox.showerror("Error", "Could not retrieve file information from database.")
            return

        # Open comparison window
        FileComparisonWindow(self.root, files_data, self.bg_color, self.fg_color,
                           self.input_bg, self.button_bg, self.refresh_results, self)

    def rename_duplicate_files(self):
        """Rename physical files with duplicate filenames and update database - FILTERED VIEW ONLY"""
        # Get currently displayed items (respects filters)
        displayed_items = self.tree.get_children()

        if not displayed_items:
            messagebox.showwarning("No Files",
                "No files in current view.\n\n"
                "Apply filters to show the files you want to rename.")
            return

        # Get program numbers from filtered view
        filtered_program_numbers = []
        for item in displayed_items:
            values = self.tree.item(item)['values']
            if values:
                filtered_program_numbers.append(values[0])

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Find files with duplicate filenames - ONLY IN FILTERED VIEW
        placeholders = ','.join('?' * len(filtered_program_numbers))
        cursor.execute(f"""
            SELECT program_number, file_path, duplicate_type, parent_file, duplicate_group,
                   validation_status
            FROM programs
            WHERE file_path IS NOT NULL
              AND program_number IN ({placeholders})
            ORDER BY file_path
        """, filtered_program_numbers)
        all_files = cursor.fetchall()

        # Group files by filename (case-insensitive)
        filename_groups = {}
        for prog_num, file_path, dup_type, parent, dup_group, val_status in all_files:
            if not file_path or not os.path.exists(file_path):
                continue

            basename = os.path.basename(file_path).lower()
            if basename not in filename_groups:
                filename_groups[basename] = []
            filename_groups[basename].append((prog_num, file_path, dup_type, parent, dup_group, val_status))

        # Find files that need renaming (duplicates)
        duplicates_to_rename = []
        for basename, files in filename_groups.items():
            if len(files) > 1:
                # Sort to determine parent (keep oldest with best validation)
                def sort_key(f):
                    val_priority = {'PASS': 0, 'WARNING': 1, 'DIMENSIONAL': 2, 'BORE_WARNING': 3, 'CRITICAL': 4, 'REPEAT': 5}
                    return val_priority.get(f[5], 9)

                sorted_files = sorted(files, key=sort_key)
                parent_file = sorted_files[0]
                child_files = sorted_files[1:]

                for child in child_files:
                    duplicates_to_rename.append({
                        'prog_num': child[0],
                        'old_path': child[1],
                        'parent_prog': parent_file[0],
                        'basename': basename
                    })

        conn.close()

        if not duplicates_to_rename:
            messagebox.showinfo("No Duplicates",
                "No duplicate filenames found.\n\n"
                "All files have unique names.")
            return

        # Show mode selection: Preview or Execute
        mode_choice = [None]

        mode_window = tk.Toplevel(self.root)
        mode_window.title("Rename Duplicate Files")
        mode_window.geometry("500x200")
        mode_window.configure(bg=self.bg_color)
        mode_window.transient(self.root)
        mode_window.grab_set()

        tk.Label(mode_window,
                text=f"Found {len(duplicates_to_rename)} file(s) with duplicate names",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 12, "bold")).pack(pady=15)

        tk.Label(mode_window,
                text="This will assign new O-numbers (o59000+) to duplicate files.",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 10)).pack(pady=5)

        tk.Label(mode_window,
                text="Updates: physical file, internal O-number, and database.",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 10)).pack(pady=5)

        tk.Label(mode_window,
                text="Choose mode:",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 10)).pack(pady=5)

        def preview_mode():
            mode_choice[0] = "preview"
            mode_window.destroy()

        def execute_mode():
            mode_choice[0] = "execute"
            mode_window.destroy()

        def cancel_mode():
            mode_window.destroy()

        btn_frame = tk.Frame(mode_window, bg=self.bg_color)
        btn_frame.pack(pady=15)

        tk.Button(btn_frame, text="🔍 Preview (Dry Run)",
                 command=preview_mode,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=20).pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="▶️ Execute (Rename Files)",
                 command=execute_mode,
                 bg=self.accent_color, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=20).pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="❌ Cancel",
                 command=cancel_mode,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10), width=10).pack(side=tk.LEFT, padx=5)

        self.root.wait_window(mode_window)

        if mode_choice[0] is None:
            return

        dry_run = (mode_choice[0] == "preview")

        # Create backup if executing
        if not dry_run:
            if not self.backup_database():
                messagebox.showerror("Backup Failed",
                    "Database backup failed. Rename operation canceled for safety.")
                return

        # Show progress window
        progress_window = tk.Toplevel(self.root)
        if dry_run:
            progress_window.title("Rename Preview (Dry Run)")
        else:
            progress_window.title("Assigning New O-Numbers to Duplicates")
        progress_window.geometry("900x600")
        progress_window.configure(bg=self.bg_color)

        tk.Label(progress_window,
                text="Preview - Files that WOULD be assigned new O-numbers:" if dry_run else "Assigning new O-numbers...",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 12, "bold")).pack(pady=10)

        progress_text = scrolledtext.ScrolledText(progress_window,
                                                  bg=self.input_bg, fg=self.fg_color,
                                                  font=("Courier", 9),
                                                  wrap=tk.WORD)
        progress_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        if dry_run:
            progress_text.insert(tk.END, f"=== DRY RUN MODE - NO CHANGES WILL BE MADE ===\n\n")

        progress_text.insert(tk.END, f"Found {len(duplicates_to_rename)} duplicate file(s) in filtered view.\n")
        progress_text.insert(tk.END, f"Assigning available O-numbers (avoiding reserved 00000-01000)...\n\n")

        renamed_count = 0
        errors = 0

        # Get all existing program numbers to find available O-numbers
        conn_temp = sqlite3.connect(self.db_path)
        cursor_temp = conn_temp.cursor()
        cursor_temp.execute("SELECT program_number FROM programs")
        existing_program_numbers = set()
        for row in cursor_temp.fetchall():
            prog = row[0]
            # Extract numeric part
            match = re.search(r'o?(\d+)', prog, re.IGNORECASE)
            if match:
                existing_program_numbers.add(int(match.group(1)))
        conn_temp.close()

        # Reserved ranges (DO NOT USE for production files)
        reserved_ranges = [
            (0, 1000)  # Machine needed files (00000-01000)
        ]

        # Helper function to check if number is in reserved range
        def is_reserved(num):
            for reserved_start, reserved_end in reserved_ranges:
                if reserved_start <= num <= reserved_end:
                    return True
            return False

        # Find available O-numbers starting from 1001 (avoiding reserved 0-1000)
        def find_next_available_onumber(start=1001):
            """Find the next available O-number, skipping reserved ranges"""
            current = start
            while current < 100000:
                if current not in existing_program_numbers and not is_reserved(current):
                    existing_program_numbers.add(current)  # Reserve it
                    return current
                current += 1
            return None  # No available numbers found

        # Track used names to avoid collisions
        used_names = set()

        for idx, dup_info in enumerate(duplicates_to_rename, 1):
            prog_num = dup_info['prog_num']
            old_path = dup_info['old_path']
            parent_prog = dup_info['parent_prog']

            # Get available O-number
            new_onumber = find_next_available_onumber()

            # Check if we ran out of numbers
            if new_onumber is None:
                progress_text.insert(tk.END, f"[{idx}] ❌ ERROR: No available O-numbers in entire range (00000-99999)!\n\n")
                errors += 1
                continue

            # Format with leading zeros (always 5 digits: o01057, not o1057)
            new_prog_num = f"o{new_onumber:05d}"

            # Generate new filename
            old_dir = os.path.dirname(old_path)
            old_basename = os.path.basename(old_path)
            name_without_ext, ext = os.path.splitext(old_basename)

            # New filename with proper O-number (5 digits with leading zeros)
            new_basename = f"o{new_onumber:05d}{ext}"
            new_path = os.path.join(old_dir, new_basename)

            # Safety check - ensure new path doesn't exist
            if os.path.exists(new_path):
                progress_text.insert(tk.END, f"[{idx}] ⚠️  SKIP: {new_basename} already exists\n\n")
                continue

            used_names.add(new_basename.lower())

            try:
                if dry_run:
                    progress_text.insert(tk.END, f"[{idx}] WOULD RENAME:\n")
                    progress_text.insert(tk.END, f"    Old File: {old_basename}\n")
                    progress_text.insert(tk.END, f"    New File: {new_basename}\n")
                    progress_text.insert(tk.END, f"    Old Program: {prog_num} (parent: {parent_prog})\n")
                    progress_text.insert(tk.END, f"    New Program: {new_prog_num}\n\n")
                    renamed_count += 1
                else:
                    # Actually rename the physical file
                    os.rename(old_path, new_path)

                    # Update internal G-code program number
                    if not self.update_gcode_program_number(new_path, new_prog_num):
                        progress_text.insert(tk.END, f"    ⚠️  Warning: Could not update internal O-number\n")

                    # Update database with new path AND new program number
                    conn = sqlite3.connect(self.db_path)
                    cursor = conn.cursor()
                    cursor.execute("""
                        UPDATE programs
                        SET file_path = ?,
                            program_number = ?
                        WHERE program_number = ?
                    """, (new_path, new_prog_num, prog_num))
                    conn.commit()
                    conn.close()

                    progress_text.insert(tk.END, f"[{idx}] ✓ RENAMED:\n")
                    progress_text.insert(tk.END, f"    Old File: {old_basename}\n")
                    progress_text.insert(tk.END, f"    New File: {new_basename}\n")
                    progress_text.insert(tk.END, f"    Old Program: {prog_num}\n")
                    progress_text.insert(tk.END, f"    New Program: {new_prog_num}\n\n")
                    renamed_count += 1

                progress_text.see(tk.END)
                self.root.update()

            except Exception as e:
                progress_text.insert(tk.END, f"[{idx}] ❌ ERROR: {prog_num}\n")
                progress_text.insert(tk.END, f"    {str(e)}\n\n")
                errors += 1
                progress_text.see(tk.END)
                self.root.update()

        # Summary
        progress_text.insert(tk.END, f"\n{'='*70}\n")
        if dry_run:
            progress_text.insert(tk.END, f"PREVIEW COMPLETE - No actual changes were made.\n")
            progress_text.insert(tk.END, f"Would assign new O-numbers to: {renamed_count} file(s)\n")
            progress_text.insert(tk.END, f"\nClick 'Execute' to apply these changes.\n")
        else:
            progress_text.insert(tk.END, f"OPERATION COMPLETE!\n")
            progress_text.insert(tk.END, f"Successfully assigned new O-numbers to: {renamed_count} file(s)\n")
            progress_text.insert(tk.END, f"Errors: {errors}\n")
            progress_text.insert(tk.END, f"\nAll files now have unique O-numbers (o59000+)\n")
            progress_text.insert(tk.END, f"Physical files renamed, internal O-numbers updated, database updated.\n")

            # Refresh display
            self.refresh_filter_values()
            self.refresh_results()

        progress_text.see(tk.END)

        # Close button
        close_btn = tk.Button(progress_window, text="Close",
                             command=progress_window.destroy,
                             bg=self.button_bg, fg=self.fg_color,
                             font=("Arial", 10, "bold"))
        close_btn.pack(pady=10)

    def organize_files_by_od(self):
        """Copy all database files to organized folder structure by OD"""
        import shutil

        # Ask user for destination folder
        dest_folder = filedialog.askdirectory(title="Select Destination Folder for Organized Files")

        if not dest_folder:
            return

        # Show progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Organizing Files by OD...")
        progress_window.geometry("600x400")
        progress_window.configure(bg=self.bg_color)

        progress_label = tk.Label(progress_window, text="Organizing files...",
                                 bg=self.bg_color, fg=self.fg_color,
                                 font=("Arial", 12))
        progress_label.pack(pady=20)

        progress_text = scrolledtext.ScrolledText(progress_window,
                                                 bg=self.input_bg, fg=self.fg_color,
                                                 width=70, height=15)
        progress_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        self.root.update()

        # Get all files from database with OD info
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT program_number, file_path, outer_diameter
            FROM programs
            WHERE file_path IS NOT NULL
        ''')
        all_files = cursor.fetchall()
        conn.close()

        progress_text.insert(tk.END, f"Found {len(all_files)} files in database.\n")
        progress_text.insert(tk.END, f"Destination: {dest_folder}\n\n")
        progress_text.see(tk.END)
        self.root.update()

        # OD folder mapping (round to standard sizes)
        od_folders = {
            5.75: "5.75 Round",
            6.00: "6.00 Round",
            6.25: "6.25 Round",
            6.50: "6.50 Round",
            7.00: "7.00 Round",
            7.50: "7.50 Round",
            8.00: "8.00 Round",
            8.50: "8.50 Round",
            9.50: "9.50 Round",
            10.25: "10.25 Round",
            10.50: "10.50 Round",
            13.00: "13.00 Round"
        }

        copied = 0
        skipped = 0
        errors = 0

        for idx, (prog_num, file_path, od) in enumerate(all_files, 1):
            progress_label.config(text=f"Processing {idx}/{len(all_files)}: {prog_num}")
            self.root.update()

            # Check if file exists
            if not os.path.exists(file_path):
                progress_text.insert(tk.END, f"SKIP: {prog_num} - file not found: {file_path}\n")
                progress_text.see(tk.END)
                skipped += 1
                continue

            # Determine OD folder
            if od is None:
                folder_name = "Unknown OD"
            else:
                # Find closest standard OD
                closest_od = min(od_folders.keys(), key=lambda x: abs(x - od))
                if abs(closest_od - od) <= 0.1:  # Within tolerance
                    folder_name = od_folders[closest_od]
                else:
                    folder_name = f"Other ({od:.2f})"

            # Create destination folder if needed
            od_folder_path = os.path.join(dest_folder, folder_name)
            os.makedirs(od_folder_path, exist_ok=True)

            # Copy file
            filename = os.path.basename(file_path)
            dest_path = os.path.join(od_folder_path, filename)

            try:
                shutil.copy2(file_path, dest_path)
                progress_text.insert(tk.END, f"COPY: {prog_num} -> {folder_name}/{filename}\n")
                progress_text.see(tk.END)
                copied += 1
            except Exception as e:
                progress_text.insert(tk.END, f"ERROR: {prog_num} - {str(e)[:100]}\n")
                progress_text.see(tk.END)
                errors += 1

        # Show results
        progress_label.config(text="Complete!")
        progress_text.insert(tk.END, f"\n{'='*60}\n")
        progress_text.insert(tk.END, f"Total files: {len(all_files)}\n")
        progress_text.insert(tk.END, f"Copied: {copied}\n")
        progress_text.insert(tk.END, f"Skipped: {skipped}\n")
        progress_text.insert(tk.END, f"Errors: {errors}\n")
        progress_text.see(tk.END)

        close_btn = tk.Button(progress_window, text="Close",
                             command=progress_window.destroy,
                             bg=self.button_bg, fg=self.fg_color,
                             font=("Arial", 10, "bold"))
        close_btn.pack(pady=10)

    def copy_filtered_view(self):
        """Copy currently filtered/displayed files to folder with OD subfolders and auto-rename"""
        import shutil

        # Get currently displayed items from treeview
        displayed_items = self.tree.get_children()
        if not displayed_items:
            messagebox.showwarning("No Results", "No files in current view to copy.\n\nPlease search/filter first.")
            return

        # Ask user: Preview or Execute?
        preview_window = tk.Toplevel(self.root)
        preview_window.title("Copy Mode")
        preview_window.geometry("400x180")
        preview_window.configure(bg=self.bg_color)

        tk.Label(preview_window,
                text=f"Ready to copy {len(displayed_items)} files",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 12, "bold")).pack(pady=15)

        tk.Label(preview_window,
                text="Choose operation mode:",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 10)).pack(pady=5)

        mode_choice = [None]  # Use list to modify in nested function

        def preview_mode():
            mode_choice[0] = "preview"
            preview_window.destroy()

        def execute_mode():
            mode_choice[0] = "execute"
            preview_window.destroy()

        btn_frame = tk.Frame(preview_window, bg=self.bg_color)
        btn_frame.pack(pady=15)

        tk.Button(btn_frame, text="🔍 Preview (Dry Run)", command=preview_mode,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=18, height=2).pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="▶️ Execute (Copy Files)", command=execute_mode,
                 bg=self.accent_color, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=18, height=2).pack(side=tk.LEFT, padx=5)

        # Wait for user choice
        self.root.wait_window(preview_window)

        if not mode_choice[0]:
            return  # User closed window

        dry_run = (mode_choice[0] == "preview")

        # Create database backup before executing (not for preview)
        if not dry_run:
            if not self.backup_database():
                messagebox.showerror("Backup Failed",
                    "Database backup failed. Operation canceled for safety.\n\n"
                    "Please check the error and try again.")
                return

        # Ask user for destination folder
        dest_folder = filedialog.askdirectory(title="Select Destination Folder for Filtered Files")
        if not dest_folder:
            return

        # Show progress window
        progress_window = tk.Toplevel(self.root)
        title_text = "Preview: Copy Filtered View" if dry_run else "Copying Filtered View..."
        progress_window.title(title_text)
        progress_window.geometry("700x500")
        progress_window.configure(bg=self.bg_color)

        label_text = "Preview Mode (No files will be copied)" if dry_run else "Copying files..."
        progress_label = tk.Label(progress_window, text=label_text,
                                 bg=self.bg_color, fg=self.fg_color,
                                 font=("Arial", 12))
        progress_label.pack(pady=20)

        progress_text = scrolledtext.ScrolledText(progress_window,
                                                 bg=self.input_bg, fg=self.fg_color,
                                                 width=80, height=20)
        progress_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        self.root.update()

        # Get file info from database for displayed items
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Extract program numbers from treeview
        program_numbers = []
        for item in displayed_items:
            values = self.tree.item(item)['values']
            if values:
                program_numbers.append(values[0])  # First column is Program #

        progress_text.insert(tk.END, f"Found {len(program_numbers)} files in current view.\n")
        progress_text.insert(tk.END, f"Destination: {dest_folder}\n\n")
        progress_text.see(tk.END)
        self.root.update()

        # Get file details from database
        placeholders = ','.join('?' * len(program_numbers))
        cursor.execute(f'''
            SELECT program_number, file_path, outer_diameter
            FROM programs
            WHERE program_number IN ({placeholders})
        ''', program_numbers)
        files_to_copy = cursor.fetchall()
        conn.close()

        # OD folder mapping
        od_folders = {
            5.75: "5.75 Round",
            6.00: "6.00 Round",
            6.25: "6.25 Round",
            6.50: "6.50 Round",
            7.00: "7.00 Round",
            7.50: "7.50 Round",
            8.00: "8.00 Round",
            8.50: "8.50 Round",
            9.50: "9.50 Round",
            10.25: "10.25 Round",
            10.50: "10.50 Round",
            13.00: "13.00 Round"
        }

        copied = 0
        skipped = 0
        errors = 0
        renamed = 0

        # Track filenames to handle collisions with auto-rename
        filename_tracker = {}  # {(folder, basename): occurrence_count}

        for idx, (prog_num, file_path, od) in enumerate(files_to_copy, 1):
            progress_label.config(text=f"Processing {idx}/{len(files_to_copy)}: {prog_num}")
            self.root.update()

            # Check if file exists
            if not file_path:
                progress_text.insert(tk.END, f"⚠️  SKIP: {prog_num} - no file path in database\n")
                progress_text.see(tk.END)
                skipped += 1
                continue

            if not os.path.exists(file_path):
                progress_text.insert(tk.END, f"⚠️  SKIP: {prog_num} - file not found at: {file_path}\n")
                progress_text.see(tk.END)
                skipped += 1
                continue

            # Determine OD folder
            if od is None:
                folder_name = "Unknown OD"
            else:
                closest_od = min(od_folders.keys(), key=lambda x: abs(x - od))
                if abs(closest_od - od) <= 0.1:
                    folder_name = od_folders[closest_od]
                else:
                    folder_name = f"Other ({od:.2f})"

            # Create destination folder
            od_folder_path = os.path.join(dest_folder, folder_name)
            os.makedirs(od_folder_path, exist_ok=True)

            # Get filename and handle collisions
            base_filename = os.path.basename(file_path)
            name_without_ext, ext = os.path.splitext(base_filename)

            # Check for collision and auto-rename with (1), (2), etc.
            folder_file_key = (od_folder_path, base_filename)
            if folder_file_key in filename_tracker:
                # Collision detected - add suffix
                filename_tracker[folder_file_key] += 1
                occurrence = filename_tracker[folder_file_key]
                new_filename = f"{name_without_ext}({occurrence}){ext}"
                progress_text.insert(tk.END, f"COLLISION: {base_filename} -> {new_filename}\n")
                renamed += 1
            else:
                # First occurrence
                filename_tracker[folder_file_key] = 1
                new_filename = base_filename

            dest_path = os.path.join(od_folder_path, new_filename)

            try:
                if dry_run:
                    # Preview mode - just show what would happen
                    progress_text.insert(tk.END, f"WOULD COPY: {prog_num} -> {folder_name}/{new_filename}\n")
                    progress_text.see(tk.END)
                    copied += 1
                else:
                    # Actually copy the file
                    shutil.copy2(file_path, dest_path)
                    progress_text.insert(tk.END, f"COPIED: {prog_num} -> {folder_name}/{new_filename}\n")
                    progress_text.see(tk.END)
                    copied += 1
            except Exception as e:
                progress_text.insert(tk.END, f"ERROR: {prog_num} - {str(e)[:100]}\n")
                progress_text.see(tk.END)
                errors += 1

        # Show results
        if dry_run:
            progress_label.config(text="Preview Complete (No files copied)")
            progress_text.insert(tk.END, f"\n{'='*70}\n")
            progress_text.insert(tk.END, f"PREVIEW MODE - No files were actually copied\n")
            progress_text.insert(tk.END, f"Total files: {len(files_to_copy)}\n")
            progress_text.insert(tk.END, f"Would copy: {copied}\n")
            progress_text.insert(tk.END, f"Would auto-rename (collisions): {renamed}\n")
            progress_text.insert(tk.END, f"Would skip: {skipped}\n")
            progress_text.insert(tk.END, f"Potential errors: {errors}\n")
        else:
            progress_label.config(text="Complete!")
            progress_text.insert(tk.END, f"\n{'='*70}\n")
            progress_text.insert(tk.END, f"Total files: {len(files_to_copy)}\n")
            progress_text.insert(tk.END, f"Copied: {copied}\n")
            progress_text.insert(tk.END, f"Auto-renamed (collisions): {renamed}\n")
            progress_text.insert(tk.END, f"Skipped: {skipped}\n")
            progress_text.insert(tk.END, f"Errors: {errors}\n")

        if skipped > 0:
            progress_text.insert(tk.END, f"\n⚠️  NOTE: {skipped} file(s) were skipped because the file paths in the database\n")
            progress_text.insert(tk.END, f"    are invalid or the files no longer exist at those locations.\n")
            progress_text.insert(tk.END, f"    You may need to use 'Scan Folder' to update the database with current file locations.\n")

        progress_text.see(tk.END)

        close_btn = tk.Button(progress_window, text="Close",
                             command=progress_window.destroy,
                             bg=self.button_bg, fg=self.fg_color,
                             font=("Arial", 10, "bold"))
        close_btn.pack(pady=10)

    def delete_filtered_view(self):
        """Delete currently filtered/displayed files from database (NOT from filesystem)"""
        # Get currently displayed items from treeview
        displayed_items = self.tree.get_children()
        if not displayed_items:
            messagebox.showwarning("No Results", "No files in current view to delete.\n\nPlease search/filter first.")
            return

        # Extract program numbers and details
        program_numbers = []
        program_details = []  # For preview
        for item in displayed_items:
            values = self.tree.item(item)['values']
            if values:
                program_numbers.append(values[0])  # First column is Program #
                # Store prog_num, filename, OD for preview
                prog_num = values[0]
                filename = os.path.basename(values[1]) if len(values) > 1 and values[1] else "Unknown"
                od = values[6] if len(values) > 6 else "N/A"
                program_details.append((prog_num, filename, od))

        count = len(program_numbers)

        # Ask user: Preview or Execute?
        mode_choice = [None]  # Use list to modify in nested function

        mode_window = tk.Toplevel(self.root)
        mode_window.title("Delete Mode Selection")
        mode_window.geometry("450x180")
        mode_window.configure(bg=self.bg_color)
        mode_window.transient(self.root)
        mode_window.grab_set()

        tk.Label(mode_window,
                text=f"Delete {count} record(s) from database",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 12, "bold")).pack(pady=15)

        tk.Label(mode_window,
                text="Choose mode:",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 10)).pack(pady=5)

        def preview_mode():
            mode_choice[0] = "preview"
            mode_window.destroy()

        def execute_mode():
            mode_choice[0] = "execute"
            mode_window.destroy()

        def cancel_mode():
            mode_window.destroy()

        btn_frame = tk.Frame(mode_window, bg=self.bg_color)
        btn_frame.pack(pady=15)

        tk.Button(btn_frame, text="🔍 Preview (Dry Run)",
                 command=preview_mode,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=20).pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="▶️ Execute (Delete Records)",
                 command=execute_mode,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=20).pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="❌ Cancel",
                 command=cancel_mode,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10), width=10).pack(side=tk.LEFT, padx=5)

        # Wait for user choice
        self.root.wait_window(mode_window)

        if mode_choice[0] is None:
            return  # User canceled

        dry_run = (mode_choice[0] == "preview")

        # If executing (not preview), require "DELETE" confirmation
        if not dry_run:
            confirm_window = tk.Toplevel(self.root)
            confirm_window.title("Final Confirmation")
            confirm_window.geometry("400x200")
            confirm_window.configure(bg=self.bg_color)
            confirm_window.transient(self.root)
            confirm_window.grab_set()

            tk.Label(confirm_window,
                    text=f"You are about to delete {count} records",
                    bg=self.bg_color, fg=self.fg_color,
                    font=("Arial", 12, "bold")).pack(pady=10)

            tk.Label(confirm_window,
                    text='Type "DELETE" to confirm:',
                    bg=self.bg_color, fg=self.fg_color,
                    font=("Arial", 10)).pack(pady=10)

            confirm_entry = tk.Entry(confirm_window, bg=self.input_bg, fg=self.fg_color,
                                    font=("Arial", 12), width=20)
            confirm_entry.pack(pady=10)
            confirm_entry.focus()

            confirmed = [False]  # Use list to modify in nested function

            def check_confirmation():
                if confirm_entry.get() == "DELETE":
                    confirmed[0] = True
                    confirm_window.destroy()
                else:
                    messagebox.showerror("Invalid", 'You must type "DELETE" exactly to confirm.')

            def cancel_delete():
                confirm_window.destroy()

            btn_frame = tk.Frame(confirm_window, bg=self.bg_color)
            btn_frame.pack(pady=10)

            tk.Button(btn_frame, text="Confirm", command=check_confirmation,
                     bg=self.button_bg, fg=self.fg_color,
                     font=("Arial", 10, "bold"), width=10).pack(side=tk.LEFT, padx=5)

            tk.Button(btn_frame, text="Cancel", command=cancel_delete,
                     bg=self.button_bg, fg=self.fg_color,
                     font=("Arial", 10, "bold"), width=10).pack(side=tk.LEFT, padx=5)

            # Bind Enter key to confirm
            confirm_entry.bind('<Return>', lambda e: check_confirmation())

            # Wait for window to close
            self.root.wait_window(confirm_window)

            if not confirmed[0]:
                return

            # Create database backup before executing delete
            if not self.backup_database():
                messagebox.showerror("Backup Failed",
                    "Database backup failed. Delete operation canceled for safety.\n\n"
                    "Please check the error and try again.")
                return

        # Show progress window
        progress_window = tk.Toplevel(self.root)
        if dry_run:
            progress_window.title("Delete Preview (Dry Run)")
        else:
            progress_window.title("Deleting Records")
        progress_window.geometry("700x500")
        progress_window.configure(bg=self.bg_color)

        tk.Label(progress_window,
                text="Preview - Records that WOULD be deleted:" if dry_run else "Deleting database records...",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 12, "bold")).pack(pady=10)

        progress_text = scrolledtext.ScrolledText(progress_window,
                                                  bg=self.input_bg, fg=self.fg_color,
                                                  font=("Courier", 9),
                                                  wrap=tk.WORD)
        progress_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Process deletions
        deleted_count = 0
        try:
            if dry_run:
                # Preview mode - just show what would be deleted
                progress_text.insert(tk.END, f"=== DRY RUN MODE - NO CHANGES WILL BE MADE ===\n\n")
                progress_text.insert(tk.END, f"The following {count} record(s) WOULD be deleted from the database:\n\n")

                for prog_num, filename, od in program_details:
                    progress_text.insert(tk.END, f"WOULD DELETE: {prog_num} - {filename} (OD: {od})\n")
                    self.root.update()

                deleted_count = count  # For summary message
                progress_text.insert(tk.END, f"\n{'='*60}\n")
                progress_text.insert(tk.END, f"Preview complete. {count} record(s) WOULD be deleted.\n")
                progress_text.insert(tk.END, f"No actual changes were made to the database.\n")
            else:
                # Execute mode - actually delete from database
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()

                placeholders = ','.join('?' * len(program_numbers))
                cursor.execute(f'''
                    DELETE FROM programs
                    WHERE program_number IN ({placeholders})
                ''', program_numbers)

                deleted_count = cursor.rowcount
                conn.commit()
                conn.close()

                # Show deleted records
                for prog_num, filename, od in program_details:
                    progress_text.insert(tk.END, f"DELETED: {prog_num} - {filename} (OD: {od})\n")
                    self.root.update()

                progress_text.insert(tk.END, f"\n{'='*60}\n")
                progress_text.insert(tk.END, f"Successfully deleted {deleted_count} record(s) from the database.\n")
                progress_text.insert(tk.END, f"⚠️  The files on disk were NOT deleted.\n")

                # Refresh the display
                self.refresh_filter_values()
                self.refresh_results()

        except Exception as e:
            progress_text.insert(tk.END, f"\n\n❌ ERROR: {str(e)}\n")
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")

        # Add close button
        close_btn = tk.Button(progress_window, text="Close",
                             command=progress_window.destroy,
                             bg=self.button_bg, fg=self.fg_color,
                             font=("Arial", 10, "bold"))
        close_btn.pack(pady=10)

    def show_context_menu(self, event):
        """Show right-click context menu"""
        # Select row under mouse
        row_id = self.tree.identify_row(event.y)
        if row_id:
            self.tree.selection_set(row_id)

            menu = tk.Menu(self.root, tearoff=0, bg=self.input_bg, fg=self.fg_color)
            menu.add_command(label="Open File", command=self.open_file)
            menu.add_command(label="Edit Entry", command=self.edit_entry)
            menu.add_command(label="View Details", command=self.view_details)
            menu.add_separator()
            menu.add_command(label="View Version History", command=self.show_version_history_window)
            menu.add_separator()
            menu.add_command(label="Delete Entry", command=self.delete_entry)

            menu.post(event.x_root, event.y_root)

    def show_statistics(self):
        """Show comprehensive database statistics with filtering support"""
        stats_window = tk.Toplevel(self.root)
        stats_window.title("Database Statistics")
        stats_window.geometry("1200x800")
        stats_window.configure(bg=self.bg_color)

        # Create notebook for different stat views
        notebook = ttk.Notebook(stats_window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Get current filters to show filtered vs total stats
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Build filter query
        filter_query = "SELECT COUNT(*) FROM programs WHERE 1=1"
        filter_params = []

        # Apply current filters
        if hasattr(self, 'filter_od_min') and self.filter_od_min.get():
            filter_query += " AND outer_diameter >= ?"
            filter_params.append(float(self.filter_od_min.get()))
        if hasattr(self, 'filter_od_max') and self.filter_od_max.get():
            filter_query += " AND outer_diameter <= ?"
            filter_params.append(float(self.filter_od_max.get()))
        if hasattr(self, 'filter_thickness_min') and self.filter_thickness_min.get():
            filter_query += " AND thickness >= ?"
            filter_params.append(float(self.filter_thickness_min.get()))
        if hasattr(self, 'filter_thickness_max') and self.filter_thickness_max.get():
            filter_query += " AND thickness <= ?"
            filter_params.append(float(self.filter_thickness_max.get()))
        if hasattr(self, 'filter_cb_min') and self.filter_cb_min.get():
            filter_query += " AND center_bore >= ?"
            filter_params.append(float(self.filter_cb_min.get()))
        if hasattr(self, 'filter_cb_max') and self.filter_cb_max.get():
            filter_query += " AND center_bore <= ?"
            filter_params.append(float(self.filter_cb_max.get()))

        # Check if any filters are active
        filters_active = len(filter_params) > 0

        # TAB 1: Overall Statistics
        tab_overall = tk.Frame(notebook, bg=self.bg_color)
        notebook.add(tab_overall, text='📊 Overall')

        overall_text = scrolledtext.ScrolledText(tab_overall, bg=self.input_bg, fg=self.fg_color,
                                                 font=("Courier", 10), wrap=tk.NONE, padx=10, pady=10)
        overall_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Get total count
        cursor.execute("SELECT COUNT(*) FROM programs")
        total_count = cursor.fetchone()[0]

        # Get filtered count
        cursor.execute(filter_query, filter_params)
        filtered_count = cursor.fetchone()[0]

        overall_text.insert(tk.END, "="*100 + "\n")
        overall_text.insert(tk.END, "DATABASE STATISTICS - OVERALL SUMMARY\n")
        overall_text.insert(tk.END, "="*100 + "\n\n")

        if filters_active:
            overall_text.insert(tk.END, f"Total Programs in Database: {total_count:,}\n")
            overall_text.insert(tk.END, f"Filtered View: {filtered_count:,} ({filtered_count/total_count*100:.1f}%)\n\n")
            overall_text.insert(tk.END, "NOTE: Statistics below show FILTERED data only\n")
            overall_text.insert(tk.END, "="*100 + "\n\n")
        else:
            overall_text.insert(tk.END, f"Total Programs: {total_count:,}\n")
            overall_text.insert(tk.END, "="*100 + "\n\n")

        # Validation Status Breakdown
        if filter_params:
            status_query = filter_query.replace("SELECT COUNT(*)", "SELECT validation_status, COUNT(*)") + " GROUP BY validation_status ORDER BY validation_status"
        else:
            status_query = "SELECT validation_status, COUNT(*) FROM programs GROUP BY validation_status ORDER BY validation_status"

        cursor.execute(status_query, filter_params)
        status_results = cursor.fetchall()

        overall_text.insert(tk.END, "VALIDATION STATUS BREAKDOWN:\n")
        overall_text.insert(tk.END, "-"*100 + "\n")
        overall_text.insert(tk.END, f"{'Status':<20} {'Count':>10} {'Percentage':>12}\n")
        overall_text.insert(tk.END, "-"*100 + "\n")

        status_total = sum(r[1] for r in status_results)
        for status, count in status_results:
            percentage = (count / status_total * 100) if status_total > 0 else 0
            overall_text.insert(tk.END, f"{status or 'NULL':<20} {count:>10,} {percentage:>11.1f}%\n")

        overall_text.insert(tk.END, "-"*100 + "\n")
        overall_text.insert(tk.END, f"{'TOTAL':<20} {status_total:>10,} {100.0:>11.1f}%\n\n\n")

        # Spacer Type Breakdown
        if filter_params:
            type_query = filter_query.replace("SELECT COUNT(*)", "SELECT spacer_type, COUNT(*)") + " GROUP BY spacer_type ORDER BY COUNT(*) DESC"
        else:
            type_query = "SELECT spacer_type, COUNT(*) FROM programs GROUP BY spacer_type ORDER BY COUNT(*) DESC"

        cursor.execute(type_query, filter_params)
        type_results = cursor.fetchall()

        overall_text.insert(tk.END, "SPACER TYPE BREAKDOWN:\n")
        overall_text.insert(tk.END, "-"*100 + "\n")
        overall_text.insert(tk.END, f"{'Type':<25} {'Count':>10} {'Percentage':>12}\n")
        overall_text.insert(tk.END, "-"*100 + "\n")

        type_total = sum(r[1] for r in type_results)
        for stype, count in type_results:
            percentage = (count / type_total * 100) if type_total > 0 else 0
            overall_text.insert(tk.END, f"{stype or 'NULL':<25} {count:>10,} {percentage:>11.1f}%\n")

        overall_text.insert(tk.END, "-"*100 + "\n")
        overall_text.insert(tk.END, f"{'TOTAL':<25} {type_total:>10,} {100.0:>11.1f}%\n\n\n")

        # Material Breakdown
        if filter_params:
            material_query = filter_query.replace("SELECT COUNT(*)", "SELECT material, COUNT(*)") + " GROUP BY material ORDER BY COUNT(*) DESC"
        else:
            material_query = "SELECT material, COUNT(*) FROM programs GROUP BY material ORDER BY COUNT(*) DESC"

        cursor.execute(material_query, filter_params)
        material_results = cursor.fetchall()

        overall_text.insert(tk.END, "MATERIAL BREAKDOWN:\n")
        overall_text.insert(tk.END, "-"*100 + "\n")
        overall_text.insert(tk.END, f"{'Material':<25} {'Count':>10} {'Percentage':>12}\n")
        overall_text.insert(tk.END, "-"*100 + "\n")

        material_total = sum(r[1] for r in material_results)
        for material, count in material_results:
            percentage = (count / material_total * 100) if material_total > 0 else 0
            overall_text.insert(tk.END, f"{material or 'NULL':<25} {count:>10,} {percentage:>11.1f}%\n")

        overall_text.insert(tk.END, "-"*100 + "\n")
        overall_text.insert(tk.END, f"{'TOTAL':<25} {material_total:>10,} {100.0:>11.1f}%\n")

        overall_text.config(state=tk.DISABLED)

        # TAB 2: By OD Size
        tab_od = tk.Frame(notebook, bg=self.bg_color)
        notebook.add(tab_od, text='📏 By OD Size')

        od_text = scrolledtext.ScrolledText(tab_od, bg=self.input_bg, fg=self.fg_color,
                                           font=("Courier", 9), wrap=tk.NONE, padx=10, pady=10)
        od_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Define OD ranges
        od_ranges = [
            (5.50, 6.00, '5.75"'),
            (6.00, 6.25, '6.00"'),
            (6.25, 6.50, '6.25"'),
            (6.50, 7.00, '6.50"'),
            (7.00, 7.50, '7.00"'),
            (7.50, 8.00, '7.50"'),
            (8.00, 8.50, '8.00"'),
            (8.50, 9.00, '8.50"'),
            (9.00, 10.00, '9.50"'),
            (10.00, 10.50, '10.25"'),
            (10.50, 11.00, '10.50"'),
            (11.00, 12.00, '11.00"'),
            (12.00, 13.50, '13.00"'),
        ]

        od_text.insert(tk.END, "="*140 + "\n")
        od_text.insert(tk.END, "STATISTICS BY OD SIZE\n")
        od_text.insert(tk.END, "="*140 + "\n\n")

        if filters_active:
            od_text.insert(tk.END, "NOTE: Showing FILTERED data only\n\n")

        od_text.insert(tk.END, f"{'OD Size':<10} {'Total':>8} {'PASS':>8} {'CRITICAL':>10} {'DIMENSIONAL':>13} {'WARNING':>9} {'BORE_WARN':>11} {'% PASS':>9}\n")
        od_text.insert(tk.END, "-"*140 + "\n")

        grand_total = 0
        grand_pass = 0
        grand_critical = 0
        grand_dimensional = 0
        grand_warning = 0
        grand_bore = 0

        for min_od, max_od, label in od_ranges:
            # Build query with OD range and optional filters
            od_filter_query = filter_query + " AND outer_diameter >= ? AND outer_diameter < ?"
            od_params = filter_params + [min_od, max_od]

            cursor.execute(od_filter_query, od_params)
            range_total = cursor.fetchone()[0]

            if range_total == 0:
                continue

            # Get status breakdown for this range
            status_query = od_filter_query.replace("SELECT COUNT(*)", "SELECT validation_status, COUNT(*)") + " GROUP BY validation_status"
            cursor.execute(status_query, od_params)
            status_breakdown = dict(cursor.fetchall())

            pass_count = status_breakdown.get('PASS', 0)
            critical_count = status_breakdown.get('CRITICAL', 0)
            dimensional_count = status_breakdown.get('DIMENSIONAL', 0)
            warning_count = status_breakdown.get('WARNING', 0)
            bore_count = status_breakdown.get('BORE_WARNING', 0)
            pass_pct = (pass_count / range_total * 100) if range_total > 0 else 0

            od_text.insert(tk.END, f"{label:<10} {range_total:>8,} {pass_count:>8,} {critical_count:>10,} {dimensional_count:>13,} {warning_count:>9,} {bore_count:>11,} {pass_pct:>8.1f}%\n")

            grand_total += range_total
            grand_pass += pass_count
            grand_critical += critical_count
            grand_dimensional += dimensional_count
            grand_warning += warning_count
            grand_bore += bore_count

        od_text.insert(tk.END, "-"*140 + "\n")
        grand_pass_pct = (grand_pass / grand_total * 100) if grand_total > 0 else 0
        od_text.insert(tk.END, f"{'TOTAL':<10} {grand_total:>8,} {grand_pass:>8,} {grand_critical:>10,} {grand_dimensional:>13,} {grand_warning:>9,} {grand_bore:>11,} {grand_pass_pct:>8.1f}%\n")

        od_text.config(state=tk.DISABLED)

        # TAB 3: Status by OD (matrix view)
        tab_matrix = tk.Frame(notebook, bg=self.bg_color)
        notebook.add(tab_matrix, text='🔢 Status Matrix')

        matrix_text = scrolledtext.ScrolledText(tab_matrix, bg=self.input_bg, fg=self.fg_color,
                                               font=("Courier", 9), wrap=tk.NONE, padx=10, pady=10)
        matrix_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        matrix_text.insert(tk.END, "="*140 + "\n")
        matrix_text.insert(tk.END, "VALIDATION STATUS BY OD SIZE - DETAILED MATRIX\n")
        matrix_text.insert(tk.END, "="*140 + "\n\n")

        if filters_active:
            matrix_text.insert(tk.END, "NOTE: Showing FILTERED data only\n\n")

        # Get all validation statuses
        status_list = ['PASS', 'CRITICAL', 'DIMENSIONAL', 'WARNING', 'BORE_WARNING']

        # Create matrix
        for status in status_list:
            matrix_text.insert(tk.END, f"\n{'='*140}\n")
            matrix_text.insert(tk.END, f"{status} BY OD SIZE:\n")
            matrix_text.insert(tk.END, f"{'='*140}\n")
            matrix_text.insert(tk.END, f"{'OD Size':<12} {'Count':>10} {'% of Size':>12} {'% of Status':>14}\n")
            matrix_text.insert(tk.END, "-"*140 + "\n")

            # Get total for this status
            status_filter_query = filter_query + " AND validation_status = ?"
            cursor.execute(status_filter_query, filter_params + [status])
            status_total = cursor.fetchone()[0]

            if status_total == 0:
                matrix_text.insert(tk.END, f"No programs with {status} status\n")
                continue

            status_subtotal = 0
            for min_od, max_od, label in od_ranges:
                od_status_query = filter_query + " AND outer_diameter >= ? AND outer_diameter < ? AND validation_status = ?"
                cursor.execute(od_status_query, filter_params + [min_od, max_od, status])
                count = cursor.fetchone()[0]

                if count == 0:
                    continue

                # Get total for this OD size
                od_total_query = filter_query + " AND outer_diameter >= ? AND outer_diameter < ?"
                cursor.execute(od_total_query, filter_params + [min_od, max_od])
                od_total = cursor.fetchone()[0]

                pct_of_size = (count / od_total * 100) if od_total > 0 else 0
                pct_of_status = (count / status_total * 100) if status_total > 0 else 0

                matrix_text.insert(tk.END, f"{label:<12} {count:>10,} {pct_of_size:>11.1f}% {pct_of_status:>13.1f}%\n")
                status_subtotal += count

            matrix_text.insert(tk.END, "-"*140 + "\n")
            matrix_text.insert(tk.END, f"{'TOTAL':<12} {status_subtotal:>10,} {'':>11} {100.0:>13.1f}%\n")

        matrix_text.config(state=tk.DISABLED)

        # TAB 4: Top Error Types
        tab_errors = tk.Frame(notebook, bg=self.bg_color)
        notebook.add(tab_errors, text='⚠️ Error Types')

        error_text = scrolledtext.ScrolledText(tab_errors, bg=self.input_bg, fg=self.fg_color,
                                              font=("Courier", 9), wrap=tk.NONE, padx=10, pady=10)
        error_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        error_text.insert(tk.END, "="*100 + "\n")
        error_text.insert(tk.END, "TOP ERROR TYPES\n")
        error_text.insert(tk.END, "="*100 + "\n\n")

        if filters_active:
            error_text.insert(tk.END, "NOTE: Showing FILTERED data only\n\n")

        # Count different error types from validation_issues
        issues_query = filter_query.replace("SELECT COUNT(*)", "SELECT validation_issues") + " AND validation_issues IS NOT NULL"
        cursor.execute(issues_query, filter_params)
        all_issues = cursor.fetchall()

        # Parse and count error types
        error_counts = {}
        for (issues_str,) in all_issues:
            if not issues_str:
                continue
            # Split by pipe delimiter
            issues = [i.strip() for i in issues_str.split('|') if i.strip()]
            for issue in issues:
                # Extract error type (first part before colon)
                error_type = issue.split(':')[0].strip() if ':' in issue else issue.strip()
                error_counts[error_type] = error_counts.get(error_type, 0) + 1

        # Sort by count
        sorted_errors = sorted(error_counts.items(), key=lambda x: x[1], reverse=True)

        error_text.insert(tk.END, f"{'Error Type':<60} {'Count':>10} {'Percentage':>12}\n")
        error_text.insert(tk.END, "-"*100 + "\n")

        total_errors = sum(error_counts.values())
        for error_type, count in sorted_errors[:20]:  # Top 20 errors
            percentage = (count / total_errors * 100) if total_errors > 0 else 0
            error_text.insert(tk.END, f"{error_type:<60} {count:>10,} {percentage:>11.1f}%\n")

        error_text.insert(tk.END, "-"*100 + "\n")
        error_text.insert(tk.END, f"{'TOTAL ERRORS':<60} {total_errors:>10,} {100.0:>11.1f}%\n\n")

        error_text.insert(tk.END, f"\nTotal programs with errors: {len(all_issues):,}\n")

        error_text.config(state=tk.DISABLED)

        # TAB 5: Status by Type (Detailed breakdown)
        tab_type_status = tk.Frame(notebook, bg=self.bg_color)
        notebook.add(tab_type_status, text='📋 Type & Status')

        type_status_text = scrolledtext.ScrolledText(tab_type_status, bg=self.input_bg, fg=self.fg_color,
                                                     font=("Courier", 9), wrap=tk.NONE, padx=10, pady=10)
        type_status_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        type_status_text.insert(tk.END, "="*140 + "\n")
        type_status_text.insert(tk.END, "VALIDATION STATUS BY SPACER TYPE\n")
        type_status_text.insert(tk.END, "="*140 + "\n\n")

        if filters_active:
            type_status_text.insert(tk.END, "NOTE: Showing FILTERED data only\n\n")

        # Get all spacer types
        type_query = "SELECT DISTINCT spacer_type FROM programs WHERE spacer_type IS NOT NULL ORDER BY spacer_type"
        cursor.execute(type_query)
        all_types = [row[0] for row in cursor.fetchall()]

        status_list = ['PASS', 'CRITICAL', 'DIMENSIONAL', 'WARNING', 'BORE_WARNING']

        for spacer_type in all_types:
            type_status_text.insert(tk.END, f"\n{spacer_type.upper()}:\n")
            type_status_text.insert(tk.END, "-"*140 + "\n")
            type_status_text.insert(tk.END, f"{'Status':<20} {'Count':>10} {'% of Type':>12}\n")
            type_status_text.insert(tk.END, "-"*140 + "\n")

            # Get total for this type
            type_total_query = filter_query + " AND spacer_type = ?"
            cursor.execute(type_total_query, filter_params + [spacer_type])
            type_total = cursor.fetchone()[0]

            if type_total == 0:
                continue

            # Get status breakdown
            type_status_query = type_total_query.replace("SELECT COUNT(*)", "SELECT validation_status, COUNT(*)") + " GROUP BY validation_status"
            cursor.execute(type_status_query, filter_params + [spacer_type])
            status_breakdown = dict(cursor.fetchall())

            for status in status_list:
                count = status_breakdown.get(status, 0)
                pct = (count / type_total * 100) if type_total > 0 else 0
                type_status_text.insert(tk.END, f"{status:<20} {count:>10,} {pct:>11.1f}%\n")

            type_status_text.insert(tk.END, "-"*140 + "\n")
            type_status_text.insert(tk.END, f"{'TOTAL':<20} {type_total:>10,} {100.0:>11.1f}%\n")

        type_status_text.config(state=tk.DISABLED)

        # TAB 6: Round Size & Type Matrix
        tab_size_type = tk.Frame(notebook, bg=self.bg_color)
        notebook.add(tab_size_type, text='🔍 Size & Type')

        size_type_text = scrolledtext.ScrolledText(tab_size_type, bg=self.input_bg, fg=self.fg_color,
                                                   font=("Courier", 9), wrap=tk.NONE, padx=10, pady=10)
        size_type_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        size_type_text.insert(tk.END, "="*140 + "\n")
        size_type_text.insert(tk.END, "SPACER TYPE BY ROUND SIZE\n")
        size_type_text.insert(tk.END, "="*140 + "\n\n")

        if filters_active:
            size_type_text.insert(tk.END, "NOTE: Showing FILTERED data only\n\n")

        size_type_text.insert(tk.END, f"{'OD Size':<10}")
        for stype in ['hub_centric', 'standard', '2PC LUG', '2PC STUD', 'step', 'STEP']:
            size_type_text.insert(tk.END, f"{stype[:8]:>10}")
        size_type_text.insert(tk.END, f"{'Total':>10}\n")
        size_type_text.insert(tk.END, "-"*140 + "\n")

        for min_od, max_od, label in od_ranges:
            od_filter_query = filter_query + " AND outer_diameter >= ? AND outer_diameter < ?"
            od_params = filter_params + [min_od, max_od]

            cursor.execute(od_filter_query, od_params)
            range_total = cursor.fetchone()[0]

            if range_total == 0:
                continue

            size_type_text.insert(tk.END, f"{label:<10}")

            # Get type breakdown for this OD range
            type_breakdown_query = od_filter_query.replace("SELECT COUNT(*)", "SELECT spacer_type, COUNT(*)") + " GROUP BY spacer_type"
            cursor.execute(type_breakdown_query, od_params)
            type_breakdown = dict(cursor.fetchall())

            for stype in ['hub_centric', 'standard', '2PC LUG', '2PC STUD', 'step', 'STEP']:
                count = type_breakdown.get(stype, 0)
                size_type_text.insert(tk.END, f"{count:>10,}")

            size_type_text.insert(tk.END, f"{range_total:>10,}\n")

        size_type_text.config(state=tk.DISABLED)

        # TAB 7: Errors/Warnings by Round Size
        tab_size_errors = tk.Frame(notebook, bg=self.bg_color)
        notebook.add(tab_size_errors, text='⚠️ Errors by Size')

        size_errors_text = scrolledtext.ScrolledText(tab_size_errors, bg=self.input_bg, fg=self.fg_color,
                                                     font=("Courier", 9), wrap=tk.NONE, padx=10, pady=10)
        size_errors_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        size_errors_text.insert(tk.END, "="*140 + "\n")
        size_errors_text.insert(tk.END, "TOP ERROR TYPES BY ROUND SIZE\n")
        size_errors_text.insert(tk.END, "="*140 + "\n\n")

        if filters_active:
            size_errors_text.insert(tk.END, "NOTE: Showing FILTERED data only\n\n")

        # For each OD range, show top error types
        for min_od, max_od, label in od_ranges:
            od_filter_query = filter_query + " AND outer_diameter >= ? AND outer_diameter < ?"
            od_params = filter_params + [min_od, max_od]

            # Get programs with issues in this range
            issues_query = od_filter_query.replace("SELECT COUNT(*)", "SELECT validation_issues") + " AND validation_issues IS NOT NULL"
            cursor.execute(issues_query, od_params)
            range_issues = cursor.fetchall()

            if not range_issues:
                continue

            size_errors_text.insert(tk.END, f"\n{label} ROUND:\n")
            size_errors_text.insert(tk.END, "-"*140 + "\n")

            # Parse error types for this range
            range_error_counts = {}
            for (issues_str,) in range_issues:
                if not issues_str:
                    continue
                issues = [i.strip() for i in issues_str.split('|') if i.strip()]
                for issue in issues:
                    error_type = issue.split(':')[0].strip() if ':' in issue else issue.strip()
                    range_error_counts[error_type] = range_error_counts.get(error_type, 0) + 1

            # Sort and show top 5 errors
            sorted_range_errors = sorted(range_error_counts.items(), key=lambda x: x[1], reverse=True)

            size_errors_text.insert(tk.END, f"{'Error Type':<60} {'Count':>10}\n")
            size_errors_text.insert(tk.END, "-"*140 + "\n")

            for error_type, count in sorted_range_errors[:5]:
                size_errors_text.insert(tk.END, f"{error_type:<60} {count:>10,}\n")

            total_range_errors = sum(range_error_counts.values())
            size_errors_text.insert(tk.END, "-"*140 + "\n")
            size_errors_text.insert(tk.END, f"Total programs with errors in {label}: {len(range_issues):,} ({total_range_errors:,} total errors)\n")

        size_errors_text.config(state=tk.DISABLED)

        conn.close()

        # Add close button
        close_btn = tk.Button(stats_window, text="Close", command=stats_window.destroy,
                             bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold"))
        close_btn.pack(pady=10)

    def show_legend(self):
        """Show validation status legend and help"""
        legend_window = tk.Toplevel(self.root)
        legend_window.title("G-Code Database Manager - Help & Workflow Guide")
        legend_window.geometry("900x900")
        legend_window.configure(bg=self.bg_color)

        # Create scrolled text widget
        text = scrolledtext.ScrolledText(legend_window, bg=self.input_bg, fg=self.fg_color,
                                        font=("Courier", 10), wrap=tk.WORD, padx=15, pady=15)
        text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Legend content
        legend_content = """
═══════════════════════════════════════════════════════════════════════════════
                      G-CODE DATABASE MANAGER - HELP GUIDE
═══════════════════════════════════════════════════════════════════════════════

TABLE OF CONTENTS:
  1. Production-Ready Workflow (Recommended Steps)
  2. Function Descriptions (What Each Button Does)
  3. Duplicate Handling Strategies
  4. Validation Status Legend
  5. Filtering Tips

═══════════════════════════════════════════════════════════════════════════════
                    1. PRODUCTION-READY WORKFLOW (START HERE!)
═══════════════════════════════════════════════════════════════════════════════

GOAL: Clean catalog with NO duplicate names, NO _dup suffixes, all unique files

STEP 1: INITIAL SCAN
───────────────────────────────────────────────────────────────────────────────
📂 Scan Folder
  • Click "📂 Scan Folder" button (Data tab)
  • Select your main G-code directory
  • This scans ALL files recursively and adds them to database
  • Automatically detects and SKIPS exact duplicates
  • Warns about name collisions (same name, different content)

What Happens:
  ✓ Exact duplicates (same name + content) → Automatically skipped
  ✓ New unique files → Added to database
  ✓ Existing files → Updated if changed
  ⚠️  Name collisions → Warned (you must rename manually before scan)

STEP 2: DETECT DUPLICATES
───────────────────────────────────────────────────────────────────────────────
🔍 Find Repeats
  • Click "🔍 Find Repeats" button (Tools tab)
  • Analyzes ALL files in database for duplicates
  • Classifies duplicates into categories

What Happens:
  • SOLID duplicates: Exact same content, different names
  • NAME_COLLISION: Same name, different content
  • CONTENT_DUP: Same content as parent file
  • Parent/child relationships established

STEP 3: REVIEW DUPLICATES
───────────────────────────────────────────────────────────────────────────────
⚖️ Compare Files
  • Hold Ctrl and click multiple files in the list
  • Click "⚖️ Compare Files" button (Tools tab)
  • View files side-by-side with metadata and G-code preview
  • Decide which to keep, rename, or delete

Decision Process:
  1. If EXACT SAME content but different names:
     → Keep the one with better name (more descriptive)
     → Delete the others from database (file stays on disk)

  2. If SAME NAME but DIFFERENT content:
     → Keep both files BUT rename one
     → Use the rename function to give it a unique name

  3. If similar but slightly different:
     → Compare the G-code to see actual differences
     → Keep the correct/latest version
     → Delete or rename the outdated version

STEP 4: AUTO-ASSIGN NEW O-NUMBERS TO DUPLICATES
───────────────────────────────────────────────────────────────────────────────
📝 Rename Duplicates
  • Filter view first (e.g., Duplicate Type = NAME_COLLISION)
  • Click "📝 Rename Duplicates" button (Tools tab)
  • Finds files with duplicate filenames IN FILTERED VIEW ONLY
  • Assigns new available O-numbers (o59000+)
  • NO _dup suffix - assigns proper clean O-numbers
  • Creates backup before making changes
  • Updates: physical file, internal O-number, database

⚠️  IMPORTANT: Works on FILTERED VIEW ONLY!
  • Filter for the duplicates you want to rename
  • Parent file keeps original name
  • Child files get new O-numbers (o59000, o59001, etc.)
  • All changes are automatic - no manual renaming needed

Preview Mode:
  • Shows exactly what O-numbers will be assigned
  • Review the plan before executing
  • Click Execute to apply changes

STEP 5: DELETE UNWANTED DUPLICATES
───────────────────────────────────────────────────────────────────────────────
🗑️ Delete Filtered View
  • Use filters to show only duplicates
  • Apply filters: Duplicate Type = SOLID, CONTENT_DUP, etc.
  • Click "🗑️ Delete Filtered View" (Tools tab)
  • Deletes from DATABASE ONLY (files stay on disk for safety)

Safety Features:
  • Preview mode shows what will be deleted
  • Database backup created automatically
  • Physical files are NOT deleted (safer)

STEP 6: VERIFY PRODUCTION-READY STATE
───────────────────────────────────────────────────────────────────────────────
✅ Final Checks:
  1. Run Find Repeats again
     → Should find ZERO duplicates

  2. Filter by Validation Status = PASS
     → These are production-ready files

  3. Check total file count
     → All files should have unique O-numbers

  4. Export to CSV for documentation
     → Click "📊 Export CSV" (Reports tab)

═══════════════════════════════════════════════════════════════════════════════
                    2. FUNCTION DESCRIPTIONS (WHAT EACH BUTTON DOES)
═══════════════════════════════════════════════════════════════════════════════

DATA TAB - File Import/Management
───────────────────────────────────────────────────────────────────────────────
📂 Scan Folder
  • Recursively scans a directory for G-code files
  • Parses dimensions, P-codes, and metadata
  • Adds new files, updates existing files
  • Auto-detects exact duplicates (skips them)
  • Warns about name collisions (same name, different content)
  • Shows live progress with file counts

🆕 Scan New Only
  • Scans directory but ONLY adds NEW files
  • Skips files already in database
  • Faster for adding new batches
  • Still detects duplicates and name collisions
  • Use after initial full scan

➕ Add Entry
  • Manually add a single program
  • Fill in all fields manually
  • Use when you need precise control
  • Links to physical file on disk

✏️ Edit Entry
  • Double-click any row OR click this button
  • Modify any field in database
  • Changes save to database
  • Does NOT modify physical file

TOOLS TAB - Duplicate Management
───────────────────────────────────────────────────────────────────────────────
🔍 Find Repeats
  • Analyzes ALL files for duplicates
  • Compares file content (not just names)
  • Creates parent/child relationships
  • Classifies duplicates: SOLID, NAME_COLLISION, CONTENT_DUP
  • Populates duplicate_type column
  • Required before using duplicate filters

⚖️ Compare Files
  • Select 2+ files (Ctrl+Click)
  • Shows side-by-side comparison
  • Displays metadata and G-code content
  • Highlights differences
  • Actions: Keep, Rename, Delete
  • Updates internal O-number when renaming

📝 Rename Duplicates
  • Works on FILTERED VIEW ONLY (not all files)
  • Finds files with same filename in current view
  • Assigns proper available O-numbers (o59000+)
  • NO _dup suffix - assigns clean O-numbers
  • Renames physical files
  • Updates internal O-number in G-code file
  • Updates database program_number
  • Preview mode: See changes before applying
  • Execute mode: Actually renames files
  • Creates automatic backup

📋 Copy Filtered View
  • Copies currently filtered files to new folder
  • Preserves directory structure
  • Only copies files matching active filters
  • Use to extract specific file groups
  • Physical file copy operation

🗑️ Delete Filtered View
  • Deletes currently filtered files from DATABASE
  • Does NOT delete physical files (safer)
  • Preview mode available
  • Creates backup before deletion
  • Use to clean up duplicate entries

REPORTS TAB - Export/Analysis
───────────────────────────────────────────────────────────────────────────────
📊 Export CSV
  • Exports current filtered view to CSV
  • Includes all columns
  • Use for Excel analysis, documentation
  • Respects active filters

📄 Export Unused Numbers
  • Finds gaps in program number sequence
  • Exports available O-numbers
  • Helps when creating new programs
  • Useful for number assignment

🗂️ Organize by OD
  • Groups files by Outer Diameter
  • Creates folders for each OD size
  • Copies files to organized structure
  • Preview mode available

❓ Help/Legend
  • Opens this help guide
  • Workflow documentation
  • Function descriptions
  • Color-coded validation system

═══════════════════════════════════════════════════════════════════════════════
                    3. DUPLICATE HANDLING STRATEGIES
═══════════════════════════════════════════════════════════════════════════════

SCENARIO A: Same Name, Same Content (Exact Duplicates)
───────────────────────────────────────────────────────────────────────────────
Example: O57001.nc in two different folders, identical content

What to do:
  1. Scan Folder → Auto-skips exact duplicates
  2. Find Repeats → Marks as SOLID duplicate
  3. Compare Files → Verify they're identical
  4. Delete one from database (keep the one in primary location)

Result: Only ONE file in database, no confusion

SCENARIO B: Same Name, Different Content (Name Collision)
───────────────────────────────────────────────────────────────────────────────
Example: O57001.nc exists twice but programs are different

THIS IS THE CRITICAL SCENARIO - REQUIRES MANUAL DECISION!

What to do:
  1. Find Repeats → Marks as NAME_COLLISION
  2. Compare Files → View both side-by-side
  3. Decide which is correct:

     Option A - Keep BOTH (they're both valid):
       • Select one file in Compare window
       • Click Rename action
       • Give it a new unique O-number (e.g., O57001A or O59999)
       • System updates filename AND internal O-number
       • Now you have O57001.nc and O59999.nc (both unique)

     Option B - Keep ONE (one is wrong/outdated):
       • Select the wrong one
       • Click Delete action
       • Removes from database (file stays on disk for safety)
       • Correct file remains as O57001.nc

⚠️  NEVER let two different programs share the same O-number!

SCENARIO C: Different Names, Same Content (SOLID Duplicate)
───────────────────────────────────────────────────────────────────────────────
Example: O57001.nc and O57001_backup.nc are identical

What to do:
  1. Find Repeats → Marks as SOLID duplicate
  2. Compare Files → Confirm they're identical
  3. Delete the backup from database
  4. Keep the one with the cleaner name

Result: Single entry with best filename

SCENARIO D: Using Rename Duplicates for Automatic Assignment
───────────────────────────────────────────────────────────────────────────────
Example: You have multiple files named O57001.nc in different folders

Automatic workflow:
  1. Filter: Duplicate Type = NAME_COLLISION
  2. Click "📝 Rename Duplicates" (Tools tab)
  3. Preview mode shows the new O-numbers to be assigned
  4. Execute: Child files automatically renamed to o59000, o59001, etc.
  5. System updates:
     • Physical filename (O57001.nc → o59000.nc)
     • Internal O-number (O57001 → O59000)
     • Database program_number

Result: All files have unique O-numbers (NO _dup suffix!)
  • Parent keeps original: O57001.nc
  • Child 1: o59000.nc
  • Child 2: o59001.nc
  • All are production-ready with proper O-numbers

═══════════════════════════════════════════════════════════════════════════════
                    4. VALIDATION STATUS LEGEND
═══════════════════════════════════════════════════════════════════════════════

The database uses a 5-color validation system to categorize issues by severity.

═══════════════════════════════════════════════════════════════════════════════

🔴 RED - CRITICAL (Highest Priority)
═══════════════════════════════════════════════════════════════════
Status: CRITICAL

What It Means:
  • Critical dimensional errors that will produce WRONG parts
  • CB or OB dimensions way outside tolerance (>±0.2-0.3mm)
  • Thickness errors beyond ±0.02"
  • Part CANNOT be used without G-code correction

Examples:
  • CB TOO SMALL: Spec=71.0mm, G-code=66.0mm (-4.96mm)
  • OB TOO LARGE: Spec=64.1mm, G-code=66.5mm (+2.40mm)
  • THICKNESS ERROR: Spec=0.75", Calculated=0.80" (+0.05")

Action Required: ⚠️ IMMEDIATE FIX NEEDED
  • Stop production
  • Fix G-code before running
  • These errors will cause part failures or machine crashes

───────────────────────────────────────────────────────────────────

🟠 ORANGE - BORE WARNING (High Priority)
═══════════════════════════════════════════════════════════════════
Status: BORE_WARNING

What It Means:
  • Bore dimensions at tolerance limits (CB/OB within ±0.1-0.2mm)
  • Part is still within spec but close to the edge
  • May cause fit issues in assembly

Examples:
  • CB at tolerance limit: Spec=38.1mm, G-code=38.2mm (+0.10mm)
  • OB at tolerance limit: Spec=64.1mm, G-code=64.0mm (-0.10mm)

Action Required: 🔍 VERIFY CAREFULLY
  • Check dimensions carefully during setup
  • Verify first article measurement
  • Part is technically acceptable but borderline
  • Consider adjusting G-code for better margin

───────────────────────────────────────────────────────────────────

🟣 PURPLE - DIMENSIONAL (Medium Priority)
═══════════════════════════════════════════════════════════════════
Status: DIMENSIONAL

What It Means:
  • P-code and thickness mismatches (setup dimension issues)
  • Wrong work offset for part thickness
  • Drill depth calculations slightly off (±0.01-0.02")
  • Work offsets don't match part thickness

Examples:
  • P-CODE MISMATCH: Thickness 1.00" expects P15/P16, but found [17,18]
  • Thickness mismatch: Spec=0.75", Calculated=0.76" (+0.01")

Action Required: 🔧 REVIEW SETUP
  • Check work offset (P-code) settings
  • Verify drill depth matches part thickness
  • Part may run but could have setup issues
  • Update P-codes or drill depth to match

───────────────────────────────────────────────────────────────────

🟡 YELLOW - WARNING (Low Priority)
═══════════════════════════════════════════════════════════════════
Status: WARNING

What It Means:
  • General warnings that don't affect critical dimensions
  • P-code pairing issues (missing OP1 or OP2)
  • Multiple P-codes found (possible copy/paste error)
  • OD slightly off (non-critical)

Examples:
  • P-CODE PAIRING: P15 found but P16 missing
  • MULTIPLE P-CODES: Found [15,16,17,18] - should only have one pair
  • OD tolerance check: Spec=5.75", G-code=5.71" (-0.04")

Action Required: 📋 REVIEW WHEN CONVENIENT
  • Check for copy/paste errors
  • Verify P-code pairing
  • Not urgent but should be fixed eventually

───────────────────────────────────────────────────────────────────

🟢 GREEN - PASS (No Issues)
═══════════════════════════════════════════════════════════════════
Status: PASS

What It Means:
  • All validations passed ✓
  • Dimensions within spec ✓
  • P-codes match thickness ✓
  • Ready to run ✓

Action Required: ✅ None - Good to go!

Color    Status         Severity    Production Impact    Action
------   ------------   ---------   ------------------   ---------
🔴 RED    CRITICAL       CRITICAL    Part failure/crash   IMMEDIATE
🟠 ORANGE BORE_WARNING   HIGH        Possible fit issues  Before 1st
🟣 PURPLE DIMENSIONAL    MEDIUM      Setup errors         Before run
🟡 YELLOW WARNING        LOW         Minor issues         When ready
🟢 GREEN  PASS           NONE        None                 N/A

TOLERANCE REFERENCE
───────────────────────────────────────────────────────────────────────────────
CB (Center Bore):
  • Acceptable:  title_cb to (title_cb + 0.1mm)
  • Orange:      ±0.1 to ±0.2mm or ±0.3mm
  • Red:         < -0.2mm or > +0.3mm

OB (Outer Bore / Hub Diameter):
  • Acceptable:  (title_ob - 0.1mm) to title_ob
  • Orange:      ±0.1 to ±0.2mm or ±0.3mm
  • Red:         < -0.3mm or > +0.2mm

Thickness:
  • Acceptable:  ±0.01"
  • Purple:      ±0.01 to ±0.02"
  • Red:         > ±0.02"

OD (Outer Diameter):
  • Acceptable:  ±0.05"
  • Yellow:      ±0.05 to ±0.1"
  • Red:         > ±0.1"

═══════════════════════════════════════════════════════════════════════════════
                    5. FILTERING TIPS
═══════════════════════════════════════════════════════════════════════════════

Multi-Select Filters
───────────────────────────────────────────────────────────────────────────────
  • Click Type/Material/Status/Duplicate Type dropdowns
  • Check multiple boxes to filter
  • Shows "3 selected" when multiple items chosen
  • Click "Apply Filters" button to filter results
  • Click "Reset Filters" to clear all filters

Common Filter Combinations
───────────────────────────────────────────────────────────────────────────────
Focus on Issues Only:
  • Status: Select CRITICAL + BORE_WARNING + DIMENSIONAL
  • Result: See only files needing attention

Production-Ready Files:
  • Status: Select only PASS
  • Duplicate Type: Leave empty or uncheck all
  • Result: Clean, validated files ready to run

View All Duplicates:
  • Duplicate Type: Select SOLID + NAME_COLLISION + CONTENT_DUP
  • Result: See all duplicate files for cleanup

View Only SOLID Duplicates (same content, different names):
  • Duplicate Type: Select only SOLID
  • Result: Files you can safely delete (keep one copy)

View Name Collisions (CRITICAL - different content, same name):
  • Duplicate Type: Select only NAME_COLLISION
  • Result: Files that MUST be manually reviewed and renamed

View by Material:
  • Material: Select specific materials (e.g., 1018, 4140)
  • Result: See files for specific material types

View by Outer Diameter:
  • OD Range: Enter min/max values
  • Result: See files within specific size range

═══════════════════════════════════════════════════════════════════════════════
                    QUICK REFERENCE - RECOMMENDED WORKFLOW
═══════════════════════════════════════════════════════════════════════════════

INITIAL SETUP (First Time):
  1. Scan Folder → Load all files
  2. Find Repeats → Detect duplicates
  3. Review & Compare → Use Compare Files
  4. Clean Up → Rename or Delete duplicates
  5. Verify → Check for _dup files and duplicates

DAILY PRODUCTION USE:
  1. Scan New Only → Add new files
  2. Filter by Status → Focus on PASS files
  3. Export CSV → Document production files
  4. Organize by OD → Group files for easy access

TROUBLESHOOTING:
  • If Copy Filtered not working → Run Find Repeats first
  • If duplicates showing → Use Rename Duplicates or Compare Files
  • If Rename Duplicates does nothing → Check that view is filtered
  • If validation errors → Fix G-code and re-scan

═══════════════════════════════════════════════════════════════════════════════

For more documentation, see project README files in the application directory.

═══════════════════════════════════════════════════════════════════════════════
"""

        text.insert("1.0", legend_content)

        # Configure text tags for colored examples
        text.tag_configure('red', foreground='#ff6b6b')
        text.tag_configure('orange', foreground='#ffa500')
        text.tag_configure('purple', foreground='#da77f2')
        text.tag_configure('yellow', foreground='#ffd43b')
        text.tag_configure('green', foreground='#69db7c')

        text.config(state=tk.DISABLED)

        # Close button
        btn_close = tk.Button(legend_window, text="Close", command=legend_window.destroy,
                             bg=self.button_bg, fg=self.fg_color,
                             font=("Arial", 11, "bold"), width=20, height=2)
        btn_close.pack(pady=10)

    def create_manual_backup(self):
        """Create a manual backup with timestamp"""
        import shutil
        from datetime import datetime

        backups_dir = os.path.join(os.path.dirname(self.db_path), "database_backups")
        os.makedirs(backups_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        backup_filename = f"gcode_db_backup_{timestamp}.db"
        backup_path = os.path.join(backups_dir, backup_filename)

        try:
            shutil.copy2(self.db_path, backup_path)
            messagebox.showinfo("Backup Created",
                f"Database backed up successfully!\n\n"
                f"Backup saved to:\n{backup_filename}\n\n"
                f"Location: {backups_dir}")
        except Exception as e:
            messagebox.showerror("Backup Failed",
                f"Failed to create backup:\n{str(e)}")

    def create_full_backup(self):
        """Create a full backup including database and all repository files"""
        import shutil
        from datetime import datetime
        import zipfile

        # Let user choose backup location
        backup_location = filedialog.askdirectory(
            title="Select Location for Full Backup",
            initialdir=os.path.expanduser("~")
        )

        if not backup_location:
            return

        # Create timestamped backup folder name
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        backup_name = f"GCode_Full_Backup_{timestamp}"
        backup_folder = os.path.join(backup_location, backup_name)

        try:
            # Create progress window
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Creating Full Backup")
            progress_window.geometry("500x200")
            progress_window.configure(bg=self.bg_color)
            progress_window.transient(self.root)
            progress_window.grab_set()

            # Center window
            progress_window.update_idletasks()
            x = (progress_window.winfo_screenwidth() // 2) - 250
            y = (progress_window.winfo_screenheight() // 2) - 100
            progress_window.geometry(f"500x200+{x}+{y}")

            status_label = tk.Label(progress_window,
                                   text="Preparing backup...",
                                   bg=self.bg_color, fg=self.fg_color,
                                   font=("Arial", 11))
            status_label.pack(pady=20)

            progress_bar = ttk.Progressbar(progress_window, length=400, mode='indeterminate')
            progress_bar.pack(pady=10)
            progress_bar.start(10)

            details_label = tk.Label(progress_window,
                                    text="",
                                    bg=self.bg_color, fg=self.fg_color,
                                    font=("Arial", 9))
            details_label.pack(pady=10)

            progress_window.update()

            # Create backup folder
            os.makedirs(backup_folder, exist_ok=True)

            # 1. Backup database
            status_label.config(text="Backing up database...")
            progress_window.update()

            db_backup_path = os.path.join(backup_folder, "gcode_database.db")
            shutil.copy2(self.db_path, db_backup_path)

            # 2. Get all repository files from database
            status_label.config(text="Collecting repository files...")
            progress_window.update()

            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT file_path FROM programs WHERE file_path IS NOT NULL")
            file_paths = [row[0] for row in cursor.fetchall() if row[0]]
            conn.close()

            # 3. Copy repository files maintaining structure
            status_label.config(text="Copying files...")
            details_label.config(text=f"Total files: {len(file_paths)}")
            progress_window.update()

            files_copied = 0
            files_missing = 0

            repository_backup = os.path.join(backup_folder, "repository")
            os.makedirs(repository_backup, exist_ok=True)

            for i, file_path in enumerate(file_paths):
                if os.path.exists(file_path):
                    # Maintain folder structure relative to repository
                    if hasattr(self, 'repository_path') and self.repository_path:
                        try:
                            rel_path = os.path.relpath(file_path, self.repository_path)
                        except:
                            rel_path = os.path.basename(file_path)
                    else:
                        rel_path = os.path.basename(file_path)

                    dest_path = os.path.join(repository_backup, rel_path)
                    os.makedirs(os.path.dirname(dest_path), exist_ok=True)

                    try:
                        shutil.copy2(file_path, dest_path)
                        files_copied += 1
                    except Exception as e:
                        files_missing += 1
                else:
                    files_missing += 1

                # Update progress every 10 files
                if i % 10 == 0:
                    details_label.config(text=f"Copied: {files_copied} | Missing: {files_missing} | Total: {len(file_paths)}")
                    progress_window.update()

            # 4. Create backup info file
            status_label.config(text="Creating backup info...")
            progress_window.update()

            info_path = os.path.join(backup_folder, "BACKUP_INFO.txt")
            with open(info_path, 'w') as f:
                f.write(f"GCode Database Full Backup\n")
                f.write(f"="*50 + "\n\n")
                f.write(f"Created: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Database records: {len(file_paths)}\n")
                f.write(f"Files copied: {files_copied}\n")
                f.write(f"Files missing: {files_missing}\n\n")
                f.write(f"Contents:\n")
                f.write(f"  - gcode_database.db (database file)\n")
                f.write(f"  - repository/ (all .nc files)\n\n")
                f.write(f"To Restore:\n")
                f.write(f"  1. Copy gcode_database.db to your program folder\n")
                f.write(f"  2. Copy repository/ contents to your repository location\n")
                f.write(f"  3. Update file paths in database if repository location changed\n")

            progress_bar.stop()
            progress_window.destroy()

            # Show success message
            messagebox.showinfo("Full Backup Complete",
                f"Full backup created successfully!\n\n"
                f"Location: {backup_folder}\n\n"
                f"Database: Backed up ✓\n"
                f"Files copied: {files_copied}\n"
                f"Files missing: {files_missing}\n\n"
                f"Backup includes:\n"
                f"  • Database file\n"
                f"  • All repository .nc files\n"
                f"  • Backup info file")

        except Exception as e:
            if 'progress_window' in locals():
                progress_window.destroy()
            messagebox.showerror("Full Backup Failed",
                f"Failed to create full backup:\n{str(e)}")

    def restore_from_backup(self):
        """Restore database from a backup file"""
        import shutil
        from datetime import datetime

        backups_dir = os.path.join(os.path.dirname(self.db_path), "database_backups")

        # Let user select backup file
        backup_file = filedialog.askopenfilename(
            title="Select Backup to Restore",
            initialdir=backups_dir if os.path.exists(backups_dir) else os.path.dirname(self.db_path),
            filetypes=[("Database files", "*.db"), ("All files", "*.*")]
        )

        if not backup_file:
            return

        # Confirm restoration
        result = messagebox.askyesno(
            "Confirm Restore",
            f"Restore database from:\n{os.path.basename(backup_file)}\n\n"
            f"⚠️ WARNING ⚠️\n"
            f"This will replace your current database!\n"
            f"All current data will be lost!\n\n"
            f"A backup of the current database will be created first.\n\n"
            f"Continue?",
            icon='warning'
        )

        if not result:
            return

        try:
            # Create backup of current database first
            current_backup_dir = os.path.join(os.path.dirname(self.db_path), "database_backups")
            os.makedirs(current_backup_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            current_backup = os.path.join(current_backup_dir, f"before_restore_{timestamp}.db")
            shutil.copy2(self.db_path, current_backup)

            # Restore from selected backup
            shutil.copy2(backup_file, self.db_path)

            messagebox.showinfo("Restore Complete",
                f"Database restored successfully!\n\n"
                f"Restored from: {os.path.basename(backup_file)}\n\n"
                f"Previous database saved as:\n{os.path.basename(current_backup)}\n\n"
                f"Please restart the application to load the restored database.")

        except Exception as e:
            messagebox.showerror("Restore Failed",
                f"Failed to restore database:\n{str(e)}")

    def view_backups(self):
        """View and manage database backups"""
        import shutil
        from datetime import datetime

        backups_dir = os.path.join(os.path.dirname(self.db_path), "database_backups")
        os.makedirs(backups_dir, exist_ok=True)

        # Create window
        backup_window = tk.Toplevel(self.root)
        backup_window.title("Database Backups")
        backup_window.geometry("900x600")
        backup_window.configure(bg=self.bg_color)

        tk.Label(backup_window,
                text="Database Backup Manager",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 14, "bold")).pack(pady=15)

        # Treeview for backup list
        tree_frame = tk.Frame(backup_window, bg=self.bg_color)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        columns = ('filename', 'records', 'date', 'size')
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)

        tree.heading('filename', text='Backup File')
        tree.heading('records', text='Records')
        tree.heading('date', text='Date Created')
        tree.heading('size', text='Size (MB)')

        tree.column('filename', width=350)
        tree.column('records', width=100)
        tree.column('date', width=200)
        tree.column('size', width=100)

        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscroll=scrollbar.set)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        def refresh_backup_list():
            # Clear existing items
            for item in tree.get_children():
                tree.delete(item)

            # Get backup files
            if not os.path.exists(backups_dir):
                tree.insert('', tk.END, values=("No backups found", "", "", ""))
                return

            backup_files = [f for f in os.listdir(backups_dir) if f.endswith('.db')]

            if not backup_files:
                tree.insert('', tk.END, values=("No backups found - use 'Backup Now' to create one", "", "", ""))
                return

            for backup_file in sorted(backup_files, reverse=True):
                backup_path = os.path.join(backups_dir, backup_file)

                # Get stats
                stat = os.stat(backup_path)
                size_mb = f"{stat.st_size / (1024 * 1024):.2f}"
                created = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")

                # Get record count
                try:
                    conn = sqlite3.connect(backup_path)
                    cursor = conn.cursor()
                    cursor.execute("SELECT COUNT(*) FROM programs")
                    count = cursor.fetchone()[0]
                    conn.close()
                except:
                    count = "Error"

                tree.insert('', tk.END, values=(backup_file, count, created, size_mb))

        def delete_backup():
            selection = tree.selection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a backup to delete.")
                return

            item = selection[0]
            filename = tree.item(item)['values'][0]

            if "No backups found" in filename:
                return

            result = messagebox.askyesno(
                "Confirm Delete",
                f"Delete backup '{filename}'?\n\n"
                f"This cannot be undone!",
                icon='warning'
            )

            if result:
                try:
                    backup_path = os.path.join(backups_dir, filename)
                    os.remove(backup_path)
                    refresh_backup_list()
                    messagebox.showinfo("Deleted", f"Backup '{filename}' deleted successfully.")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to delete backup:\n{str(e)}")

        def restore_selected():
            selection = tree.selection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a backup to restore.")
                return

            item = selection[0]
            filename = tree.item(item)['values'][0]

            if "No backups found" in filename:
                return

            backup_path = os.path.join(backups_dir, filename)

            result = messagebox.askyesno(
                "Confirm Restore",
                f"Restore database from:\n{filename}\n\n"
                f"⚠️ WARNING ⚠️\n"
                f"This will replace your current database!\n\n"
                f"A backup of current database will be created first.\n\n"
                f"Continue?",
                icon='warning'
            )

            if result:
                try:
                    # Backup current database
                    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                    current_backup = os.path.join(backups_dir, f"before_restore_{timestamp}.db")
                    shutil.copy2(self.db_path, current_backup)

                    # Restore
                    shutil.copy2(backup_path, self.db_path)

                    messagebox.showinfo("Restore Complete",
                        f"Database restored from {filename}!\n\n"
                        f"Current database backed up as:\n{os.path.basename(current_backup)}\n\n"
                        f"Please restart the application.")
                    backup_window.destroy()

                except Exception as e:
                    messagebox.showerror("Error", f"Failed to restore:\n{str(e)}")

        # Initial load
        refresh_backup_list()

        # Info label
        info_text = f"Backup location: {backups_dir}"
        tk.Label(backup_window, text=info_text,
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 9)).pack(pady=5)

        # Button frame
        btn_frame = tk.Frame(backup_window, bg=self.bg_color)
        btn_frame.pack(pady=15)

        tk.Button(btn_frame, text="💾 Backup Now", command=lambda: [self.create_manual_backup(), refresh_backup_list()],
                 bg="#1976D2", fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=14).pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="📂 Restore", command=restore_selected,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=14).pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="🗑️ Delete", command=delete_backup,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=14).pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="🔄 Refresh", command=refresh_backup_list,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=14).pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="❌ Close", command=backup_window.destroy,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=14).pack(side=tk.LEFT, padx=5)

    # ===== Repository Management Methods =====

    def show_repository_stats(self):
        """Show ONLY repository (managed) files statistics"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Only count managed files (is_managed = 1)
        cursor.execute("SELECT COUNT(*) FROM programs WHERE is_managed = 1")
        managed_count = cursor.fetchone()[0]

        # Count versions
        cursor.execute("SELECT COUNT(*) FROM program_versions")
        versions_count = cursor.fetchone()[0]

        # Repository size
        repo_size = 0
        if os.path.exists(self.repository_path):
            for file in os.listdir(self.repository_path):
                file_path = os.path.join(self.repository_path, file)
                if os.path.isfile(file_path):
                    repo_size += os.path.getsize(file_path)

        # Versions size
        versions_size = 0
        if os.path.exists(self.versions_path):
            for root, _, files in os.walk(self.versions_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    versions_size += os.path.getsize(file_path)

        conn.close()

        # Create stats dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Repository Statistics")
        dialog.geometry("500x350")
        dialog.configure(bg=self.bg_color)
        dialog.transient(self.root)
        dialog.grab_set()

        # Title
        tk.Label(dialog, text="📊 Repository Statistics",
                font=("Arial", 14, "bold"), bg=self.bg_color, fg=self.fg_color).pack(pady=10)

        tk.Label(dialog, text="(Managed files in repository/ folder only)",
                font=("Arial", 9, "italic"), bg=self.bg_color, fg="#888888").pack(pady=2)

        # Stats frame
        stats_frame = tk.Frame(dialog, bg=self.bg_color)
        stats_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # Display stats
        stats_data = [
            ("Repository Files:", f"{managed_count}"),
            ("", ""),
            ("Total Versions:", f"{versions_count}"),
            ("", ""),
            ("Repository Size:", f"{repo_size / (1024 * 1024):.2f} MB"),
            ("Versions Size:", f"{versions_size / (1024 * 1024):.2f} MB"),
            ("Total Storage:", f"{(repo_size + versions_size) / (1024 * 1024):.2f} MB"),
        ]

        for label, value in stats_data:
            if label:  # Skip empty rows for spacing
                row = tk.Frame(stats_frame, bg=self.bg_color)
                row.pack(fill=tk.X, pady=5)

                tk.Label(row, text=label, font=("Arial", 11), bg=self.bg_color, fg=self.fg_color,
                        anchor='w').pack(side=tk.LEFT)
                tk.Label(row, text=value, font=("Arial", 11, "bold"), bg=self.bg_color, fg=self.accent_color,
                        anchor='e').pack(side=tk.RIGHT)

        # Buttons
        btn_frame = tk.Frame(dialog, bg=self.bg_color)
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="🔄 Refresh", command=lambda: [dialog.destroy(), self.show_repository_stats()],
                 bg=self.accent_color, fg=self.fg_color, font=("Arial", 10, "bold"),
                 width=12, height=2).pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="Close", command=dialog.destroy,
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold"),
                 width=12, height=2).pack(side=tk.LEFT, padx=5)

    def show_all_programs_stats(self):
        """Show ALL programs statistics (repository + external combined)"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Count all programs
        cursor.execute("SELECT COUNT(*) FROM programs WHERE file_path IS NOT NULL")
        total_count = cursor.fetchone()[0]

        # Count managed files
        cursor.execute("SELECT COUNT(*) FROM programs WHERE is_managed = 1")
        managed_count = cursor.fetchone()[0]

        # Count external files
        external_count = total_count - managed_count

        # Count versions
        cursor.execute("SELECT COUNT(*) FROM program_versions")
        versions_count = cursor.fetchone()[0]

        # Repository size
        repo_size = 0
        if os.path.exists(self.repository_path):
            for file in os.listdir(self.repository_path):
                file_path = os.path.join(self.repository_path, file)
                if os.path.isfile(file_path):
                    repo_size += os.path.getsize(file_path)

        # Versions size
        versions_size = 0
        if os.path.exists(self.versions_path):
            for root, _, files in os.walk(self.versions_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    versions_size += os.path.getsize(file_path)

        conn.close()

        # Create stats dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("All Programs Statistics")
        dialog.geometry("500x400")
        dialog.configure(bg=self.bg_color)
        dialog.transient(self.root)
        dialog.grab_set()

        # Title
        tk.Label(dialog, text="📊 All Programs Statistics",
                font=("Arial", 14, "bold"), bg=self.bg_color, fg=self.fg_color).pack(pady=10)

        tk.Label(dialog, text="(Repository + External files combined)",
                font=("Arial", 9, "italic"), bg=self.bg_color, fg="#888888").pack(pady=2)

        # Stats frame
        stats_frame = tk.Frame(dialog, bg=self.bg_color)
        stats_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # Display stats
        managed_pct = (managed_count / max(total_count, 1)) * 100
        external_pct = (external_count / max(total_count, 1)) * 100

        stats_data = [
            ("Total Programs:", f"{total_count}"),
            ("", ""),
            ("Managed Files (Repository):", f"{managed_count} ({managed_pct:.1f}%)"),
            ("External Files:", f"{external_count} ({external_pct:.1f}%)"),
            ("", ""),
            ("Total Versions:", f"{versions_count}"),
            ("", ""),
            ("Repository Size:", f"{repo_size / (1024 * 1024):.2f} MB"),
            ("Versions Size:", f"{versions_size / (1024 * 1024):.2f} MB"),
            ("Total Storage:", f"{(repo_size + versions_size) / (1024 * 1024):.2f} MB"),
        ]

        for label, value in stats_data:
            if label:  # Skip empty rows for spacing
                row = tk.Frame(stats_frame, bg=self.bg_color)
                row.pack(fill=tk.X, pady=5)

                tk.Label(row, text=label, font=("Arial", 11), bg=self.bg_color, fg=self.fg_color,
                        anchor='w').pack(side=tk.LEFT)
                tk.Label(row, text=value, font=("Arial", 11, "bold"), bg=self.bg_color, fg=self.accent_color,
                        anchor='e').pack(side=tk.RIGHT)

        # Buttons
        btn_frame = tk.Frame(dialog, bg=self.bg_color)
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="🔄 Refresh", command=lambda: [dialog.destroy(), self.show_all_programs_stats()],
                 bg=self.accent_color, fg=self.fg_color, font=("Arial", 10, "bold"),
                 width=12, height=2).pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="Close", command=dialog.destroy,
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold"),
                 width=12, height=2).pack(side=tk.LEFT, padx=5)

    def show_external_stats(self):
        """Show ONLY external (non-managed) files statistics"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Only count external files (is_managed = 0 or NULL)
        cursor.execute("SELECT COUNT(*) FROM programs WHERE (is_managed = 0 OR is_managed IS NULL) AND file_path IS NOT NULL")
        external_count = cursor.fetchone()[0]

        conn.close()

        # Create stats dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("External Files Statistics")
        dialog.geometry("500x300")
        dialog.configure(bg=self.bg_color)
        dialog.transient(self.root)
        dialog.grab_set()

        # Title
        tk.Label(dialog, text="📊 External Files Statistics",
                font=("Arial", 14, "bold"), bg=self.bg_color, fg=self.fg_color).pack(pady=10)

        tk.Label(dialog, text="(Scanned files NOT in repository)",
                font=("Arial", 9, "italic"), bg=self.bg_color, fg="#888888").pack(pady=2)

        # Stats frame
        stats_frame = tk.Frame(dialog, bg=self.bg_color)
        stats_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # Display stats
        stats_data = [
            ("External Files:", f"{external_count}"),
            ("", ""),
            ("Note:", "External files remain in their original locations"),
            ("", "(USB drives, network folders, etc.)"),
        ]

        for label, value in stats_data:
            if label:  # Skip empty rows for spacing
                row = tk.Frame(stats_frame, bg=self.bg_color)
                row.pack(fill=tk.X, pady=5)

                if label == "Note:":
                    tk.Label(row, text=label, font=("Arial", 11, "italic"), bg=self.bg_color, fg="#888888",
                            anchor='w').pack(side=tk.LEFT)
                    tk.Label(row, text=value, font=("Arial", 11, "italic"), bg=self.bg_color, fg="#888888",
                            anchor='e').pack(side=tk.RIGHT)
                else:
                    tk.Label(row, text=label, font=("Arial", 11), bg=self.bg_color, fg=self.fg_color,
                            anchor='w').pack(side=tk.LEFT)
                    tk.Label(row, text=value, font=("Arial", 11, "bold"), bg=self.bg_color, fg=self.accent_color,
                            anchor='e').pack(side=tk.RIGHT)

        # Buttons
        btn_frame = tk.Frame(dialog, bg=self.bg_color)
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="🔄 Refresh", command=lambda: [dialog.destroy(), self.show_external_stats()],
                 bg=self.accent_color, fg=self.fg_color, font=("Arial", 10, "bold"),
                 width=12, height=2).pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="Close", command=dialog.destroy,
                 bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold"),
                 width=12, height=2).pack(side=tk.LEFT, padx=5)

    def delete_from_repository(self):
        """Delete selected program from repository (and optionally from database)"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a program to delete")
            return

        item = self.tree.item(selection[0])
        values = item['values']
        program_number = values[0]

        # Get program info
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT file_path, is_managed, title FROM programs WHERE program_number = ?",
                      (program_number,))
        result = cursor.fetchone()
        conn.close()

        if not result:
            messagebox.showerror("Error", "Program not found in database")
            return

        file_path, is_managed, title = result

        if not is_managed:
            messagebox.showinfo("Not in Repository",
                              f"{program_number} is not in the repository.\n\n"
                              "Use 'Remove from Database' in the External tab instead.")
            return

        # Confirm deletion
        response = messagebox.askyesnocancel(
            "Delete from Repository",
            f"Delete {program_number} from repository?\n\n"
            f"Title: {title}\n"
            f"File: {file_path}\n\n"
            "Choose:\n"
            "• YES - Delete file from repository AND remove from database\n"
            "• NO - Delete file from repository but keep database entry\n"
            "• CANCEL - Don't delete anything"
        )

        if response is None:  # Cancel
            return

        try:
            # Delete the file from repository
            if file_path and os.path.exists(file_path):
                os.remove(file_path)
                print(f"[Repository] Deleted file: {file_path}")

            if response:  # YES - also remove from database
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute("DELETE FROM programs WHERE program_number = ?", (program_number,))
                conn.commit()
                conn.close()

                self.log_activity('delete_from_repository', program_number, {
                    'file_path': file_path,
                    'removed_from_db': True
                })

                messagebox.showinfo("Success", f"{program_number} deleted from repository and database")
            else:  # NO - just update is_managed flag
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute("UPDATE programs SET is_managed = 0, file_path = NULL WHERE program_number = ?",
                             (program_number,))
                conn.commit()
                conn.close()

                self.log_activity('delete_from_repository', program_number, {
                    'file_path': file_path,
                    'removed_from_db': False
                })

                messagebox.showinfo("Success", f"{program_number} deleted from repository (database entry kept)")

            # Refresh the view
            current_tab = self.view_notebook.index(self.view_notebook.select())
            if current_tab == 1:
                self.refresh_results(view_mode='repository')
            else:
                self.refresh_results(view_mode='all')

        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete file:\n{e}")

    def add_selected_to_repository(self):
        """Add selected external program to repository"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a program to add to repository")
            return

        item = self.tree.item(selection[0])
        values = item['values']
        program_number = values[0]

        # Migrate to repository
        success = self.migrate_file_to_repository(program_number)

        if success:
            messagebox.showinfo("Success", f"{program_number} added to repository")
            # Refresh view
            current_tab = self.view_notebook.index(self.view_notebook.select())
            if current_tab == 2:
                self.refresh_results(view_mode='external')
            else:
                self.refresh_results(view_mode='all')
        else:
            messagebox.showerror("Error", f"Failed to add {program_number} to repository")

    def remove_from_database(self):
        """Remove selected external program from database (doesn't delete file)"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a program to remove")
            return

        item = self.tree.item(selection[0])
        values = item['values']
        program_number = values[0]

        # Get program info
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT file_path, is_managed, title FROM programs WHERE program_number = ?",
                      (program_number,))
        result = cursor.fetchone()
        conn.close()

        if not result:
            messagebox.showerror("Error", "Program not found in database")
            return

        file_path, is_managed, title = result

        if is_managed:
            messagebox.showinfo("Repository File",
                              f"{program_number} is in the repository.\n\n"
                              "Use 'Delete from Repository' in the Repository tab instead.")
            return

        # Confirm removal
        response = messagebox.askyesno(
            "Remove from Database",
            f"Remove {program_number} from database?\n\n"
            f"Title: {title}\n"
            f"File: {file_path}\n\n"
            "The external file will NOT be deleted,\n"
            "only the database entry will be removed."
        )

        if not response:
            return

        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM programs WHERE program_number = ?", (program_number,))
            conn.commit()
            conn.close()

            self.log_activity('remove_from_database', program_number, {
                'file_path': file_path,
                'was_external': True
            })

            messagebox.showinfo("Success", f"{program_number} removed from database")

            # Refresh view
            current_tab = self.view_notebook.index(self.view_notebook.select())
            if current_tab == 2:
                self.refresh_results(view_mode='external')
            else:
                self.refresh_results(view_mode='all')

        except Exception as e:
            messagebox.showerror("Error", f"Failed to remove from database:\n{e}")

    def export_selected_file(self):
        """Export selected file from repository to a location"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a file to export")
            return

        item = self.tree.item(selection[0])
        values = item['values']
        program_number = values[0]

        # Get file path
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT file_path, title FROM programs WHERE program_number = ?", (program_number,))
        result = cursor.fetchone()
        conn.close()

        if not result or not result[0]:
            messagebox.showerror("Error", "File path not found")
            return

        source_file = result[0]

        if not os.path.exists(source_file):
            messagebox.showerror("Error", f"Source file not found:\n{source_file}")
            return

        # Ask for destination
        default_filename = os.path.basename(source_file)
        dest_file = filedialog.asksaveasfilename(
            title="Export File",
            initialfile=default_filename,
            defaultextension=os.path.splitext(default_filename)[1],
            filetypes=[
                ("G-Code files", "*.nc *.gcode"),
                ("All files", "*.*")
            ]
        )

        if not dest_file:
            return

        try:
            shutil.copy2(source_file, dest_file)
            self.log_activity('export_file', program_number, {
                'source': source_file,
                'destination': dest_file
            })
            messagebox.showinfo("Success", f"File exported to:\n{dest_file}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export file:\n{e}")

    def manage_duplicates(self):
        """Master duplicate management dialog - choose which type to manage"""
        # Create management dialog
        mgmt_window = tk.Toplevel(self.root)
        mgmt_window.title("Duplicate Management")
        mgmt_window.geometry("750x600")
        mgmt_window.configure(bg=self.bg_color)
        mgmt_window.transient(self.root)

        # Header
        tk.Label(mgmt_window, text="🔍 Duplicate Management",
                font=("Arial", 14, "bold"), bg=self.bg_color, fg=self.fg_color).pack(pady=8)

        tk.Label(mgmt_window, text="Choose the type of duplicates to manage:",
                font=("Arial", 9), bg=self.bg_color, fg=self.fg_color).pack(pady=3)

        # Create scrollable canvas
        canvas = tk.Canvas(mgmt_window, bg=self.bg_color, highlightthickness=0)
        scrollbar = ttk.Scrollbar(mgmt_window, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=self.bg_color)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Main options frame (inside scrollable canvas)
        options_frame = scrollable_frame

        # Option 1: Content Duplicates (Type 2 & 3)
        content_frame = tk.LabelFrame(options_frame, text="STEP 1: Content Duplicates (Type 2 & 3) - DO THIS FIRST",
                                     bg=self.bg_color, fg=self.fg_color,
                                     font=("Arial", 10, "bold"), relief=tk.RIDGE, bd=1)
        content_frame.pack(fill=tk.X, pady=5, padx=10)

        tk.Label(content_frame, text="Files with identical content (same OR different names)",
                bg=self.bg_color, fg=self.accent_color,
                font=("Arial", 8)).pack(anchor=tk.W, padx=8, pady=2)

        tk.Label(content_frame,
                text="• Type 2: o12345, o12345, o12345 (same name) • Type 3: o62000, o62500 (different names)",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 8), justify=tk.LEFT).pack(anchor=tk.W, padx=15, pady=2)

        tk.Label(content_frame,
                text="• SHA256 hash match • Keeps parent/lowest # • Moves to deleted/",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 8), justify=tk.LEFT).pack(anchor=tk.W, padx=15, pady=2)

        tk.Button(content_frame, text="🗑️ Delete Content Duplicates (Type 2 & 3)",
                 command=lambda: (mgmt_window.destroy(), self.delete_content_duplicates()),
                 bg="#E65100", fg=self.fg_color,
                 font=("Arial", 9, "bold"), width=35).pack(pady=5)

        # Option 2: Name Duplicates (Type 1)
        name_frame = tk.LabelFrame(options_frame, text="STEP 2: Name Conflicts (Type 1) - DO THIS SECOND",
                                  bg=self.bg_color, fg=self.fg_color,
                                  font=("Arial", 10, "bold"), relief=tk.RIDGE, bd=1)
        name_frame.pack(fill=tk.X, pady=5, padx=10)

        tk.Label(name_frame, text="Same base name but different content (versions/revisions)",
                bg=self.bg_color, fg=self.accent_color,
                font=("Arial", 8)).pack(anchor=tk.W, padx=8, pady=2)

        tk.Label(name_frame,
                text="• Type 1: o12345, o12345(1), o12345(2) with different dimensions",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 8), justify=tk.LEFT).pack(anchor=tk.W, padx=15, pady=2)

        tk.Label(name_frame,
                text="• Keeps 1st file • Renames others to correct range • Updates file + database + registry",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 8), justify=tk.LEFT).pack(anchor=tk.W, padx=15, pady=2)

        tk.Button(name_frame, text="✏️ Rename Name Conflicts (Type 1)",
                 command=lambda: (mgmt_window.destroy(), self.rename_name_duplicates()),
                 bg="#7B1FA2", fg=self.fg_color,
                 font=("Arial", 9, "bold"), width=35).pack(pady=5)

        # Option 3: Fix Underscore Suffixes
        underscore_frame = tk.LabelFrame(options_frame, text="STEP 3: Fix Underscore Suffixes - CLEANUP",
                                        bg=self.bg_color, fg=self.fg_color,
                                        font=("Arial", 10, "bold"), relief=tk.RIDGE, bd=1)
        underscore_frame.pack(fill=tk.X, pady=5, padx=10)

        tk.Label(underscore_frame, text="Files with underscore suffixes (o12345_1.nc, o12345_2.nc)",
                bg=self.bg_color, fg=self.accent_color,
                font=("Arial", 8)).pack(anchor=tk.W, padx=8, pady=2)

        tk.Label(underscore_frame,
                text="• Finds: o#####_#.nc patterns • Renames to correct range • Updates database + registry",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 8), justify=tk.LEFT).pack(anchor=tk.W, padx=15, pady=2)

        tk.Button(underscore_frame, text="🔧 Fix Underscore Suffix Files",
                 command=lambda: (mgmt_window.destroy(), self.fix_underscore_suffix_files()),
                 bg="#FF6B00", fg=self.fg_color,
                 font=("Arial", 9, "bold"), width=35).pack(pady=5)

        # Option 4: Scan/Report Only
        scan_frame = tk.LabelFrame(options_frame, text="Scan for Duplicates",
                                  bg=self.bg_color, fg=self.fg_color,
                                  font=("Arial", 10, "bold"), relief=tk.RIDGE, bd=1)
        scan_frame.pack(fill=tk.X, pady=5, padx=10)

        tk.Label(scan_frame, text="Report all duplicates (no changes)",
                bg=self.bg_color, fg=self.accent_color,
                font=("Arial", 8)).pack(anchor=tk.W, padx=8, pady=2)

        tk.Label(scan_frame,
                text="• Read-only • Content + Name • Safe preview of duplicates",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 8), justify=tk.LEFT).pack(anchor=tk.W, padx=15, pady=2)

        tk.Button(scan_frame, text="🔍 Scan for Duplicates (Report Only)",
                 command=lambda: (mgmt_window.destroy(), self.scan_for_duplicates()),
                 bg="#1976D2", fg=self.fg_color,
                 font=("Arial", 9, "bold"), width=28).pack(pady=5)

        # Option 4: Delete REPEAT Status Files
        repeat_frame = tk.LabelFrame(options_frame, text="REPEAT Status Files",
                                    bg=self.bg_color, fg=self.fg_color,
                                    font=("Arial", 10, "bold"), relief=tk.RIDGE, bd=1)
        repeat_frame.pack(fill=tk.X, pady=5, padx=10)

        tk.Label(repeat_frame, text="Files marked REPEAT by 'Find Repeats'",
                bg=self.bg_color, fg=self.accent_color,
                font=("Arial", 8)).pack(anchor=tk.W, padx=8, pady=2)

        tk.Label(repeat_frame,
                text="• validation_status='REPEAT' • Keeps parents • Moves to deleted/",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 8), justify=tk.LEFT).pack(anchor=tk.W, padx=15, pady=2)

        button_frame = tk.Frame(repeat_frame, bg=self.bg_color)
        button_frame.pack(pady=5)

        tk.Button(button_frame, text="🔍 Re-scan",
                 command=lambda: (mgmt_window.destroy(), self.rescan_repository_duplicates()),
                 bg="#1976D2", fg=self.fg_color,
                 font=("Arial", 8, "bold"), width=18).pack(side=tk.LEFT, padx=3)

        tk.Button(button_frame, text="🗑️ Clean",
                 command=lambda: (mgmt_window.destroy(), self.delete_repeat_status_files()),
                 bg="#C62828", fg=self.fg_color,
                 font=("Arial", 8, "bold"), width=18).pack(side=tk.LEFT, padx=3)

        # Close button (outside scrollable frame)
        close_frame = tk.Frame(mgmt_window, bg=self.bg_color)
        close_frame.pack(side=tk.BOTTOM, pady=5)

        tk.Button(close_frame, text="Cancel", command=mgmt_window.destroy,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 9), width=12).pack()

    def rescan_repository_duplicates(self):
        """Re-scan all repository files and update REPEAT status based on current duplicate detection"""
        # Create progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Re-scan Repository for Duplicates")
        progress_window.geometry("1000x700")
        progress_window.configure(bg=self.bg_color)
        progress_window.transient(self.root)

        tk.Label(progress_window, text="🔍 Re-scan Repository for Duplicates",
                font=("Arial", 14, "bold"), bg=self.bg_color, fg=self.fg_color).pack(pady=10)

        # Progress text
        text_frame = tk.Frame(progress_window, bg=self.bg_color)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        progress_text = tk.Text(text_frame, bg=self.input_bg, fg=self.fg_color,
                               font=("Consolas", 9), wrap=tk.WORD)
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=progress_text.yview)
        progress_text.configure(yscrollcommand=scrollbar.set)

        progress_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        progress_text.insert(tk.END, "Re-scanning repository files for duplicates...\n\n")
        progress_text.insert(tk.END, "This will verify which files should have REPEAT status.\n\n")
        progress_text.see(tk.END)
        self.root.update()

        try:
            import hashlib
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # Get all repository files
            cursor.execute("""
                SELECT program_number, file_path, title, duplicate_type, parent_file, validation_status
                FROM programs
                WHERE is_managed = 1
                ORDER BY program_number
            """)

            all_files = cursor.fetchall()
            progress_text.insert(tk.END, f"Found {len(all_files)} files in repository\n\n")
            progress_text.see(tk.END)
            self.root.update()

            # Calculate hashes for all files to detect content duplicates
            progress_text.insert(tk.END, "Step 1: Calculating file hashes for content analysis...\n")
            progress_text.see(tk.END)
            self.root.update()

            file_hashes = {}  # hash -> [(prog_num, title, file_path, old_status), ...]
            missing_files = []

            for idx, (prog_num, file_path, title, dup_type, parent_file, val_status) in enumerate(all_files, 1):
                if idx % 50 == 0:
                    progress_text.insert(tk.END, f"  Processed {idx}/{len(all_files)} files...\n")
                    progress_text.see(tk.END)
                    self.root.update()

                if file_path and os.path.exists(file_path):
                    try:
                        sha256_hash = hashlib.sha256()
                        with open(file_path, 'rb') as f:
                            for byte_block in iter(lambda: f.read(4096), b""):
                                sha256_hash.update(byte_block)
                        file_hash = sha256_hash.hexdigest()

                        if file_hash not in file_hashes:
                            file_hashes[file_hash] = []
                        file_hashes[file_hash].append((prog_num, title, file_path, val_status, dup_type, parent_file))
                    except Exception as e:
                        progress_text.insert(tk.END, f"  ⚠️ Error reading {prog_num}: {e}\n")
                else:
                    missing_files.append(prog_num)

            progress_text.insert(tk.END, f"\nCompleted hash calculation for {len(all_files) - len(missing_files)} files\n")
            if missing_files:
                progress_text.insert(tk.END, f"⚠️ Skipped {len(missing_files)} missing files\n")
            progress_text.insert(tk.END, f"\n")
            progress_text.see(tk.END)
            self.root.update()

            # Analyze duplicates
            progress_text.insert(tk.END, "Step 2: Analyzing duplicates...\n\n")
            progress_text.see(tk.END)
            self.root.update()

            content_duplicates = {h: files for h, files in file_hashes.items() if len(files) > 1}

            progress_text.insert(tk.END, f"Found {len(content_duplicates)} groups with duplicate content\n")
            progress_text.insert(tk.END, f"Total duplicate files: {sum(len(files) - 1 for files in content_duplicates.values())}\n\n")

            # Determine which files should be marked as REPEAT
            files_to_mark_repeat = []
            files_to_clear_repeat = []

            for file_hash, duplicates in content_duplicates.items():
                # Sort by: parent files first, then by program number
                parent_files = [f for f in duplicates if f[4] == 'parent']  # dup_type == 'parent'
                child_files = [f for f in duplicates if f[4] != 'parent']

                if parent_files:
                    # Keep first parent, mark others as REPEAT
                    keeper = sorted(parent_files, key=lambda x: x[0])[0]  # Lowest program number parent

                    for dup in parent_files[1:] if len(parent_files) > 1 else []:
                        if dup[3] != 'REPEAT':  # val_status
                            files_to_mark_repeat.append((dup[0], keeper[0], 'duplicate parent'))

                    for dup in child_files:
                        if dup[3] != 'REPEAT':
                            files_to_mark_repeat.append((dup[0], keeper[0], 'duplicate child'))
                else:
                    # No parent files, keep lowest program number
                    keeper = sorted(duplicates, key=lambda x: x[0])[0]

                    for dup in duplicates[1:]:
                        if dup[3] != 'REPEAT':
                            files_to_mark_repeat.append((dup[0], keeper[0], 'content duplicate'))

            # Find files marked as REPEAT but are NOT duplicates anymore
            for file_hash, files in file_hashes.items():
                if len(files) == 1:  # Only one file with this hash
                    prog_num, title, file_path, val_status, dup_type, parent_file = files[0]
                    if val_status == 'REPEAT':
                        files_to_clear_repeat.append(prog_num)

            progress_text.insert(tk.END, f"{'='*90}\n")
            progress_text.insert(tk.END, f"ANALYSIS RESULTS\n")
            progress_text.insert(tk.END, f"{'='*90}\n\n")
            progress_text.insert(tk.END, f"Files to mark as REPEAT: {len(files_to_mark_repeat)}\n")
            progress_text.insert(tk.END, f"Files to clear REPEAT status: {len(files_to_clear_repeat)}\n\n")

            # Show preview
            if files_to_mark_repeat:
                progress_text.insert(tk.END, f"Files to mark as REPEAT (first 30):\n")
                for prog_num, parent, reason in files_to_mark_repeat[:30]:
                    progress_text.insert(tk.END, f"  → {prog_num} ({reason} of {parent})\n")
                if len(files_to_mark_repeat) > 30:
                    progress_text.insert(tk.END, f"  ... and {len(files_to_mark_repeat) - 30} more\n")
                progress_text.insert(tk.END, f"\n")

            if files_to_clear_repeat:
                progress_text.insert(tk.END, f"Files to clear REPEAT status (not duplicates):\n")
                for prog_num in files_to_clear_repeat[:30]:
                    progress_text.insert(tk.END, f"  ✓ {prog_num} (unique file)\n")
                if len(files_to_clear_repeat) > 30:
                    progress_text.insert(tk.END, f"  ... and {len(files_to_clear_repeat) - 30} more\n")

            progress_text.insert(tk.END, f"\n")
            progress_text.see(tk.END)
            self.root.update()

            # Confirmation buttons
            def confirm_update():
                progress_text.insert(tk.END, f"\n{'='*90}\n")
                progress_text.insert(tk.END, f"UPDATING VALIDATION STATUS\n")
                progress_text.insert(tk.END, f"{'='*90}\n\n")
                progress_text.see(tk.END)

                marked_count = 0
                cleared_count = 0

                # Mark files as REPEAT
                for prog_num, parent, reason in files_to_mark_repeat:
                    try:
                        cursor.execute("""
                            UPDATE programs
                            SET validation_status = 'REPEAT',
                                parent_file = ?,
                                duplicate_type = 'child'
                            WHERE program_number = ?
                        """, (parent, prog_num))
                        marked_count += 1

                        if marked_count % 50 == 0:
                            progress_text.insert(tk.END, f"  Marked {marked_count} files as REPEAT...\n")
                            progress_text.see(tk.END)
                            self.root.update()
                    except Exception as e:
                        progress_text.insert(tk.END, f"  ✗ ERROR marking {prog_num}: {e}\n")

                # Clear REPEAT status from unique files
                for prog_num in files_to_clear_repeat:
                    try:
                        # Reset to PASS status (or you could keep original status)
                        cursor.execute("""
                            UPDATE programs
                            SET validation_status = 'PASS',
                                parent_file = NULL,
                                duplicate_type = NULL
                            WHERE program_number = ?
                        """, (prog_num,))
                        cleared_count += 1

                        if cleared_count % 50 == 0:
                            progress_text.insert(tk.END, f"  Cleared {cleared_count} REPEAT statuses...\n")
                            progress_text.see(tk.END)
                            self.root.update()
                    except Exception as e:
                        progress_text.insert(tk.END, f"  ✗ ERROR clearing {prog_num}: {e}\n")

                conn.commit()

                progress_text.insert(tk.END, f"\n{'='*90}\n")
                progress_text.insert(tk.END, f"COMPLETE\n")
                progress_text.insert(tk.END, f"{'='*90}\n\n")
                progress_text.insert(tk.END, f"✓ Marked as REPEAT: {marked_count} files\n")
                progress_text.insert(tk.END, f"✓ Cleared REPEAT status: {cleared_count} files\n")
                progress_text.see(tk.END)

                # Log activity
                self.log_activity('rescan_repository_duplicates', 'batch', {
                    'marked_repeat': marked_count,
                    'cleared_repeat': cleared_count
                })

                # Refresh the view
                self.refresh_results()

                # Close database connection
                conn.close()

                # Update button
                btn_frame.pack_forget()
                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)

            def cancel_update():
                conn.close()
                progress_window.destroy()

            # Button frame
            btn_frame = tk.Frame(progress_window, bg=self.bg_color)
            btn_frame.pack(pady=10)

            tk.Button(btn_frame, text="✓ Update Status", command=confirm_update,
                     bg="#1976D2", fg=self.fg_color, font=("Arial", 10, "bold"),
                     width=18, height=2).pack(side=tk.LEFT, padx=10)

            tk.Button(btn_frame, text="✗ Cancel", command=cancel_update,
                     bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold"),
                     width=18, height=2).pack(side=tk.LEFT, padx=10)

        except Exception as e:
            progress_text.insert(tk.END, f"\n\nERROR: {e}\n")
            import traceback
            progress_text.insert(tk.END, traceback.format_exc())
            progress_text.see(tk.END)

            tk.Button(progress_window, text="Close", command=progress_window.destroy,
                     bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)

    def delete_repeat_status_files(self):
        """Remove files with REPEAT validation status from repository (moves to deleted folder)"""
        # Create progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Clean REPEAT Status Files")
        progress_window.geometry("900x700")
        progress_window.configure(bg=self.bg_color)
        progress_window.transient(self.root)

        tk.Label(progress_window, text="🗑️ Clean REPEAT Status Files",
                font=("Arial", 14, "bold"), bg=self.bg_color, fg=self.fg_color).pack(pady=10)

        # Progress text
        text_frame = tk.Frame(progress_window, bg=self.bg_color)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        progress_text = tk.Text(text_frame, bg=self.input_bg, fg=self.fg_color,
                               font=("Consolas", 9), wrap=tk.WORD)
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=progress_text.yview)
        progress_text.configure(yscrollcommand=scrollbar.set)

        progress_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        progress_text.insert(tk.END, "Analyzing REPEAT status files in repository...\n\n")
        progress_text.see(tk.END)
        self.root.update()

        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            # Get all REPEAT status files from repository
            cursor.execute("""
                SELECT program_number, file_path, title, parent_file
                FROM programs
                WHERE is_managed = 1 AND validation_status = 'REPEAT'
                ORDER BY program_number
            """)

            repeat_files = cursor.fetchall()
            progress_text.insert(tk.END, f"Found {len(repeat_files)} REPEAT status files in repository\n\n")
            progress_text.see(tk.END)
            self.root.update()

            if not repeat_files:
                progress_text.insert(tk.END, "✓ No REPEAT status files found!\n\n")
                progress_text.insert(tk.END, "All repository files have other validation statuses.\n")
                progress_text.see(tk.END)

                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)
                return

            # Show preview of what will be removed
            progress_text.insert(tk.END, f"{'='*80}\n")
            progress_text.insert(tk.END, f"REPEAT FILES TO BE REMOVED\n")
            progress_text.insert(tk.END, f"{'='*80}\n\n")

            for prog_num, file_path, title, parent_file in repeat_files[:50]:
                progress_text.insert(tk.END, f"  ✗ REMOVE: {prog_num}")
                if parent_file:
                    progress_text.insert(tk.END, f" (child of {parent_file})")
                progress_text.insert(tk.END, f"\n     Title: {title[:60]}\n")

            if len(repeat_files) > 50:
                progress_text.insert(tk.END, f"\n... and {len(repeat_files) - 50} more files\n")

            progress_text.insert(tk.END, f"\n{'='*80}\n")
            progress_text.insert(tk.END, f"SUMMARY\n")
            progress_text.insert(tk.END, f"{'='*80}\n\n")
            progress_text.insert(tk.END, f"Files to remove: {len(repeat_files)}\n")
            progress_text.insert(tk.END, f"Action: Move to deleted/ folder\n\n")
            progress_text.see(tk.END)
            self.root.update()

            # Confirmation buttons
            def confirm_clean():
                progress_text.insert(tk.END, f"\n{'='*80}\n")
                progress_text.insert(tk.END, f"REMOVING REPEAT FILES\n")
                progress_text.insert(tk.END, f"{'='*80}\n\n")
                progress_text.see(tk.END)

                removed_count = 0
                error_count = 0

                for prog_num, file_path, title, parent_file in repeat_files:
                    try:
                        # Move physical file to deleted folder
                        if file_path and os.path.exists(file_path):
                            import shutil, time
                            filename = os.path.basename(file_path)
                            deleted_file_path = os.path.join(self.deleted_path, filename)

                            # If file already exists in deleted folder, add timestamp
                            if os.path.exists(deleted_file_path):
                                timestamp = time.strftime("%Y%m%d_%H%M%S")
                                name, ext = os.path.splitext(filename)
                                deleted_file_path = os.path.join(self.deleted_path, f"{name}_{timestamp}{ext}")

                            shutil.move(file_path, deleted_file_path)

                        # Remove from database
                        cursor.execute("DELETE FROM programs WHERE program_number = ?", (prog_num,))

                        progress_text.insert(tk.END, f"  ✓ Removed: {prog_num}\n")
                        removed_count += 1

                        if removed_count % 10 == 0:
                            progress_text.see(tk.END)
                            self.root.update()

                    except Exception as e:
                        progress_text.insert(tk.END, f"  ✗ ERROR removing {prog_num}: {e}\n")
                        error_count += 1

                conn.commit()

                progress_text.insert(tk.END, f"\n{'='*80}\n")
                progress_text.insert(tk.END, f"COMPLETE\n")
                progress_text.insert(tk.END, f"{'='*80}\n\n")
                progress_text.insert(tk.END, f"Successfully removed: {removed_count} files\n")
                if error_count > 0:
                    progress_text.insert(tk.END, f"Errors: {error_count} files\n")
                progress_text.insert(tk.END, f"\nFiles moved to: deleted/\n")
                progress_text.see(tk.END)

                # Log activity
                self.log_activity('delete_repeat_status_files', 'batch', {
                    'removed_count': removed_count,
                    'error_count': error_count
                })

                # Refresh the view
                self.refresh_results()

                # Close database connection
                conn.close()

                # Update button
                btn_frame.pack_forget()
                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)

            def cancel_clean():
                conn.close()
                progress_window.destroy()

            # Button frame
            btn_frame = tk.Frame(progress_window, bg=self.bg_color)
            btn_frame.pack(pady=10)

            tk.Button(btn_frame, text="✓ Confirm Clean", command=confirm_clean,
                     bg="#C62828", fg=self.fg_color, font=("Arial", 10, "bold"),
                     width=18, height=2).pack(side=tk.LEFT, padx=10)

            tk.Button(btn_frame, text="✗ Cancel", command=cancel_clean,
                     bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold"),
                     width=18, height=2).pack(side=tk.LEFT, padx=10)

        except Exception as e:
            progress_text.insert(tk.END, f"\n\nERROR: {e}\n")
            import traceback
            progress_text.insert(tk.END, traceback.format_exc())
            progress_text.see(tk.END)

            tk.Button(progress_window, text="Close", command=progress_window.destroy,
                     bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)

    def delete_content_duplicates(self):
        """Find and delete files with duplicate content (same hash), keeping parent files"""
        import hashlib

        # Create progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Delete Content Duplicates")
        progress_window.geometry("900x700")
        progress_window.configure(bg=self.bg_color)
        progress_window.transient(self.root)

        tk.Label(progress_window, text="🧹 Delete Content Duplicates",
                font=("Arial", 14, "bold"), bg=self.bg_color, fg=self.fg_color).pack(pady=10)

        # Progress text
        text_frame = tk.Frame(progress_window, bg=self.bg_color)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        progress_text = tk.Text(text_frame, bg=self.input_bg, fg=self.fg_color,
                               font=("Consolas", 9), wrap=tk.WORD)
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=progress_text.yview)
        progress_text.configure(yscrollcommand=scrollbar.set)

        progress_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        progress_text.insert(tk.END, "Analyzing repository files for content duplicates...\n\n")
        progress_text.see(tk.END)
        self.root.update()

        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            # Get all managed files from repository
            cursor.execute("""
                SELECT program_number, file_path, duplicate_type, parent_file
                FROM programs
                WHERE is_managed = 1
                ORDER BY program_number
            """)

            all_files = cursor.fetchall()
            progress_text.insert(tk.END, f"Found {len(all_files)} files in repository\n\n")
            progress_text.see(tk.END)
            self.root.update()

            # Calculate hashes for all files
            progress_text.insert(tk.END, "Calculating file hashes...\n")
            progress_text.see(tk.END)
            self.root.update()

            file_hashes = {}  # hash -> list of (program_number, file_path, duplicate_type, parent_file)
            missing_files = []

            for idx, (prog_num, file_path, dup_type, parent_file) in enumerate(all_files, 1):
                if idx % 50 == 0:
                    progress_text.insert(tk.END, f"  Processed {idx}/{len(all_files)} files...\n")
                    progress_text.see(tk.END)
                    self.root.update()

                if not file_path or not os.path.exists(file_path):
                    missing_files.append((prog_num, file_path))
                    continue

                try:
                    # Calculate SHA256 hash
                    sha256_hash = hashlib.sha256()
                    with open(file_path, 'rb') as f:
                        for byte_block in iter(lambda: f.read(4096), b""):
                            sha256_hash.update(byte_block)
                    file_hash = sha256_hash.hexdigest()

                    if file_hash not in file_hashes:
                        file_hashes[file_hash] = []
                    file_hashes[file_hash].append((prog_num, file_path, dup_type, parent_file))
                except Exception as e:
                    progress_text.insert(tk.END, f"  ERROR reading {prog_num}: {e}\n")

            progress_text.insert(tk.END, f"\nHash calculation complete!\n\n")
            progress_text.see(tk.END)
            self.root.update()

            # Find duplicate groups (hashes with multiple files)
            duplicate_groups = {h: files for h, files in file_hashes.items() if len(files) > 1}

            progress_text.insert(tk.END, f"{'='*80}\n")
            progress_text.insert(tk.END, f"DUPLICATE ANALYSIS\n")
            progress_text.insert(tk.END, f"{'='*80}\n\n")
            progress_text.insert(tk.END, f"Total files analyzed: {len(all_files)}\n")
            progress_text.insert(tk.END, f"Unique content: {len(file_hashes)}\n")
            progress_text.insert(tk.END, f"Duplicate groups: {len(duplicate_groups)}\n")

            if missing_files:
                progress_text.insert(tk.END, f"Missing files: {len(missing_files)}\n")

            progress_text.insert(tk.END, f"\n")
            progress_text.see(tk.END)
            self.root.update()

            if not duplicate_groups:
                progress_text.insert(tk.END, "✓ No content duplicates found!\n\n")
                progress_text.insert(tk.END, "All files in repository have unique content.\n")
                progress_text.see(tk.END)
                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)
                conn.close()
                return

            # Analyze each duplicate group and determine what to delete
            files_to_delete = []  # List of (program_number, file_path, reason)

            progress_text.insert(tk.END, f"{'='*80}\n")
            progress_text.insert(tk.END, f"DUPLICATE GROUPS\n")
            progress_text.insert(tk.END, f"{'='*80}\n\n")

            for hash_val, duplicates in duplicate_groups.items():
                progress_text.insert(tk.END, f"Group ({len(duplicates)} files with identical content):\n")

                # Sort by: parent files first, then by program number
                # Parent files (duplicate_type = 'parent' or parent_file IS NULL) should be kept
                parent_files = [f for f in duplicates if f[2] == 'parent' or not f[3]]
                child_files = [f for f in duplicates if f[2] != 'parent' and f[3]]

                if parent_files:
                    # Keep the first parent (or lowest program number if multiple parents)
                    parent_files_sorted = sorted(parent_files, key=lambda x: x[0])
                    keeper = parent_files_sorted[0]
                    progress_text.insert(tk.END, f"  ✓ KEEP: {keeper[0]} (parent/original)\n")

                    # Mark other parents as duplicates to delete
                    for dup in parent_files_sorted[1:]:
                        files_to_delete.append((dup[0], dup[1], f"duplicate of parent {keeper[0]}"))
                        progress_text.insert(tk.END, f"  ✗ DELETE: {dup[0]} (duplicate parent)\n")

                    # Mark all children as duplicates to delete
                    for dup in child_files:
                        files_to_delete.append((dup[0], dup[1], f"duplicate of parent {keeper[0]}"))
                        progress_text.insert(tk.END, f"  ✗ DELETE: {dup[0]} (duplicate child)\n")
                else:
                    # No parent files, keep the one with lowest program number
                    duplicates_sorted = sorted(duplicates, key=lambda x: x[0])
                    keeper = duplicates_sorted[0]
                    progress_text.insert(tk.END, f"  ✓ KEEP: {keeper[0]} (lowest program number)\n")

                    for dup in duplicates_sorted[1:]:
                        files_to_delete.append((dup[0], dup[1], f"duplicate of {keeper[0]}"))
                        progress_text.insert(tk.END, f"  ✗ DELETE: {dup[0]}\n")

                progress_text.insert(tk.END, f"\n")
                progress_text.see(tk.END)
                self.root.update()

            # Summary
            progress_text.insert(tk.END, f"{'='*80}\n")
            progress_text.insert(tk.END, f"SUMMARY\n")
            progress_text.insert(tk.END, f"{'='*80}\n\n")
            progress_text.insert(tk.END, f"Files to delete: {len(files_to_delete)}\n")
            progress_text.insert(tk.END, f"Files to keep: {len(file_hashes)}\n\n")

            if files_to_delete:
                progress_text.insert(tk.END, f"Files marked for deletion:\n")
                for prog_num, file_path, reason in files_to_delete[:20]:
                    progress_text.insert(tk.END, f"  • {prog_num} - {reason}\n")
                if len(files_to_delete) > 20:
                    progress_text.insert(tk.END, f"  ... and {len(files_to_delete) - 20} more\n")

            progress_text.insert(tk.END, f"\n")
            progress_text.see(tk.END)
            self.root.update()

            # Confirmation buttons
            def confirm_delete():
                progress_text.insert(tk.END, f"\n{'='*80}\n")
                progress_text.insert(tk.END, f"DELETING FILES\n")
                progress_text.insert(tk.END, f"{'='*80}\n\n")
                progress_text.see(tk.END)

                deleted_count = 0
                error_count = 0

                for prog_num, file_path, reason in files_to_delete:
                    try:
                        # Move physical file to deleted folder instead of deleting
                        if os.path.exists(file_path):
                            import shutil
                            filename = os.path.basename(file_path)
                            deleted_file_path = os.path.join(self.deleted_path, filename)

                            # If file already exists in deleted folder, add timestamp
                            if os.path.exists(deleted_file_path):
                                import time
                                timestamp = time.strftime("%Y%m%d_%H%M%S")
                                name, ext = os.path.splitext(filename)
                                deleted_file_path = os.path.join(self.deleted_path, f"{name}_{timestamp}{ext}")

                            shutil.move(file_path, deleted_file_path)

                        # Remove from database
                        cursor.execute("DELETE FROM programs WHERE program_number = ?", (prog_num,))

                        progress_text.insert(tk.END, f"  ✓ Moved to deleted/: {prog_num}\n")
                        deleted_count += 1

                        if deleted_count % 10 == 0:
                            progress_text.see(tk.END)
                            self.root.update()

                    except Exception as e:
                        progress_text.insert(tk.END, f"  ✗ ERROR moving {prog_num}: {e}\n")
                        error_count += 1

                conn.commit()

                progress_text.insert(tk.END, f"\n{'='*80}\n")
                progress_text.insert(tk.END, f"COMPLETE\n")
                progress_text.insert(tk.END, f"{'='*80}\n\n")
                progress_text.insert(tk.END, f"Successfully deleted: {deleted_count} files\n")
                if error_count > 0:
                    progress_text.insert(tk.END, f"Errors: {error_count} files\n")
                progress_text.see(tk.END)

                # Log activity
                self.log_activity('delete_content_duplicates', 'batch', {
                    'deleted_count': deleted_count,
                    'error_count': error_count
                })

                # Refresh the view
                self.refresh_results()

                # Close database connection
                conn.close()

                # Update button
                btn_frame.pack_forget()
                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)

            def cancel_delete():
                conn.close()
                progress_window.destroy()

            # Button frame
            btn_frame = tk.Frame(progress_window, bg=self.bg_color)
            btn_frame.pack(pady=10)

            tk.Button(btn_frame, text="✓ Confirm Delete", command=confirm_delete,
                     bg="#D32F2F", fg=self.fg_color, font=("Arial", 10, "bold"),
                     width=18, height=2).pack(side=tk.LEFT, padx=10)

            tk.Button(btn_frame, text="✗ Cancel", command=cancel_delete,
                     bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold"),
                     width=18, height=2).pack(side=tk.LEFT, padx=10)

        except Exception as e:
            progress_text.insert(tk.END, f"\n\nERROR: {e}\n")
            import traceback
            progress_text.insert(tk.END, traceback.format_exc())
            progress_text.see(tk.END)

            tk.Button(progress_window, text="Close", command=progress_window.destroy,
                     bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)

    def scan_for_duplicates(self):
        """Scan repository for all types of duplicates and show report"""
        import hashlib

        # Create progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Duplicate Scan Report")
        progress_window.geometry("1000x700")
        progress_window.configure(bg=self.bg_color)
        progress_window.transient(self.root)

        tk.Label(progress_window, text="🔍 Duplicate Scan Report",
                font=("Arial", 14, "bold"), bg=self.bg_color, fg=self.fg_color).pack(pady=10)

        # Progress text
        text_frame = tk.Frame(progress_window, bg=self.bg_color)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        progress_text = tk.Text(text_frame, bg=self.input_bg, fg=self.fg_color,
                               font=("Consolas", 9), wrap=tk.WORD)
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=progress_text.yview)
        progress_text.configure(yscrollcommand=scrollbar.set)

        progress_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        progress_text.insert(tk.END, "Scanning repository for duplicates...\n\n")
        progress_text.see(tk.END)
        self.root.update()

        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            # Get all managed files from repository
            cursor.execute("""
                SELECT program_number, title, file_path
                FROM programs
                WHERE is_managed = 1
                ORDER BY program_number
            """)

            all_files = cursor.fetchall()
            progress_text.insert(tk.END, f"Found {len(all_files)} files in repository\n\n")
            progress_text.see(tk.END)
            self.root.update()

            # Group by program number (name duplicates)
            name_groups = {}
            for prog_num, title, file_path in all_files:
                # Extract base program number (remove suffixes like (1), (2))
                import re
                base_num = re.sub(r'\(\d+\)$', '', prog_num)

                if base_num not in name_groups:
                    name_groups[base_num] = []
                name_groups[base_num].append((prog_num, title, file_path))

            # Find name duplicates
            name_duplicates = {k: v for k, v in name_groups.items() if len(v) > 1}

            # Calculate hashes for content duplicates
            progress_text.insert(tk.END, "Calculating file hashes for content analysis...\n")
            progress_text.see(tk.END)
            self.root.update()

            file_hashes = {}
            for idx, (prog_num, title, file_path) in enumerate(all_files, 1):
                if idx % 50 == 0:
                    progress_text.insert(tk.END, f"  Processed {idx}/{len(all_files)} files...\n")
                    progress_text.see(tk.END)
                    self.root.update()

                if file_path and os.path.exists(file_path):
                    try:
                        sha256_hash = hashlib.sha256()
                        with open(file_path, 'rb') as f:
                            for byte_block in iter(lambda: f.read(4096), b""):
                                sha256_hash.update(byte_block)
                        file_hash = sha256_hash.hexdigest()

                        if file_hash not in file_hashes:
                            file_hashes[file_hash] = []
                        file_hashes[file_hash].append((prog_num, title, file_path))
                    except:
                        pass

            # Find content duplicates
            content_duplicates = {h: files for h, files in file_hashes.items() if len(files) > 1}

            # Generate report
            progress_text.insert(tk.END, f"\n{'='*90}\n")
            progress_text.insert(tk.END, f"DUPLICATE SCAN REPORT\n")
            progress_text.insert(tk.END, f"{'='*90}\n\n")

            progress_text.insert(tk.END, f"Total repository files: {len(all_files)}\n")
            progress_text.insert(tk.END, f"Name duplicate groups: {len(name_duplicates)}\n")
            progress_text.insert(tk.END, f"Content duplicate groups: {len(content_duplicates)}\n\n")

            # Section 1: Name Duplicates
            if name_duplicates:
                progress_text.insert(tk.END, f"{'='*90}\n")
                progress_text.insert(tk.END, f"NAME DUPLICATES (Same base program number, different titles/content)\n")
                progress_text.insert(tk.END, f"{'='*90}\n\n")

                for base_num, files in sorted(name_duplicates.items())[:20]:
                    progress_text.insert(tk.END, f"Base: {base_num} ({len(files)} files)\n")
                    for prog_num, title, file_path in files:
                        progress_text.insert(tk.END, f"  • {prog_num}: {title[:50]}\n")
                    progress_text.insert(tk.END, f"\n")

                if len(name_duplicates) > 20:
                    progress_text.insert(tk.END, f"... and {len(name_duplicates) - 20} more groups\n\n")
            else:
                progress_text.insert(tk.END, "✓ No name duplicates found\n\n")

            # Section 2: Content Duplicates
            if content_duplicates:
                progress_text.insert(tk.END, f"{'='*90}\n")
                progress_text.insert(tk.END, f"CONTENT DUPLICATES (Exact file content match)\n")
                progress_text.insert(tk.END, f"{'='*90}\n\n")

                for hash_val, files in list(content_duplicates.items())[:20]:
                    progress_text.insert(tk.END, f"Group ({len(files)} files with identical content):\n")
                    for prog_num, title, file_path in files:
                        progress_text.insert(tk.END, f"  • {prog_num}: {title[:50]}\n")
                    progress_text.insert(tk.END, f"\n")

                if len(content_duplicates) > 20:
                    progress_text.insert(tk.END, f"... and {len(content_duplicates) - 20} more groups\n\n")
            else:
                progress_text.insert(tk.END, "✓ No content duplicates found\n\n")

            # Summary and recommendations
            progress_text.insert(tk.END, f"{'='*90}\n")
            progress_text.insert(tk.END, f"RECOMMENDATIONS\n")
            progress_text.insert(tk.END, f"{'='*90}\n\n")

            if name_duplicates:
                progress_text.insert(tk.END, f"→ Name Duplicates: Use '✏️ Rename Name Duplicates' to consolidate\n")
            if content_duplicates:
                progress_text.insert(tk.END, f"→ Content Duplicates: Use '🧹 Delete Content Duplicates' to remove\n")

            if not name_duplicates and not content_duplicates:
                progress_text.insert(tk.END, f"✓ Repository is clean - no duplicates found!\n")

            progress_text.see(tk.END)
            conn.close()

            # Close button
            tk.Button(progress_window, text="Close", command=progress_window.destroy,
                     bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)

        except Exception as e:
            progress_text.insert(tk.END, f"\n\nERROR: {e}\n")
            import traceback
            progress_text.insert(tk.END, traceback.format_exc())
            progress_text.see(tk.END)

            tk.Button(progress_window, text="Close", command=progress_window.destroy,
                     bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)

    def rename_name_duplicates(self):
        """Rename files with same base name but different content/title"""
        # Create progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Rename Name Duplicates")
        progress_window.geometry("900x700")
        progress_window.configure(bg=self.bg_color)
        progress_window.transient(self.root)

        tk.Label(progress_window, text="✏️ Rename Name Duplicates",
                font=("Arial", 14, "bold"), bg=self.bg_color, fg=self.fg_color).pack(pady=10)

        # Progress text
        text_frame = tk.Frame(progress_window, bg=self.bg_color)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        progress_text = tk.Text(text_frame, bg=self.input_bg, fg=self.fg_color,
                               font=("Consolas", 9), wrap=tk.WORD)
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=progress_text.yview)
        progress_text.configure(yscrollcommand=scrollbar.set)

        progress_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        progress_text.insert(tk.END, "Analyzing name duplicates...\n\n")
        progress_text.see(tk.END)
        self.root.update()

        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            # Get all managed files from repository
            cursor.execute("""
                SELECT program_number, title, file_path
                FROM programs
                WHERE is_managed = 1
                ORDER BY program_number
            """)

            all_files = cursor.fetchall()
            progress_text.insert(tk.END, f"Found {len(all_files)} files in repository\n\n")
            progress_text.see(tk.END)
            self.root.update()

            # Group by base program number (remove suffixes like (1), (2))
            import re
            name_groups = {}
            for prog_num, title, file_path in all_files:
                base_num = re.sub(r'\(\d+\)$', '', prog_num)

                if base_num not in name_groups:
                    name_groups[base_num] = []
                name_groups[base_num].append((prog_num, title, file_path))

            # Find name duplicates (multiple files with same base number)
            name_duplicates = {k: v for k, v in name_groups.items() if len(v) > 1}

            progress_text.insert(tk.END, f"{'='*80}\n")
            progress_text.insert(tk.END, f"NAME DUPLICATE ANALYSIS\n")
            progress_text.insert(tk.END, f"{'='*80}\n\n")
            progress_text.insert(tk.END, f"Name duplicate groups found: {len(name_duplicates)}\n\n")

            if not name_duplicates:
                progress_text.insert(tk.END, "✓ No name duplicates found!\n\n")
                progress_text.insert(tk.END, "All files have unique base program numbers.\n")
                progress_text.see(tk.END)
                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)
                conn.close()
                return

            # Show duplicates and prepare rename suggestions
            renames_to_apply = []  # (old_prog_num, new_prog_num, round_size)
            assigned_numbers = set()  # Track numbers we've already assigned in this session

            progress_text.insert(tk.END, f"{'='*80}\n")
            progress_text.insert(tk.END, f"DUPLICATE GROUPS (Will keep first, rename others)\n")
            progress_text.insert(tk.END, f"{'='*80}\n\n")

            for base_num, files in sorted(name_duplicates.items())[:30]:
                progress_text.insert(tk.END, f"Base: {base_num} ({len(files)} files)\n")

                # Keep first file, rename others
                for idx, (prog_num, title, file_path) in enumerate(files):
                    if idx == 0:
                        progress_text.insert(tk.END, f"  ✓ KEEP: {prog_num} - {title[:40]}\n")
                    else:
                        # Get round size for this duplicate file
                        cursor.execute("SELECT round_size FROM programs WHERE program_number = ?", (prog_num,))
                        round_result = cursor.fetchone()
                        round_size = round_result[0] if round_result and round_result[0] else None

                        if round_size:
                            # Find next available number in correct range for this round size
                            # Keep trying until we find one not already assigned in this session
                            new_prog_num = None
                            attempts = 0
                            max_attempts = 100  # Prevent infinite loop

                            while attempts < max_attempts:
                                candidate = self.find_next_available_number(round_size)
                                if not candidate:
                                    break  # No more available numbers

                                # Check if already exists in database OR already assigned in this session
                                cursor.execute("SELECT COUNT(*) FROM programs WHERE program_number = ?", (candidate,))
                                exists = cursor.fetchone()[0] > 0

                                if not exists and candidate not in assigned_numbers:
                                    new_prog_num = candidate
                                    assigned_numbers.add(new_prog_num)  # Mark as assigned
                                    break
                                else:
                                    # This number is taken, temporarily mark it in registry to get next one
                                    cursor.execute("UPDATE program_number_registry SET status = 'IN_USE' WHERE program_number = ?", (candidate,))
                                    conn.commit()  # Commit immediately to avoid locking
                                    attempts += 1

                            if new_prog_num:
                                # Get the range info to show in output
                                range_info = self.get_range_for_round_size(round_size)
                                if range_info:
                                    range_start, range_end = range_info
                                    renames_to_apply.append((prog_num, new_prog_num, round_size))
                                    progress_text.insert(tk.END, f"  ✏️ RENAME: {prog_num} → {new_prog_num} ({round_size}\" range: o{range_start}-o{range_end}) - {title[:40]}\n")
                                else:
                                    renames_to_apply.append((prog_num, new_prog_num, round_size))
                                    progress_text.insert(tk.END, f"  ✏️ RENAME: {prog_num} → {new_prog_num} ({round_size}\") - {title[:40]}\n")
                            else:
                                progress_text.insert(tk.END, f"  ⚠️ SKIP: {prog_num} - No available numbers for {round_size}\" - {title[:40]}\n")
                        else:
                            # No round size detected - use free range (o1000-o9999)
                            # Keep trying until we find one not already assigned
                            new_prog_num = None
                            attempts = 0
                            max_attempts = 100

                            while attempts < max_attempts:
                                cursor.execute("""
                                    SELECT MIN(program_number)
                                    FROM program_number_registry
                                    WHERE status = 'AVAILABLE'
                                    AND CAST(REPLACE(program_number, 'o', '') AS INTEGER) BETWEEN 1000 AND 9999
                                """)
                                free_result = cursor.fetchone()
                                if not free_result or not free_result[0]:
                                    break  # No more available numbers

                                candidate = free_result[0]

                                # Check if already exists in database OR already assigned in this session
                                cursor.execute("SELECT COUNT(*) FROM programs WHERE program_number = ?", (candidate,))
                                exists = cursor.fetchone()[0] > 0

                                if not exists and candidate not in assigned_numbers:
                                    new_prog_num = candidate
                                    assigned_numbers.add(new_prog_num)  # Mark as assigned
                                    break
                                else:
                                    # This number is taken, temporarily mark it in registry to get next one
                                    cursor.execute("UPDATE program_number_registry SET status = 'IN_USE' WHERE program_number = ?", (candidate,))
                                    conn.commit()  # Commit immediately to avoid locking
                                    attempts += 1

                            if new_prog_num:
                                renames_to_apply.append((prog_num, new_prog_num, None))
                                progress_text.insert(tk.END, f"  ✏️ RENAME: {prog_num} → {new_prog_num} (free range) - {title[:40]}\n")
                            else:
                                progress_text.insert(tk.END, f"  ⚠️ SKIP: {prog_num} - No round size, no free numbers - {title[:40]}\n")

                progress_text.insert(tk.END, f"\n")
                progress_text.see(tk.END)
                self.root.update()

            if len(name_duplicates) > 30:
                progress_text.insert(tk.END, f"... and {len(name_duplicates) - 30} more groups\n\n")

            # Summary
            progress_text.insert(tk.END, f"{'='*80}\n")
            progress_text.insert(tk.END, f"SUMMARY\n")
            progress_text.insert(tk.END, f"{'='*80}\n\n")
            progress_text.insert(tk.END, f"Files to rename: {len(renames_to_apply)}\n\n")

            if renames_to_apply:
                progress_text.insert(tk.END, f"Rename list (first 20):\n")
                for item in renames_to_apply[:20]:
                    old_num, new_num = item[0], item[1]
                    round_size = item[2] if len(item) > 2 else None
                    if round_size:
                        progress_text.insert(tk.END, f"  • {old_num} → {new_num} ({round_size}\")\n")
                    else:
                        progress_text.insert(tk.END, f"  • {old_num} → {new_num} (free range)\n")
                if len(renames_to_apply) > 20:
                    progress_text.insert(tk.END, f"  ... and {len(renames_to_apply) - 20} more\n")

            progress_text.insert(tk.END, f"\n")
            progress_text.see(tk.END)
            self.root.update()

            # Confirmation buttons
            def confirm_rename():
                progress_text.insert(tk.END, f"\n{'='*80}\n")
                progress_text.insert(tk.END, f"RENAMING FILES\n")
                progress_text.insert(tk.END, f"{'='*80}\n\n")
                progress_text.see(tk.END)

                renamed_count = 0
                error_count = 0

                for item in renames_to_apply:
                    old_num, new_num = item[0], item[1]
                    try:
                        # Get current file path
                        cursor.execute("SELECT file_path FROM programs WHERE program_number = ?", (old_num,))
                        result = cursor.fetchone()
                        if not result or not result[0]:
                            progress_text.insert(tk.END, f"  ✗ ERROR: No file path for {old_num}\n")
                            error_count += 1
                            continue

                        old_file_path = result[0]

                        # Skip if file doesn't exist
                        if not os.path.exists(old_file_path):
                            progress_text.insert(tk.END, f"  ⚠️ SKIP: File not found for {old_num}\n")
                            error_count += 1
                            continue

                        # Generate new file path
                        old_dir = os.path.dirname(old_file_path)
                        old_filename = os.path.basename(old_file_path)

                        # Create new filename - match exactly the program number
                        # Example: program_number = o85000 -> filename = o85000.nc
                        new_filename = f"{new_num}.nc"
                        new_file_path = os.path.join(old_dir, new_filename)

                        # Read file content and update internal program number
                        with open(old_file_path, 'r', encoding='utf-8', errors='ignore') as f:
                            content = f.read()

                        # Replace first line program number (e.g., O12345 -> O12346)
                        import re
                        # Match O-number at start of line (case insensitive)
                        old_num_numeric = old_num.replace('o', '').replace('O', '')
                        new_num_numeric = new_num.replace('o', '').replace('O', '')

                        # Replace O-number in first line
                        lines = content.split('\n')
                        if lines and re.match(r'^[oO]\d+', lines[0].strip()):
                            lines[0] = re.sub(r'^[oO]\d+', f'O{new_num_numeric}', lines[0].strip())
                            content = '\n'.join(lines)

                        # Write to new file
                        with open(new_file_path, 'w', encoding='utf-8') as f:
                            f.write(content)

                        # Delete old file
                        os.remove(old_file_path)

                        # Update database with new program number and file path
                        cursor.execute("UPDATE programs SET program_number = ?, file_path = ? WHERE program_number = ?",
                                     (new_num, new_file_path, old_num))

                        # Update registry: mark old number as AVAILABLE, new number as IN_USE
                        cursor.execute("UPDATE program_number_registry SET status = 'AVAILABLE', file_path = NULL WHERE program_number = ?", (old_num,))
                        cursor.execute("UPDATE program_number_registry SET status = 'IN_USE', file_path = ? WHERE program_number = ?", (new_file_path, new_num))

                        progress_text.insert(tk.END, f"  ✓ Renamed: {old_num} → {new_num} (file + internal + registry)\n")
                        renamed_count += 1

                        if renamed_count % 10 == 0:
                            progress_text.see(tk.END)
                            self.root.update()

                    except Exception as e:
                        progress_text.insert(tk.END, f"  ✗ ERROR renaming {old_num}: {e}\n")
                        error_count += 1

                conn.commit()

                progress_text.insert(tk.END, f"\n{'='*80}\n")
                progress_text.insert(tk.END, f"COMPLETE\n")
                progress_text.insert(tk.END, f"{'='*80}\n\n")
                progress_text.insert(tk.END, f"Successfully renamed: {renamed_count} files\n")
                if error_count > 0:
                    progress_text.insert(tk.END, f"Errors: {error_count} files\n")
                progress_text.see(tk.END)

                # Log activity
                self.log_activity('rename_name_duplicates', 'batch', {
                    'renamed_count': renamed_count,
                    'error_count': error_count
                })

                # Refresh the view
                self.refresh_results()

                # Close database connection
                conn.close()

                # Update button
                btn_frame.pack_forget()
                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)

            def cancel_rename():
                # Rollback temporary IN_USE markings in registry
                for num in assigned_numbers:
                    cursor.execute("UPDATE program_number_registry SET status = 'AVAILABLE' WHERE program_number = ?", (num,))
                conn.commit()
                conn.close()
                progress_window.destroy()

            # Set window close handler to ensure cleanup even if user closes window with X button
            progress_window.protocol("WM_DELETE_WINDOW", cancel_rename)

            # Button frame
            btn_frame = tk.Frame(progress_window, bg=self.bg_color)
            btn_frame.pack(pady=10)

            tk.Button(btn_frame, text="✓ Confirm Rename", command=confirm_rename,
                     bg="#7B1FA2", fg=self.fg_color, font=("Arial", 10, "bold"),
                     width=18, height=2).pack(side=tk.LEFT, padx=10)

            tk.Button(btn_frame, text="✗ Cancel", command=cancel_rename,
                     bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold"),
                     width=18, height=2).pack(side=tk.LEFT, padx=10)

        except Exception as e:
            # Rollback temporary IN_USE markings if they exist
            if 'assigned_numbers' in locals() and 'cursor' in locals() and 'conn' in locals():
                try:
                    for num in assigned_numbers:
                        cursor.execute("UPDATE program_number_registry SET status = 'AVAILABLE' WHERE program_number = ?", (num,))
                    conn.commit()
                except:
                    pass  # Ignore errors during cleanup

            progress_text.insert(tk.END, f"\n\nERROR: {e}\n")
            import traceback
            progress_text.insert(tk.END, traceback.format_exc())
            progress_text.see(tk.END)

            # Close connection if it exists
            if 'conn' in locals():
                try:
                    conn.close()
                except:
                    pass

            tk.Button(progress_window, text="Close", command=progress_window.destroy,
                     bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)

    def fix_underscore_suffix_files(self):
        """Fix files with underscore suffix patterns (o12345_1.nc) by renaming to correct ranges"""

        # Create progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Fix Underscore Suffix Files")
        progress_window.geometry("800x600")
        progress_window.configure(bg=self.bg_color)
        progress_window.transient(self.root)
        progress_window.grab_set()

        tk.Label(progress_window, text="🔧 Fix Underscore Suffix Files",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 14, "bold")).pack(pady=10)

        # Info frame
        info_frame = tk.Frame(progress_window, bg=self.bg_color)
        info_frame.pack(fill=tk.X, padx=20, pady=5)

        tk.Label(info_frame,
                text="This will find and rename files with underscore suffix patterns (o12345_1.nc, o12345_2.nc)\n"
                     "to available numbers in their correct round size ranges.",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 9), justify=tk.LEFT).pack(anchor=tk.W)

        # Progress text
        text_frame = tk.Frame(progress_window, bg=self.bg_color)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        scrollbar = ttk.Scrollbar(text_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        progress_text = tk.Text(text_frame, height=25, yscrollcommand=scrollbar.set,
                               bg="#2B2B2B", fg="#FFFFFF", font=("Consolas", 9), wrap=tk.WORD)
        scrollbar.config(command=progress_text.yview)
        progress_text.pack(fill=tk.BOTH, expand=True)

        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            progress_text.insert(tk.END, "Scanning for files with underscore suffixes...\n")
            progress_text.insert(tk.END, "="*80 + "\n\n")
            progress_text.see(tk.END)
            self.root.update()

            # Find all repository files with underscore suffix patterns in program_number or filename
            # Pattern: o#####_# (5 digits followed by underscore and more digits)
            cursor.execute("""
                SELECT program_number, file_path, round_size, title
                FROM programs
                WHERE is_managed = 1
                AND (
                    program_number LIKE 'o%_%'
                    OR program_number LIKE 'o%(%'
                )
                ORDER BY program_number
            """)

            underscore_files = cursor.fetchall()

            if not underscore_files:
                progress_text.insert(tk.END, "No files with underscore suffixes found.\n")
                progress_text.see(tk.END)

                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)
                return

            progress_text.insert(tk.END, f"Found {len(underscore_files)} files with underscore/parenthesis patterns\n\n")

            # Group and analyze
            renames_to_apply = []
            assigned_numbers = set()  # Track numbers already assigned in this session

            for prog_num, file_path, round_size, title in underscore_files:
                progress_text.insert(tk.END, f"Analyzing: {prog_num}\n")
                progress_text.insert(tk.END, f"  File: {os.path.basename(file_path)}\n")
                progress_text.insert(tk.END, f"  Title: {title or 'N/A'}\n")

                if round_size:
                    progress_text.insert(tk.END, f"  Round Size: {round_size}\"\n")

                    # Find next available number - keep trying until we get one not already assigned
                    new_prog_num = None
                    attempts = 0
                    max_attempts = 100

                    while attempts < max_attempts:
                        candidate = self.find_next_available_number(round_size)
                        if not candidate:
                            break

                        # Check if already exists in database OR already assigned in this session
                        cursor.execute("SELECT COUNT(*) FROM programs WHERE program_number = ?", (candidate,))
                        exists = cursor.fetchone()[0] > 0

                        if not exists and candidate not in assigned_numbers:
                            new_prog_num = candidate
                            assigned_numbers.add(new_prog_num)
                            break
                        else:
                            # Temporarily mark as IN_USE to get next number
                            cursor.execute("UPDATE program_number_registry SET status = 'IN_USE' WHERE program_number = ?", (candidate,))
                            conn.commit()  # Commit immediately to avoid locking and ensure next query sees updated status
                            attempts += 1

                    if new_prog_num:
                        renames_to_apply.append((prog_num, new_prog_num, round_size, file_path, title))
                        progress_text.insert(tk.END, f"  ✓ Will rename to: {new_prog_num} ({round_size}\" range)\n")
                    else:
                        progress_text.insert(tk.END, f"  ✗ No available numbers in {round_size}\" range\n")
                else:
                    progress_text.insert(tk.END, f"  Round Size: Not detected\n")

                    # Use free range - keep trying until we get one not already assigned
                    new_prog_num = None
                    attempts = 0
                    max_attempts = 100

                    while attempts < max_attempts:
                        cursor.execute("""
                            SELECT MIN(program_number)
                            FROM program_number_registry
                            WHERE status = 'AVAILABLE'
                            AND CAST(REPLACE(program_number, 'o', '') AS INTEGER) BETWEEN 1000 AND 9999
                        """)
                        free_result = cursor.fetchone()
                        if not free_result or not free_result[0]:
                            break

                        candidate = free_result[0]

                        # Check if already exists in database OR already assigned in this session
                        cursor.execute("SELECT COUNT(*) FROM programs WHERE program_number = ?", (candidate,))
                        exists = cursor.fetchone()[0] > 0

                        if not exists and candidate not in assigned_numbers:
                            new_prog_num = candidate
                            assigned_numbers.add(new_prog_num)
                            break
                        else:
                            # Temporarily mark as IN_USE to get next number
                            cursor.execute("UPDATE program_number_registry SET status = 'IN_USE' WHERE program_number = ?", (candidate,))
                            conn.commit()  # Commit immediately to avoid locking and ensure next query sees updated status
                            attempts += 1

                    if new_prog_num:
                        renames_to_apply.append((prog_num, new_prog_num, None, file_path, title))
                        progress_text.insert(tk.END, f"  ✓ Will rename to: {new_prog_num} (free range)\n")
                    else:
                        progress_text.insert(tk.END, f"  ✗ No available numbers in free range\n")

                progress_text.insert(tk.END, "\n")
                progress_text.see(tk.END)
                self.root.update()

            if not renames_to_apply:
                progress_text.insert(tk.END, "\n" + "="*80 + "\n")
                progress_text.insert(tk.END, "No renames possible (no available program numbers)\n")
                progress_text.see(tk.END)

                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)
                return

            # Show summary
            progress_text.insert(tk.END, "="*80 + "\n")
            progress_text.insert(tk.END, f"PREVIEW - {len(renames_to_apply)} files ready to rename:\n")
            progress_text.insert(tk.END, "="*80 + "\n\n")

            for old_num, new_num, round_size, file_path, title in renames_to_apply:
                size_str = f"({round_size}\")" if round_size else "(free range)"
                title_str = title or "UNKNOWN"
                progress_text.insert(tk.END, f"✏️ RENAME: {old_num} → {new_num} {size_str} - {title_str}\n")

            progress_text.insert(tk.END, "\n" + "="*80 + "\n")
            progress_text.insert(tk.END, "Click 'Confirm Rename' to proceed or 'Cancel' to abort.\n")
            progress_text.see(tk.END)

            def confirm_rename():
                progress_text.insert(tk.END, "\n" + "="*80 + "\n")
                progress_text.insert(tk.END, "STARTING RENAME OPERATIONS...\n")
                progress_text.insert(tk.END, "="*80 + "\n\n")
                progress_text.see(tk.END)

                renamed_count = 0
                error_count = 0

                for old_num, new_num, round_size, old_file_path, title in renames_to_apply:
                    try:
                        progress_text.insert(tk.END, f"Processing: {old_num} → {new_num}\n")

                        # Read the file content
                        if not os.path.exists(old_file_path):
                            progress_text.insert(tk.END, f"  ✗ ERROR: File not found: {old_file_path}\n")
                            error_count += 1
                            continue

                        with open(old_file_path, 'r', encoding='utf-8', errors='ignore') as f:
                            content = f.read()

                        # Replace the O-number in the content (case-insensitive)
                        # Remove suffix before replacing (o12345_1 → o12345)
                        base_old_num = old_num.split('_')[0].split('(')[0]
                        new_content = re.sub(
                            rf'(?i)({base_old_num})',
                            new_num,
                            content
                        )

                        # Create new filename - exactly the program number
                        old_dir = os.path.dirname(old_file_path)
                        new_filename = f"{new_num}.nc"
                        new_file_path = os.path.join(old_dir, new_filename)

                        # Rename the file
                        os.rename(old_file_path, new_file_path)
                        progress_text.insert(tk.END, f"  ✓ File renamed: {os.path.basename(old_file_path)} → {new_filename}\n")

                        # Write the updated content
                        with open(new_file_path, 'w', encoding='utf-8') as f:
                            f.write(new_content)
                        progress_text.insert(tk.END, f"  ✓ Internal O-number updated: {base_old_num} → {new_num}\n")

                        # Update database
                        cursor.execute("""
                            UPDATE programs
                            SET program_number = ?,
                                file_path = ?
                            WHERE program_number = ?
                        """, (new_num, new_file_path, old_num))
                        progress_text.insert(tk.END, f"  ✓ Database updated\n")

                        # Update registry: mark old number as AVAILABLE (if it exists), new number as IN_USE
                        cursor.execute("UPDATE program_number_registry SET status = 'AVAILABLE', file_path = NULL WHERE program_number = ?", (old_num,))
                        cursor.execute("UPDATE program_number_registry SET status = 'IN_USE', file_path = ? WHERE program_number = ?", (new_file_path, new_num))

                        progress_text.insert(tk.END, f"  ✓ Registry updated\n")
                        progress_text.insert(tk.END, f"  ✅ Complete: {old_num} → {new_num}\n\n")
                        renamed_count += 1

                        if renamed_count % 5 == 0:
                            progress_text.see(tk.END)
                            self.root.update()

                    except Exception as e:
                        progress_text.insert(tk.END, f"  ✗ ERROR: {e}\n\n")
                        error_count += 1

                conn.commit()

                progress_text.insert(tk.END, "="*80 + "\n")
                progress_text.insert(tk.END, "COMPLETE\n")
                progress_text.insert(tk.END, "="*80 + "\n\n")
                progress_text.insert(tk.END, f"Successfully renamed: {renamed_count} files\n")
                if error_count > 0:
                    progress_text.insert(tk.END, f"Errors: {error_count} files\n")
                progress_text.see(tk.END)

                # Log activity
                self.log_activity('fix_underscore_suffixes', 'batch', {
                    'renamed_count': renamed_count,
                    'error_count': error_count
                })

                # Refresh the view
                self.refresh_results()

                # Close database connection
                conn.close()

                # Update button
                btn_frame.pack_forget()
                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)

            def cancel_rename():
                # Rollback temporary IN_USE markings in registry
                for num in assigned_numbers:
                    cursor.execute("UPDATE program_number_registry SET status = 'AVAILABLE' WHERE program_number = ?", (num,))
                conn.commit()
                conn.close()
                progress_window.destroy()

            # Set window close handler to ensure cleanup even if user closes window with X button
            progress_window.protocol("WM_DELETE_WINDOW", cancel_rename)

            # Button frame
            btn_frame = tk.Frame(progress_window, bg=self.bg_color)
            btn_frame.pack(pady=10)

            tk.Button(btn_frame, text="✓ Confirm Rename", command=confirm_rename,
                     bg="#7B1FA2", fg=self.fg_color, font=("Arial", 10, "bold"),
                     width=18, height=2).pack(side=tk.LEFT, padx=10)

            tk.Button(btn_frame, text="✗ Cancel", command=cancel_rename,
                     bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold"),
                     width=18, height=2).pack(side=tk.LEFT, padx=10)

        except Exception as e:
            # Rollback temporary IN_USE markings if they exist
            if 'assigned_numbers' in locals() and 'cursor' in locals() and 'conn' in locals():
                try:
                    for num in assigned_numbers:
                        cursor.execute("UPDATE program_number_registry SET status = 'AVAILABLE' WHERE program_number = ?", (num,))
                    conn.commit()
                except:
                    pass  # Ignore errors during cleanup

            progress_text.insert(tk.END, f"\n\nERROR: {e}\n")
            import traceback
            progress_text.insert(tk.END, traceback.format_exc())
            progress_text.see(tk.END)

            # Close connection if it exists
            if 'conn' in locals():
                try:
                    conn.close()
                except:
                    pass

            tk.Button(progress_window, text="Close", command=progress_window.destroy,
                     bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)

    def sync_filenames_with_database(self):
        """Synchronize filenames with their program numbers in the database"""

        # Create progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Sync Filenames with Database")
        progress_window.geometry("800x600")
        progress_window.configure(bg=self.bg_color)
        progress_window.transient(self.root)

        tk.Label(progress_window, text="🔄 Sync Filenames with Program Numbers",
                font=("Arial", 14, "bold"), bg=self.bg_color, fg=self.fg_color).pack(pady=10)

        # Progress text
        text_frame = tk.Frame(progress_window, bg=self.bg_color)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        progress_text = tk.Text(text_frame, bg=self.input_bg, fg=self.fg_color,
                               font=("Consolas", 9), wrap=tk.WORD)
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=progress_text.yview)
        progress_text.configure(yscrollcommand=scrollbar.set)

        progress_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            progress_text.insert(tk.END, "Scanning for filename mismatches...\n")
            progress_text.insert(tk.END, "="*80 + "\n\n")
            progress_text.see(tk.END)
            self.root.update()

            # Find all managed files where filename doesn't match program number
            cursor.execute("""
                SELECT program_number, file_path, title
                FROM programs
                WHERE is_managed = 1
                ORDER BY program_number
            """)

            all_files = cursor.fetchall()
            mismatches = []

            for prog_num, file_path, title in all_files:
                if not file_path or not os.path.exists(file_path):
                    continue

                # Get current filename without extension
                current_filename = os.path.basename(file_path)
                current_base = os.path.splitext(current_filename)[0]

                # Expected filename from program number
                expected_base = prog_num

                # Check if they match
                if current_base != expected_base:
                    mismatches.append((prog_num, file_path, current_base, expected_base, title))

            if not mismatches:
                progress_text.insert(tk.END, "✓ No filename mismatches found!\n\n")
                progress_text.insert(tk.END, "All filenames match their program numbers.\n")
                progress_text.see(tk.END)

                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)
                conn.close()
                return

            progress_text.insert(tk.END, f"Found {len(mismatches)} filename mismatches\n\n")
            progress_text.insert(tk.END, "="*80 + "\n")
            progress_text.insert(tk.END, "FILENAME MISMATCHES (will rename files to match program numbers)\n")
            progress_text.insert(tk.END, "="*80 + "\n\n")

            # Show preview
            for prog_num, file_path, current_base, expected_base, title in mismatches[:50]:
                progress_text.insert(tk.END, f"Program: {prog_num}\n")
                progress_text.insert(tk.END, f"  Current filename: {current_base}.nc\n")
                progress_text.insert(tk.END, f"  Will rename to:   {expected_base}.nc\n")
                if title:
                    progress_text.insert(tk.END, f"  Title: {title}\n")
                progress_text.insert(tk.END, "\n")

            if len(mismatches) > 50:
                progress_text.insert(tk.END, f"... and {len(mismatches) - 50} more\n\n")

            progress_text.insert(tk.END, "="*80 + "\n")
            progress_text.insert(tk.END, "Click 'Confirm Rename' to rename files or 'Cancel' to abort.\n")
            progress_text.see(tk.END)

            def confirm_rename():
                progress_text.insert(tk.END, "\n" + "="*80 + "\n")
                progress_text.insert(tk.END, "STARTING FILENAME SYNC...\n")
                progress_text.insert(tk.END, "="*80 + "\n\n")
                progress_text.see(tk.END)

                renamed_count = 0
                error_count = 0

                for prog_num, old_file_path, current_base, expected_base, title in mismatches:
                    try:
                        progress_text.insert(tk.END, f"Processing: {current_base}.nc → {expected_base}.nc\n")

                        # Generate new file path
                        old_dir = os.path.dirname(old_file_path)
                        new_filename = f"{expected_base}.nc"
                        new_file_path = os.path.join(old_dir, new_filename)

                        # Check if this is just a case change (Windows is case-insensitive)
                        if old_file_path.lower() == new_file_path.lower():
                            # Case-only change - use temporary rename
                            import tempfile
                            temp_name = os.path.join(old_dir, f"temp_{os.path.basename(old_file_path)}")
                            os.rename(old_file_path, temp_name)
                            os.rename(temp_name, new_file_path)
                            progress_text.insert(tk.END, f"  ✓ File renamed (case change)\n")
                        else:
                            # Different filename - check if target already exists
                            if os.path.exists(new_file_path):
                                progress_text.insert(tk.END, f"  ⚠️ SKIP: {new_filename} already exists\n\n")
                                error_count += 1
                                continue

                            # Rename the file
                            os.rename(old_file_path, new_file_path)
                            progress_text.insert(tk.END, f"  ✓ File renamed\n")

                        # Update database with new file path
                        cursor.execute("""
                            UPDATE programs
                            SET file_path = ?
                            WHERE program_number = ?
                        """, (new_file_path, prog_num))
                        progress_text.insert(tk.END, f"  ✓ Database updated\n")

                        # Update registry with new file path
                        cursor.execute("""
                            UPDATE program_number_registry
                            SET file_path = ?
                            WHERE program_number = ?
                        """, (new_file_path, prog_num))
                        progress_text.insert(tk.END, f"  ✓ Registry updated\n")

                        progress_text.insert(tk.END, f"  ✅ Complete: {current_base}.nc → {expected_base}.nc\n\n")
                        renamed_count += 1

                        if renamed_count % 10 == 0:
                            progress_text.see(tk.END)
                            self.root.update()

                    except Exception as e:
                        progress_text.insert(tk.END, f"  ✗ ERROR: {e}\n\n")
                        error_count += 1

                conn.commit()

                progress_text.insert(tk.END, "="*80 + "\n")
                progress_text.insert(tk.END, "COMPLETE\n")
                progress_text.insert(tk.END, "="*80 + "\n\n")
                progress_text.insert(tk.END, f"Successfully renamed: {renamed_count} files\n")
                if error_count > 0:
                    progress_text.insert(tk.END, f"Errors: {error_count} files\n")
                progress_text.see(tk.END)

                # Log activity
                self.log_activity('sync_filenames', 'batch', {
                    'renamed_count': renamed_count,
                    'error_count': error_count
                })

                # Refresh the view
                self.refresh_results()

                # Close database connection
                conn.close()

                # Update button
                btn_frame.pack_forget()
                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)

            def cancel_rename():
                conn.close()
                progress_window.destroy()

            # Button frame
            btn_frame = tk.Frame(progress_window, bg=self.bg_color)
            btn_frame.pack(pady=10)

            tk.Button(btn_frame, text="✓ Confirm Rename", command=confirm_rename,
                     bg="#7B1FA2", fg=self.fg_color, font=("Arial", 10, "bold"),
                     width=18, height=2).pack(side=tk.LEFT, padx=10)

            tk.Button(btn_frame, text="✗ Cancel", command=cancel_rename,
                     bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold"),
                     width=18, height=2).pack(side=tk.LEFT, padx=10)

        except Exception as e:
            progress_text.insert(tk.END, f"\n\nERROR: {e}\n")
            import traceback
            progress_text.insert(tk.END, traceback.format_exc())
            progress_text.see(tk.END)

            # Close connection if it exists
            if 'conn' in locals():
                try:
                    conn.close()
                except:
                    pass

            tk.Button(progress_window, text="Close", command=progress_window.destroy,
                     bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)

    def export_repository_by_round_size(self):
        """Export repository files organized by round size folders"""

        # Ask user to select export destination
        export_root = filedialog.askdirectory(
            title="Select Export Destination Folder",
            initialdir=os.path.expanduser("~")
        )

        if not export_root:
            return  # User cancelled

        # Create progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Export Repository by Round Size")
        progress_window.geometry("900x700")
        progress_window.configure(bg=self.bg_color)
        progress_window.transient(self.root)

        tk.Label(progress_window, text="📦 Export Repository by Round Size",
                font=("Arial", 14, "bold"), bg=self.bg_color, fg=self.fg_color).pack(pady=10)

        # Progress text
        text_frame = tk.Frame(progress_window, bg=self.bg_color)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        progress_text = tk.Text(text_frame, bg=self.input_bg, fg=self.fg_color,
                               font=("Consolas", 9), wrap=tk.WORD)
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=progress_text.yview)
        progress_text.configure(yscrollcommand=scrollbar.set)

        progress_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            progress_text.insert(tk.END, f"Export Destination: {export_root}\n")
            progress_text.insert(tk.END, "="*80 + "\n\n")
            progress_text.see(tk.END)
            self.root.update()

            # Define standard round size folders
            # Map detected sizes to standard folder names
            standard_folders = {
                # Exact matches
                5.75: "5.75",
                6.0: "6.0",
                6.25: "6.25",
                6.5: "6.5",
                7.0: "7.0",
                7.5: "7.5",
                8.0: "8.0",
                8.5: "8.5",
                9.5: "9.5",
                10.25: "10.25",
                10.5: "10.5",
                13.0: "13.0",
            }

            # Function to map any round size to nearest standard folder
            def get_folder_for_round_size(round_size):
                if not round_size:
                    return "NO_ROUND_SIZE"

                # Check for exact match
                if round_size in standard_folders:
                    return standard_folders[round_size]

                # Find nearest standard size
                nearest = min(standard_folders.keys(), key=lambda x: abs(x - round_size))
                return standard_folders[nearest]

            # Get repository path
            repo_path = self.repository_path
            if not repo_path:
                progress_text.insert(tk.END, "ERROR: No repository path configured\n")
                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)
                conn.close()
                return

            # Get ALL repository files:
            # - In repository folder
            # - Has file_path set
            # - Any extension (or no extension)
            # - Include files with or without round size
            cursor.execute("""
                SELECT program_number, file_path, round_size, title
                FROM programs
                WHERE is_managed = 1
                  AND file_path IS NOT NULL
                  AND file_path LIKE ?
                ORDER BY round_size, program_number
            """, (f"{repo_path}%",))

            all_files = cursor.fetchall()

            if not all_files:
                progress_text.insert(tk.END, "No repository files found to export.\n")
                progress_text.see(tk.END)
                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)
                conn.close()
                return

            # Filter to only files that actually exist
            verified_files = []
            skipped_count = 0
            for prog_num, file_path, round_size, title in all_files:
                if file_path and os.path.exists(file_path):
                    verified_files.append((prog_num, file_path, round_size, title))
                else:
                    skipped_count += 1

            all_files = verified_files

            if not all_files:
                progress_text.insert(tk.END, "No files found that actually exist.\n")
                progress_text.insert(tk.END, f"(Skipped {skipped_count} database entries with missing files)\n")
                progress_text.see(tk.END)
                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)
                conn.close()
                return

            progress_text.insert(tk.END, f"Found {len(all_files)} repository files to export\n")
            if skipped_count > 0:
                progress_text.insert(tk.END, f"(Skipped {skipped_count} entries with missing files)\n")
            progress_text.insert(tk.END, "\nExporting ALL files from repository:\n")
            progress_text.insert(tk.END, "  ✓ All extensions (.nc, .txt, no extension)\n")
            progress_text.insert(tk.END, "  ✓ Organized by round size\n")
            progress_text.insert(tk.END, "  ✓ Only files that actually exist\n\n")
            progress_text.insert(tk.END, "Organizing files by round size...\n")
            progress_text.insert(tk.END, "="*80 + "\n\n")
            progress_text.see(tk.END)
            self.root.update()

            # Group files by folder
            files_by_folder = {}
            for prog_num, file_path, round_size, title in all_files:
                folder_name = get_folder_for_round_size(round_size)

                if folder_name not in files_by_folder:
                    files_by_folder[folder_name] = []

                files_by_folder[folder_name].append((prog_num, file_path, round_size, title))

            # Show organization summary
            progress_text.insert(tk.END, "EXPORT ORGANIZATION:\n")
            progress_text.insert(tk.END, "="*80 + "\n\n")

            for folder_name in sorted(files_by_folder.keys()):
                file_count = len(files_by_folder[folder_name])
                progress_text.insert(tk.END, f"📁 {folder_name}/ ({file_count} files)\n")

            progress_text.insert(tk.END, f"\n{'='*80}\n")
            progress_text.insert(tk.END, f"Total: {len(all_files)} files in {len(files_by_folder)} folders\n\n")
            progress_text.see(tk.END)
            self.root.update()

            # Start export
            progress_text.insert(tk.END, "="*80 + "\n")
            progress_text.insert(tk.END, "STARTING EXPORT...\n")
            progress_text.insert(tk.END, "="*80 + "\n\n")
            progress_text.see(tk.END)

            exported_count = 0
            error_count = 0
            created_folders = set()

            for folder_name in sorted(files_by_folder.keys()):
                # Create folder
                folder_path = os.path.join(export_root, folder_name)
                if not os.path.exists(folder_path):
                    os.makedirs(folder_path)
                    created_folders.add(folder_name)
                    progress_text.insert(tk.END, f"📁 Created folder: {folder_name}/\n")

                progress_text.insert(tk.END, f"\nExporting to {folder_name}/:\n")

                # Copy files to this folder
                for prog_num, file_path, round_size, title in files_by_folder[folder_name]:
                    try:
                        # Check if file_path is None or empty
                        if not file_path:
                            progress_text.insert(tk.END, f"  ⚠️ SKIP: {prog_num} - No file path in database (run Repair File Paths)\n")
                            error_count += 1
                            continue

                        if not os.path.exists(file_path):
                            progress_text.insert(tk.END, f"  ⚠️ SKIP: {prog_num} - File not found: {file_path}\n")
                            error_count += 1
                            continue

                        # Copy file
                        filename = os.path.basename(file_path)
                        dest_path = os.path.join(folder_path, filename)

                        import shutil
                        shutil.copy2(file_path, dest_path)

                        progress_text.insert(tk.END, f"  ✓ {prog_num} - {filename}\n")
                        exported_count += 1

                        if exported_count % 50 == 0:
                            progress_text.see(tk.END)
                            self.root.update()

                    except Exception as e:
                        progress_text.insert(tk.END, f"  ✗ ERROR copying {prog_num}: {e}\n")
                        error_count += 1

                progress_text.see(tk.END)
                self.root.update()

            # Summary
            progress_text.insert(tk.END, f"\n{'='*80}\n")
            progress_text.insert(tk.END, "EXPORT COMPLETE\n")
            progress_text.insert(tk.END, "="*80 + "\n\n")
            progress_text.insert(tk.END, f"Export Location: {export_root}\n\n")
            progress_text.insert(tk.END, f"Folders Created: {len(created_folders)}\n")
            progress_text.insert(tk.END, f"Files Exported: {exported_count}\n")
            if error_count > 0:
                progress_text.insert(tk.END, f"Errors: {error_count}\n")
            progress_text.insert(tk.END, f"\nTotal Size: {len(files_by_folder)} folders, {exported_count} files\n")
            progress_text.see(tk.END)

            # Log activity
            self.log_activity('export_repository', 'export', {
                'export_root': export_root,
                'folders_created': len(created_folders),
                'files_exported': exported_count,
                'error_count': error_count
            })

            conn.close()

            # Show completion message
            messagebox.showinfo(
                "Export Complete",
                f"Repository exported successfully!\n\n"
                f"Location: {export_root}\n"
                f"Folders: {len(created_folders)}\n"
                f"Files: {exported_count}\n"
                f"Errors: {error_count}"
            )

            tk.Button(progress_window, text="Close", command=progress_window.destroy,
                     bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)

        except Exception as e:
            progress_text.insert(tk.END, f"\n\nERROR: {e}\n")
            import traceback
            progress_text.insert(tk.END, traceback.format_exc())
            progress_text.see(tk.END)

            # Close connection if it exists
            if 'conn' in locals():
                try:
                    conn.close()
                except:
                    pass

            tk.Button(progress_window, text="Close", command=progress_window.destroy,
                     bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)

    def fix_program_number_formatting(self):
        """Fix program numbers that are missing leading zeros (e.g., o1000 -> o01000)"""

        # Create progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Fix Program Number Formatting")
        progress_window.geometry("900x700")
        progress_window.configure(bg=self.bg_color)

        # Title
        title_label = tk.Label(progress_window,
                              text="🔢 Fix Program Number Formatting",
                              font=("Arial", 14, "bold"),
                              bg=self.bg_color, fg=self.fg_color)
        title_label.pack(pady=10)

        # Info
        info_label = tk.Label(progress_window,
                             text="This will add leading zeros to program numbers (o1000 → o01000).\n"
                                  "All program numbers should be in the format o##### (5 digits).",
                             font=("Arial", 10),
                             bg=self.bg_color, fg=self.fg_color,
                             justify=tk.LEFT)
        info_label.pack(pady=5, padx=20, anchor=tk.W)

        # Scrolled text for output
        output_frame = tk.Frame(progress_window, bg=self.bg_color)
        output_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        output_text = scrolledtext.ScrolledText(output_frame,
                                                wrap=tk.WORD,
                                                width=100, height=35,
                                                font=("Courier New", 9),
                                                bg="#1e1e1e", fg="#ffffff")
        output_text.pack(fill=tk.BOTH, expand=True)

        progress_window.update()

        conn = None
        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            output_text.insert(tk.END, "=" * 80 + "\n")
            output_text.insert(tk.END, "SCANNING FOR INCORRECTLY FORMATTED PROGRAM NUMBERS\n")
            output_text.insert(tk.END, "=" * 80 + "\n\n")

            # Find all programs with incorrect format (missing leading zeros)
            cursor.execute("""
                SELECT program_number, file_path, title
                FROM programs
                WHERE LENGTH(program_number) < 6
                ORDER BY CAST(REPLACE(program_number, 'o', '') AS INTEGER)
            """)

            incorrect_programs = cursor.fetchall()
            output_text.insert(tk.END, f"Found {len(incorrect_programs)} programs with incorrect formatting\n\n")

            if not incorrect_programs:
                output_text.insert(tk.END, "✅ All program numbers are correctly formatted!\n")
                conn.close()
                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)
                return

            output_text.insert(tk.END, "=" * 80 + "\n")
            output_text.insert(tk.END, "FIX PREVIEW\n")
            output_text.insert(tk.END, "=" * 80 + "\n\n")

            fixes = []
            for old_num, file_path, title in incorrect_programs[:50]:  # Show first 50
                new_num = self.format_program_number(old_num)

                output_text.insert(tk.END, f"Program: {old_num} → {new_num}\n")
                output_text.insert(tk.END, f"  Title: {title}\n")
                output_text.insert(tk.END, f"  File: {os.path.basename(file_path) if file_path else 'None'}\n\n")

                fixes.append((old_num, new_num, file_path, title))
                progress_window.update()

            if len(incorrect_programs) > 50:
                output_text.insert(tk.END, f"... and {len(incorrect_programs) - 50} more programs\n\n")

            # Include all programs for fixing, not just the displayed ones
            for old_num, file_path, title in incorrect_programs[50:]:
                new_num = self.format_program_number(old_num)
                fixes.append((old_num, new_num, file_path, title))

            output_text.insert(tk.END, "=" * 80 + "\n")
            output_text.insert(tk.END, f"Total programs to fix: {len(fixes)}\n")
            output_text.insert(tk.END, "=" * 80 + "\n\n")

            # Ask user to confirm
            def apply_fixes():
                output_text.insert(tk.END, "\n" + "=" * 80 + "\n")
                output_text.insert(tk.END, "APPLYING FIXES\n")
                output_text.insert(tk.END, "=" * 80 + "\n\n")

                fixed_count = 0
                error_count = 0

                for old_num, new_num, file_path, title in fixes:
                    try:
                        # Check if new number already exists
                        cursor.execute("SELECT COUNT(*) FROM programs WHERE program_number = ?", (new_num,))
                        if cursor.fetchone()[0] > 0 and new_num != old_num:
                            output_text.insert(tk.END, f"⚠️ SKIP: {old_num} → {new_num} (target already exists)\n")
                            error_count += 1
                            continue

                        # Rename file if it exists
                        if file_path and os.path.exists(file_path):
                            old_dir = os.path.dirname(file_path)
                            new_filename = f"{new_num}.nc"
                            new_file_path = os.path.join(old_dir, new_filename)

                            # Only rename if filenames are different
                            if file_path != new_file_path:
                                os.rename(file_path, new_file_path)
                                file_path = new_file_path

                        # Update programs table
                        cursor.execute("""
                            UPDATE programs
                            SET program_number = ?, file_path = ?
                            WHERE program_number = ?
                        """, (new_num, file_path, old_num))

                        # Update registry table
                        cursor.execute("""
                            UPDATE program_number_registry
                            SET program_number = ?, file_path = ?
                            WHERE program_number = ?
                        """, (new_num, file_path, old_num))

                        conn.commit()
                        fixed_count += 1
                        output_text.insert(tk.END, f"✓ Fixed: {old_num} → {new_num}\n")
                    except Exception as e:
                        error_count += 1
                        output_text.insert(tk.END, f"❌ Error fixing {old_num}: {e}\n")

                    progress_window.update()

                output_text.insert(tk.END, "\n" + "=" * 80 + "\n")
                output_text.insert(tk.END, "FIX COMPLETE\n")
                output_text.insert(tk.END, "=" * 80 + "\n\n")
                output_text.insert(tk.END, f"✅ Fixed: {fixed_count}\n")
                output_text.insert(tk.END, f"❌ Errors: {error_count}\n")
                output_text.insert(tk.END, "\nAll program numbers now have proper leading zeros!\n")

                confirm_btn.pack_forget()
                cancel_btn.pack_forget()

                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)

            def cancel_fix():
                conn.close()
                progress_window.destroy()

            # Buttons
            button_frame = tk.Frame(progress_window, bg=self.bg_color)
            button_frame.pack(pady=10)

            confirm_btn = tk.Button(button_frame, text="✓ Apply Fixes",
                                    command=apply_fixes,
                                    bg="#4CAF50", fg="white", font=("Arial", 11, "bold"),
                                    width=15)
            confirm_btn.pack(side=tk.LEFT, padx=5)

            cancel_btn = tk.Button(button_frame, text="✗ Cancel",
                                   command=cancel_fix,
                                   bg="#f44336", fg="white", font=("Arial", 11, "bold"),
                                   width=15)
            cancel_btn.pack(side=tk.LEFT, padx=5)

        except Exception as e:
            if 'output_text' in locals():
                output_text.insert(tk.END, f"\n❌ ERROR: {str(e)}\n")
                import traceback
                output_text.insert(tk.END, f"\n{traceback.format_exc()}\n")

            if 'conn' in locals():
                try:
                    conn.close()
                except:
                    pass

            tk.Button(progress_window, text="Close", command=progress_window.destroy,
                     bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)

    def move_to_correct_range(self):
        """
        Move programs to their correct range based on round size.

        This function prioritizes using proper round size ranges before free ranges.
        It finds and moves two types of programs:
        1. Programs outside their correct range (e.g., o01010 with 7.0" → o70000-o79999)
        2. Programs in free ranges (1000-9999, 14000-49999) that have real round sizes
           and can be consolidated into their proper round size ranges

        This ensures proper round size ranges are filled before using overflow free ranges.
        """
        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            # Create progress window
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Move to Correct Range")
            progress_window.geometry("1000x700")
            progress_window.configure(bg=self.bg_color)
            progress_window.transient(self.root)

            tk.Label(progress_window, text="🎯 Move Programs to Correct Range",
                    font=("Arial", 14, "bold"), bg=self.bg_color, fg=self.fg_color).pack(pady=10)

            # Progress text
            text_frame = tk.Frame(progress_window, bg=self.bg_color)
            text_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

            progress_text = scrolledtext.ScrolledText(text_frame, bg=self.input_bg, fg=self.fg_color,
                                                     font=("Consolas", 9), wrap=tk.WORD)
            progress_text.pack(fill=tk.BOTH, expand=True)

            progress_text.insert(tk.END, "=" * 80 + "\n")
            progress_text.insert(tk.END, "SCANNING FOR PROGRAMS TO MOVE\n")
            progress_text.insert(tk.END, "=" * 80 + "\n")
            progress_text.insert(tk.END, "Looking for:\n")
            progress_text.insert(tk.END, "  1. Programs outside their correct round size range\n")
            progress_text.insert(tk.END, "  2. Programs in free ranges that can move to proper ranges\n")
            progress_text.insert(tk.END, "=" * 80 + "\n\n")
            progress_text.see(tk.END)
            self.root.update()

            # Get programs that need to be moved
            # A program is in wrong range if:
            # 1. It has a round_size OR outer_diameter (OD column in GUI)
            # 2. Its program number doesn't match the range for that round size

            cursor.execute("""
                SELECT program_number, file_path, round_size, outer_diameter, title
                FROM programs
                WHERE is_managed = 1
                  AND file_path IS NOT NULL
                  AND (round_size IS NOT NULL OR outer_diameter IS NOT NULL)
                ORDER BY COALESCE(round_size, outer_diameter), program_number
            """)

            all_programs = cursor.fetchall()

            needs_move = []

            # Define free ranges
            FREE_RANGE_1 = (1000, 9999)    # round_size 0.0
            FREE_RANGE_2 = (14000, 49999)  # round_size -1.0

            for prog_num, file_path, round_size, outer_diameter, title in all_programs:
                # Use round_size if available, otherwise use outer_diameter (OD column)
                effective_round_size = round_size if round_size is not None else outer_diameter

                if effective_round_size is None:
                    continue  # Skip if both are NULL

                # Determine correct range for this round size
                correct_range = self.get_range_for_round_size(effective_round_size)

                if not correct_range:
                    continue  # Unknown round size

                range_start, range_end = correct_range

                # Extract number from program number
                num_str = prog_num.replace('o', '').replace('O', '')
                try:
                    prog_int = int(num_str)
                except:
                    continue

                # Check if program needs to be moved:
                # 1. Program is outside its correct range (wrong range)
                # 2. Program is in a free range BUT has a real round size (not 0.0 or -1.0)
                #    This prioritizes using proper round size ranges before free ranges

                in_free_range_1 = (FREE_RANGE_1[0] <= prog_int <= FREE_RANGE_1[1])
                in_free_range_2 = (FREE_RANGE_2[0] <= prog_int <= FREE_RANGE_2[1])
                in_any_free_range = in_free_range_1 or in_free_range_2

                # Has a real round size (not free range sizes 0.0 or -1.0)
                has_real_round_size = effective_round_size not in (0.0, -1.0)

                # Check if outside correct range
                outside_correct_range = (prog_int < range_start or prog_int > range_end)

                # Move if: outside range OR (in free range AND has real round size)
                if outside_correct_range or (in_any_free_range and has_real_round_size):
                    needs_move.append((prog_num, file_path, effective_round_size, title, range_start, range_end))

            progress_text.insert(tk.END, f"Found {len(needs_move)} programs to move\n\n")

            if not needs_move:
                progress_text.insert(tk.END, "✓ All programs are already in correct ranges!\n")
                progress_text.insert(tk.END, "✓ No programs in free ranges need consolidation!\n")
                progress_text.see(tk.END)
                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)
                conn.close()
                return

            progress_text.insert(tk.END, "=" * 80 + "\n")
            progress_text.insert(tk.END, "PREVIEW - Programs to Move\n")
            progress_text.insert(tk.END, "=" * 80 + "\n\n")

            # Define free ranges for status reporting
            FREE_RANGE_1 = (1000, 9999)
            FREE_RANGE_2 = (14000, 49999)

            # Show preview (first 50)
            for prog_num, file_path, round_size, title, range_start, range_end in needs_move[:50]:
                num_str = prog_num.replace('o', '').replace('O', '')
                prog_int = int(num_str)

                # Determine reason for move
                in_free_range_1 = (FREE_RANGE_1[0] <= prog_int <= FREE_RANGE_1[1])
                in_free_range_2 = (FREE_RANGE_2[0] <= prog_int <= FREE_RANGE_2[1])

                if in_free_range_1:
                    reason = "📦 Currently in Free Range 1 (o1000-o9999)"
                elif in_free_range_2:
                    reason = "📦 Currently in Free Range 2 (o14000-o49999)"
                else:
                    reason = "⚠️ Currently in wrong range"

                progress_text.insert(tk.END, f"Program: {prog_num}\n")
                progress_text.insert(tk.END, f"  {reason}\n")
                progress_text.insert(tk.END, f"  Round size: {round_size}\"\n")
                progress_text.insert(tk.END, f"  Correct range: o{range_start:05d} - o{range_end:05d}\n")
                progress_text.insert(tk.END, f"  Title: {title[:50] if title else '(no title)'}\n")
                progress_text.insert(tk.END, f"  → Will move to first available in correct range\n\n")

            if len(needs_move) > 50:
                progress_text.insert(tk.END, f"... and {len(needs_move) - 50} more programs\n\n")

            progress_text.insert(tk.END, "=" * 80 + "\n")
            progress_text.insert(tk.END, f"Total: {len(needs_move)} programs will be moved\n")
            progress_text.insert(tk.END, "=" * 80 + "\n")
            progress_text.see(tk.END)

            def apply_move():
                progress_text.insert(tk.END, "\n" + "=" * 80 + "\n")
                progress_text.insert(tk.END, "MOVING PROGRAMS\n")
                progress_text.insert(tk.END, "=" * 80 + "\n\n")
                progress_text.see(tk.END)
                self.root.update()

                moved_count = 0
                error_count = 0
                assigned_numbers = set()

                for prog_num, file_path, round_size, title, range_start, range_end in needs_move:
                    try:
                        # Find next available number in correct range
                        new_number = self.find_next_available_number_in_range(
                            range_start, range_end, assigned_numbers
                        )

                        if not new_number:
                            progress_text.insert(tk.END, f"✗ {prog_num}: No available numbers in range\n")
                            error_count += 1
                            continue

                        assigned_numbers.add(new_number)

                        # Check if file exists
                        if not os.path.exists(file_path):
                            progress_text.insert(tk.END, f"✗ {prog_num}: File not found\n")
                            error_count += 1
                            continue

                        # Read file content
                        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                            content = f.read()

                        # Update internal O-number in file
                        old_num_plain = prog_num.replace('o', '').replace('O', '')
                        new_num_plain = new_number.replace('o', '').replace('O', '')

                        # Update O-number at start of line
                        updated_content = re.sub(
                            rf'^[oO]{old_num_plain}\b',
                            new_number.upper(),
                            content,
                            flags=re.MULTILINE
                        )

                        # Update O-number in comments
                        updated_content = re.sub(
                            rf'\b[oO]{old_num_plain}\b',
                            new_number.upper(),
                            updated_content
                        )

                        # Add legacy comment
                        from datetime import datetime
                        today = datetime.now().strftime("%Y-%m-%d")
                        legacy_comment = f"(MOVED FROM {prog_num.upper()} ON {today} - RANGE CORRECTION)\n"

                        # Insert after first O-number line
                        lines = updated_content.split('\n')
                        for i, line in enumerate(lines):
                            if re.match(rf'^[oO]{new_num_plain}\b', line):
                                lines.insert(i + 1, legacy_comment)
                                break
                        updated_content = '\n'.join(lines)

                        # Create new file path
                        old_dir = os.path.dirname(file_path)
                        new_file_path = os.path.join(old_dir, f"{new_number}.nc")

                        # Write to new file
                        with open(new_file_path, 'w', encoding='utf-8') as f:
                            f.write(updated_content)

                        # Delete old file
                        if os.path.exists(file_path) and file_path != new_file_path:
                            os.remove(file_path)

                        # Update database
                        cursor.execute("""
                            UPDATE programs
                            SET program_number = ?, file_path = ?
                            WHERE program_number = ?
                        """, (new_number, new_file_path, prog_num))

                        # Update registry - mark old as AVAILABLE
                        cursor.execute("""
                            UPDATE program_number_registry
                            SET status = 'AVAILABLE', file_path = NULL
                            WHERE program_number = ?
                        """, (prog_num,))

                        # Update registry - mark new as IN_USE
                        cursor.execute("""
                            UPDATE program_number_registry
                            SET status = 'IN_USE', file_path = ?
                            WHERE program_number = ?
                        """, (new_file_path, new_number))

                        conn.commit()
                        moved_count += 1

                        progress_text.insert(tk.END, f"✓ {prog_num} → {new_number} (Round {round_size}\")\n")

                        if moved_count % 10 == 0:
                            progress_text.see(tk.END)
                            self.root.update()

                    except Exception as e:
                        error_count += 1
                        progress_text.insert(tk.END, f"✗ Error moving {prog_num}: {e}\n")

                progress_text.insert(tk.END, "\n" + "=" * 80 + "\n")
                progress_text.insert(tk.END, "COMPLETE\n")
                progress_text.insert(tk.END, "=" * 80 + "\n\n")
                progress_text.insert(tk.END, f"Successfully moved: {moved_count} programs\n")
                if error_count > 0:
                    progress_text.insert(tk.END, f"Errors: {error_count}\n")
                progress_text.see(tk.END)

                # Refresh the view
                self.refresh_results()

                # Close database
                conn.close()

                # Update button
                btn_frame.pack_forget()
                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)

            def cancel_move():
                conn.close()
                progress_window.destroy()

            # Buttons
            btn_frame = tk.Frame(progress_window, bg=self.bg_color)
            btn_frame.pack(pady=10)

            tk.Button(btn_frame, text="✓ Move Programs", command=apply_move,
                     bg="#E91E63", fg=self.fg_color, font=("Arial", 10, "bold"),
                     width=20, height=2).pack(side=tk.LEFT, padx=10)

            tk.Button(btn_frame, text="✗ Cancel", command=cancel_move,
                     bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold"),
                     width=20, height=2).pack(side=tk.LEFT, padx=10)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to scan programs: {str(e)}")
            import traceback
            traceback.print_exc()

    def find_next_available_number_in_range(self, range_start, range_end, assigned_numbers):
        """Find next available program number in specified range"""
        conn = sqlite3.connect(self.db_path, timeout=30.0)
        cursor = conn.cursor()

        for num in range(range_start, range_end + 1):
            prog_num = self.format_program_number(num)

            if prog_num in assigned_numbers:
                continue

            # Check registry
            cursor.execute("""
                SELECT status FROM program_number_registry
                WHERE program_number = ?
            """, (prog_num,))

            result = cursor.fetchone()
            if result and result[0] == 'AVAILABLE':
                conn.close()
                return prog_num

        conn.close()
        return None  # No available numbers found

    def repair_file_paths(self):
        """Repair database file_path entries by scanning repository folder and matching files"""

        # Create progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Repair File Paths")
        progress_window.geometry("900x700")
        progress_window.configure(bg=self.bg_color)

        # Title
        title_label = tk.Label(progress_window,
                              text="🔧 Repair Database File Paths",
                              font=("Arial", 14, "bold"),
                              bg=self.bg_color, fg=self.fg_color)
        title_label.pack(pady=10)

        # Info
        info_label = tk.Label(progress_window,
                             text="This will scan the repository folder and fix file_path entries in the database.\n"
                                  "It matches files on disk to their database entries and updates incorrect paths.",
                             font=("Arial", 10),
                             bg=self.bg_color, fg=self.fg_color,
                             justify=tk.LEFT)
        info_label.pack(pady=5, padx=20, anchor=tk.W)

        # Scrolled text for output
        output_frame = tk.Frame(progress_window, bg=self.bg_color)
        output_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        output_text = scrolledtext.ScrolledText(output_frame,
                                                wrap=tk.WORD,
                                                width=100, height=35,
                                                font=("Courier New", 9),
                                                bg="#1e1e1e", fg="#ffffff")
        output_text.pack(fill=tk.BOTH, expand=True)

        progress_window.update()

        conn = None
        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            output_text.insert(tk.END, "=" * 80 + "\n")
            output_text.insert(tk.END, "SCANNING REPOSITORY FOLDER\n")
            output_text.insert(tk.END, "=" * 80 + "\n\n")

            # Get repository path from class instance
            repo_path = self.repository_path
            if not repo_path:
                output_text.insert(tk.END, "❌ ERROR: No repository path configured\n")
                conn.close()
                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)
                return

            output_text.insert(tk.END, f"Repository Path: {repo_path}\n\n")

            if not os.path.exists(repo_path):
                output_text.insert(tk.END, f"❌ ERROR: Repository path does not exist\n")
                conn.close()
                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)
                return

            # Scan repository for all .nc files
            output_text.insert(tk.END, "Scanning for .nc files...\n")
            progress_window.update()

            nc_files = {}  # program_number -> full_path
            for root, dirs, files in os.walk(repo_path):
                for filename in files:
                    if filename.lower().endswith('.nc'):
                        full_path = os.path.join(root, filename)
                        base_name = os.path.splitext(filename)[0].lower()
                        nc_files[base_name] = full_path

            output_text.insert(tk.END, f"Found {len(nc_files)} .nc files in repository\n\n")
            output_text.insert(tk.END, "=" * 80 + "\n")
            output_text.insert(tk.END, "CHECKING DATABASE ENTRIES\n")
            output_text.insert(tk.END, "=" * 80 + "\n\n")
            progress_window.update()

            # Get all managed programs
            cursor.execute("""
                SELECT program_number, file_path, title
                FROM programs
                WHERE is_managed = 1
                ORDER BY program_number
            """)

            all_programs = cursor.fetchall()
            output_text.insert(tk.END, f"Found {len(all_programs)} managed programs in database\n\n")

            # Check each program's file_path
            needs_repair = []
            for prog_num, file_path, title in all_programs:
                # Check if file exists at stored path
                if file_path and os.path.exists(file_path):
                    continue  # Path is correct

                # File doesn't exist at stored path - try to find it
                # Strip any suffixes like (1), _1, etc. to get base program number
                import re
                # Remove suffix patterns: (1), (2), _1, _2, etc.
                base_prog = re.sub(r'[\(_]\d+[\)]?$', '', prog_num).lower()

                # Try multiple variations
                possible_names = [
                    prog_num.lower(),  # Exact match: o00801(2)
                    base_prog,  # Without suffixes: o00801
                    prog_num.replace('(', '_').replace(')', '').lower(),  # Convert () to _: o00801_2
                ]

                found_path = None
                for name in possible_names:
                    if name in nc_files:
                        found_path = nc_files[name]
                        break

                # Only add to needs_repair if we're actually changing something
                # (Either found a new path, or old path was None/invalid)
                if found_path or not file_path or not os.path.exists(file_path):
                    needs_repair.append((prog_num, file_path, found_path, title))

            output_text.insert(tk.END, f"Found {len(needs_repair)} entries that need repair\n\n")

            if not needs_repair:
                output_text.insert(tk.END, "✅ All file paths are correct! No repairs needed.\n")
                conn.close()
                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)
                return

            output_text.insert(tk.END, "=" * 80 + "\n")
            output_text.insert(tk.END, "REPAIR PREVIEW\n")
            output_text.insert(tk.END, "=" * 80 + "\n\n")

            fixed_count = 0
            missing_count = 0

            for prog_num, old_path, new_path, title in needs_repair[:50]:  # Show first 50
                output_text.insert(tk.END, f"Program: {prog_num}\n")
                output_text.insert(tk.END, f"  Title: {title}\n")
                output_text.insert(tk.END, f"  Current DB path: {old_path if old_path else '(empty)'}\n")

                if new_path:
                    output_text.insert(tk.END, f"  ✓ Found at: {new_path}\n")
                    fixed_count += 1
                else:
                    output_text.insert(tk.END, f"  ❌ File not found in repository\n")
                    missing_count += 1

                output_text.insert(tk.END, "\n")
                progress_window.update()

            if len(needs_repair) > 50:
                output_text.insert(tk.END, f"... and {len(needs_repair) - 50} more entries\n\n")

            output_text.insert(tk.END, "=" * 80 + "\n")
            output_text.insert(tk.END, f"SUMMARY:\n")
            output_text.insert(tk.END, f"  • Can be fixed: {fixed_count}\n")
            output_text.insert(tk.END, f"  • Missing files: {missing_count}\n")
            output_text.insert(tk.END, "=" * 80 + "\n\n")

            # Ask user to confirm
            def apply_repairs():
                output_text.insert(tk.END, "\n" + "=" * 80 + "\n")
                output_text.insert(tk.END, "APPLYING REPAIRS\n")
                output_text.insert(tk.END, "=" * 80 + "\n\n")

                repaired = 0
                failed = 0

                for prog_num, old_path, new_path, title in needs_repair:
                    if new_path:
                        try:
                            # Update programs table
                            cursor.execute("""
                                UPDATE programs
                                SET file_path = ?
                                WHERE program_number = ?
                            """, (new_path, prog_num))

                            # Update registry table
                            cursor.execute("""
                                UPDATE program_number_registry
                                SET file_path = ?
                                WHERE program_number = ?
                            """, (new_path, prog_num))

                            conn.commit()
                            repaired += 1
                            output_text.insert(tk.END, f"✓ Fixed: {prog_num}\n")
                        except Exception as e:
                            failed += 1
                            output_text.insert(tk.END, f"❌ Error fixing {prog_num}: {e}\n")
                    else:
                        # File not found - clear the file_path
                        try:
                            cursor.execute("""
                                UPDATE programs
                                SET file_path = NULL
                                WHERE program_number = ?
                            """, (prog_num,))

                            cursor.execute("""
                                UPDATE program_number_registry
                                SET file_path = NULL
                                WHERE program_number = ?
                            """, (prog_num,))

                            conn.commit()
                            failed += 1
                            output_text.insert(tk.END, f"⚠️ Cleared path for missing file: {prog_num}\n")
                        except Exception as e:
                            failed += 1
                            output_text.insert(tk.END, f"❌ Error clearing {prog_num}: {e}\n")

                    progress_window.update()

                output_text.insert(tk.END, "\n" + "=" * 80 + "\n")
                output_text.insert(tk.END, "REPAIR COMPLETE\n")
                output_text.insert(tk.END, "=" * 80 + "\n\n")
                output_text.insert(tk.END, f"✅ Repaired: {repaired}\n")
                output_text.insert(tk.END, f"⚠️ Missing files: {failed}\n")
                output_text.insert(tk.END, "\nYou can now try batch rename again!\n")

                confirm_btn.pack_forget()
                cancel_btn.pack_forget()

                tk.Button(progress_window, text="Close", command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)

            def cancel_repair():
                conn.close()
                progress_window.destroy()

            # Buttons
            button_frame = tk.Frame(progress_window, bg=self.bg_color)
            button_frame.pack(pady=10)

            confirm_btn = tk.Button(button_frame, text="✓ Apply Repairs",
                                    command=apply_repairs,
                                    bg="#4CAF50", fg="white", font=("Arial", 11, "bold"),
                                    width=15)
            confirm_btn.pack(side=tk.LEFT, padx=5)

            cancel_btn = tk.Button(button_frame, text="✗ Cancel",
                                   command=cancel_repair,
                                   bg="#f44336", fg="white", font=("Arial", 11, "bold"),
                                   width=15)
            cancel_btn.pack(side=tk.LEFT, padx=5)

        except Exception as e:
            if 'output_text' in locals():
                output_text.insert(tk.END, f"\n❌ ERROR: {str(e)}\n")
                import traceback
                output_text.insert(tk.END, f"\n{traceback.format_exc()}\n")

            if 'conn' in locals():
                try:
                    conn.close()
                except:
                    pass

            tk.Button(progress_window, text="Close", command=progress_window.destroy,
                     bg=self.button_bg, fg=self.fg_color, font=("Arial", 10, "bold")).pack(pady=10)

    # ===== Program Number Registry Management =====

    def show_registry_window(self):
        """Show the program number registry statistics window"""
        RegistryStatisticsWindow(self.root, self)

    def show_out_of_range_window(self):
        """Show the out-of-range programs window"""
        OutOfRangeWindow(self.root, self)

    def show_batch_rename_window(self):
        """Show the batch rename resolution window"""
        BatchRenameWindow(self.root, self)

    def show_batch_operations(self):
        """Show batch operations window for selected programs"""
        # Get selected items
        selected_items = self.tree.selection()

        if not selected_items:
            messagebox.showwarning(
                "No Selection",
                "Please select one or more programs first.\n\n"
                "Use Ctrl+Click to select multiple items\n"
                "Use Shift+Click to select a range"
            )
            return

        # Get program numbers for selected items
        selected_programs = []
        for item in selected_items:
            values = self.tree.item(item, 'values')
            if values:
                selected_programs.append(values[0])  # program_number is first column

        if not selected_programs:
            messagebox.showwarning("No Programs", "No valid programs selected.")
            return

        # Create batch operations window
        batch_window = tk.Toplevel(self.root)
        batch_window.title(f"Batch Operations - {len(selected_programs)} Selected")
        batch_window.geometry("700x600")
        batch_window.configure(bg=self.bg_color)
        batch_window.transient(self.root)

        # Header
        header = tk.Label(batch_window,
                         text=f"📦 Batch Operations - {len(selected_programs)} Programs Selected",
                         bg=self.bg_color, fg=self.fg_color,
                         font=("Arial", 14, "bold"))
        header.pack(pady=15)

        # Selected programs list
        list_frame = tk.LabelFrame(batch_window, text="Selected Programs",
                                   bg=self.bg_color, fg=self.fg_color,
                                   font=("Arial", 10, "bold"))
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        list_scroll = ttk.Scrollbar(list_frame)
        list_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        programs_list = tk.Listbox(list_frame, yscrollcommand=list_scroll.set,
                                   bg=self.input_bg, fg=self.fg_color,
                                   font=("Consolas", 9), height=10)
        list_scroll.config(command=programs_list.yview)
        programs_list.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        for prog in selected_programs:
            programs_list.insert(tk.END, prog)

        # Operations frame
        ops_frame = tk.LabelFrame(batch_window, text="Available Operations",
                                 bg=self.bg_color, fg=self.fg_color,
                                 font=("Arial", 10, "bold"))
        ops_frame.pack(fill=tk.X, padx=20, pady=10)

        # Operation buttons - 2 columns
        btn_frame = tk.Frame(ops_frame, bg=self.bg_color)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)

        # Left column
        left_col = tk.Frame(btn_frame, bg=self.bg_color)
        left_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)

        tk.Button(left_col, text="📤 Export Selected",
                 command=lambda: self.batch_export(selected_programs, batch_window),
                 bg="#2E7D32", fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=20, height=2).pack(pady=5)

        tk.Button(left_col, text="🗑️ Delete Selected",
                 command=lambda: self.batch_delete(selected_programs, batch_window),
                 bg="#D32F2F", fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=20, height=2).pack(pady=5)

        tk.Button(left_col, text="📁 Move to Repository",
                 command=lambda: self.batch_move_to_repository(selected_programs, batch_window),
                 bg="#1976D2", fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=20, height=2).pack(pady=5)

        # Right column
        right_col = tk.Frame(btn_frame, bg=self.bg_color)
        right_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)

        tk.Button(right_col, text="🔧 Update Material",
                 command=lambda: self.batch_update_material(selected_programs, batch_window),
                 bg="#FF6F00", fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=20, height=2).pack(pady=5)

        tk.Button(right_col, text="🔄 Re-parse Selected",
                 command=lambda: self.batch_reparse(selected_programs, batch_window),
                 bg="#7B1FA2", fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=20, height=2).pack(pady=5)

        tk.Button(right_col, text="📊 View Statistics",
                 command=lambda: self.batch_statistics(selected_programs, batch_window),
                 bg="#00796B", fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=20, height=2).pack(pady=5)

        # Close button
        tk.Button(batch_window, text="❌ Close", command=batch_window.destroy,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=15).pack(pady=10)

    def batch_export(self, program_numbers, parent_window):
        """Export selected programs to a folder"""
        export_dir = filedialog.askdirectory(title="Select Export Directory")
        if not export_dir:
            return

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        success_count = 0
        failed = []

        for prog_num in program_numbers:
            try:
                cursor.execute("SELECT file_path FROM programs WHERE program_number = ?", (prog_num,))
                result = cursor.fetchone()
                if result and result[0] and os.path.exists(result[0]):
                    src = result[0]
                    dst = os.path.join(export_dir, os.path.basename(src))
                    import shutil
                    shutil.copy2(src, dst)
                    success_count += 1
                else:
                    failed.append(f"{prog_num} (file not found)")
            except Exception as e:
                failed.append(f"{prog_num} ({str(e)})")

        conn.close()

        msg = f"✅ Exported {success_count} of {len(program_numbers)} programs"
        if failed:
            msg += f"\n\n❌ Failed:\n" + "\n".join(failed[:10])
            if len(failed) > 10:
                msg += f"\n... and {len(failed)-10} more"

        messagebox.showinfo("Export Complete", msg)

    def batch_delete(self, program_numbers, parent_window):
        """Delete selected programs from database"""
        confirm = messagebox.askyesno(
            "Confirm Batch Delete",
            f"Are you sure you want to delete {len(program_numbers)} programs from the database?\n\n"
            "This will NOT delete the physical files, only the database entries."
        )
        if not confirm:
            return

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        for prog_num in program_numbers:
            cursor.execute("DELETE FROM programs WHERE program_number = ?", (prog_num,))

        conn.commit()
        conn.close()

        messagebox.showinfo("Delete Complete", f"✅ Deleted {len(program_numbers)} programs from database")
        parent_window.destroy()
        self.refresh_results()

    def batch_move_to_repository(self, program_numbers, parent_window):
        """Move selected programs to repository"""
        moved = 0
        failed = []

        for prog_num in program_numbers:
            try:
                if self.add_to_repository(prog_num):
                    moved += 1
                else:
                    failed.append(prog_num)
            except Exception as e:
                failed.append(f"{prog_num} ({str(e)})")

        msg = f"✅ Moved {moved} of {len(program_numbers)} programs to repository"
        if failed:
            msg += f"\n\n❌ Failed:\n" + "\n".join(str(f) for f in failed[:10])
            if len(failed) > 10:
                msg += f"\n... and {len(failed)-10} more"

        messagebox.showinfo("Move Complete", msg)
        self.refresh_results()

    def batch_update_material(self, program_numbers, parent_window):
        """Update material for selected programs"""
        # Create dialog to select material
        dialog = tk.Toplevel(parent_window)
        dialog.title("Select Material")
        dialog.geometry("300x200")
        dialog.configure(bg=self.bg_color)
        dialog.transient(parent_window)
        dialog.grab_set()

        tk.Label(dialog, text="Select Material:",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 11, "bold")).pack(pady=10)

        material_var = tk.StringVar()
        material_combo = ttk.Combobox(dialog, textvariable=material_var,
                                     values=self.config.get("material_list", []),
                                     state="readonly", width=25)
        material_combo.pack(pady=10)

        def apply_material():
            material = material_var.get()
            if not material:
                messagebox.showwarning("No Material", "Please select a material")
                return

            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            for prog_num in program_numbers:
                cursor.execute("UPDATE programs SET material = ? WHERE program_number = ?",
                             (material, prog_num))

            conn.commit()
            conn.close()

            messagebox.showinfo("Update Complete",
                              f"✅ Updated material to '{material}' for {len(program_numbers)} programs")
            dialog.destroy()
            self.refresh_results()

        tk.Button(dialog, text="Apply", command=apply_material,
                 bg=self.accent_color, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=15).pack(pady=10)

        tk.Button(dialog, text="Cancel", command=dialog.destroy,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=15).pack(pady=5)

    def batch_reparse(self, program_numbers, parent_window):
        """Re-parse selected programs"""
        confirm = messagebox.askyesno(
            "Confirm Re-parse",
            f"Re-parse {len(program_numbers)} selected programs?\n\n"
            "This will update dimensions, validation status, and tool/safety analysis."
        )
        if not confirm:
            return

        from improved_gcode_parser import ImprovedGCodeParser
        parser = ImprovedGCodeParser()

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        success = 0
        failed = []

        for prog_num in program_numbers:
            try:
                cursor.execute("SELECT file_path FROM programs WHERE program_number = ?", (prog_num,))
                result = cursor.fetchone()
                if result and result[0] and os.path.exists(result[0]):
                    parse_result = parser.parse_file(result[0])

                    # Calculate validation status
                    # NOTE: Safety and tool validation disabled for now - too many false positives
                    validation_status = "PASS"
                    if parse_result.validation_issues:
                        validation_status = "CRITICAL"
                    # elif parse_result.safety_blocks_status == "MISSING":
                    #     validation_status = "SAFETY_ERROR"  # DISABLED - too strict
                    # elif parse_result.tool_validation_status == "ERROR":
                    #     validation_status = "TOOL_ERROR"  # DISABLED - needs tuning
                    elif parse_result.bore_warnings:
                        validation_status = "BORE_WARNING"
                    elif parse_result.dimensional_issues:
                        validation_status = "DIMENSIONAL"
                    # elif parse_result.tool_validation_status == "WARNING":
                    #     validation_status = "TOOL_WARNING"  # DISABLED - needs tuning
                    elif parse_result.validation_warnings:
                        validation_status = "WARNING"

                    # Update database
                    import json
                    cursor.execute("""
                        UPDATE programs SET
                            title = ?, spacer_type = ?, outer_diameter = ?, thickness = ?,
                            thickness_display = ?, center_bore = ?, hub_height = ?, hub_diameter = ?,
                            counter_bore_diameter = ?, counter_bore_depth = ?, material = ?,
                            validation_status = ?, validation_issues = ?, validation_warnings = ?,
                            bore_warnings = ?, dimensional_issues = ?,
                            tools_used = ?, tool_sequence = ?, tool_validation_status = ?,
                            tool_validation_issues = ?, safety_blocks_status = ?, safety_blocks_issues = ?
                        WHERE program_number = ?
                    """, (
                        parse_result.title, parse_result.spacer_type, parse_result.outer_diameter,
                        parse_result.thickness, parse_result.thickness_display, parse_result.center_bore,
                        parse_result.hub_height, parse_result.hub_diameter, parse_result.counter_bore_diameter,
                        parse_result.counter_bore_depth, parse_result.material, validation_status,
                        '|'.join(parse_result.validation_issues) if parse_result.validation_issues else None,
                        '|'.join(parse_result.validation_warnings) if parse_result.validation_warnings else None,
                        '|'.join(parse_result.bore_warnings) if parse_result.bore_warnings else None,
                        '|'.join(parse_result.dimensional_issues) if parse_result.dimensional_issues else None,
                        json.dumps(parse_result.tools_used) if parse_result.tools_used else None,
                        json.dumps(parse_result.tool_sequence) if parse_result.tool_sequence else None,
                        None,  # tool_validation_status - DISABLED
                        None,  # tool_validation_issues - DISABLED
                        None,  # safety_blocks_status - DISABLED
                        None,  # safety_blocks_issues - DISABLED
                        prog_num
                    ))
                    success += 1
                else:
                    failed.append(f"{prog_num} (file not found)")
            except Exception as e:
                failed.append(f"{prog_num} ({str(e)})")

        conn.commit()
        conn.close()

        msg = f"✅ Re-parsed {success} of {len(program_numbers)} programs"
        if failed:
            msg += f"\n\n❌ Failed:\n" + "\n".join(failed[:10])
            if len(failed) > 10:
                msg += f"\n... and {len(failed)-10} more"

        messagebox.showinfo("Re-parse Complete", msg)
        self.refresh_results()

    def batch_statistics(self, program_numbers, parent_window):
        """Show statistics for selected programs"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Gather statistics
        stats = {
            'total': len(program_numbers),
            'types': {},
            'materials': {},
            'statuses': {},
            'avg_od': 0,
            'avg_thickness': 0,
            'avg_cb': 0
        }

        od_values = []
        thickness_values = []
        cb_values = []

        for prog_num in program_numbers:
            cursor.execute("""
                SELECT spacer_type, material, validation_status,
                       outer_diameter, thickness, center_bore
                FROM programs WHERE program_number = ?
            """, (prog_num,))
            result = cursor.fetchone()
            if result:
                stype, mat, status, od, thick, cb = result
                stats['types'][stype] = stats['types'].get(stype, 0) + 1
                stats['materials'][mat or 'Unknown'] = stats['materials'].get(mat or 'Unknown', 0) + 1
                stats['statuses'][status or 'N/A'] = stats['statuses'].get(status or 'N/A', 0) + 1

                if od:
                    od_values.append(od)
                if thick:
                    thickness_values.append(thick)
                if cb:
                    cb_values.append(cb)

        conn.close()

        if od_values:
            stats['avg_od'] = sum(od_values) / len(od_values)
        if thickness_values:
            stats['avg_thickness'] = sum(thickness_values) / len(thickness_values)
        if cb_values:
            stats['avg_cb'] = sum(cb_values) / len(cb_values)

        # Display statistics window
        stats_window = tk.Toplevel(parent_window)
        stats_window.title("Batch Statistics")
        stats_window.geometry("500x600")
        stats_window.configure(bg=self.bg_color)
        stats_window.transient(parent_window)

        tk.Label(stats_window, text=f"📊 Statistics for {stats['total']} Selected Programs",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 13, "bold")).pack(pady=15)

        # Stats text
        stats_frame = tk.Frame(stats_window, bg=self.bg_color)
        stats_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        stats_scroll = ttk.Scrollbar(stats_frame)
        stats_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        stats_text = tk.Text(stats_frame, yscrollcommand=stats_scroll.set,
                            bg=self.input_bg, fg=self.fg_color,
                            font=("Consolas", 10), wrap=tk.WORD)
        stats_scroll.config(command=stats_text.yview)
        stats_text.pack(fill=tk.BOTH, expand=True)

        # Build stats text
        output = f"Total Programs: {stats['total']}\n\n"

        output += "SPACER TYPES:\n"
        for stype, count in sorted(stats['types'].items()):
            pct = (count / stats['total']) * 100
            output += f"  {stype}: {count} ({pct:.1f}%)\n"

        output += "\nMATERIALS:\n"
        for mat, count in sorted(stats['materials'].items()):
            pct = (count / stats['total']) * 100
            output += f"  {mat}: {count} ({pct:.1f}%)\n"

        output += "\nVALIDATION STATUS:\n"
        for status, count in sorted(stats['statuses'].items()):
            pct = (count / stats['total']) * 100
            output += f"  {status}: {count} ({pct:.1f}%)\n"

        output += "\nAVERAGE DIMENSIONS:\n"
        output += f"  Outer Diameter: {stats['avg_od']:.2f} mm\n"
        output += f"  Thickness: {stats['avg_thickness']:.3f} mm\n"
        output += f"  Center Bore: {stats['avg_cb']:.2f} mm\n"

        stats_text.insert(tk.END, output)
        stats_text.config(state=tk.DISABLED)

        tk.Button(stats_window, text="Close", command=stats_window.destroy,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=15).pack(pady=10)

    # ===== Workflow UI Methods =====

    def sync_registry_ui(self):
        """UI wrapper for syncing the program number registry"""
        # Confirm with user
        confirm = messagebox.askyesno(
            "Sync Program Number Registry",
            "This will update the program number registry with all current programs.\n\n"
            "What it does:\n"
            "  • Marks all existing program numbers as IN_USE\n"
            "  • Marks all unused numbers as AVAILABLE\n"
            "  • Takes about 0.4 seconds\n\n"
            "When to run:\n"
            "  • After scanning new folders\n"
            "  • After deleting programs\n"
            "  • Before batch rename operations\n\n"
            "Do you want to sync the registry now?",
            icon='question'
        )

        if not confirm:
            return

        # Create progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Syncing Registry")
        progress_window.geometry("600x400")
        progress_window.configure(bg=self.bg_color)
        progress_window.transient(self.root)
        progress_window.grab_set()

        tk.Label(progress_window, text="🔄 Syncing Program Number Registry...",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 12, "bold")).pack(pady=10)

        # Log text
        log_frame = tk.Frame(progress_window, bg=self.bg_color)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        log_scroll = ttk.Scrollbar(log_frame)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        log_text = tk.Text(log_frame, height=20, yscrollcommand=log_scroll.set,
                          bg="#2B2B2B", fg="#FFFFFF", font=("Consolas", 9))
        log_scroll.config(command=log_text.yview)
        log_text.pack(fill=tk.BOTH, expand=True)

        def log(message):
            log_text.insert(tk.END, message + "\n")
            log_text.see(tk.END)
            progress_window.update()

        # Run sync
        try:
            log("Starting registry sync...")
            log("-" * 60)

            stats = self.populate_program_registry()

            if stats:
                log("-" * 60)
                log("✅ REGISTRY SYNC COMPLETE")
                log(f"Total numbers: {stats['total_generated']:,}")
                log(f"In use: {stats['in_use']:,}")
                log(f"Available: {stats['available']:,}")
                log(f"Duplicates: {stats['duplicates']}")
                log("")
                log("Registry is now up to date!")

                # Add close button
                tk.Button(progress_window, text="Close",
                         command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color,
                         font=("Arial", 10, "bold")).pack(pady=10)

                messagebox.showinfo(
                    "Sync Complete",
                    f"Registry synced successfully!\n\n"
                    f"In use: {stats['in_use']:,}\n"
                    f"Available: {stats['available']:,}\n"
                    f"Total: {stats['total_generated']:,}"
                )
            else:
                log("❌ Sync failed!")
                tk.Button(progress_window, text="Close",
                         command=progress_window.destroy,
                         bg=self.button_bg, fg=self.fg_color,
                         font=("Arial", 10, "bold")).pack(pady=10)

        except Exception as e:
            log(f"❌ Error: {str(e)}")
            tk.Button(progress_window, text="Close",
                     command=progress_window.destroy,
                     bg=self.button_bg, fg=self.fg_color,
                     font=("Arial", 10, "bold")).pack(pady=10)
            messagebox.showerror("Sync Error", f"Failed to sync registry:\n{str(e)}")

    def detect_round_sizes_ui(self):
        """UI wrapper for detecting round sizes in all programs"""
        # Confirm with user
        confirm = messagebox.askyesno(
            "Detect Round Sizes",
            "This will detect round sizes for all programs in the database.\n\n"
            "Detection methods (in priority order):\n"
            "  1. Title - looks for patterns like '6.25 OD', '10.5 rnd'\n"
            "  2. G-code - uses ob_from_gcode field\n"
            "  3. Dimensions - uses outer_diameter field\n\n"
            "This will update the database with:\n"
            "  • round_size\n"
            "  • round_size_confidence (HIGH/MEDIUM/LOW)\n"
            "  • round_size_source (title/gcode/dimension)\n"
            "  • in_correct_range (0 or 1)\n\n"
            "Do you want to detect round sizes now?",
            icon='question'
        )

        if not confirm:
            return

        # Create progress window
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Detecting Round Sizes")
        progress_window.geometry("700x500")
        progress_window.configure(bg=self.bg_color)
        progress_window.transient(self.root)
        progress_window.grab_set()

        tk.Label(progress_window, text="🎯 Detecting Round Sizes...",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 12, "bold")).pack(pady=10)

        # Progress bar
        progress_var = tk.DoubleVar()
        progress_bar = ttk.Progressbar(progress_window, variable=progress_var,
                                      maximum=100, length=600)
        progress_bar.pack(pady=10, padx=20)

        # Status label
        status_label = tk.Label(progress_window, text="Starting...",
                               bg=self.bg_color, fg=self.fg_color,
                               font=("Arial", 10))
        status_label.pack(pady=5)

        # Log text
        log_frame = tk.Frame(progress_window, bg=self.bg_color)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        log_scroll = ttk.Scrollbar(log_frame)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        log_text = tk.Text(log_frame, height=20, yscrollcommand=log_scroll.set,
                          bg="#2B2B2B", fg="#FFFFFF", font=("Consolas", 9))
        log_scroll.config(command=log_text.yview)
        log_text.pack(fill=tk.BOTH, expand=True)

        def log(message):
            log_text.insert(tk.END, message + "\n")
            log_text.see(tk.END)
            progress_window.update()

        # Run detection
        try:
            import time
            start_time = time.time()

            log("Fetching all programs...")
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT program_number, title, ob_from_gcode, outer_diameter FROM programs")
            programs = cursor.fetchall()

            total = len(programs)
            log(f"Found {total:,} programs to process")
            log("-" * 60)

            results = {
                'processed': 0,
                'detected_title': 0,
                'detected_gcode': 0,
                'detected_dimension': 0,
                'manual_needed': 0,
                'in_correct_range': 0,
                'out_of_range': 0,
                'errors': 0
            }

            for i, (program_number, title, ob_from_gcode, outer_diameter) in enumerate(programs):
                try:
                    # Update progress
                    if i % 100 == 0:
                        progress_var.set((i / total) * 100)
                        status_label.config(text=f"Processing {i+1}/{total}: {program_number}")
                        progress_window.update()

                    # Detect round size using priority order
                    round_size = None
                    confidence = 'NONE'
                    source = 'MANUAL'

                    # Method 1: Title
                    if title:
                        title_match = self.detect_round_size_from_title(title)
                        if title_match:
                            round_size = title_match
                            confidence = 'HIGH'
                            source = 'title'
                            results['detected_title'] += 1

                    # Method 2: G-code (if title didn't work)
                    if not round_size and ob_from_gcode:
                        if 5.0 <= ob_from_gcode <= 15.0:
                            round_size = ob_from_gcode
                            confidence = 'MEDIUM'
                            source = 'gcode'
                            results['detected_gcode'] += 1

                    # Method 3: Dimensions (if others didn't work)
                    if not round_size and outer_diameter:
                        if 5.0 <= outer_diameter <= 15.0:
                            round_size = outer_diameter
                            confidence = 'LOW'
                            source = 'dimension'
                            results['detected_dimension'] += 1

                    # Check if in correct range
                    in_correct_range = 1 if round_size and self.is_in_correct_range(program_number, round_size) else 0

                    if not round_size:
                        results['manual_needed'] += 1
                    elif in_correct_range:
                        results['in_correct_range'] += 1
                    else:
                        results['out_of_range'] += 1

                    # Update database
                    cursor.execute("""
                        UPDATE programs
                        SET round_size = ?,
                            round_size_confidence = ?,
                            round_size_source = ?,
                            in_correct_range = ?
                        WHERE program_number = ?
                    """, (round_size, confidence, source, in_correct_range, program_number))

                    results['processed'] += 1

                except Exception as e:
                    results['errors'] += 1
                    if results['errors'] <= 5:  # Only log first 5 errors
                        log(f"Error processing {program_number}: {str(e)}")

            conn.commit()
            conn.close()

            elapsed = time.time() - start_time

            # Show results
            progress_var.set(100)
            status_label.config(text="Complete!")
            log("-" * 60)
            log("✅ DETECTION COMPLETE")
            log(f"Total processed: {results['processed']:,}")
            log(f"Detected from title: {results['detected_title']:,} ({results['detected_title']/total*100:.1f}%)")
            log(f"Detected from G-code: {results['detected_gcode']:,} ({results['detected_gcode']/total*100:.1f}%)")
            log(f"Detected from dimensions: {results['detected_dimension']:,} ({results['detected_dimension']/total*100:.1f}%)")
            log(f"Manual input needed: {results['manual_needed']:,} ({results['manual_needed']/total*100:.1f}%)")
            log("")
            log(f"In correct range: {results['in_correct_range']:,}")
            log(f"Out of range: {results['out_of_range']:,}")
            log(f"Errors: {results['errors']}")
            log(f"Time elapsed: {elapsed:.2f} seconds")

            # Add close button
            tk.Button(progress_window, text="Close",
                     command=progress_window.destroy,
                     bg=self.button_bg, fg=self.fg_color,
                     font=("Arial", 10, "bold")).pack(pady=10)

            messagebox.showinfo(
                "Detection Complete",
                f"Round size detection completed!\n\n"
                f"Detected: {results['detected_title'] + results['detected_gcode'] + results['detected_dimension']:,}\n"
                f"Manual needed: {results['manual_needed']:,}\n"
                f"Out of range: {results['out_of_range']:,}\n\n"
                f"Time: {elapsed:.2f} seconds"
            )

            # Refresh the view to show updated data
            self.refresh_results()

        except Exception as e:
            log(f"❌ Error: {str(e)}")
            tk.Button(progress_window, text="Close",
                     command=progress_window.destroy,
                     bg=self.button_bg, fg=self.fg_color,
                     font=("Arial", 10, "bold")).pack(pady=10)
            messagebox.showerror("Detection Error", f"Failed to detect round sizes:\n{str(e)}")

    def show_round_size_stats(self):
        """Show statistics about round size detection"""
        try:
            conn = sqlite3.connect(self.db_path, timeout=30.0)
            cursor = conn.cursor()

            # Get overall stats
            cursor.execute("""
                SELECT
                    COUNT(*) as total,
                    SUM(CASE WHEN round_size IS NOT NULL THEN 1 ELSE 0 END) as detected,
                    SUM(CASE WHEN round_size IS NULL THEN 1 ELSE 0 END) as not_detected,
                    SUM(CASE WHEN in_correct_range = 1 THEN 1 ELSE 0 END) as in_range,
                    SUM(CASE WHEN in_correct_range = 0 AND round_size IS NOT NULL THEN 1 ELSE 0 END) as out_of_range
                FROM programs
            """)
            overall = cursor.fetchone()

            # Get stats by source
            cursor.execute("""
                SELECT round_size_source, COUNT(*) as count
                FROM programs
                WHERE round_size IS NOT NULL
                GROUP BY round_size_source
            """)
            by_source = cursor.fetchall()

            # Get stats by round size
            cursor.execute("""
                SELECT
                    round_size,
                    COUNT(*) as count,
                    SUM(CASE WHEN in_correct_range = 1 THEN 1 ELSE 0 END) as in_range,
                    SUM(CASE WHEN in_correct_range = 0 THEN 1 ELSE 0 END) as out_of_range
                FROM programs
                WHERE round_size IS NOT NULL
                GROUP BY round_size
                ORDER BY round_size
            """)
            by_size = cursor.fetchall()

            conn.close()

            # Create window
            stats_window = tk.Toplevel(self.root)
            stats_window.title("Round Size Statistics")
            stats_window.geometry("800x700")
            stats_window.configure(bg=self.bg_color)

            tk.Label(stats_window, text="📊 Round Size Detection Statistics",
                    bg=self.bg_color, fg="#9B59B6",
                    font=("Arial", 14, "bold")).pack(pady=10)

            # Create notebook for tabs
            notebook = ttk.Notebook(stats_window)
            notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            # Tab 1: Overall Stats
            overall_frame = tk.Frame(notebook, bg=self.bg_color)
            notebook.add(overall_frame, text="Overall")

            total, detected, not_detected, in_range, out_of_range = overall

            stats_text = f"""
Total Programs: {total:,}

Detection Results:
  ✅ Detected: {detected:,} ({detected/total*100:.1f}%)
  ❌ Not detected: {not_detected:,} ({not_detected/total*100:.1f}%)

Range Validation:
  ✅ In correct range: {in_range:,} ({in_range/total*100:.1f}%)
  ⚠️ Out of range: {out_of_range:,} ({out_of_range/total*100:.1f}%)
  ❓ No round size: {not_detected:,}
"""

            tk.Label(overall_frame, text=stats_text,
                    bg=self.bg_color, fg=self.fg_color,
                    font=("Consolas", 11), justify=tk.LEFT).pack(pady=20, padx=20)

            # Tab 2: By Source
            source_frame = tk.Frame(notebook, bg=self.bg_color)
            notebook.add(source_frame, text="By Source")

            source_text = "Detection Sources:\n\n"
            for source, count in by_source:
                source_text += f"  {source or 'MANUAL'}: {count:,}\n"

            tk.Label(source_frame, text=source_text,
                    bg=self.bg_color, fg=self.fg_color,
                    font=("Consolas", 11), justify=tk.LEFT).pack(pady=20, padx=20)

            # Tab 3: By Round Size
            size_frame = tk.Frame(notebook, bg=self.bg_color)
            notebook.add(size_frame, text="By Round Size")

            # Create treeview
            tree_frame = tk.Frame(size_frame, bg=self.bg_color)
            tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            vsb = ttk.Scrollbar(tree_frame, orient="vertical")
            vsb.pack(side=tk.RIGHT, fill=tk.Y)

            columns = ("Round Size", "Total", "In Range", "Out of Range", "% Out of Range")
            tree = ttk.Treeview(tree_frame, columns=columns, show="headings",
                               yscrollcommand=vsb.set)
            vsb.config(command=tree.yview)

            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, width=120)

            for size, count, in_range, out in by_size:
                pct_out = (out / count * 100) if count > 0 else 0
                tree.insert("", tk.END, values=(
                    f"{size:.2f}\"",
                    count,
                    in_range,
                    out,
                    f"{pct_out:.1f}%"
                ))

            tree.pack(fill=tk.BOTH, expand=True)

            tk.Button(stats_window, text="Close",
                     command=stats_window.destroy,
                     bg=self.button_bg, fg=self.fg_color,
                     font=("Arial", 10, "bold")).pack(pady=10)

        except Exception as e:
            messagebox.showerror("Statistics Error", f"Failed to get statistics:\n{str(e)}")

    def show_workflow_guide(self):
        """Show the workflow guide"""
        try:
            # Try to open RECOMMENDED_WORKFLOW.md or DETAILED_USER_WORKFLOW.md
            workflow_files = [
                "RECOMMENDED_WORKFLOW.md",
                "DETAILED_USER_WORKFLOW.md",
                "README.md"
            ]

            for filename in workflow_files:
                if os.path.exists(filename):
                    # Open in default markdown viewer or text editor
                    if os.name == 'nt':  # Windows
                        os.startfile(filename)
                    else:  # macOS and Linux
                        import subprocess
                        subprocess.call(['open' if sys.platform == 'darwin' else 'xdg-open', filename])
                    return

            # If no file found, show a basic guide
            messagebox.showinfo(
                "Workflow Guide",
                "Recommended Workflow:\n\n"
                "1. Scan Folder (Files tab)\n"
                "   → Import G-code files from external folders\n\n"
                "2. Sync Registry (Workflow tab)\n"
                "   → Update program number registry\n\n"
                "3. Detect Round Sizes (Workflow tab)\n"
                "   → Auto-detect round sizes from titles/G-code\n\n"
                "4. Add to Repository (External tab)\n"
                "   → Move files to managed repository\n\n"
                "5. Resolve Out-of-Range (Repository tab)\n"
                "   → Batch rename programs to correct ranges\n\n"
                "6. Manage Duplicates (Repository tab)\n"
                "   → Clean up duplicate programs\n\n"
                "See RECOMMENDED_WORKFLOW.md for detailed guide."
            )

        except Exception as e:
            messagebox.showerror("Error", f"Failed to open workflow guide:\n{str(e)}")


class EditEntryWindow:
    """Window for adding/editing entries"""
    def __init__(self, parent, program_number, callback):
        self.parent = parent
        self.program_number = program_number
        self.callback = callback
        
        self.window = tk.Toplevel(parent.root)
        self.window.title("Edit Entry" if program_number else "Add Entry")
        self.window.geometry("600x700")
        self.window.configure(bg=parent.bg_color)
        
        # Load existing data if editing
        self.record = None
        if program_number:
            conn = sqlite3.connect(parent.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM programs WHERE program_number = ?", (program_number,))
            self.record = cursor.fetchone()
            conn.close()
        
        self.create_form()
        
    def create_form(self):
        """Create entry form"""
        # Container
        container = tk.Frame(self.window, bg=self.parent.bg_color)
        container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        row = 0
        
        # Program Number
        tk.Label(container, text="Program Number:", bg=self.parent.bg_color, 
                fg=self.parent.fg_color).grid(row=row, column=0, sticky="w", pady=5)
        self.entry_program = tk.Entry(container, bg=self.parent.input_bg, 
                                     fg=self.parent.fg_color, width=30)
        self.entry_program.grid(row=row, column=1, sticky="ew", pady=5)
        if self.record:
            self.entry_program.insert(0, self.record[0])
            self.entry_program.config(state="readonly")  # Can't change primary key
        row += 1
        
        # Spacer Type
        tk.Label(container, text="Spacer Type:", bg=self.parent.bg_color, 
                fg=self.parent.fg_color).grid(row=row, column=0, sticky="w", pady=5)
        self.entry_type = ttk.Combobox(container, values=self.parent.config["spacer_types"], 
                                      state="readonly", width=28)
        self.entry_type.grid(row=row, column=1, sticky="ew", pady=5)
        if self.record and self.record[1]:
            self.entry_type.set(self.record[1])
        row += 1
        
        # Dimensions
        dims = [
            ("Outer Diameter:", 2, "outer_diameter"),
            ("Thickness:", 3, "thickness"),
            ("Center Bore:", 4, "center_bore"),
            ("Hub Height:", 5, "hub_height"),
            ("Hub Diameter:", 6, "hub_diameter"),
            ("Counter Bore Diameter:", 7, "cb_diameter"),
            ("Counter Bore Depth:", 8, "cb_depth"),
        ]
        
        self.entries = {}
        for label, idx, key in dims:
            tk.Label(container, text=label, bg=self.parent.bg_color, 
                    fg=self.parent.fg_color).grid(row=row, column=0, sticky="w", pady=5)
            entry = tk.Entry(container, bg=self.parent.input_bg, 
                           fg=self.parent.fg_color, width=30)
            entry.grid(row=row, column=1, sticky="ew", pady=5)
            if self.record and self.record[idx]:
                entry.insert(0, str(self.record[idx]))
            self.entries[key] = entry
            row += 1
        
        # Paired Program
        tk.Label(container, text="Paired Program:", bg=self.parent.bg_color, 
                fg=self.parent.fg_color).grid(row=row, column=0, sticky="w", pady=5)
        self.entry_paired = tk.Entry(container, bg=self.parent.input_bg, 
                                     fg=self.parent.fg_color, width=30)
        self.entry_paired.grid(row=row, column=1, sticky="ew", pady=5)
        if self.record and self.record[9]:
            self.entry_paired.insert(0, self.record[9])
        row += 1
        
        # Material
        tk.Label(container, text="Material:", bg=self.parent.bg_color, 
                fg=self.parent.fg_color).grid(row=row, column=0, sticky="w", pady=5)
        self.entry_material = ttk.Combobox(container, 
                                          values=self.parent.config["material_list"], 
                                          state="readonly", width=28)
        self.entry_material.grid(row=row, column=1, sticky="ew", pady=5)
        if self.record and self.record[10]:
            self.entry_material.set(self.record[10])
        row += 1
        
        # Notes
        tk.Label(container, text="Notes:", bg=self.parent.bg_color, 
                fg=self.parent.fg_color).grid(row=row, column=0, sticky="nw", pady=5)
        self.entry_notes = tk.Text(container, bg=self.parent.input_bg, 
                                  fg=self.parent.fg_color, width=30, height=4)
        self.entry_notes.grid(row=row, column=1, sticky="ew", pady=5)
        if self.record and self.record[11]:
            self.entry_notes.insert("1.0", self.record[11])
        row += 1
        
        # File Path
        tk.Label(container, text="File Path:", bg=self.parent.bg_color, 
                fg=self.parent.fg_color).grid(row=row, column=0, sticky="w", pady=5)
        path_frame = tk.Frame(container, bg=self.parent.bg_color)
        path_frame.grid(row=row, column=1, sticky="ew", pady=5)
        
        self.entry_filepath = tk.Entry(path_frame, bg=self.parent.input_bg, 
                                      fg=self.parent.fg_color)
        self.entry_filepath.pack(side=tk.LEFT, fill=tk.X, expand=True)
        if self.record and self.record[14]:
            self.entry_filepath.insert(0, self.record[14])
        
        btn_browse = tk.Button(path_frame, text="Browse", command=self.browse_file,
                              bg=self.parent.button_bg, fg=self.parent.fg_color)
        btn_browse.pack(side=tk.LEFT, padx=5)
        row += 1
        
        # Buttons
        button_frame = tk.Frame(container, bg=self.parent.bg_color)
        button_frame.grid(row=row, column=0, columnspan=2, pady=20)
        
        btn_save = tk.Button(button_frame, text="💾 Save", command=self.save_entry,
                            bg=self.parent.accent_color, fg=self.parent.fg_color,
                            font=("Arial", 10, "bold"), width=15)
        btn_save.pack(side=tk.LEFT, padx=5)
        
        btn_cancel = tk.Button(button_frame, text="❌ Cancel", command=self.window.destroy,
                              bg=self.parent.button_bg, fg=self.parent.fg_color,
                              font=("Arial", 10, "bold"), width=15)
        btn_cancel.pack(side=tk.LEFT, padx=5)
        
        container.columnconfigure(1, weight=1)
        
    def browse_file(self):
        """Browse for file"""
        filepath = filedialog.askopenfilename(
            title="Select G-Code File",
            filetypes=[("G-Code files", "*.nc *.gcode *.cnc"), ("All files", "*.*")]
        )
        if filepath:
            self.entry_filepath.delete(0, tk.END)
            self.entry_filepath.insert(0, filepath)
            
    def save_entry(self):
        """Save entry to database"""
        # Validate
        program_number = self.entry_program.get().strip()
        if not program_number:
            messagebox.showerror("Error", "Program number is required")
            return
        
        spacer_type = self.entry_type.get()
        if not spacer_type:
            messagebox.showerror("Error", "Spacer type is required")
            return
        
        # Get values
        def get_float(entry):
            try:
                val = entry.get().strip()
                return float(val) if val else None
            except:
                return None
        
        outer_diameter = get_float(self.entries["outer_diameter"])
        thickness = get_float(self.entries["thickness"])
        center_bore = get_float(self.entries["center_bore"])
        hub_height = get_float(self.entries["hub_height"])
        hub_diameter = get_float(self.entries["hub_diameter"])
        cb_diameter = get_float(self.entries["cb_diameter"])
        cb_depth = get_float(self.entries["cb_depth"])
        
        paired_program = self.entry_paired.get().strip() or None
        material = self.entry_material.get() or None
        notes = self.entry_notes.get("1.0", tk.END).strip() or None
        file_path = self.entry_filepath.get().strip() or None
        
        # Save to database
        conn = sqlite3.connect(self.parent.db_path)
        cursor = conn.cursor()
        
        try:
            if self.program_number:  # Update
                cursor.execute('''
                    UPDATE programs SET
                        spacer_type = ?, outer_diameter = ?, thickness = ?,
                        center_bore = ?, hub_height = ?, hub_diameter = ?,
                        counter_bore_diameter = ?, counter_bore_depth = ?,
                        paired_program = ?, material = ?, notes = ?,
                        last_modified = ?, file_path = ?
                    WHERE program_number = ?
                ''', (spacer_type, outer_diameter, thickness, center_bore,
                     hub_height, hub_diameter, cb_diameter, cb_depth,
                     paired_program, material, notes, 
                     datetime.now().isoformat(), file_path, program_number))
            else:  # Insert
                cursor.execute('''
                    INSERT INTO programs VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (program_number, None,  # title
                     spacer_type, outer_diameter, thickness, None,  # thickness_display
                     center_bore, hub_height, hub_diameter, cb_diameter, cb_depth,
                     paired_program, material, notes, datetime.now().isoformat(),
                     datetime.now().isoformat(), file_path,
                     None, None, None, None, None,  # detection_confidence, detection_method, validation_status, validation_issues, validation_warnings
                     None, None,  # bore_warnings, dimensional_issues
                     None, None, None,  # cb_from_gcode, ob_from_gcode, lathe
                     None, None, None,  # duplicate_type, parent_file, duplicate_group
                     1, self.parent.current_username, 0,  # current_version, modified_by, is_managed
                     None, None, None, None,  # round_size, round_size_confidence, round_size_source, in_correct_range
                     None, None, None,  # legacy_names, last_renamed_date, rename_reason
                     None, None, None, None, None, None))  # tools_used, tool_sequence, tool_validation_status, tool_validation_issues, safety_blocks_status, safety_blocks_issues
            
            conn.commit()
            messagebox.showinfo("Success", "Entry saved successfully")
            self.window.destroy()
            self.callback()
            
        except sqlite3.IntegrityError:
            messagebox.showerror("Error", "Program number already exists")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save: {e}")
        finally:
            conn.close()


class DetailsWindow:
    """Window to view full details of an entry"""
    def __init__(self, parent, record):
        self.parent = parent
        self.record = record
        
        self.window = tk.Toplevel(parent.root)
        self.window.title(f"Details - {record[0]}")
        self.window.geometry("500x600")
        self.window.configure(bg=parent.bg_color)
        
        self.create_view()
        
    def create_view(self):
        """Create details view"""
        # Scrolled text
        text = scrolledtext.ScrolledText(self.window, bg=self.parent.input_bg, 
                                        fg=self.parent.fg_color,
                                        font=("Courier", 10), wrap=tk.WORD)
        text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Format details
        fields = [
            ("Program Number", 0),
            ("Title", 1),
            ("Spacer Type", 2),
            ("Outer Diameter", 3),
            ("Thickness", 4),
            ("Thickness Display", 5),
            ("Center Bore", 6),
            ("Hub Height", 7),
            ("Hub Diameter", 8),
            ("Counter Bore Diameter", 9),
            ("Counter Bore Depth", 10),
            ("Paired Program", 11),
            ("Material", 12),
            ("Notes", 13),
            ("Date Created", 14),
            ("Last Modified", 15),
            ("File Path", 16),
        ]

        text.insert(tk.END, "="*50 + "\n")
        text.insert(tk.END, f"  PROGRAM DETAILS - {self.record[0]}\n")
        text.insert(tk.END, "="*50 + "\n\n")

        for label, idx in fields:
            value = self.record[idx] if self.record[idx] else "-"
            text.insert(tk.END, f"{label}:\n  {value}\n\n")

        # Add validation information if available
        if len(self.record) > 18:
            text.insert(tk.END, "="*50 + "\n")
            text.insert(tk.END, "  VALIDATION RESULTS\n")
            text.insert(tk.END, "="*50 + "\n\n")

            # Detection info
            if self.record[17]:  # detection_confidence
                text.insert(tk.END, f"Detection Confidence: {self.record[17]}\n")
            if self.record[18]:  # detection_method
                text.insert(tk.END, f"Detection Method: {self.record[18]}\n")

            # Lathe assignment
            if len(self.record) > 26 and self.record[26]:  # lathe
                text.insert(tk.END, f"Lathe: {self.record[26]}\n")

            # Validation status
            status = self.record[19] if self.record[19] else "N/A"
            text.insert(tk.END, f"\nValidation Status: {status}\n")

            # If REPEAT status, show parent file information
            if status == 'REPEAT':
                text.insert(tk.END, "\n" + "="*50 + "\n")
                text.insert(tk.END, "  DUPLICATE INFORMATION\n")
                text.insert(tk.END, "="*50 + "\n\n")

                # Get parent_file and duplicate_type from record
                parent_file = self.record[28] if len(self.record) > 28 and self.record[28] else None
                duplicate_type = self.record[27] if len(self.record) > 27 and self.record[27] else None

                if duplicate_type:
                    text.insert(tk.END, f"Duplicate Type: {duplicate_type}\n")

                if parent_file:
                    text.insert(tk.END, f"Parent Program: {parent_file}\n\n")

                    # Fetch parent program details from database
                    try:
                        conn = sqlite3.connect(self.parent.db_path)
                        cursor = conn.cursor()
                        cursor.execute("""
                            SELECT program_number, title, outer_diameter, thickness, center_bore,
                                   file_path, validation_status
                            FROM programs
                            WHERE program_number = ?
                        """, (parent_file,))
                        parent_record = cursor.fetchone()
                        conn.close()

                        if parent_record:
                            text.insert(tk.END, "Parent Program Details:\n")
                            text.insert(tk.END, f"  Program: {parent_record[0]}\n")
                            if parent_record[1]:
                                text.insert(tk.END, f"  Title: {parent_record[1]}\n")
                            if parent_record[2]:
                                text.insert(tk.END, f"  OD: {parent_record[2]}in\n")
                            if parent_record[3]:
                                text.insert(tk.END, f"  Thickness: {parent_record[3]}in\n")
                            if parent_record[4]:
                                text.insert(tk.END, f"  CB: {parent_record[4]}mm\n")
                            if parent_record[5]:
                                text.insert(tk.END, f"  Location: {parent_record[5]}\n")
                            if parent_record[6]:
                                text.insert(tk.END, f"  Status: {parent_record[6]}\n")
                        else:
                            text.insert(tk.END, f"  ⚠️ Parent program not found in database\n")
                    except Exception as e:
                        text.insert(tk.END, f"  ⚠️ Error loading parent details: {e}\n")
                else:
                    text.insert(tk.END, "  ⚠️ No parent program recorded\n")

                text.insert(tk.END, "\n")
            else:
                text.insert(tk.END, "\n")

            # G-code dimensions
            if len(self.record) > 22 and self.record[22]:  # cb_from_gcode (index 22)
                text.insert(tk.END, f"CB from G-code: {self.record[22]:.1f}mm\n")
            if len(self.record) > 23 and self.record[23]:  # ob_from_gcode (index 23)
                text.insert(tk.END, f"OB from G-code: {self.record[23]:.1f}mm\n")

            # Validation issues (CRITICAL - RED)
            if self.record[20]:  # validation_issues
                try:
                    # Try JSON format first (backward compatibility)
                    issues = json.loads(self.record[20])
                except:
                    # Fall back to pipe-delimited format
                    issues = [i.strip() for i in self.record[20].split('|') if i.strip()]

                if issues:
                    text.insert(tk.END, "\nCRITICAL ISSUES:\n")
                    for issue in issues:
                        text.insert(tk.END, f"  - {issue}\n")

            # Bore warnings (ORANGE)
            if len(self.record) > 24 and self.record[24]:  # bore_warnings (index 24)
                try:
                    # Try JSON format first (backward compatibility)
                    bore_warns = json.loads(self.record[24])
                except:
                    # Fall back to pipe-delimited format
                    bore_warns = [i.strip() for i in self.record[24].split('|') if i.strip()]

                if bore_warns:
                    text.insert(tk.END, "\nBORE WARNINGS:\n")
                    for warning in bore_warns:
                        text.insert(tk.END, f"  - {warning}\n")

            # Dimensional issues (PURPLE)
            if len(self.record) > 25 and self.record[25]:  # dimensional_issues (index 25)
                try:
                    # Try JSON format first (backward compatibility)
                    dim_issues = json.loads(self.record[25])
                except:
                    # Fall back to pipe-delimited format
                    dim_issues = [i.strip() for i in self.record[25].split('|') if i.strip()]

                if dim_issues:
                    text.insert(tk.END, "\nDIMENSIONAL ISSUES:\n")
                    for issue in dim_issues:
                        text.insert(tk.END, f"  - {issue}\n")

            # Validation warnings (YELLOW)
            if self.record[21]:  # validation_warnings
                try:
                    # Try JSON format first (backward compatibility)
                    warnings = json.loads(self.record[21])
                except:
                    # Fall back to pipe-delimited format
                    warnings = [i.strip() for i in self.record[21].split('|') if i.strip()]

                if warnings:
                    text.insert(tk.END, "\nWARNINGS:\n")
                    for warning in warnings:
                        text.insert(tk.END, f"  - {warning}\n")

        text.config(state=tk.DISABLED)
        
        # Close button
        btn_close = tk.Button(self.window, text="Close", command=self.window.destroy,
                             bg=self.parent.button_bg, fg=self.parent.fg_color,
                             font=("Arial", 10, "bold"), width=15)
        btn_close.pack(pady=10)


class FileComparisonWindow:
    """Window to compare multiple files side-by-side"""
    def __init__(self, parent, files_data, bg_color, fg_color, input_bg, button_bg, refresh_callback, manager=None):
        self.parent = parent
        self.files_data = files_data
        self.bg_color = bg_color
        self.fg_color = fg_color
        self.input_bg = input_bg
        self.button_bg = button_bg
        self.refresh_callback = refresh_callback
        self.manager = manager  # Reference to GCodeDatabaseManager instance

        # Create lookup dictionary: {program_number: (all file data)}
        self.files_lookup = {file_info[0]: file_info for file_info in files_data}

        # Create window
        self.window = tk.Toplevel(parent)
        self.window.title(f"Compare {len(files_data)} Files")
        self.window.geometry("1400x900")
        self.window.configure(bg=bg_color)

        # Track actions for each file
        self.file_actions = {}  # {program_number: 'keep'/'delete'/'rename'}

        self.setup_ui()

    def setup_ui(self):
        """Setup the comparison UI"""
        # Title
        title_label = tk.Label(self.window,
                              text=f"File Comparison - {len(self.files_data)} Files Selected",
                              bg=self.bg_color, fg=self.fg_color,
                              font=("Arial", 14, "bold"))
        title_label.pack(pady=10)

        # Main comparison area with scrollbar
        main_frame = tk.Frame(self.window, bg=self.bg_color)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Canvas with scrollbar for horizontal scrolling
        canvas = tk.Canvas(main_frame, bg=self.bg_color)
        h_scrollbar = tk.Scrollbar(main_frame, orient=tk.HORIZONTAL, command=canvas.xview)
        v_scrollbar = tk.Scrollbar(main_frame, orient=tk.VERTICAL, command=canvas.yview)

        scrollable_frame = tk.Frame(canvas, bg=self.bg_color)
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(xscrollcommand=h_scrollbar.set, yscrollcommand=v_scrollbar.set)

        # Pack scrollbars and canvas
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Create side-by-side panels for each file
        panels_frame = tk.Frame(scrollable_frame, bg=self.bg_color)
        panels_frame.pack(fill=tk.BOTH, expand=True)

        # Load file contents
        file_contents = {}
        for file_info in self.files_data:
            prog_num = file_info[0]
            file_path = file_info[2]

            if file_path and os.path.exists(file_path):
                try:
                    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                        file_contents[prog_num] = f.read()
                except:
                    file_contents[prog_num] = "[Error reading file]"
            else:
                file_contents[prog_num] = "[File not found]"

        # Create a panel for each file
        for idx, file_info in enumerate(self.files_data):
            self.create_file_panel(panels_frame, file_info, file_contents, idx)

        # Bottom action buttons
        bottom_frame = tk.Frame(self.window, bg=self.bg_color)
        bottom_frame.pack(fill=tk.X, padx=10, pady=10)

        tk.Button(bottom_frame, text="Apply Actions & Close",
                 command=self.apply_actions,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 11, "bold"), width=20).pack(side=tk.LEFT, padx=5)

        tk.Button(bottom_frame, text="Cancel",
                 command=self.window.destroy,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 11), width=15).pack(side=tk.LEFT, padx=5)

        # Summary label
        self.summary_label = tk.Label(bottom_frame,
                                     text="Select actions for each file",
                                     bg=self.bg_color, fg=self.fg_color,
                                     font=("Arial", 10))
        self.summary_label.pack(side=tk.RIGHT, padx=10)

    def create_file_panel(self, parent, file_info, file_contents, index):
        """Create a panel for one file"""
        prog_num, title, file_path, spacer_type, od, thickness = file_info[:6]
        cb, hub_h, hub_d, cb_d, cb_dep, val_status, confidence = file_info[6:13]
        dup_type, parent_file, modified, created = file_info[13:17]

        # Panel container
        panel = tk.Frame(parent, bg=self.input_bg, relief=tk.RAISED, borderwidth=2)
        panel.grid(row=0, column=index, padx=5, pady=5, sticky="nsew")

        # Make columns expand equally
        parent.grid_columnconfigure(index, weight=1, uniform="col")

        # Header with file number and status
        header_bg = self.get_status_color(val_status)
        header = tk.Frame(panel, bg=header_bg)
        header.pack(fill=tk.X, pady=(0, 5))

        tk.Label(header, text=f"File {index + 1}: {prog_num}",
                bg=header_bg, fg="white",
                font=("Arial", 12, "bold")).pack(pady=5)

        if val_status:
            tk.Label(header, text=f"Status: {val_status}",
                    bg=header_bg, fg="white",
                    font=("Arial", 9)).pack()

        # Metadata section
        meta_frame = tk.Frame(panel, bg=self.input_bg)
        meta_frame.pack(fill=tk.X, padx=5, pady=5)

        metadata = [
            ("Title", title or "N/A"),
            ("Type", spacer_type or "N/A"),
            ("OD", f"{od:.3f}\"" if od else "N/A"),
            ("Thickness", f"{thickness:.3f}\"" if thickness else "N/A"),
            ("Center Bore", f"{cb:.3f}\"" if cb else "N/A"),
            ("Confidence", confidence or "N/A"),
        ]

        if dup_type:
            metadata.append(("Dup Type", dup_type))
        if parent_file:
            metadata.append(("Parent", parent_file))

        for label, value in metadata:
            row = tk.Frame(meta_frame, bg=self.input_bg)
            row.pack(fill=tk.X, pady=1)
            tk.Label(row, text=f"{label}:", bg=self.input_bg, fg=self.fg_color,
                    font=("Arial", 9, "bold"), width=12, anchor='w').pack(side=tk.LEFT)
            tk.Label(row, text=value, bg=self.input_bg, fg=self.fg_color,
                    font=("Arial", 9), anchor='w').pack(side=tk.LEFT, fill=tk.X, expand=True)

        # File content preview with syntax highlighting
        tk.Label(panel, text="G-Code Preview:", bg=self.input_bg, fg=self.fg_color,
                font=("Arial", 10, "bold")).pack(pady=(10, 5))

        content_frame = tk.Frame(panel, bg=self.input_bg)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        content_text = scrolledtext.ScrolledText(content_frame,
                                                bg="#1e1e1e", fg="#d4d4d4",
                                                font=("Courier New", 8),
                                                height=20, wrap=tk.NONE)
        content_text.pack(fill=tk.BOTH, expand=True)

        # Insert content with basic syntax highlighting
        content = file_contents.get(prog_num, "[No content]")
        self.insert_with_highlighting(content_text, content)
        content_text.config(state=tk.DISABLED)

        # Highlight differences if multiple files
        if len(self.files_data) > 1 and index > 0:
            self.highlight_differences(content_text, content, file_contents[self.files_data[0][0]])

        # Action buttons
        action_frame = tk.Frame(panel, bg=self.input_bg)
        action_frame.pack(fill=tk.X, padx=5, pady=10)

        tk.Label(action_frame, text="Action:", bg=self.input_bg, fg=self.fg_color,
                font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)

        action_var = tk.StringVar(value="keep")
        self.file_actions[prog_num] = action_var

        tk.Radiobutton(action_frame, text="✓ Keep", variable=action_var, value="keep",
                      bg=self.input_bg, fg=self.fg_color, selectcolor=self.button_bg,
                      font=("Arial", 9), command=self.update_summary).pack(side=tk.LEFT, padx=3)

        tk.Radiobutton(action_frame, text="✏️ Rename", variable=action_var, value="rename",
                      bg=self.input_bg, fg=self.fg_color, selectcolor=self.button_bg,
                      font=("Arial", 9), command=self.update_summary).pack(side=tk.LEFT, padx=3)

        tk.Radiobutton(action_frame, text="🗑️ Delete", variable=action_var, value="delete",
                      bg=self.input_bg, fg=self.fg_color, selectcolor=self.button_bg,
                      font=("Arial", 9), command=self.update_summary).pack(side=tk.LEFT, padx=3)

    def get_status_color(self, status):
        """Get color based on validation status"""
        colors = {
            'CRITICAL': '#d32f2f',          # RED - Critical errors
            'SAFETY_ERROR': '#b71c1c',      # DARK RED - Missing safety blocks
            'TOOL_ERROR': '#ff5722',        # ORANGE-RED - Wrong/missing tools
            'DIMENSIONAL': '#7b1fa2',       # PURPLE - P-code mismatches
            'BORE_WARNING': '#f57c00',      # ORANGE - Bore warnings
            'TOOL_WARNING': '#fbc02d',      # AMBER - Tool suggestions
            'WARNING': '#fdd835',           # YELLOW - General warnings
            'PASS': '#388e3c',              # GREEN - All good
            'REPEAT': '#757575'             # GRAY - Duplicate
        }
        return colors.get(status, '#455a64')

    def insert_with_highlighting(self, text_widget, content):
        """Insert content with basic G-code syntax highlighting"""
        text_widget.tag_configure("gcode", foreground="#569cd6")  # Blue for G/M codes
        text_widget.tag_configure("comment", foreground="#6a9955")  # Green for comments
        text_widget.tag_configure("number", foreground="#b5cea8")  # Light green for numbers

        lines = content.split('\n')
        for line in lines:
            if '(' in line:
                # Line has comments
                before_comment = line.split('(')[0]
                comment_part = '(' + '('.join(line.split('(')[1:])
                text_widget.insert(tk.END, before_comment)
                text_widget.insert(tk.END, comment_part, "comment")
                text_widget.insert(tk.END, "\n")
            elif line.strip().startswith('G') or line.strip().startswith('M'):
                text_widget.insert(tk.END, line, "gcode")
                text_widget.insert(tk.END, "\n")
            else:
                text_widget.insert(tk.END, line + "\n")

    def highlight_differences(self, text_widget, content1, content2):
        """Highlight differences between two file contents with color coding"""
        import difflib

        # Configure tags for different types of changes
        text_widget.tag_configure("changed", background="#5a3a00", foreground="#ffcc80")  # Orange for changed lines
        text_widget.tag_configure("added", background="#1b5e20", foreground="#a5d6a7")     # Green for added lines
        text_widget.tag_configure("deleted", background="#7f0000", foreground="#ef9a9a")  # Red for deleted lines
        text_widget.tag_configure("diff_marker", foreground="#ff9800", font=("Arial", 9, "bold"))

        if content1 == content2:
            return  # No differences

        lines1 = content1.split('\n')
        lines2 = content2.split('\n')

        # Use SequenceMatcher to find differences
        matcher = difflib.SequenceMatcher(None, lines2, lines1)

        # Track changes by line number
        changed_lines = set()
        added_lines = set()
        deleted_lines = set()

        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            if tag == 'replace':
                # Lines were changed
                for line_num in range(j1, j2):
                    changed_lines.add(line_num)
            elif tag == 'insert':
                # Lines were added (exist in content1 but not content2)
                for line_num in range(j1, j2):
                    added_lines.add(line_num)
            elif tag == 'delete':
                # Lines were deleted (exist in content2 but not content1)
                # Mark where they would have been
                if j1 < len(lines1):
                    deleted_lines.add(j1)

        # Apply highlighting to the text widget
        # We need to re-enable the widget temporarily
        text_widget.config(state=tk.NORMAL)

        # Add difference summary at top
        total_changes = len(changed_lines) + len(added_lines) + len(deleted_lines)
        if total_changes > 0:
            summary = f"⚠ {total_changes} difference(s): {len(changed_lines)} changed, {len(added_lines)} added, {len(deleted_lines)} removed\n"
            text_widget.insert("1.0", summary, "diff_marker")
            text_widget.insert("1.0", "\n")  # Add blank line after summary

        # Highlight changed lines
        for line_num in changed_lines:
            start_idx = f"{line_num + 3}.0"  # +3 because we added 2 lines at top (blank + summary)
            end_idx = f"{line_num + 3}.end"
            text_widget.tag_add("changed", start_idx, end_idx)

        # Highlight added lines
        for line_num in added_lines:
            start_idx = f"{line_num + 3}.0"
            end_idx = f"{line_num + 3}.end"
            text_widget.tag_add("added", start_idx, end_idx)

        # Highlight deleted line markers
        for line_num in deleted_lines:
            start_idx = f"{line_num + 3}.0"
            end_idx = f"{line_num + 3}.end"
            text_widget.tag_add("deleted", start_idx, end_idx)

        # Re-disable the widget
        text_widget.config(state=tk.DISABLED)

    def update_summary(self):
        """Update the summary label with current action counts"""
        keep_count = sum(1 for var in self.file_actions.values() if var.get() == "keep")
        rename_count = sum(1 for var in self.file_actions.values() if var.get() == "rename")
        delete_count = sum(1 for var in self.file_actions.values() if var.get() == "delete")

        summary = f"Keep: {keep_count} | Rename: {rename_count} | Delete: {delete_count}"
        self.summary_label.config(text=summary)

    def apply_actions(self):
        """Apply the selected actions to files"""
        actions_to_apply = {prog: var.get() for prog, var in self.file_actions.items()}

        # Validate at least one file is kept
        if not any(action == "keep" for action in actions_to_apply.values()):
            messagebox.showerror("Invalid Action",
                "You must keep at least one file.\n\n"
                "Please select 'Keep' for at least one file.")
            return

        # Confirm actions
        delete_count = sum(1 for action in actions_to_apply.values() if action == "delete")
        rename_count = sum(1 for action in actions_to_apply.values() if action == "rename")

        msg = f"Apply the following actions?\n\n"
        msg += f"Delete: {delete_count} file(s)\n"
        msg += f"Rename: {rename_count} file(s)\n\n"
        msg += f"⚠️  Files will be deleted from DATABASE only (not disk)"

        result = messagebox.askyesno("Confirm Actions", msg)
        if not result:
            return

        # Apply actions
        conn = sqlite3.connect("gcode_database.db")
        cursor = conn.cursor()

        deleted = 0
        renamed = 0

        for prog_num, action in actions_to_apply.items():
            if action == "delete":
                cursor.execute("DELETE FROM programs WHERE program_number = ?", (prog_num,))
                deleted += 1
            elif action == "rename":
                # Prompt for new name
                new_name = self.prompt_rename(prog_num)
                if new_name and new_name != prog_num:
                    # Get file_path for this program
                    file_info = self.files_lookup.get(prog_num)
                    if file_info and len(file_info) > 2:
                        file_path = file_info[2]  # file_path is at index 2

                        # Update internal G-code program number if we have manager access
                        if self.manager and file_path and os.path.exists(file_path):
                            if not self.manager.update_gcode_program_number(file_path, new_name):
                                print(f"Warning: Could not update internal O-number in {file_path}")

                    # Update database
                    cursor.execute("""
                        UPDATE programs
                        SET program_number = ?
                        WHERE program_number = ?
                    """, (new_name, prog_num))
                    renamed += 1

        conn.commit()
        conn.close()

        # Show results
        messagebox.showinfo("Actions Applied",
            f"Successfully applied actions:\n\n"
            f"Deleted: {deleted} file(s)\n"
            f"Renamed: {renamed} file(s)")

        # Refresh parent and close
        if self.refresh_callback:
            self.refresh_callback()
        self.window.destroy()

    def prompt_rename(self, old_name):
        """Prompt user for new filename"""
        dialog = tk.Toplevel(self.window)
        dialog.title(f"Rename {old_name}")
        dialog.geometry("400x150")
        dialog.configure(bg=self.bg_color)
        dialog.transient(self.window)
        dialog.grab_set()

        tk.Label(dialog, text=f"Rename {old_name} to:",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 11)).pack(pady=15)

        entry = tk.Entry(dialog, bg=self.input_bg, fg=self.fg_color,
                        font=("Arial", 11), width=30)
        entry.insert(0, old_name)
        entry.pack(pady=10)
        entry.focus()
        entry.select_range(0, tk.END)

        result = [old_name]

        def confirm():
            new_name = entry.get().strip()
            if new_name:
                result[0] = new_name
            dialog.destroy()

        def cancel():
            dialog.destroy()

        btn_frame = tk.Frame(dialog, bg=self.bg_color)
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="OK", command=confirm,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10), width=10).pack(side=tk.LEFT, padx=5)

        tk.Button(btn_frame, text="Cancel", command=cancel,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10), width=10).pack(side=tk.LEFT, padx=5)

        entry.bind('<Return>', lambda e: confirm())
        entry.bind('<Escape>', lambda e: cancel())

        self.window.wait_window(dialog)
        return result[0]

    # ===== Multi-File Comparison Tool =====

    def compare_files(self):
        """Launch multi-file comparison tool with difference highlighting"""
        import difflib

        # Get selected programs from tree
        selected_items = self.tree.selection()

        if len(selected_items) < 2:
            messagebox.showwarning("Selection Required",
                                 "Please select at least 2 files to compare.\n\n"
                                 "Tip: Hold Ctrl and click to select multiple files.")
            return

        if len(selected_items) > 10:
            messagebox.showwarning("Too Many Files",
                                 "Please select no more than 10 files to compare.\n"
                                 f"Currently selected: {len(selected_items)}")
            return

        # Get file information
        files_to_compare = []
        for item in selected_items:
            values = self.tree.item(item)['values']
            if values:
                program_number = values[0]
                title = values[1] if len(values) > 1 else ""

                # Get file path
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute("SELECT file_path FROM programs WHERE program_number = ?", (program_number,))
                result = cursor.fetchone()
                conn.close()

                if result and result[0] and os.path.exists(result[0]):
                    files_to_compare.append({
                        'program_number': program_number,
                        'title': title,
                        'file_path': result[0]
                    })
                else:
                    messagebox.showerror("File Not Found",
                                       f"Could not find file for {program_number}")
                    return

        # Create comparison window
        comp_window = tk.Toplevel(self.root)
        comp_window.title(f"Compare {len(files_to_compare)} Files - Difference Viewer")
        comp_window.geometry("1400x900")
        comp_window.configure(bg=self.bg_color)

        # Header frame
        header_frame = tk.Frame(comp_window, bg=self.bg_color)
        header_frame.pack(fill=tk.X, padx=10, pady=10)

        tk.Label(header_frame, text=f"Comparing {len(files_to_compare)} Files",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 14, "bold")).pack()

        # Control frame
        control_frame = tk.Frame(comp_window, bg=self.bg_color)
        control_frame.pack(fill=tk.X, padx=10, pady=5)

        tk.Label(control_frame, text="Select files to compare:",
                bg=self.bg_color, fg=self.fg_color,
                font=("Arial", 10)).pack(side=tk.LEFT, padx=5)

        # File selection dropdowns
        file1_var = tk.StringVar(value=files_to_compare[0]['program_number'])
        file2_var = tk.StringVar(value=files_to_compare[1]['program_number'] if len(files_to_compare) > 1 else files_to_compare[0]['program_number'])

        file_options = [f"{f['program_number']} - {f['title']}" for f in files_to_compare]

        tk.Label(control_frame, text="File 1:",
                bg=self.bg_color, fg=self.fg_color).pack(side=tk.LEFT, padx=5)
        file1_combo = ttk.Combobox(control_frame, textvariable=file1_var,
                                  values=file_options, width=40, state='readonly')
        file1_combo.pack(side=tk.LEFT, padx=5)

        tk.Label(control_frame, text="File 2:",
                bg=self.bg_color, fg=self.fg_color).pack(side=tk.LEFT, padx=5)
        file2_combo = ttk.Combobox(control_frame, textvariable=file2_var,
                                  values=file_options, width=40, state='readonly')
        file2_combo.pack(side=tk.LEFT, padx=5)

        # Stats frame
        stats_frame = tk.Frame(comp_window, bg=self.bg_color)
        stats_frame.pack(fill=tk.X, padx=10, pady=5)

        stats_label = tk.Label(stats_frame, text="", bg=self.bg_color,
                              fg=self.accent_color, font=("Arial", 10, "bold"))
        stats_label.pack()

        # Main comparison frame (side-by-side)
        main_frame = tk.Frame(comp_window, bg=self.bg_color)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Left panel
        left_frame = tk.LabelFrame(main_frame, text="File 1", bg=self.bg_color,
                                  fg=self.fg_color, font=("Arial", 10, "bold"))
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)

        left_text = tk.Text(left_frame, wrap=tk.NONE, bg=self.input_bg,
                           fg=self.fg_color, font=("Courier New", 9))
        left_scroll_y = tk.Scrollbar(left_frame, command=left_text.yview)
        left_scroll_x = tk.Scrollbar(left_frame, orient=tk.HORIZONTAL, command=left_text.xview)
        left_text.config(yscrollcommand=left_scroll_y.set, xscrollcommand=left_scroll_x.set)

        left_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        left_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        left_text.pack(fill=tk.BOTH, expand=True)

        # Right panel
        right_frame = tk.LabelFrame(main_frame, text="File 2", bg=self.bg_color,
                                   fg=self.fg_color, font=("Arial", 10, "bold"))
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5)

        right_text = tk.Text(right_frame, wrap=tk.NONE, bg=self.input_bg,
                            fg=self.fg_color, font=("Courier New", 9))
        right_scroll_y = tk.Scrollbar(right_frame, command=right_text.yview)
        right_scroll_x = tk.Scrollbar(right_frame, orient=tk.HORIZONTAL, command=right_text.xview)
        right_text.config(yscrollcommand=right_scroll_y.set, xscrollcommand=right_scroll_x.set)

        right_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        right_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        right_text.pack(fill=tk.BOTH, expand=True)

        # Configure tags for highlighting
        # Green for additions, Red for deletions, Yellow for changes
        left_text.tag_config("delete", background="#5D1F1A", foreground="#FF6B6B")
        left_text.tag_config("change", background="#5D4A1F", foreground="#FFD93D")
        left_text.tag_config("equal", background=self.input_bg, foreground=self.fg_color)

        right_text.tag_config("insert", background="#1F4D2A", foreground="#6BCF7F")
        right_text.tag_config("change", background="#5D4A1F", foreground="#FFD93D")
        right_text.tag_config("equal", background=self.input_bg, foreground=self.fg_color)

        # Sync scrolling
        def sync_scroll(*args):
            left_text.yview_moveto(args[0])
            right_text.yview_moveto(args[0])

        left_scroll_y.config(command=sync_scroll)
        right_scroll_y.config(command=sync_scroll)

        def perform_comparison():
            """Perform the actual file comparison"""
            # Get selected file indices
            file1_idx = file1_combo.current()
            file2_idx = file2_combo.current()

            if file1_idx == file2_idx:
                messagebox.showwarning("Same File", "Please select two different files to compare.")
                return

            file1_info = files_to_compare[file1_idx]
            file2_info = files_to_compare[file2_idx]

            # Read files
            try:
                with open(file1_info['file_path'], 'r', encoding='utf-8', errors='ignore') as f:
                    file1_lines = f.readlines()
                with open(file2_info['file_path'], 'r', encoding='utf-8', errors='ignore') as f:
                    file2_lines = f.readlines()
            except Exception as e:
                messagebox.showerror("Read Error", f"Could not read files: {e}")
                return

            # Clear previous content
            left_text.delete('1.0', tk.END)
            right_text.delete('1.0', tk.END)

            # Update labels
            left_frame.config(text=f"File 1: {file1_info['program_number']} - {file1_info['title']}")
            right_frame.config(text=f"File 2: {file2_info['program_number']} - {file2_info['title']}")

            # Generate diff using difflib
            differ = difflib.Differ()
            diff = list(differ.compare(file1_lines, file2_lines))

            # Count differences
            additions = sum(1 for line in diff if line.startswith('+ '))
            deletions = sum(1 for line in diff if line.startswith('- '))
            changes = min(additions, deletions)

            stats_label.config(text=f"Differences: {deletions} deletions, {additions} additions, ~{changes} changes")

            # Process diff and display with highlighting
            left_line_num = 0
            right_line_num = 0

            i = 0
            while i < len(diff):
                line = diff[i]

                if line.startswith('  '):  # Equal line
                    # Show in both panels
                    left_line_num += 1
                    right_line_num += 1
                    content = line[2:]
                    left_text.insert(tk.END, f"{left_line_num:4d} | {content}", "equal")
                    right_text.insert(tk.END, f"{right_line_num:4d} | {content}", "equal")
                    i += 1

                elif line.startswith('- '):  # Deletion (in file1, not in file2)
                    left_line_num += 1
                    content = line[2:]
                    left_text.insert(tk.END, f"{left_line_num:4d} - {content}", "delete")

                    # Check if next line is an addition (change)
                    if i + 1 < len(diff) and diff[i + 1].startswith('+ '):
                        right_line_num += 1
                        right_content = diff[i + 1][2:]
                        right_text.insert(tk.END, f"{right_line_num:4d} + {right_content}", "change")
                        left_text.delete(f"{left_text.index('end-1c linestart')}", f"{left_text.index('end-1c lineend')}")
                        left_text.insert(f"{left_text.index('end-1c linestart')}", f"{left_line_num:4d} ~ {content}", "change")
                        i += 2
                    else:
                        right_text.insert(tk.END, f"     | \n", "equal")
                        i += 1

                elif line.startswith('+ '):  # Addition (in file2, not in file1)
                    right_line_num += 1
                    content = line[2:]
                    right_text.insert(tk.END, f"{right_line_num:4d} + {content}", "insert")
                    left_text.insert(tk.END, f"     | \n", "equal")
                    i += 1

                elif line.startswith('? '):  # Diff hint (ignore)
                    i += 1

                else:
                    i += 1

            # Make read-only
            left_text.config(state=tk.DISABLED)
            right_text.config(state=tk.DISABLED)

            # Re-enable for next comparison
            comp_window.after(100, lambda: (left_text.config(state=tk.NORMAL), right_text.config(state=tk.NORMAL)))

        # Compare button
        tk.Button(control_frame, text="🔄 Compare", command=perform_comparison,
                 bg=self.accent_color, fg=self.fg_color,
                 font=("Arial", 10, "bold"), width=12).pack(side=tk.LEFT, padx=10)

        # Legend frame
        legend_frame = tk.Frame(comp_window, bg=self.bg_color)
        legend_frame.pack(fill=tk.X, padx=10, pady=5)

        tk.Label(legend_frame, text="Legend:", bg=self.bg_color,
                fg=self.fg_color, font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Label(legend_frame, text="  Green = Added  ", bg="#1F4D2A",
                fg="#6BCF7F", font=("Arial", 9)).pack(side=tk.LEFT, padx=2)
        tk.Label(legend_frame, text="  Red = Deleted  ", bg="#5D1F1A",
                fg="#FF6B6B", font=("Arial", 9)).pack(side=tk.LEFT, padx=2)
        tk.Label(legend_frame, text="  Yellow = Changed  ", bg="#5D4A1F",
                fg="#FFD93D", font=("Arial", 9)).pack(side=tk.LEFT, padx=2)

        # Close button
        tk.Button(comp_window, text="Close", command=comp_window.destroy,
                 bg=self.button_bg, fg=self.fg_color,
                 font=("Arial", 10), width=15).pack(pady=10)

        # Perform initial comparison
        perform_comparison()


class VersionHistoryWindow:
    """Window to display and compare version history of a program"""

    def __init__(self, parent, db_manager, program_number, current_file_path):
        self.window = tk.Toplevel(parent)
        self.db_manager = db_manager
        self.program_number = program_number
        self.current_file_path = current_file_path

        # Get colors from parent
        self.bg_color = db_manager.bg_color
        self.fg_color = db_manager.fg_color
        self.input_bg = db_manager.input_bg
        self.button_bg = db_manager.button_bg
        self.accent_color = db_manager.accent_color

        self.window.title(f"Version History - {program_number}")
        self.window.geometry("900x600")
        self.window.configure(bg=self.bg_color)

        self.setup_ui()
        self.load_versions()

    def setup_ui(self):
        """Setup the UI components"""
        # Title
        title_label = tk.Label(self.window,
                              text=f"Version History for {self.program_number}",
                              font=("Arial", 14, "bold"),
                              bg=self.bg_color, fg=self.fg_color)
        title_label.pack(pady=10)

        # Info label
        info_label = tk.Label(self.window,
                             text="Select a version and click 'Compare to Current' to see differences",
                             font=("Arial", 9, "italic"),
                             bg=self.bg_color, fg="#888888")
        info_label.pack(pady=(0, 10))

        # Frame for treeview
        tree_frame = tk.Frame(self.window, bg=self.bg_color)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Scrollbars
        vsb = tk.Scrollbar(tree_frame, orient="vertical")
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        hsb = tk.Scrollbar(tree_frame, orient="horizontal")
        hsb.pack(side=tk.BOTTOM, fill=tk.X)

        # Treeview for versions
        columns = ("version", "tag", "date", "created_by", "summary")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings",
                                yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)

        # Configure columns
        self.tree.heading("version", text="Version")
        self.tree.heading("tag", text="Tag")
        self.tree.heading("date", text="Date Created")
        self.tree.heading("created_by", text="Created By")
        self.tree.heading("summary", text="Change Summary")

        self.tree.column("version", width=80, anchor="center")
        self.tree.column("tag", width=100)
        self.tree.column("date", width=150)
        self.tree.column("created_by", width=120)
        self.tree.column("summary", width=350)

        self.tree.pack(fill=tk.BOTH, expand=True)

        # Button frame
        btn_frame = tk.Frame(self.window, bg=self.bg_color)
        btn_frame.pack(fill=tk.X, pady=10, padx=10)

        # Compare button
        compare_btn = tk.Button(btn_frame, text="Compare to Current",
                               command=self.compare_to_current,
                               bg=self.accent_color, fg=self.fg_color,
                               font=("Arial", 10, "bold"), width=20)
        compare_btn.pack(side=tk.LEFT, padx=5)

        # Restore version button
        restore_btn = tk.Button(btn_frame, text="Restore This Version",
                               command=self.restore_version,
                               bg=self.button_bg, fg=self.fg_color,
                               font=("Arial", 10, "bold"), width=20)
        restore_btn.pack(side=tk.LEFT, padx=5)

        # Close button
        close_btn = tk.Button(btn_frame, text="Close",
                             command=self.window.destroy,
                             bg=self.button_bg, fg=self.fg_color,
                             font=("Arial", 10, "bold"), width=15)
        close_btn.pack(side=tk.RIGHT, padx=5)

    def load_versions(self):
        """Load version history into the treeview"""
        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Get versions from database
        versions = self.db_manager.get_version_history(self.program_number)

        if not versions:
            # Show message if no versions found
            self.tree.insert("", tk.END, values=("No versions found", "", "", "", ""))
            return

        # Insert versions into tree
        for version in versions:
            version_id, version_number, version_tag, date_created, created_by, change_summary = version

            # Format date
            if date_created:
                try:
                    from datetime import datetime
                    date_obj = datetime.fromisoformat(date_created)
                    date_str = date_obj.strftime("%Y-%m-%d %H:%M:%S")
                except:
                    date_str = date_created
            else:
                date_str = ""

            # Insert into tree
            self.tree.insert("", tk.END,
                           values=(f"v{version_number}",
                                  version_tag or "",
                                  date_str,
                                  created_by or "",
                                  change_summary or ""),
                           tags=(version_id,))

    def compare_to_current(self):
        """Compare selected version to current file"""
        # Get selected item
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a version to compare.")
            return

        # Get version_id from tags
        item = selection[0]
        tags = self.tree.item(item, "tags")
        if not tags:
            messagebox.showerror("Error", "No version data found.")
            return

        version_id = tags[0]
        version_number = self.tree.item(item, "values")[0]

        try:
            # Get version content from database
            conn = sqlite3.connect(self.db_manager.db_path)
            cursor = conn.cursor()

            cursor.execute("SELECT file_content FROM program_versions WHERE version_id = ?",
                          (version_id,))
            result = cursor.fetchone()
            conn.close()

            if not result or not result[0]:
                messagebox.showerror("Error", "Version content not found in database.")
                return

            version_content = result[0]

            # Read current file content
            if not os.path.exists(self.current_file_path):
                messagebox.showerror("Error",
                    f"Current file not found:\n{self.current_file_path}")
                return

            with open(self.current_file_path, 'r', encoding='utf-8', errors='ignore') as f:
                current_content = f.read()

            # Create comparison data structure
            files_to_compare = [
                (f"{self.program_number} (Current)", current_content),
                (f"{self.program_number} ({version_number})", version_content)
            ]

            # Open file comparison window
            FileComparisonWindow(self.window, self.db_manager, files_to_compare)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to compare versions:\n{str(e)}")

    def restore_version(self):
        """Restore selected version as current file"""
        # Get selected item
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a version to restore.")
            return

        # Get version_id from tags
        item = selection[0]
        tags = self.tree.item(item, "tags")
        if not tags:
            messagebox.showerror("Error", "No version data found.")
            return

        version_id = tags[0]
        version_number = self.tree.item(item, "values")[0]

        # Confirm restoration
        confirm = messagebox.askyesno("Confirm Restore",
            f"Are you sure you want to restore {version_number}?\n\n"
            f"This will:\n"
            f"1. Create a backup of the current file as a new version\n"
            f"2. Replace the current file with {version_number}\n\n"
            f"This action cannot be undone.",
            icon='warning')

        if not confirm:
            return

        try:
            # Get version content from database
            conn = sqlite3.connect(self.db_manager.db_path)
            cursor = conn.cursor()

            cursor.execute("SELECT file_content FROM program_versions WHERE version_id = ?",
                          (version_id,))
            result = cursor.fetchone()
            conn.close()

            if not result or not result[0]:
                messagebox.showerror("Error", "Version content not found in database.")
                return

            version_content = result[0]

            # Create backup of current file first
            self.db_manager.create_version(self.program_number,
                                          f"Backup before restoring {version_number}")

            # Write restored content to current file
            with open(self.current_file_path, 'w', encoding='utf-8') as f:
                f.write(version_content)

            messagebox.showinfo("Success",
                f"Successfully restored {version_number}\n\n"
                f"A backup of the previous version has been saved.")

            # Reload versions to show the new backup
            self.load_versions()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to restore version:\n{str(e)}")


class RegistryStatisticsWindow:
    """Window to display program number registry statistics"""

    def __init__(self, parent, db_manager):
        self.window = tk.Toplevel(parent)
        self.db_manager = db_manager
        self.window.title("Program Number Registry Statistics")
        self.window.geometry("900x700")
        self.window.configure(bg=db_manager.bg_color)

        # Title
        tk.Label(self.window, text="📋 Program Number Registry Statistics",
                bg=db_manager.bg_color, fg=db_manager.fg_color,
                font=("Arial", 14, "bold")).pack(pady=10)

        # Get registry statistics
        stats = db_manager.get_registry_statistics()

        if not stats:
            tk.Label(self.window, text="Failed to load registry statistics",
                    bg=db_manager.bg_color, fg="red",
                    font=("Arial", 12)).pack(pady=20)
            return

        # Overall statistics frame
        overall_frame = tk.LabelFrame(self.window, text="Overall Statistics",
                                     bg=db_manager.bg_color, fg=db_manager.fg_color,
                                     font=("Arial", 11, "bold"))
        overall_frame.pack(fill=tk.X, padx=20, pady=10)

        stats_text = f"""
Total Program Numbers: {stats['total_numbers']:,}
In Use: {stats['in_use']:,} ({stats['in_use']/stats['total_numbers']*100:.2f}%)
Available: {stats['available']:,} ({stats['available']/stats['total_numbers']*100:.2f}%)
Reserved: {stats['reserved']:,}
Duplicates: {stats['duplicates']:,}
"""

        tk.Label(overall_frame, text=stats_text,
                bg=db_manager.bg_color, fg=db_manager.fg_color,
                font=("Arial", 10), justify=tk.LEFT).pack(padx=10, pady=10)

        # Range statistics
        range_frame = tk.LabelFrame(self.window, text="Range Statistics",
                                   bg=db_manager.bg_color, fg=db_manager.fg_color,
                                   font=("Arial", 11, "bold"))
        range_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # Create treeview for ranges
        tree_frame = tk.Frame(range_frame, bg=db_manager.bg_color)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # Create treeview
        columns = ("Range", "Round Size", "Total", "In Use", "Available", "Usage %", "Duplicates")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings",
                           yscrollcommand=vsb.set, height=15)
        vsb.config(command=tree.yview)

        # Configure columns
        tree.heading("Range", text="Range")
        tree.heading("Round Size", text="Round Size")
        tree.heading("Total", text="Total")
        tree.heading("In Use", text="In Use")
        tree.heading("Available", text="Available")
        tree.heading("Usage %", text="Usage %")
        tree.heading("Duplicates", text="Duplicates")

        tree.column("Range", width=150)
        tree.column("Round Size", width=80)
        tree.column("Total", width=80)
        tree.column("In Use", width=80)
        tree.column("Available", width=100)
        tree.column("Usage %", width=80)
        tree.column("Duplicates", width=80)

        # Populate tree
        for range_name, range_stats in stats['by_range'].items():
            values = (
                range_stats['range'],
                range_stats['round_size'],
                f"{range_stats['total']:,}",
                f"{range_stats['in_use']:,}",
                f"{range_stats['available']:,}",
                f"{range_stats['usage_percent']:.1f}%",
                range_stats['duplicates']
            )
            tree.insert("", tk.END, values=values)

        tree.pack(fill=tk.BOTH, expand=True)

        # Buttons
        button_frame = tk.Frame(self.window, bg=db_manager.bg_color)
        button_frame.pack(fill=tk.X, padx=20, pady=10)

        tk.Button(button_frame, text="Refresh",
                 command=lambda: self.refresh_stats(),
                 bg=db_manager.accent_color, fg=db_manager.fg_color,
                 font=("Arial", 10, "bold"), width=15).pack(side=tk.LEFT, padx=5)

        tk.Button(button_frame, text="Close",
                 command=self.window.destroy,
                 bg=db_manager.button_bg, fg=db_manager.fg_color,
                 font=("Arial", 10, "bold"), width=15).pack(side=tk.RIGHT, padx=5)

    def refresh_stats(self):
        """Refresh the statistics display"""
        self.window.destroy()
        RegistryStatisticsWindow(self.window.master, self.db_manager)


class OutOfRangeWindow:
    """Window to display programs that are in wrong ranges for their round size"""

    def __init__(self, parent, db_manager):
        self.window = tk.Toplevel(parent)
        self.db_manager = db_manager
        self.window.title("Out-of-Range Programs")
        self.window.geometry("1100x700")
        self.window.configure(bg=db_manager.bg_color)

        # Title
        tk.Label(self.window, text="⚠️ Programs in Wrong Ranges",
                bg=db_manager.bg_color, fg="#FF6B6B",
                font=("Arial", 14, "bold")).pack(pady=10)

        # Info
        info_text = """These programs have a detected round size but their program number
is in the wrong range. They should be renamed to match their round size."""

        tk.Label(self.window, text=info_text,
                bg=db_manager.bg_color, fg=db_manager.fg_color,
                font=("Arial", 10), justify=tk.CENTER).pack(pady=5)

        # Get out-of-range programs
        out_of_range = db_manager.get_out_of_range_programs()

        # Count label
        tk.Label(self.window, text=f"Found {len(out_of_range):,} programs in wrong ranges",
                bg=db_manager.bg_color, fg="#FFA500",
                font=("Arial", 11, "bold")).pack(pady=5)

        # Create treeview
        tree_frame = tk.Frame(self.window, bg=db_manager.bg_color)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        hsb.pack(side=tk.BOTTOM, fill=tk.X)

        # Create treeview
        columns = ("Program #", "Round Size", "Current Range", "Correct Range", "Title")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings",
                                yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)

        # Configure columns
        self.tree.heading("Program #", text="Program #")
        self.tree.heading("Round Size", text="Round Size")
        self.tree.heading("Current Range", text="Current Range")
        self.tree.heading("Correct Range", text="Correct Range")
        self.tree.heading("Title", text="Title")

        self.tree.column("Program #", width=100)
        self.tree.column("Round Size", width=100)
        self.tree.column("Current Range", width=200)
        self.tree.column("Correct Range", width=150)
        self.tree.column("Title", width=400)

        # Populate tree
        for prog_num, round_size, current_range, correct_range, title in out_of_range:
            values = (
                prog_num,
                round_size,
                current_range,
                correct_range,
                title or "(No title)"
            )
            self.tree.insert("", tk.END, values=values)

        self.tree.pack(fill=tk.BOTH, expand=True)

        # Buttons
        button_frame = tk.Frame(self.window, bg=db_manager.bg_color)
        button_frame.pack(fill=tk.X, padx=20, pady=10)

        tk.Button(button_frame, text="Export to CSV",
                 command=lambda: self.export_to_csv(out_of_range),
                 bg=db_manager.accent_color, fg=db_manager.fg_color,
                 font=("Arial", 10, "bold"), width=15).pack(side=tk.LEFT, padx=5)

        tk.Button(button_frame, text="Refresh",
                 command=lambda: self.refresh_data(),
                 bg=db_manager.accent_color, fg=db_manager.fg_color,
                 font=("Arial", 10, "bold"), width=15).pack(side=tk.LEFT, padx=5)

        tk.Button(button_frame, text="Close",
                 command=self.window.destroy,
                 bg=db_manager.button_bg, fg=db_manager.fg_color,
                 font=("Arial", 10, "bold"), width=15).pack(side=tk.RIGHT, padx=5)

    def export_to_csv(self, data):
        """Export out-of-range programs to CSV"""
        try:
            filepath = filedialog.asksaveasfilename(
                title="Export Out-of-Range Programs",
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )

            if filepath:
                import csv
                with open(filepath, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(["Program Number", "Round Size", "Current Range", "Correct Range", "Title"])
                    for row in data:
                        writer.writerow(row)

                messagebox.showinfo("Success", f"Exported {len(data):,} programs to:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export:\n{str(e)}")

    def refresh_data(self):
        """Refresh the display"""
        self.window.destroy()
        OutOfRangeWindow(self.window.master, self.db_manager)


class BatchRenameWindow:
    """Window to preview and execute batch rename of out-of-range programs"""

    def __init__(self, parent, db_manager):
        self.window = tk.Toplevel(parent)
        self.db_manager = db_manager
        self.window.title("Batch Rename Out-of-Range Programs")
        self.window.geometry("1200x800")
        self.window.configure(bg=db_manager.bg_color)

        # Title
        tk.Label(self.window, text="🔧 Batch Rename Resolution - Type 1 Duplicates",
                bg=db_manager.bg_color, fg="#9B59B6",
                font=("Arial", 14, "bold")).pack(pady=10)

        # Info
        info_frame = tk.Frame(self.window, bg=db_manager.bg_color)
        info_frame.pack(fill=tk.X, padx=20, pady=5)

        info_text = """This will rename programs that are in wrong ranges to match their round size.
Each program will be renamed to the next available number in the correct range.
Legacy names will be tracked in the database and added as comments in the files."""

        tk.Label(info_frame, text=info_text,
                bg=db_manager.bg_color, fg=db_manager.fg_color,
                font=("Arial", 10), justify=tk.LEFT).pack()

        # Preview section
        preview_frame = tk.LabelFrame(self.window, text="Rename Preview",
                                     bg=db_manager.bg_color, fg=db_manager.fg_color,
                                     font=("Arial", 11, "bold"))
        preview_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # Generate preview button
        preview_button_frame = tk.Frame(preview_frame, bg=db_manager.bg_color)
        preview_button_frame.pack(fill=tk.X, padx=10, pady=5)

        tk.Button(preview_button_frame, text="Generate Preview",
                 command=self.generate_preview,
                 bg=db_manager.accent_color, fg=db_manager.fg_color,
                 font=("Arial", 10, "bold"), width=20).pack(side=tk.LEFT, padx=5)

        self.preview_limit_var = tk.StringVar(value="50")
        tk.Label(preview_button_frame, text="Limit:",
                bg=db_manager.bg_color, fg=db_manager.fg_color,
                font=("Arial", 9)).pack(side=tk.LEFT, padx=5)
        tk.Entry(preview_button_frame, textvariable=self.preview_limit_var,
                font=("Arial", 9), width=8).pack(side=tk.LEFT)

        tk.Label(preview_button_frame, text="(use 'all' for no limit)",
                bg=db_manager.bg_color, fg=db_manager.fg_color,
                font=("Arial", 8, "italic")).pack(side=tk.LEFT, padx=5)

        # Create treeview for preview
        tree_frame = tk.Frame(preview_frame, bg=db_manager.bg_color)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        hsb.pack(side=tk.BOTTOM, fill=tk.X)

        # Create treeview
        columns = ("Old #", "New #", "Round Size", "Current Range", "Correct Range", "Title", "Status")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings",
                                yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)

        # Configure columns
        self.tree.heading("Old #", text="Old #")
        self.tree.heading("New #", text="New #")
        self.tree.heading("Round Size", text="Round Size")
        self.tree.heading("Current Range", text="Current Range")
        self.tree.heading("Correct Range", text="Correct Range")
        self.tree.heading("Title", text="Title")
        self.tree.heading("Status", text="Status")

        self.tree.column("Old #", width=80)
        self.tree.column("New #", width=80)
        self.tree.column("Round Size", width=90)
        self.tree.column("Current Range", width=150)
        self.tree.column("Correct Range", width=120)
        self.tree.column("Title", width=350)
        self.tree.column("Status", width=100)

        self.tree.pack(fill=tk.BOTH, expand=True)

        # Stats label
        self.stats_label = tk.Label(self.window, text="No preview generated yet",
                                   bg=db_manager.bg_color, fg="#FFA500",
                                   font=("Arial", 10, "bold"))
        self.stats_label.pack(pady=5)

        # Action buttons
        button_frame = tk.Frame(self.window, bg=db_manager.bg_color)
        button_frame.pack(fill=tk.X, padx=20, pady=10)

        self.execute_button = tk.Button(button_frame, text="⚠️ EXECUTE BATCH RENAME ⚠️",
                                        command=self.execute_batch_rename,
                                        bg="#C41E3A", fg="white",
                                        font=("Arial", 11, "bold"), width=30,
                                        state=tk.DISABLED)
        self.execute_button.pack(side=tk.LEFT, padx=5)

        tk.Button(button_frame, text="Export Preview to CSV",
                 command=self.export_preview,
                 bg=db_manager.accent_color, fg=db_manager.fg_color,
                 font=("Arial", 10, "bold"), width=20).pack(side=tk.LEFT, padx=5)

        tk.Button(button_frame, text="Close",
                 command=self.window.destroy,
                 bg=db_manager.button_bg, fg=db_manager.fg_color,
                 font=("Arial", 10, "bold"), width=15).pack(side=tk.RIGHT, padx=5)

        # Store preview data
        self.preview_data = []

    def generate_preview(self):
        """Generate preview of what will be renamed"""
        try:
            # Clear existing tree
            for item in self.tree.get_children():
                self.tree.delete(item)

            # Get limit
            limit_str = self.preview_limit_var.get().strip().lower()
            limit = None if limit_str == 'all' else int(limit_str)

            # Generate preview
            self.stats_label.config(text="Generating preview...", fg="orange")
            self.window.update()

            self.preview_data = self.db_manager.preview_rename_plan(limit=limit)

            # Populate tree
            for item in self.preview_data:
                values = (
                    item['old_number'],
                    item['new_number'],
                    item['round_size'],
                    item['current_range'],
                    item['correct_range'],
                    item['title'] or "(No title)",
                    item['status']
                )
                self.tree.insert("", tk.END, values=values)

            # Update stats
            total = len(self.preview_data)
            ready = sum(1 for item in self.preview_data if item['status'] == 'Ready')
            errors = total - ready

            self.stats_label.config(
                text=f"Preview: {total:,} programs | {ready:,} ready | {errors} errors",
                fg="green" if errors == 0 else "orange"
            )

            # Enable execute button if we have valid renames
            if ready > 0:
                self.execute_button.config(state=tk.NORMAL)
            else:
                self.execute_button.config(state=tk.DISABLED)

        except ValueError:
            messagebox.showerror("Invalid Input", "Limit must be a number or 'all'")
        except Exception as e:
            messagebox.showerror("Preview Error", f"Failed to generate preview:\n{str(e)}")
            self.stats_label.config(text="Preview generation failed", fg="red")

    def execute_batch_rename(self):
        """Execute the batch rename operation"""
        if not self.preview_data:
            messagebox.showwarning("No Preview", "Please generate a preview first")
            return

        # Count how many will be renamed
        ready_count = sum(1 for item in self.preview_data if item['status'] == 'Ready')

        # Confirm with user
        confirm = messagebox.askyesno(
            "Confirm Batch Rename",
            f"This will rename {ready_count:,} programs.\n\n"
            "Each program will:\n"
            "  - Get a new number in the correct range\n"
            "  - Have legacy name added to database\n"
            "  - Have comment added to file\n"
            "  - Be logged in resolution audit table\n\n"
            "This operation cannot be easily undone.\n\n"
            "Do you want to proceed?",
            icon='warning'
        )

        if not confirm:
            return

        # Create progress window
        progress_window = tk.Toplevel(self.window)
        progress_window.title("Batch Rename Progress")
        progress_window.geometry("600x400")
        progress_window.configure(bg=self.db_manager.bg_color)

        tk.Label(progress_window, text="🔧 Renaming Programs...",
                bg=self.db_manager.bg_color, fg=self.db_manager.fg_color,
                font=("Arial", 12, "bold")).pack(pady=10)

        # Progress bar
        progress_var = tk.DoubleVar()
        progress_bar = ttk.Progressbar(progress_window, variable=progress_var,
                                      maximum=100, length=500)
        progress_bar.pack(pady=10, padx=20)

        # Status label
        status_label = tk.Label(progress_window, text="Starting...",
                               bg=self.db_manager.bg_color, fg=self.db_manager.fg_color,
                               font=("Arial", 10))
        status_label.pack(pady=5)

        # Log text
        log_frame = tk.Frame(progress_window, bg=self.db_manager.bg_color)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        log_scroll = ttk.Scrollbar(log_frame)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        log_text = tk.Text(log_frame, height=15, yscrollcommand=log_scroll.set,
                          bg="#2B2B2B", fg="#FFFFFF", font=("Consolas", 9))
        log_scroll.config(command=log_text.yview)
        log_text.pack(fill=tk.BOTH, expand=True)

        def log(message):
            log_text.insert(tk.END, message + "\n")
            log_text.see(tk.END)
            progress_window.update()

        # Progress callback
        def progress_callback(current, total, prog_num):
            progress_var.set((current / total) * 100)
            status_label.config(text=f"Processing {current}/{total}: {prog_num}")
            log(f"[{current}/{total}] Processing {prog_num}...")

        # Get programs to rename (only those with 'Ready' status)
        programs_to_rename = [item['old_number'] for item in self.preview_data if item['status'] == 'Ready']

        log(f"Starting batch rename of {len(programs_to_rename):,} programs...")
        log("-" * 60)

        # Execute batch rename
        stats = self.db_manager.batch_resolve_out_of_range(
            program_numbers=programs_to_rename,
            dry_run=False,
            progress_callback=progress_callback
        )

        # Show results
        log("-" * 60)
        log("BATCH RENAME COMPLETE")
        log(f"Total: {stats['total']}")
        log(f"Successful: {stats['successful']}")
        log(f"Failed: {stats['failed']}")
        log(f"Skipped: {stats['skipped']}")

        if stats['errors']:
            log("\nErrors:")
            for error in stats['errors'][:10]:  # Show first 10 errors
                log(f"  - {error['program']}: {error['error']}")
            if len(stats['errors']) > 10:
                log(f"  ... and {len(stats['errors']) - 10} more errors")

        status_label.config(text="Complete!")
        progress_var.set(100)

        # Add close button
        tk.Button(progress_window, text="Close",
                 command=progress_window.destroy,
                 bg=self.db_manager.button_bg, fg=self.db_manager.fg_color,
                 font=("Arial", 10, "bold")).pack(pady=10)

        # Show summary message
        messagebox.showinfo(
            "Batch Rename Complete",
            f"Batch rename completed!\n\n"
            f"Successful: {stats['successful']:,}\n"
            f"Failed: {stats['failed']}\n"
            f"Skipped: {stats['skipped']}\n\n"
            f"Check the log for details."
        )

        # Refresh preview
        self.generate_preview()

    def export_preview(self):
        """Export preview data to CSV"""
        if not self.preview_data:
            messagebox.showwarning("No Preview", "Please generate a preview first")
            return

        try:
            filepath = filedialog.asksaveasfilename(
                title="Export Rename Preview",
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )

            if filepath:
                import csv
                with open(filepath, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(["Old Number", "New Number", "Round Size",
                                   "Current Range", "Correct Range", "Title", "Status"])
                    for item in self.preview_data:
                        writer.writerow([
                            item['old_number'],
                            item['new_number'],
                            item['round_size'],
                            item['current_range'],
                            item['correct_range'],
                            item['title'] or "(No title)",
                            item['status']
                        ])

                messagebox.showinfo("Success", f"Exported {len(self.preview_data):,} items to:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export:\n{str(e)}")


def main():
    # Try to use TkinterDnD for drag & drop support, fallback to regular Tk
    try:
        from tkinterdnd2 import TkinterDnD
        root = TkinterDnD.Tk()
    except ImportError:
        root = tk.Tk()

    app = GCodeDatabaseGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
